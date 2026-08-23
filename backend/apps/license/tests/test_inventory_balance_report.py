"""
Tests for the Inventory Balance Report (backend/apps/license/views/inventory_balance_report.py).

Phase 2A regression coverage — this report had zero test coverage before
(confirmed via a prior audit: no test file, and no frontend page consumes
it at all, it's backend/API-only). These tests establish the JSON-vs-Excel
equivalence baseline using the new shared helpers in
apps.core.tests.report_assertions, per the same pattern already proven in
test_license_purchase_profit_report.py.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.core.models import CompanyModel, HSCodeModel, HeadSIONNormsModel, ItemNameModel, SionNormClassModel
from apps.core.tests.report_assertions import assert_excel_rows_match_json_rows
from apps.license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel

User = get_user_model()

REPORT_URL = "/api/reports/inventory-balance/"


@pytest.fixture
def report_viewer_client(db):
    user = User.objects.create_user(
        username="inventory-balance-viewer",
        email="inventory-balance-viewer@example.com",
        password="RoleP@ssw0rd123",
    )
    group, _ = Group.objects.get_or_create(name="REPORT_VIEWER")
    user.groups.add(group)
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


@pytest.fixture
def inv_masters(db):
    head_norm = HeadSIONNormsModel.objects.create(name="Inventory Balance Test Head Norm")
    return {
        "exporter": CompanyModel.objects.create(iec="9990001111", name="Inventory Balance Exporter"),
        "hs_code": HSCodeModel.objects.create(hs_code="77778888", product_description="Inventory Balance Test Product"),
        "item_a": ItemNameModel.objects.create(name="Inventory Balance Item A"),
        "e1_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E1", is_active=True),
    }


def _make_license(number, exporter):
    return LicenseDetailsModel.objects.create(
        license_number=number,
        license_date=date.today() - timedelta(days=180),
        license_expiry_date=date.today() + timedelta(days=180),
        exporter=exporter,
    )


def _make_export_item(license_obj, norm_class, cif_fc=Decimal("50000.00")):
    return LicenseExportItemModel.objects.create(
        license=license_obj,
        description=f"Export item for {license_obj.license_number}",
        norm_class=norm_class,
        cif_fc=cif_fc,
        cif_inr=cif_fc * Decimal("84.5"),
    )


def _make_import_item(
    license_obj, hs_code, item_names,
    *,
    serial=1,
    quantity=Decimal("1000.000"),
    debited_quantity=Decimal("200.000"),
    allotted_quantity=Decimal("100.000"),
    available_quantity=Decimal("700.000"),
    cif_fc=Decimal("500.00"),
    available_value=Decimal("350.00"),
):
    item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=serial,
        description=f"Import item {serial}",
        hs_code=hs_code,
        quantity=quantity,
        debited_quantity=debited_quantity,
        allotted_quantity=allotted_quantity,
        available_quantity=available_quantity,
        cif_fc=cif_fc,
        available_value=available_value,
    )
    item.items.set(item_names)
    return item


@pytest.mark.django_db
def test_json_response_matches_envelope_shape(report_viewer_client, inv_masters):
    lic = _make_license("INV-BAL-001", inv_masters["exporter"])
    _make_export_item(lic, inv_masters["e1_norm"])
    item_obj = _make_import_item(lic, inv_masters["hs_code"], [inv_masters["item_a"]])
    # `available_value` is recalculated by a post-save signal tied to the
    # license's live balance (see apps/license/signals.py), so it won't
    # necessarily equal what was passed to .create() above — read the
    # actual persisted value back rather than asserting a hand-picked one.
    item_obj.refresh_from_db()

    response = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1"})
    assert response.status_code == 200
    data = response.json()

    assert set(data.keys()) == {"sion_norm", "summary", "items"}
    assert data["sion_norm"]["code"] == "E1"
    assert data["summary"]["total_licenses"] == 1
    assert data["summary"]["total_items"] == 1
    assert len(data["items"]) == 1
    item = data["items"][0]
    assert item["item_name"] == "Inventory Balance Item A"
    assert item["total_quantity"] == 1000.0
    assert item["debited_quantity"] == 200.0
    assert item["allotted_quantity"] == 100.0
    assert item["available_quantity"] == 700.0
    assert item["total_cif_value"] == 500.0
    assert item["available_cif_value"] == float(item_obj.available_value)
    assert item["license_count"] == 1


@pytest.mark.django_db
def test_excel_item_rows_match_json_items(report_viewer_client, inv_masters):
    lic1 = _make_license("INV-BAL-010", inv_masters["exporter"])
    _make_export_item(lic1, inv_masters["e1_norm"])
    _make_import_item(
        lic1, inv_masters["hs_code"], [inv_masters["item_a"]],
        quantity=Decimal("1000.000"), debited_quantity=Decimal("200.000"),
        allotted_quantity=Decimal("100.000"), available_quantity=Decimal("700.000"),
        cif_fc=Decimal("500.00"), available_value=Decimal("350.00"),
    )

    lic2 = _make_license("INV-BAL-011", inv_masters["exporter"])
    _make_export_item(lic2, inv_masters["e1_norm"])
    _make_import_item(
        lic2, inv_masters["hs_code"], [inv_masters["item_a"]],
        serial=2,
        quantity=Decimal("500.000"), debited_quantity=Decimal("50.000"),
        allotted_quantity=Decimal("25.000"), available_quantity=Decimal("425.000"),
        cif_fc=Decimal("250.00"), available_value=Decimal("212.50"),
    )

    json_response = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1"}).json()
    assert len(json_response["items"]) == 1
    # Both import items share the same item name (Item A) — aggregated into
    # one row: 1000+500 qty, 200+50 debited, etc. Confirms JSON aggregation
    # before asserting the Excel export renders that same aggregated figure,
    # not a per-license row.
    assert json_response["items"][0]["total_quantity"] == 1500.0
    assert json_response["items"][0]["license_count"] == 2

    excel_response = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1", "format": "excel"})
    assert excel_response.status_code == 200
    workbook = load_workbook(BytesIO(excel_response.content))
    sheet_name = workbook.sheetnames[0]

    header_row = next(
        r for r in range(1, 10)
        if workbook[sheet_name].cell(row=r, column=1).value == "Item Name"
    )

    assert_excel_rows_match_json_rows(
        workbook, json_response["items"], sheet_name,
        column_map={
            "item_name": 1, "hs_code": 2, "total_quantity": 5,
            "debited_quantity": 6, "allotted_quantity": 7, "available_quantity": 8,
            "total_cif_value": 9, "available_cif_value": 10, "license_count": 11,
        },
        header_row=header_row, key_field="item_name",
    )


@pytest.mark.django_db
def test_excel_summary_row_matches_json_summary(report_viewer_client, inv_masters):
    lic = _make_license("INV-BAL-020", inv_masters["exporter"])
    _make_export_item(lic, inv_masters["e1_norm"])
    _make_import_item(lic, inv_masters["hs_code"], [inv_masters["item_a"]])

    json_summary = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1"}).json()["summary"]

    excel_response = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1", "format": "excel"})
    workbook = load_workbook(BytesIO(excel_response.content))
    sheet = workbook[workbook.sheetnames[0]]

    total_row = next(
        r for r in range(1, 30)
        if sheet.cell(row=r, column=1).value == "Total"
    )
    assert sheet.cell(row=total_row, column=5).value == json_summary["total_quantity"]
    assert sheet.cell(row=total_row, column=6).value == json_summary["total_debited"]
    assert sheet.cell(row=total_row, column=7).value == json_summary["total_allotted"]
    assert sheet.cell(row=total_row, column=8).value == json_summary["total_available"]
    assert sheet.cell(row=total_row, column=9).value == json_summary["total_cif_value"]
    assert sheet.cell(row=total_row, column=10).value == json_summary["available_cif_value"]
    assert sheet.cell(row=total_row, column=11).value == json_summary["total_items"]


@pytest.mark.django_db
def test_export_filename_follows_standard_convention(report_viewer_client, inv_masters):
    lic = _make_license("INV-BAL-030", inv_masters["exporter"])
    _make_export_item(lic, inv_masters["e1_norm"])
    _make_import_item(lic, inv_masters["hs_code"], [inv_masters["item_a"]])

    response = report_viewer_client.get(REPORT_URL, {"sion_norm": "E1", "format": "excel"})
    disposition = response["Content-Disposition"]
    assert disposition.startswith('attachment; filename="inventory-balance-E1_')
    assert disposition.endswith('.xlsx"')
