"""Excel renderers for license balance reports.

Extracted verbatim from ``LicenseDetailsViewSet.balance_excel``,
``balance_excel_unused`` and ``bulk_balance_excel`` as part of the
clean-architecture refactor. The viewset now only resolves inputs and delegates
here; behaviour is unchanged.
"""
def build_balance_excel_unused(license_obj):
    """Original full balance Excel — kept for reference, no longer exposed."""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    # BL-LEDGER-02: `license_obj.balance_cif` is a denormalized cache that
    # can be stale -- compute the live figure once and reuse it everywhere
    # this function shows "Balance CIF", matching every other module.
    from apps.license.services.balance_calculator import LicenseBalanceCalculator
    from apps.license.services.condition_pool import available_value_bulk_map
    live_balance_cif = LicenseBalanceCalculator.calculate_financial_balance(license_obj)
    import_items = list(license_obj.import_license.all())
    available_value_map = available_value_bulk_map(import_items)


    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "License Balance"

    # Header styling
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    section_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=12)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    current_row = 1
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "License Balance Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2

    # License Header Information
    # Row 1 Headers
    headers_row1 = ['License Number', 'License Date', 'License Expiry Date', 'Exporter Name', 'Port Name']
    for col_num, header in enumerate(headers_row1, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1

    # Row 1 Values
    values_row1 = [
        license_obj.license_number or '-',
        license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-',
        license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-',
        license_obj.exporter.name if license_obj.exporter else '-',
        license_obj.port.name if license_obj.port else '-'
    ]
    for col_num, value in enumerate(values_row1, 1):
        cell = ws.cell(row=current_row, column=col_num, value=value)
        cell.fill = data_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    current_row += 2

    # Row 2 Headers
    headers_row2 = ['Purchase Status', 'Balance CIF', 'Get Norm Class', 'Latest Transfer']
    for col_num, header in enumerate(headers_row2, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1

    # Row 2 Values
    values_row2 = [
        str(license_obj.purchase_status) if license_obj.purchase_status else '-',
        f"{float(live_balance_cif):.2f}",
        license_obj.get_norm_class or '-',
        str(license_obj.latest_transfer) if license_obj.latest_transfer else '-'
    ]
    for col_num, value in enumerate(values_row2, 1):
        cell = ws.cell(row=current_row, column=col_num, value=value)
        cell.fill = data_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    current_row += 2

    # Export Items Section
    if license_obj.export_license.exists():
        # Section header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "Export Items"
        section_cell.fill = section_fill
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Export items headers
        export_headers = ['Item', 'Total CIF', 'Balance CIF']
        for col_num, header in enumerate(export_headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        current_row += 1

        # Export items data
        for item in license_obj.export_license.all():
            item_desc = item.description or (str(item.norm_class) if item.norm_class else None) or 'None'
            values = [
                item_desc,
                f"{float(item.cif_fc or item.fob_fc or 0):.2f}",
                f"{float(live_balance_cif):.2f}"
            ]
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = thin_border
            current_row += 1

        current_row += 1

    # Import Items Section
    if import_items:
        from apps.bill_of_entry.models import RowDetails, annotate_and_exclude_hidden
        from apps.allotment.models import AllotmentItems

        # Section header
        ws.merge_cells(f'A{current_row}:J{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "Import Items"
        section_cell.fill = section_fill
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        for item in import_items:
            # Item headers
            item_headers = ['Sr', 'HS Code', 'Description', 'Item', 'Total Qty',
                           'Allotted', 'Debited', 'Available', 'CIF FC', 'Bal CIF']
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            current_row += 1

            # Item data
            item_names = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else '-'
            hs_code_display = str(item.hs_code.hs_code if item.hs_code else '-')

            item_values = [
                str(item.serial_number or '-'),
                hs_code_display,
                str(item.description or '-'),
                item_names,
                f"{float(item.quantity or 0):.2f}",
                f"{float(item.allotted_quantity or 0):.2f}",
                f"{float(item.debited_quantity or 0):.2f}",
                f"{float(item.available_quantity or 0):.2f}",
                f"{float(item.cif_fc or 0):.2f}",
                f"{float(available_value_map.get(item.id, 0)):.2f}"
            ]
            from apps.license.utils.condition_excel import annotate_cell as _annotate_cond_unused
            for col_num, value in enumerate(item_values, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.fill = data_fill
                cell.border = thin_border
                # Tint the Serial Number cell when this item carries a
                # licence condition (AU / 2% / 3% / 5% / 10%).
                if col_num == 1 and item.condition_type:
                    _annotate_cond_unused(cell, item.condition_type)
            current_row += 1

            # BOE Details
            # Previous-owner "hidden" BOEs (genuinely hidden per audit
            # trail) are excluded — this report is a balance/financial
            # figure, not the Customs History audit view.
            boes = annotate_and_exclude_hidden(
                RowDetails.objects.filter(sr_number_id=item.id, transaction_type='D'),
                boe_field="bill_of_entry",
            ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')
            if boes.exists():
                current_row += 1
                ws.merge_cells(f'A{current_row}:G{current_row}')
                boe_header_cell = ws[f'A{current_row}']
                boe_header_cell.value = "BOEs"
                boe_header_cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                boe_header_cell.font = Font(bold=True, color="FFFFFF")
                current_row += 1

                boe_headers = ['BOE Number', 'Date', 'Port', 'Company', 'Qty', 'CIF $', 'CIF INR']
                for col_num, header in enumerate(boe_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.border = thin_border
                current_row += 1

                for boe in boes:
                    boe_values = [
                        boe.bill_of_entry.bill_of_entry_number if boe.bill_of_entry else '-',
                        boe.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if boe.bill_of_entry and boe.bill_of_entry.bill_of_entry_date else '-',
                        boe.bill_of_entry.port.name if boe.bill_of_entry and boe.bill_of_entry.port else '-',
                        boe.bill_of_entry.company.name if boe.bill_of_entry and boe.bill_of_entry.company else '-',
                        f"{float(boe.qty or 0):.2f}",
                        f"{float(boe.cif_fc or 0):.2f}",
                        f"{float(boe.cif_inr or 0):.2f}"
                    ]
                    for col_num, value in enumerate(boe_values, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        cell.border = thin_border
                    current_row += 1

            # Allotment Details
            # Only show allotments where bill_of_entry is NULL (not yet converted to BOE)
            allotments = AllotmentItems.objects.filter(
                item=item,
                allotment__bill_of_entry__isnull=True
            ).select_related('allotment', 'allotment__company')
            if allotments.exists():
                current_row += 1
                ws.merge_cells(f'A{current_row}:D{current_row}')
                allot_header_cell = ws[f'A{current_row}']
                allot_header_cell.value = "Allotments"
                allot_header_cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                allot_header_cell.font = Font(bold=True, color="FFFFFF")
                current_row += 1

                allot_headers = ['Company', 'Qty', 'CIF $', 'CIF INR']
                for col_num, header in enumerate(allot_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.border = thin_border
                current_row += 1

                for allot in allotments:
                    allot_values = [
                        allot.allotment.company.name if allot.allotment and allot.allotment.company else '-',
                        f"{float(allot.qty or 0):.2f}",
                        f"{float(allot.cif_fc or 0):.2f}",
                        f"{float(allot.cif_inr or 0):.2f}"
                    ]
                    for col_num, value in enumerate(allot_values, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        cell.border = thin_border
                    current_row += 1

            # Balance calculation — reads the stored, single-source-of-truth
            # `available_quantity` field (`Total − Debited − Outstanding
            # Allotted`, kept correct by `apps.core.scripts.calculate_
            # balance.update_balance_values`) rather than recombining the
            # other three columns independently here.
            current_row += 1
            balance = float(item.available_quantity or 0)
            ws.merge_cells(f'A{current_row}:J{current_row}')
            balance_cell = ws[f'A{current_row}']
            balance_cell.value = f"Balance Quantity: {balance:.2f}"
            balance_cell.fill = PatternFill(start_color="e8e8e8", end_color="e8e8e8", fill_type="solid")
            balance_cell.font = Font(bold=True, color="e74c3c")
            balance_cell.border = thin_border
            current_row += 2

    # Notes Section
    if license_obj.balance_report_notes:
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        notes_header_cell = ws[f'A{current_row}']
        notes_header_cell.value = "Notes"
        notes_header_cell.fill = section_fill
        notes_header_cell.font = section_font
        notes_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        ws.merge_cells(f'A{current_row}:J{current_row}')
        notes_cell = ws[f'A{current_row}']
        notes_cell.value = license_obj.balance_report_notes
        notes_cell.fill = PatternFill(start_color="fffacd", end_color="fffacd", fill_type="solid")
        notes_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        notes_cell.border = thin_border
        ws.row_dimensions[current_row].height = 60

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{license_obj.license_number}-balance.xlsx"'
    return response


def _write_financial_ledger_sheet(wb, rows, summary):
    """
    Adds a "Financial Ledger" worksheet — the same bank-statement rows and
    Financial Summary / Reconciliation Summary numbers as the PDF's Licence
    Financial Ledger section and the workspace UI, all read from
    `LicenseBalanceLedgerBuilder` so the three surfaces can never disagree.

    `rows`/`summary` are `LicenseBalanceLedgerBuilder.build_financial_ledger()`'s
    own return value — computed ONCE by the caller (`build_balance_excel`)
    and passed in here; this function does no calculation of its own.
    Hierarchical "trade" (Licence Trade Sold) rows' `children` (one per
    underlying BOE allocation, when the sale is matched to one) are
    rendered immediately below their parent, informational only (blank
    Credit/Debit/Running Balance — same convention as the PDF's
    `_build_financial_ledger_elements`), and grouped via openpyxl's native
    row outlining (`outline_level`) so they can be collapsed/expanded in
    Excel itself — `summaryBelow=False` puts the collapse control on the
    parent row since the parent sits ABOVE its children here, not below.

    Inserted as the FIRST sheet in the workbook.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb.create_sheet("Financial Ledger", 0)
    ws.sheet_properties.outlinePr.summaryBelow = False

    HDR_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    ROW_FILLS = {
        'opening': PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid"),
        'boe': PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),
        'allotment': PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
        'trade_purchase': PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid"),
        'trade': PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid"),
        'final': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
    }
    CHILD_FILL = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
    CHILD_FONT = Font(color="555555", size=9)
    MISMATCH_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def fmt_date(d):
        return d.strftime('%d-%m-%Y') if d else '-'

    ws.merge_cells('A1:P1')
    ws['A1'] = 'LICENCE FINANCIAL LEDGER'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = HDR_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    header_row_num = 2
    warning = summary['missing_purchase_warning']
    if warning['show_warning']:
        # Shown instead of a fabricated Opening Balance row when a Sale
        # exists with no matching Purchase — same message as the UI/PDF.
        ws.merge_cells('A2:P2')
        ws['A2'] = f"⚠ {warning['message']}"
        ws['A2'].font = Font(bold=True, color="7D6608", size=10)
        ws['A2'].fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 20
        header_row_num = 3

    headers = [
        'Sr', 'Txn Date', 'Txn Type', 'Doc Number', 'BOE Number', 'BOE Date',
        'Company (Importer)', 'Item Name', 'Invoice(s)', 'Qty',
        'BOE CIF (USD)', 'BOE INR', 'Credit (USD)', 'Debit (USD)',
        'Running Balance (USD)', 'Remarks',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row_num, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r = header_row_num + 1
    for row_data in rows:
        fill = MISMATCH_FILL if row_data.get('mismatched') else ROW_FILLS.get(row_data['row_kind'])
        font_white = row_data['row_kind'] in ('opening', 'final') and not row_data.get('mismatched')
        values = [
            row_data['sr'],
            fmt_date(row_data['date']),
            row_data['type'],
            row_data['document_number'] or '-',
            row_data['boe_number'] or '-',
            row_data.get('boe_date_display') or fmt_date(row_data['boe_date']),
            row_data['company'] or '-',
            row_data['item_name'] or '-',
            ', '.join(row_data['invoice_numbers']) if row_data['invoice_numbers'] else '-',
            float(row_data['qty']) if row_data['qty'] is not None else '-',
            float(row_data['cif_usd']) if row_data['cif_usd'] is not None else '-',
            float(row_data['cif_inr']) if row_data['cif_inr'] is not None else '-',
            float(row_data['credit']) if row_data['credit'] else '-',
            float(row_data['debit']) if row_data['debit'] else '-',
            float(row_data['running_balance']),
            row_data['remarks'],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = THIN
            if fill:
                c.fill = fill
            if font_white:
                c.font = Font(color="FFFFFF", bold=True)
            if col in (10, 11, 12, 13, 14, 15) and isinstance(val, float):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left', wrap_text=True)
        r += 1

        # Hierarchy — one row per underlying BOE allocation under a
        # consolidated "BOE Allocation" parent row (see `build_financial_
        # ledger`'s docstring). Informational only: Credit/Debit/Running
        # Balance stay blank, the parent row above already carries the
        # accounting impact. `outline_level = 1` groups these rows under
        # their parent for Excel's native collapse/expand.
        for child in (row_data.get('children') or []):
            child_values = [
                '', '', f"↳ {child.get('type', '-')}", '',
                child.get('boe_number') or '-', fmt_date(child.get('boe_date')),
                child.get('company') or '-', child.get('item_name') or '-',
                ', '.join(child.get('invoice_numbers') or []) if child.get('invoice_numbers') else '-',
                float(child['qty']) if child.get('qty') is not None else '-',
                float(child['cif_usd']) if child.get('cif_usd') is not None else '-',
                float(child['cif_inr']) if child.get('cif_inr') is not None else '-',
                '-', '-', '-',
                f"{child.get('status', '-')} — {child.get('remarks', '-')}",
            ]
            for col, val in enumerate(child_values, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = THIN
                c.fill = CHILD_FILL
                c.font = CHILD_FONT
                if col == 3:
                    c.alignment = Alignment(horizontal='left', wrap_text=True, indent=2)
                elif col in (10, 11, 12) and isinstance(val, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')
                else:
                    c.alignment = Alignment(horizontal='left', wrap_text=True)
            ws.row_dimensions[r].outline_level = 1
            r += 1

    r += 1
    ws.merge_cells(f'A{r}:P{r}')
    ws[f'A{r}'] = 'FINANCIAL SUMMARY & RECONCILIATION'
    ws[f'A{r}'].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f'A{r}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    summary_rows = [
        ('Original Licence CIF', summary['opening_balance']),
        ('Total BOE Debits', summary['total_boe_debit']),
    ]
    if summary.get('total_invoice_allocation_debit', 0) > 0:
        summary_rows.append(('Total Invoice Allocation Debits', summary['total_invoice_allocation_debit']))
    summary_rows.append(('Outstanding Active Allotments', summary['total_allotment_debit']))
    if summary.get('total_purchase_credit', 0) > 0:
        summary_rows.append(('Total Purchase Credits', summary['total_purchase_credit']))
    if summary['total_trade_debit'] > 0:
        summary_rows.append(('Total Trade (Sold) Debits', summary['total_trade_debit']))
    summary_rows += [
        ('Current Available Balance', summary['computed_balance']),
        ('Licence Balance Engine', summary['engine_balance']),
        ('Difference', summary['difference']),
        ('Tolerance', summary['tolerance']),
    ]
    for label, value in summary_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        vcell = ws.cell(row=r, column=2, value=float(value))
        vcell.number_format = '#,##0.00'
        r += 1

    r += 1
    status_text = 'FINANCIAL RECONCILIATION FAILED' if summary['mismatched'] else 'MATCHED'
    status_color = 'C0392B' if summary['mismatched'] else '1E8449'
    ws.merge_cells(f'A{r}:P{r}')
    ws[f'A{r}'] = status_text
    ws[f'A{r}'].font = Font(bold=True, color=status_color, size=12)
    ws[f'A{r}'].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 20
    for col in ['J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['P'].width = 28
    ws.freeze_panes = f'A{header_row_num + 1}'


def _write_customs_ledger_sheet(wb, rows, summary):
    """
    Adds a "Customs Ledger" worksheet — the running CUSTOMS utilisation
    statement from `LicenseBalanceLedgerBuilder.build_customs_ledger()`
    (see its docstring: every BOE debits the licence at its FULL raw CIF,
    unconditionally, unlike the Financial Ledger's allocation-adjusted
    debit — the two are expected to diverge whenever reconciliation is
    incomplete, which is the actionable signal this report exists to
    surface). Preceded by its own "Customs Summary" block, mirroring the
    PDF's `_build_customs_ledger_elements` column set and section order.

    `rows`/`summary` are computed ONCE by the caller — no calculation
    happens in this function.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb.create_sheet("Customs Ledger")

    HDR_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    ROW_FILLS = {
        'customs_opening': PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid"),
        'customs_boe': PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid"),
        'customs_pending_allotment': PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        'final': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
    }
    MISMATCH_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def fmt_date(d):
        return d.strftime('%d-%m-%Y') if d else '-'

    ws.merge_cells('A1:O1')
    ws['A1'] = 'CUSTOMS SUMMARY'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    difference = abs(summary['computed_balance'] - summary['engine_balance'])
    r = 2
    summary_rows = [
        ('Original Licence CIF', summary['opening_balance']),
        ('Total BOE CIF', summary['total_boe_cif']),
        ('Pending Allotment CIF', summary['total_pending_allotment_cif']),
        ('Available Balance', summary['computed_balance']),
        ('Balance Engine', summary['engine_balance']),
        ('Difference', difference),
    ]
    for label, value in summary_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        vcell = ws.cell(row=r, column=2, value=float(value))
        vcell.number_format = '#,##0.00'
        if summary['mismatched'] and label == 'Difference':
            ws.cell(row=r, column=1).fill = MISMATCH_FILL
            ws.cell(row=r, column=2).fill = MISMATCH_FILL
        r += 1

    r += 1
    status_text = 'RECONCILIATION FAILED' if summary['mismatched'] else 'MATCHED'
    status_color = 'C0392B' if summary['mismatched'] else '1E8449'
    ws.merge_cells(f'A{r}:O{r}')
    ws[f'A{r}'] = status_text
    ws[f'A{r}'].font = Font(bold=True, color=status_color, size=12)
    ws[f'A{r}'].alignment = Alignment(horizontal='center')
    r += 2

    ws.merge_cells(f'A{r}:O{r}')
    ws[f'A{r}'] = 'CUSTOMS LEDGER'
    ws[f'A{r}'].font = Font(bold=True, color="FFFFFF", size=14)
    ws[f'A{r}'].fill = HDR_FILL
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    headers = [
        'Sr', 'Date', 'Transaction Type', 'Document Number', 'BOE Number', 'BOE Date',
        'Company', 'Item', 'Quantity', 'CIF (USD)', 'Credit (USD)', 'Debit (USD)',
        'Running Balance (USD)', 'Status', 'Remarks',
    ]
    header_row = r
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    r += 1

    for row_data in rows:
        fill = MISMATCH_FILL if row_data.get('mismatched') else ROW_FILLS.get(row_data['row_kind'])
        font_white = row_data['row_kind'] in ('customs_opening', 'final') and not row_data.get('mismatched')
        values = [
            row_data['sr'],
            fmt_date(row_data['date']),
            row_data['type'],
            row_data['document_number'] or '-',
            row_data['boe_number'] or '-',
            fmt_date(row_data['boe_date']),
            row_data['company'] or '-',
            row_data['item_name'] or '-',
            float(row_data['qty']) if row_data['qty'] is not None else '-',
            float(row_data['cif_usd']) if row_data['cif_usd'] is not None else '-',
            float(row_data['credit']) if row_data['credit'] else '-',
            float(row_data['debit']) if row_data['debit'] else '-',
            float(row_data['running_balance']),
            row_data.get('status', '-'),
            row_data['remarks'],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = THIN
            if fill:
                c.fill = fill
            if font_white:
                c.font = Font(color="FFFFFF", bold=True)
            if col in (9, 10, 11, 12, 13) and isinstance(val, float):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left', wrap_text=True)
        r += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 22
    for col in ['I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 28
    ws.freeze_panes = f'A{header_row + 1}'


def _write_timeline_sheet(wb, events):
    """
    Adds a "Timeline" worksheet — `LicenseBalanceLedgerBuilder.build_timeline()`'s
    real, timestamped business-lifecycle events (never fabricated), with
    hierarchical children (e.g. each BOE under an "Invoice <-> BOE
    Reconciled" parent) rendered immediately below their parent and grouped
    via openpyxl row outlining, same pattern as the Financial Ledger sheet.
    If there are no events, a single explanatory row is written instead of
    an empty table (matches the PDF's empty-state handling).
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb.create_sheet("Timeline")
    ws.sheet_properties.outlinePr.summaryBelow = False

    HDR_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    TONE_FILLS = {
        'blue': PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
        'orange': PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        'green': PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        'purple': PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
        'teal': PatternFill(start_color="D1F2EB", end_color="D1F2EB", fill_type="solid"),
        'grey': PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid"),
        'red': PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    }
    CHILD_FILL = PatternFill(start_color="F7F9FA", end_color="F7F9FA", fill_type="solid")
    CHILD_FONT = Font(color="555555", size=9)
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def fmt_dt(dt):
        return dt.strftime('%d-%m-%Y %H:%M') if dt else '-'

    def fmt_money(value):
        return float(value) if value is not None else '-'

    ws.merge_cells('A1:J1')
    ws['A1'] = 'TIMELINE'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = PatternFill(start_color="0B3D59", end_color="0B3D59", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    if not events:
        ws.merge_cells('A2:J2')
        ws['A2'] = 'No timeline events recorded for this licence yet.'
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions['A'].width = 20
        return

    headers = ['Sr', 'Date', 'Event Type', 'Document Number', 'Company', 'Qty', 'CIF (USD)', 'User', 'Status', 'Remarks']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r = 3
    for e in events:
        fill = TONE_FILLS.get(e.get('color'))
        values = [
            e['sr'], fmt_dt(e['date']), e['label'],
            e.get('document_number') or '-', e.get('company') or '-',
            fmt_money(e.get('quantity')), fmt_money(e.get('cif')),
            e.get('user') or '-', e.get('status') or '-', e.get('remarks') or '-',
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = THIN
            if fill:
                c.fill = fill
            if col in (6, 7) and isinstance(val, float):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left', wrap_text=True)
        r += 1

        for child in (e.get('children') or []):
            child_values = [
                '', fmt_dt(child.get('date')), f"↳ {child.get('label', '-')}",
                child.get('document_number') or '-', child.get('company') or '-',
                fmt_money(child.get('quantity')), fmt_money(child.get('cif')),
                child.get('user') or '-', child.get('status') or '-', child.get('remarks') or '-',
            ]
            for col, val in enumerate(child_values, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = THIN
                c.fill = CHILD_FILL
                c.font = CHILD_FONT
                if col == 3:
                    c.alignment = Alignment(horizontal='left', wrap_text=True, indent=2)
                elif col in (6, 7) and isinstance(val, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')
                else:
                    c.alignment = Alignment(horizontal='left', wrap_text=True)
            ws.row_dimensions[r].outline_level = 1
            r += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 40
    ws.freeze_panes = 'A3'


def _write_reconciliation_sheet(wb, license_obj, financial_summary, customs_summary, reconciliation):
    """
    Adds a "Reconciliation" worksheet: the three-way Financial Ledger vs.
    Customs Ledger vs. Balance Engine comparison from
    `LicenseBalanceLedgerBuilder.build_reconciliation_summary()`
    (`reconciliation`, passed in — computed once by the caller), plus the
    pre-existing (real, tested) "Summary (BOE & Allotments)" table and Plan
    Utilization section that used to live on this workbook's old "Summary"
    sheet — folded in here rather than dropped.

    `financial_summary`/`customs_summary` are accepted for symmetry with
    the other `_write_X_sheet` helpers but are not directly rendered here
    (their numbers already feed `reconciliation`); kept as explicit
    parameters so a future addition to this sheet never has to recompute
    them.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import date as _date_cls
    from apps.license.services.item_usage import get_item_usage

    ws = wb.create_sheet("Reconciliation")

    HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    BOE_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    ALLOT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    MISMATCH_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
    BOLD = Font(bold=True, size=9)
    NORM = Font(size=9)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return c

    def _cell(ws, row, col, value, fill=None, bold=False, align='left', num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill: c.fill = fill
        c.font = BOLD if bold else NORM
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if num_fmt: c.number_format = num_fmt
        return c

    license_date_str = license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-'
    license_expiry_str = license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-'
    ledger_date_str = license_obj.ledger_date.strftime('%d-%m-%Y') if license_obj.ledger_date else '-'
    lic_no = license_obj.license_number or '-'

    # ── Collect summary rows (BOE & Allotments) ─────────────────────────
    summary_rows = []
    total_cif = 0.0

    for item in license_obj.import_license.all():
        item_name = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else (item.description or '-')

        _usage = get_item_usage(item)
        boes = _usage['boes']

        for rd in boes:
            qty  = float(rd.qty or 0)
            cif  = float(rd.cif_fc or 0)
            rate = cif / qty if qty else 0.0
            total_cif += cif
            boe_company = rd.bill_of_entry.company.name if rd.bill_of_entry.company else '-'
            ref_no   = rd.bill_of_entry.bill_of_entry_number or '-'
            ref_date = rd.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if rd.bill_of_entry.bill_of_entry_date else ''
            ref_str  = f"{ref_no} / {ref_date}" if ref_date else ref_no
            product  = rd.bill_of_entry.product_name or item_name
            _sort_dt = rd.bill_of_entry.bill_of_entry_date or _date_cls.min
            summary_rows.append((0, _sort_dt, {
                'item': product, 'type': 'BOE', 'company': boe_company,
                'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
            }, True))

        allotments = _usage['allotments']

        for ai in allotments:
            qty     = float(ai.qty or 0)
            cif     = float(ai.cif_fc or 0)
            rate    = cif / qty if qty else 0.0
            total_cif += cif
            company = ai.allotment.company.name if ai.allotment.company else '-'
            invoice = ai.allotment.invoice or '-'
            eta     = ai.allotment.estimated_arrival_date.strftime('%d-%m-%Y') if ai.allotment.estimated_arrival_date else ''
            ref_str = f"{invoice} / ETA: {eta}" if eta else invoice
            product = ai.allotment.item_name or item_name
            _sort_dt = ai.allotment.estimated_arrival_date or _date_cls.min
            summary_rows.append((1, _sort_dt, {
                'item': product, 'type': 'Allotment', 'company': company,
                'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
            }, False))

    # BOEs first (sorted by BOE date), then allotments (sorted by allotment date)
    summary_rows.sort(key=lambda x: (x[0], x[1]))

    _license_balance = float(license_obj.get_balance_cif or 0)
    total_license_cif = total_cif + _license_balance

    # ══════════════════════════════════════════════════════════════════════
    # Section 1: License info row
    # ══════════════════════════════════════════════════════════════════════
    r = 1
    _today = _date_cls.today()
    INFO_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    INFO_FONT = Font(bold=True, color="FFFFFF", size=9)
    if license_obj.license_expiry_date:
        _days = (license_obj.license_expiry_date - _today).days
        if _days < 0:
            EXPIRY_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        elif _days <= 90:
            EXPIRY_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        else:
            EXPIRY_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    else:
        EXPIRY_FILL = INFO_FILL
    for col, (label, val) in enumerate([
        ('License No', lic_no),
        ('License Date', license_date_str),
        ('Expiry Date', license_expiry_str),
        ('Total CIF', f"{total_license_cif:,.2f}"),
        ('Ledger Date', ledger_date_str),
    ], 1):
        c = ws.cell(row=r, column=col, value=f"{label}: {val}")
        c.fill = EXPIRY_FILL if col == 3 else INFO_FILL
        c.font = INFO_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='right' if col == 4 else 'left', vertical='center')
    r += 2

    # ══════════════════════════════════════════════════════════════════════
    # Section 2: Final Reconciliation Summary (three-way)
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f'A{r}:G{r}')
    sh = ws[f'A{r}']
    sh.value = 'FINAL RECONCILIATION SUMMARY'
    sh.fill = PatternFill(start_color="0B3D59", end_color="0B3D59", fill_type="solid")
    sh.font = Font(bold=True, color="FFFFFF", size=11)
    sh.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    rec_rows = [
        ('Financial Ledger Balance', reconciliation['financial_ledger_balance']),
        ('Customs Ledger Balance', reconciliation['customs_ledger_balance']),
        ('Licence Balance Engine', reconciliation['balance_engine']),
        ('Difference', reconciliation['difference']),
        ('Tolerance', reconciliation['tolerance']),
    ]
    diff_row = None
    for label, value in rec_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        vcell = ws.cell(row=r, column=2, value=float(value))
        vcell.number_format = '#,##0.00'
        if label == 'Difference':
            diff_row = r
        r += 1
    if not reconciliation['matched'] and diff_row:
        ws.cell(row=diff_row, column=1).fill = MISMATCH_FILL
        ws.cell(row=diff_row, column=2).fill = MISMATCH_FILL

    r += 1
    status_text = '✓ MATCHED' if reconciliation['matched'] else '⚠ DIFFERENCE FOUND'
    status_color = '1E8449' if reconciliation['matched'] else 'C0392B'
    ws.merge_cells(f'A{r}:G{r}')
    ws[f'A{r}'] = status_text
    ws[f'A{r}'].font = Font(bold=True, color=status_color, size=12)
    ws[f'A{r}'].alignment = Alignment(horizontal='center')
    r += 2

    # ══════════════════════════════════════════════════════════════════════
    # Section 3: Summary (BOE & Allotments)
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells(f'A{r}:G{r}')
    sh = ws[f'A{r}']
    sh.value = 'Summary (BOE & Allotments)'
    sh.fill = HDR_FILL; sh.font = Font(bold=True, color="FFFFFF", size=10)
    sh.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    SUMM_COLS = ['Item', 'Type', 'Company', 'Reference', 'Qty', 'Rate', 'CIF Value (FC)']
    for col, h in enumerate(SUMM_COLS, 1):
        _hdr(ws, r, col, h)
    r += 1

    for _s, _sd, row_data, is_boe in summary_rows:
        fill = BOE_FILL if is_boe else ALLOT_FILL
        _cell(ws, r, 1, row_data['item'],      fill=fill)
        _cell(ws, r, 2, row_data['type'],      fill=fill)
        _cell(ws, r, 3, row_data['company'],   fill=fill)
        _cell(ws, r, 4, row_data['reference'], fill=fill)
        _cell(ws, r, 5, row_data['qty'],       fill=fill, align='right', num_fmt='#,##0.00')
        _cell(ws, r, 6, row_data['rate'],      fill=fill, align='right', num_fmt='#,##0.00')
        _cell(ws, r, 7, row_data['cif'],       fill=fill, align='right', num_fmt='#,##0.00')
        r += 1

    # Total row
    if summary_rows:
        _cell(ws, r, 1, '', fill=TOTAL_FILL)
        _cell(ws, r, 2, '', fill=TOTAL_FILL)
        _cell(ws, r, 3, '', fill=TOTAL_FILL)
        _cell(ws, r, 4, 'TOTAL', fill=TOTAL_FILL, bold=True, align='right')
        _cell(ws, r, 5, '', fill=TOTAL_FILL)
        _cell(ws, r, 6, '', fill=TOTAL_FILL)
        _cell(ws, r, 7, total_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        r += 1

    r += 1  # blank row

    # ══════════════════════════════════════════════════════════════════════
    # Section 4: Plan Utilization  (matches the PlanTab UI exactly)
    # Shared with the bulk exporter's per-license sheet builder and the
    # "Utilization Planning Summary" sheet — see render_plan_utilization_section.
    # ══════════════════════════════════════════════════════════════════════
    from apps.license.services.plan_reporting import plan_map_for_license as _plan_map_fn
    from apps.license.services.exporters.planning_split_rows import render_plan_utilization_section
    _user_plan_map = _plan_map_fn(license_obj.id)

    r = render_plan_utilization_section(ws, r, license_obj, _license_balance, plan_map=_user_plan_map)

    # ── Column widths  (A=Item | B=HS | C=S.No | D=Status | E=Avail | F=Planned | G=Rem | H=PlannedCIF | I=RemCIF)
    ws.column_dimensions['A'].width = 38  # Item Description
    ws.column_dimensions['B'].width = 14  # HS Code
    ws.column_dimensions['C'].width = 14  # S.No
    ws.column_dimensions['D'].width = 14  # Status
    ws.column_dimensions['E'].width = 14  # Available Qty
    ws.column_dimensions['F'].width = 14  # Planned Qty
    ws.column_dimensions['G'].width = 14  # Remaining Qty
    ws.column_dimensions['H'].width = 16  # Planned CIF
    ws.column_dimensions['I'].width = 16  # Remaining CIF
    ws.column_dimensions['J'].width = 12

    ws.freeze_panes = 'A2'


def _write_audit_log_sheet(wb, license_obj):
    """
    Adds an "Audit Log" worksheet: every `ReconciliationLog` row for this
    licence's import items (the same append-only audit trail
    `LicenseBalanceLedgerBuilder.build_timeline()` reuses for its "Manual
    Adjustments" / reconciliation-action events — the
    `ReconciliationLog.objects.filter(license_item__license=license_obj)`
    query is reproduced verbatim here, see that method's docstring) plus
    every persisted `IgnoredWarning` row (ignore/restore history) for this
    licence, if any — both are real, persisted rows, never fabricated.
    """
    import json
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from apps.reconciliation.models import ReconciliationLog, IgnoredWarning

    ws = wb.create_sheet("Audit Log")

    HDR_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    ALT_FILL = PatternFill(start_color="F7F9FA", end_color="F7F9FA", fill_type="solid")
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def user_label(user):
        if not user:
            return '-'
        return getattr(user, 'get_full_name', lambda: None)() or getattr(user, 'username', None) or '-'

    def fmt_dt(dt):
        return dt.strftime('%d-%m-%Y %H:%M') if dt else '-'

    def fmt_json(value):
        if not value:
            return '-'
        try:
            return json.dumps(value, default=str)
        except TypeError:
            return str(value)

    ws.merge_cells('A1:I1')
    ws['A1'] = 'AUDIT LOG'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = HDR_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    r = 2
    ws.merge_cells(f'A{r}:I{r}')
    ws[f'A{r}'] = 'Reconciliation Actions'
    ws[f'A{r}'].font = Font(bold=True, color="FFFFFF", size=10)
    ws[f'A{r}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    logs = list(
        ReconciliationLog.objects.filter(license_item__license=license_obj)
        .select_related('user')
        .order_by('created_on')
    )
    log_headers = ['Sr', 'Date/Time', 'Action', 'User', 'Reason', 'Before', 'After']
    for col, h in enumerate(log_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    r += 1

    if not logs:
        ws.merge_cells(f'A{r}:I{r}')
        ws[f'A{r}'] = 'No reconciliation actions recorded for this licence yet.'
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
        r += 1
    else:
        for i, log in enumerate(logs, 1):
            fill = ALT_FILL if i % 2 == 0 else None
            values = [
                i, fmt_dt(log.created_on), log.get_action_display(),
                user_label(log.user), log.reason or '-',
                fmt_json(log.before), fmt_json(log.after),
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = THIN
                if fill:
                    c.fill = fill
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            r += 1

    r += 1

    ws.merge_cells(f'A{r}:I{r}')
    ws[f'A{r}'] = 'Ignored / Restored Warnings'
    ws[f'A{r}'].font = Font(bold=True, color="FFFFFF", size=10)
    ws[f'A{r}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    ignored_warnings = list(
        IgnoredWarning.objects.filter(license=license_obj)
        .select_related('ignored_by', 'restored_by')
        .order_by('-ignored_at')
    )
    warn_headers = [
        'Warning Type', 'Entity Type', 'Entity ID', 'Status',
        'Ignored By', 'Ignored At', 'Restored By', 'Restored At', 'Reason',
    ]
    for col, h in enumerate(warn_headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    r += 1

    if not ignored_warnings:
        ws.merge_cells(f'A{r}:I{r}')
        ws[f'A{r}'] = 'No ignored/restored warnings recorded for this licence yet.'
        ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
        r += 1
    else:
        for i, iw in enumerate(ignored_warnings, 1):
            fill = ALT_FILL if i % 2 == 0 else None
            values = [
                iw.warning_type, iw.entity_type, iw.entity_id,
                'Ignored' if iw.ignored else 'Restored',
                user_label(iw.ignored_by), fmt_dt(iw.ignored_at),
                user_label(iw.restored_by), fmt_dt(iw.restored_at),
                iw.reason or '-',
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = THIN
                if fill:
                    c.fill = fill
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 30
    ws.freeze_panes = 'A2'


def build_balance_excel(license_obj, show_hidden=False):
    """
    Generate the Licence Balance Workspace Excel export: five worksheets —
    "Financial Ledger", "Customs Ledger", "Timeline", "Reconciliation", and
    "Audit Log" — all built from `LicenseBalanceLedgerBuilder` data. Each
    builder method is called exactly ONCE here and its result passed to the
    relevant `_write_X_sheet` helper, so none of these worksheets can
    independently drift from the JSON API workspace or the PDF report —
    and no calculation happens anywhere in this module.

    `show_hidden` mirrors the on-screen "show hidden BOE" toggle for the
    Customs Ledger sheet — see `LicenseBalanceLedgerBuilder.build_customs_ledger`.

    Phase 4E-D: Running balance values sourced from CanonicalLedgerService,
    not independently recalculated.
    """
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO
    from apps.license.services.license_balance_ledger_builder import (
        LicenseBalanceLedgerBuilder, boe_invoice_allocation_map, boe_external_invoice_map,
    )
    from apps.license.services.canonical_ledger_service import CanonicalLedgerService

    # Fetch canonical dataset once (Phase 4E-D: single authoritative source)
    canonical_data = CanonicalLedgerService.build_canonical_ledger_dataset(
        license_id=license_obj.id,
        license_type='DFIA'
    )

    # Build balance map: transaction_id → running_balance
    canonical_balance_map = {
        txn['id']: txn['license_running_balance']
        for txn in canonical_data.get('transactions', [])
    }

    alloc_map = boe_invoice_allocation_map(license_obj)
    ext_map = boe_external_invoice_map(license_obj)
    financial_rows, financial_summary = LicenseBalanceLedgerBuilder.build_financial_ledger(
        license_obj,
        alloc_map,
        ext_map,
        canonical_balance_map=canonical_balance_map,
        canonical_data=canonical_data
    )
    customs_rows, customs_summary = LicenseBalanceLedgerBuilder.build_customs_ledger(
        license_obj,
        show_hidden=show_hidden,
        canonical_balance_map=canonical_balance_map,
        canonical_data=canonical_data
    )
    timeline_events = LicenseBalanceLedgerBuilder.build_timeline(license_obj)
    reconciliation = LicenseBalanceLedgerBuilder.build_reconciliation_summary(license_obj, financial_summary, customs_summary)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_financial_ledger_sheet(wb, financial_rows, financial_summary)
    _write_customs_ledger_sheet(wb, customs_rows, customs_summary)
    _write_timeline_sheet(wb, timeline_events)
    _write_reconciliation_sheet(wb, license_obj, financial_summary, customs_summary, reconciliation)
    _write_audit_log_sheet(wb, license_obj)

    # ── Save ──────────────────────────────────────────────────────────────
    # Recalculate formulas (e.g. the Plan Utilization TOTAL =SUM) on open so
    # viewers show computed values, not blank cached results.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{license_obj.license_number}-summary.xlsx"'
    return response


def build_bulk_balance_excel(request):
    """
    Generate a multi-sheet Excel with one sheet per license.
    Sheet name = license number. Same layout as balance_excel.
    POST body: {"license_numbers": ["3011007415", "3011007018", ...]}
    """
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from decimal import Decimal as _Dec
    from collections import defaultdict
    from apps.bill_of_entry.models import RowDetails, annotate_and_exclude_hidden
    from apps.allotment.models import AllotmentItems
    from rest_framework.response import Response
    from apps.license.models import LicenseDetailsModel

    license_numbers = request.data.get('license_numbers')
    if not isinstance(license_numbers, list):
        return Response({'error': 'license_numbers must be a non-empty list of strings.'}, status=400)

    license_numbers = [item.strip() for item in license_numbers if isinstance(item, str) and item.strip()]
    if not license_numbers:
        return Response({'error': 'license_numbers must be a non-empty list of strings.'}, status=400)
    license_numbers = list(dict.fromkeys(license_numbers))

    licenses = LicenseDetailsModel.objects.filter(
        license_number__in=license_numbers
    ).prefetch_related('import_license', 'import_license__items')

    if not licenses.exists():
        return Response({'error': 'No matching licenses found.'}, status=404)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Shared styles ──────────────────────────────────────────────────────
    HDR_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
    BOE_FILL   = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    ALLOT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    YEL_FILL   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ALT_FILL   = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    BOLD       = Font(bold=True, size=9)
    NORM       = Font(size=9)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return c

    def _cell(ws, row, col, value, fill=None, bold=False, align='left', num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill: c.fill = fill
        c.font = BOLD if bold else NORM
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if num_fmt: c.number_format = num_fmt
        return c

    def _write_license_sheet(wb, license_obj):
        from datetime import date as _date_cls
        requested_sheet_name = str(license_obj.license_number or license_obj.pk)[:31]
        ws = wb.create_sheet(title=requested_sheet_name)
        sheet_name = ws.title

        license_date_str = license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-'
        license_expiry_str = license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-'
        ledger_date_str = license_obj.ledger_date.strftime('%d-%m-%Y') if license_obj.ledger_date else '-'
        lic_no = license_obj.license_number or '-'

        summary_rows = []
        total_cif = 0.0
        total_cif_inr = 0.0

        for item in license_obj.import_license.all():
            item_name = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else (item.description or '-')

            # Previous-owner "hidden" BOEs (genuinely hidden per audit
            # trail) are excluded — this bulk export is a balance/
            # financial figure, not the Customs History audit view.
            boes = annotate_and_exclude_hidden(
                RowDetails.objects.filter(sr_number_id=item.id, transaction_type='D'),
                boe_field="bill_of_entry",
            ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')

            for rd in boes:
                qty  = float(rd.qty or 0)
                cif  = float(rd.cif_fc or 0)
                cif_inr = float(rd.cif_inr or 0)
                rate = cif / qty if qty else 0.0
                total_cif += cif
                total_cif_inr += cif_inr
                boe_company = rd.bill_of_entry.company.name if rd.bill_of_entry.company else '-'
                ref_no   = rd.bill_of_entry.bill_of_entry_number or '-'
                ref_date = rd.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if rd.bill_of_entry.bill_of_entry_date else ''
                ref_str  = f"{ref_no} / {ref_date}" if ref_date else ref_no
                product  = rd.bill_of_entry.product_name or item_name
                _sort_dt = rd.bill_of_entry.bill_of_entry_date or _date_cls.min
                summary_rows.append((0, _sort_dt, {
                    'item': product, 'type': 'BOE', 'company': boe_company,
                    'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif,
                    'cif_inr': cif_inr,
                }, True))

            allotments = AllotmentItems.objects.filter(
                item_id=item.id, allotment__bill_of_entry__isnull=True
            ).select_related('allotment', 'allotment__company')

            for ai in allotments:
                qty     = float(ai.qty or 0)
                cif     = float(ai.cif_fc or 0)
                cif_inr = float(ai.cif_inr or 0)
                rate    = cif / qty if qty else 0.0
                total_cif += cif
                total_cif_inr += cif_inr
                company = ai.allotment.company.name if ai.allotment.company else '-'
                invoice = ai.allotment.invoice or '-'
                eta     = ai.allotment.estimated_arrival_date.strftime('%d-%m-%Y') if ai.allotment.estimated_arrival_date else ''
                ref_str = f"{invoice} / ETA: {eta}" if eta else invoice
                product = ai.allotment.item_name or item_name
                _sort_dt = ai.allotment.estimated_arrival_date or _date_cls.min
                summary_rows.append((1, _sort_dt, {
                    'item': product, 'type': 'Allotment', 'company': company,
                    'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif,
                    'cif_inr': cif_inr,
                }, False))

        # BOEs first (sorted by BOE date), then allotments (sorted by allotment date)
        summary_rows.sort(key=lambda x: (x[0], x[1]))

        # New restriction model: condition_type on LicenseImportItemsModel is
        # the source of truth. Percentage conditions share a pool computed
        # by compute_condition_pools(); AU / blank conditions use the full
        # licence balance.
        from apps.license.services.condition_pool import compute_condition_pools as _ccp
        _cond_pools = _ccp(license_obj)

        _bal_agg = defaultdict(lambda: {
            'qty': 0.0, 'total_qty': 0.0, 'sr_ids': [],
            'description': '', 'hs_code': '', 'condition_type': '',
            'qty_by_cond': {},
            'plan_qty': 0.0, 'plan_cif': 0.0,
        })
        # Pre-computed once for every exported license by the caller (see
        # `_balance_by_license` in `build_bulk_balance_excel`) — reused here
        # to avoid re-triggering `get_balance_cif` for every license.
        _license_balance = float(_balance_by_license.get(license_obj.id, 0) or 0)
        # Read persisted plan from LicenseItemPlan only (NO planner calls in read path).
        from apps.license.models import LicenseItemPlan
        from decimal import Decimal

        # Fetch persisted plans for this license
        persisted_plans = (
            LicenseItemPlan.objects
            .filter(license=license_obj)
            .select_related('import_item')
            .values('import_item_id', 'planned_quantity', 'planned_cif_fc')
        )

        # Build plan map: import_item_id -> {planned_quantity, planned_cif}
        _plan_map = {}
        _plan_source = ''
        for plan in persisted_plans:
            iid = plan['import_item_id']
            if iid not in _plan_map:
                _plan_map[iid] = {
                    'planned_quantity': Decimal('0'),
                    'planned_cif': Decimal('0'),
                }
            _plan_map[iid]['planned_quantity'] += Decimal(str(plan['planned_quantity'] or 0))
            _plan_map[iid]['planned_cif'] += Decimal(str(plan['planned_cif_fc'] or 0))

        # Set source if any plans exist
        if _plan_map:
            _plan_source = 'manual'
        else:
            _plan_map = {}
            _plan_source = ''
        for _item in license_obj.import_license.all():
            _key = ', '.join(sorted([i.name for i in _item.items.all()])) if _item.items.exists() else (_item.description or '-')
            _avail = float(_item.available_quantity or 0)
            _bal_agg[_key]['qty'] += _avail
            _bal_agg[_key]['total_qty'] += float(_item.quantity or 0)
            _pl = _plan_map.get(_item.id)
            if _pl:
                # This aggregation feeds openpyxl numeric cells and the
                # surrounding balance fields are floats.  Keep the boundary
                # explicit: persisted plans are Decimal, but never mix the
                # two types in Python arithmetic.
                _bal_agg[_key]['plan_qty'] += float(_pl['planned_quantity'])
                _bal_agg[_key]['plan_cif'] += float(_pl['planned_cif'])
            _bal_agg[_key]['sr_ids'].append(_item.serial_number)
            if not _bal_agg[_key]['description']:
                _bal_agg[_key]['description'] = _item.description or _key
            if not _bal_agg[_key]['hs_code']:
                _bal_agg[_key]['hs_code'] = str(_item.hs_code.hs_code if _item.hs_code else '-')
            # Carry per-item licence-condition (AU / 2% / 3% / 5% / 10%)
            # through to the bulk-balance Excel cell.
            if _item.condition_type and not _bal_agg[_key]['condition_type']:
                _bal_agg[_key]['condition_type'] = _item.condition_type
            # Per-condition qty breakdown for E1's Display/Util-qty split.
            _ct = (_item.condition_type or '').strip()
            _bal_agg[_key]['qty_by_cond'][_ct] = _bal_agg[_key]['qty_by_cond'].get(_ct, 0.0) + _avail

        total_license_cif = total_cif + _license_balance

        r = 1
        _today = _date_cls.today()
        INFO_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        INFO_FONT = Font(bold=True, color="FFFFFF", size=9)
        if license_obj.license_expiry_date:
            _days = (license_obj.license_expiry_date - _today).days
            if _days < 0:
                EXPIRY_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            elif _days <= 90:
                EXPIRY_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            else:
                EXPIRY_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        else:
            EXPIRY_FILL = INFO_FILL
        iec_val = license_obj.exporter.iec if license_obj.exporter else '-'
        for col, (label, val) in enumerate([
            ('License No', lic_no),
            ('IEC', iec_val),
            ('License Date', license_date_str),
            ('Expiry Date', license_expiry_str),
            ('Total CIF', f"{total_license_cif:,.2f}"),
            ('Ledger Date', ledger_date_str),
        ], 1):
            c = ws.cell(row=r, column=col, value=f"{label}: {val}")
            c.fill = EXPIRY_FILL if col == 4 else INFO_FILL
            c.font = INFO_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='right' if col == 5 else 'left', vertical='center')
        r += 1

        ws.merge_cells(f'A{r}:H{r}')
        sh = ws[f'A{r}']
        sh.value = 'Summary (BOE & Allotments)'
        sh.fill = HDR_FILL; sh.font = Font(bold=True, color="FFFFFF", size=10)
        sh.alignment = Alignment(horizontal='center', vertical='center')
        r += 1

        SUMM_COLS = ['Item', 'Type', 'Company', 'Reference', 'Qty', 'Rate', 'CIF Value (FC)', 'CIF Value (INR)']
        for col, h in enumerate(SUMM_COLS, 1):
            _hdr(ws, r, col, h)
        r += 1

        for _s, _sd, row_data, is_boe in summary_rows:
            fill = BOE_FILL if is_boe else ALLOT_FILL
            _cell(ws, r, 1, row_data['item'],          fill=fill)
            _cell(ws, r, 2, row_data['type'],          fill=fill)
            _cell(ws, r, 3, row_data['company'],       fill=fill)
            _cell(ws, r, 4, row_data['reference'],     fill=fill)
            _cell(ws, r, 5, row_data['qty'],           fill=fill, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 6, row_data['rate'],          fill=fill, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 7, row_data['cif'],           fill=fill, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 8, row_data.get('cif_inr', 0), fill=fill, align='right', num_fmt='#,##0.00')
            r += 1

        if summary_rows:
            _cell(ws, r, 1, '', fill=TOTAL_FILL); _cell(ws, r, 2, '', fill=TOTAL_FILL)
            _cell(ws, r, 3, '', fill=TOTAL_FILL)
            _cell(ws, r, 4, 'TOTAL', fill=TOTAL_FILL, bold=True, align='right')
            _cell(ws, r, 5, '', fill=TOTAL_FILL); _cell(ws, r, 6, '', fill=TOTAL_FILL)
            _cell(ws, r, 7, total_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 8, total_cif_inr, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
            r += 1

        r += 1

        # ── Plan Utilization section (matches PlanTab UI) ───────────────────
        # Shared with the single-licence sheet above and with the
        # "Utilization Planning Summary" sheet — see
        # render_plan_utilization_section.
        from apps.license.services.plan_reporting import plan_map_for_license as _plan_map_fn_bulk
        from apps.license.services.exporters.planning_split_rows import render_plan_utilization_section
        from apps.license.services.plan_utilization import plan_utilization_rows as _plan_utilization_rows_bulk
        _user_plan_map_b = _plan_map_fn_bulk(license_obj.id)
        # Computed ONCE per license and reused by both this sheet's own
        # "Plan Utilization" table AND the "Utilization Planning Summary"
        # sheet's Planning Matrix pivot (`_license_pivot_data`, via
        # `_util_return['groups']` below) — that pivot used to call
        # `plan_utilization_rows()` a second, independent time per license,
        # doubling the whole plan_status_for/group_ids_of query cost for
        # every exported license.
        _groups_b = _plan_utilization_rows_bulk(license_obj, plan_map=_user_plan_map_b)

        _plan_totals_b = {}
        r = render_plan_utilization_section(
            ws, r, license_obj, _license_balance,
            plan_map=_user_plan_map_b, totals_out=_plan_totals_b, groups=_groups_b,
        )

        # _util_return — feeds the "Utilization Planning Summary" sheet. Carries
        # the already-fetched `license_obj`/`plan_map`/`groups` so that sheet
        # can build its pivot for the same license without re-querying
        # plan_map_for_license or re-running plan_utilization_rows.
        _exporter_name = license_obj.exporter.name if license_obj.exporter else ''
        _util_return = {
            'lic_no': lic_no,
            'license_obj': license_obj,
            'balance_cif': _license_balance,
            'total_license_cif': total_license_cif,
            'license_date': license_obj.license_date,
            'license_expiry_date': license_obj.license_expiry_date,
            'exporter_name': _exporter_name,
            'sheet_name': sheet_name,
            'plan_map': _user_plan_map_b,
            'plan_totals': _plan_totals_b,
            'groups': _groups_b,
        }

        ws.column_dimensions['A'].width = 38  # Item Description
        ws.column_dimensions['B'].width = 14  # HS Code
        ws.column_dimensions['C'].width = 14  # S.No
        ws.column_dimensions['D'].width = 14  # Status
        ws.column_dimensions['E'].width = 14  # Available Qty
        ws.column_dimensions['F'].width = 14  # Planned Qty
        ws.column_dimensions['G'].width = 14  # Remaining Qty
        ws.column_dimensions['H'].width = 16  # Planned CIF
        ws.column_dimensions['I'].width = 16  # Remaining CIF
        ws.column_dimensions['J'].width = 12
        ws.freeze_panes = 'A2'
        return _util_return

    def _norm_sort_key(lic):
        norms = list(lic.export_license.values_list('norm_class__norm_class', flat=True))
        norm_str = ', '.join(sorted(str(n) for n in norms if n)) or 'ZZZ'
        # Group order: E1 first, E5 second, rest alphabetically
        if any('E1' in str(n) and 'E126' not in str(n) and 'E132' not in str(n) for n in norms if n):
            return ('0_E1', norm_str)
        if any(str(n).strip() == 'E5' for n in norms if n):
            return ('1_E5', norm_str)
        return ('2_' + norm_str, norm_str)

    sorted_licenses = sorted(licenses, key=_norm_sort_key)

    # Batch-compute every exported license's final balance (== what
    # `license_obj.get_balance_cif` / `LicenseBalanceCalculator.calculate_balance`
    # would return) ONCE here, in 4 queries total, instead of `_write_license_sheet`
    # triggering `get_balance_cif` per license (4 queries: credit+debit+
    # allotment+trade). `_write_license_sheet` reads this dict via closure
    # instead of the model property. Plans are read directly from persisted
    # LicenseItemPlan (no planner calls in read path). Nothing outside this
    # bulk export is touched — `get_balance_cif` itself is unchanged.
    from apps.license.services.balance_calculator import LicenseBalanceCalculator as _LBC_bulk
    _bulk_lic_ids = [lic.id for lic in sorted_licenses]
    _balance_by_license = _LBC_bulk.calculate_balance_for_licenses(_bulk_lic_ids)

    _util_summaries = []
    for license_obj in sorted_licenses:
        _util_summaries.append(_write_license_sheet(wb, license_obj))

    # ── Create Utilization Planning Summary as first sheet ─────────────────
    # Norm-grouped planning MATRIX (third design of this sheet): one section
    # per distinct SION norm actually present among the exported licenses —
    # a Planning Matrix (one row per license, pivoted by Planning Item Name
    # into Available/Planned Qty/Planned CIF column groups), a Norm Total
    # row, and a Planning Item Summary (the same per-item totals,
    # transposed). After every norm section: a Grand Summary by Norm and a
    # Grand Total. No norm-specific or item-name-specific branching —
    # whatever norms/planning items exist in the data become
    # sections/columns automatically.
    #
    # Data source is exclusively `plan_utilization_rows()` (the same
    # function `render_plan_utilization_section()` calls for every
    # per-license sheet) plus `rows_for_splits()` for the visible-split
    # filter — no new planning calculation, no new split query.
    #
    # Two attribution rules that must NOT be conflated (see the docstrings
    # below for the derivation each one mirrors):
    #   - Planning-item pivot columns (matrix / Norm Total / Planning Item
    #     Summary): planned qty/CIF are split-exclusive (each visible split
    #     counted under exactly one item-name column, 'Unassigned' when
    #     untagged); available qty is the group's FULL available_quantity
    #     attributed to EVERY distinct item-name column with a visible split
    #     in that group — an intentional double count across columns,
    #     matching `item_pivot_report.py`'s `_build_license_row` convention
    #     (`for item in import_item.items.all(): item_quantities[item.id]
    #     ['available_quantity'] += ...` — a shared import item's
    #     availability is added under every attached item name, not
    #     divided).
    #   - Grand Summary by Norm / Grand Total: available/planned/CIF are
    #     GROUP-level totals (never derived from the pivot columns, which
    #     would inherit the available-qty double count above) — the exact
    #     same derivation `render_plan_utilization_section()` uses for its
    #     `totals_out` (Σ available_quantity across every group; Σ
    #     original_quantity/original_cif_fc across PLANNED groups only).
    from apps.license.services.plan_utilization import plan_utilization_rows
    from apps.license.services.exporters.planning_split_rows import rows_for_splits
    from apps.license.models import LicenseExportItemModel
    from apps.license.services.balance_calculator import LicenseBalanceCalculator
    from openpyxl.utils import get_column_letter

    _UNASSIGNED = 'Unassigned'

    # One batched query for every exported license's PRIMARY export norm
    # (mirrors norm_plan.detect_norm's own `export_license.first()` — i.e.
    # the lowest-pk export item — without querying per license in a loop).
    _lic_ids = _bulk_lic_ids
    _first_norm_by_license: dict = {}
    for _lic_id, _norm_code in (
        LicenseExportItemModel.objects
        .filter(license_id__in=_lic_ids)
        .order_by('license_id', 'pk')
        .values_list('license_id', 'norm_class__norm_class')
    ):
        _first_norm_by_license.setdefault(_lic_id, _norm_code)

    # One batched query each for Total CIF / Debited CIF across every
    # exported license, instead of `LicenseBalanceCalculator.calculate_credit`/
    # `calculate_boe_debit_total` being called once per license (2 queries x
    # 214 licenses in production) inside the per-license loop below. Raw,
    # unconditional BOE debit — matches the Balance CIF formula exactly
    # (`calculate_balance`), not the allocation-netted `calculate_debit`.
    _credit_by_license = LicenseBalanceCalculator.calculate_credit_for_licenses(_lic_ids)
    _debit_by_license = LicenseBalanceCalculator.calculate_boe_debit_total_for_licenses(_lic_ids)

    # Bucket licenses by norm, preserving the order they were exported in
    # (== `sorted_licenses` order, already E1-first/E5-second/alpha-rest).
    _licenses_by_norm: dict = {}
    _norm_order: list = []
    for _row in _util_summaries:
        _norm_code = (_first_norm_by_license.get(_row['license_obj'].id) or '').strip()
        _norm_label = _norm_code or 'Unclassified'
        if _norm_label not in _licenses_by_norm:
            _licenses_by_norm[_norm_label] = []
            _norm_order.append(_norm_label)
        _licenses_by_norm[_norm_label].append(_row)

    def _license_pivot_data(_lic_row):
        """One license's pivot data: `{item_name: {available, planned_qty,
        planned_cif}}` (floats, split-exclusive planned figures / doubled-up
        available per the module docstring above) plus this license's
        GROUP-level totals (`available_quantity`, `planned_quantity`,
        `planned_cif`) — derived exactly like
        `render_plan_utilization_section`'s own `_g_avail`/`_g_planned_qty`/
        `_g_planned_cif` (see that function for the proof these equal Σ of
        the group's unioned splits when the group has a plan).
        """
        _lic_obj = _lic_row['license_obj']
        # Reuse the SAME groups `_write_license_sheet` already computed for
        # this license (see `_util_return['groups']`) instead of calling
        # `plan_utilization_rows()` a second time — that used to double the
        # whole plan_status_for/group_ids_of query cost for every license.
        _groups = _lic_row.get('groups')
        if _groups is None:  # pragma: no cover - defensive; always set today
            _groups = plan_utilization_rows(_lic_obj, plan_map=_lic_row.get('plan_map'))
        _item_data: dict = {}
        _totals = {'available_quantity': 0.0, 'planned_quantity': 0.0, 'planned_cif': 0.0}
        for _grp in _groups:
            _avail = float(_grp.get('available_quantity') or 0)
            _totals['available_quantity'] += _avail
            _has_plan = bool(_grp.get('has_plan'))
            _totals['planned_quantity'] += float(_grp['original_quantity']) if _has_plan else 0.0
            _totals['planned_cif'] += float(_grp['original_cif_fc']) if _has_plan else 0.0

            _distinct_names_in_group = set()
            for _sp in rows_for_splits(_grp.get('splits') or []):
                _name = (_sp.get('item_name') or '').strip() or _UNASSIGNED
                _distinct_names_in_group.add(_name)
                _bucket = _item_data.setdefault(
                    _name, {'available': 0.0, 'planned_qty': 0.0, 'planned_cif': 0.0},
                )
                _bucket['planned_qty'] += float(_sp.get('planned_quantity') or 0)
                _bucket['planned_cif'] += float(_sp.get('planned_cif_fc') or 0)
            for _name in _distinct_names_in_group:
                _item_data.setdefault(
                    _name, {'available': 0.0, 'planned_qty': 0.0, 'planned_cif': 0.0},
                )['available'] += _avail
        return _item_data, _totals

    _sw = wb.create_sheet(title="Utilization Planning Summary")
    wb.move_sheet(_sw, offset=-(len(wb.worksheets) - 1))

    ITEM_HDR_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ITEM_HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    NORM_BANNER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    # ── Expiry traffic-light — computed once at export time (no conditional-
    # formatting formulas, so colors stay correct even with recalc disabled).
    # Applied only to a license's own Planning Matrix row (see the per-license
    # loop below) — never to section banners, header rows, Norm Total,
    # Planning Item Summary, Grand Summary by Norm, or Grand Total.
    EXPIRED_FILL      = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    EXPIRED_FONT      = Font(bold=True, color="FFFFFF", size=9)
    EXPIRING_30_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    EXPIRING_30_FONT  = Font(bold=True, color="000000", size=9)
    EXPIRING_60_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    EXPIRING_60_FONT  = Font(bold=False, color="000000", size=9)
    from datetime import date as _expiry_date_cls
    _expiry_today = _expiry_date_cls.today()

    def _expiry_highlight(expiry_date):
        """(fill, font) override for a license row, or (None, None) when the
        license isn't within the expiry window — callers keep the normal
        alternating-row fill/font in that case."""
        if not expiry_date:
            return None, None
        days = (expiry_date - _expiry_today).days
        if days < 0:
            return EXPIRED_FILL, EXPIRED_FONT
        if days <= 30:
            return EXPIRING_30_FILL, EXPIRING_30_FONT
        if days <= 60:
            return EXPIRING_60_FILL, EXPIRING_60_FONT
        return None, None

    _sr = 1
    _max_matrix_cols = 5  # widened as sections are laid out; Grand Summary needs at least 5.

    # One row per norm: (norm_label, license_count, available, planned_qty, planned_cif) —
    # GROUP-level totals only, per the module docstring above.
    _grand_summary_by_norm: list = []

    for _norm_label in _norm_order:
        _norm_rows = _licenses_by_norm[_norm_label]

        # Pass 1: this norm's per-license pivot data/totals + the set of
        # distinct planning-item names present anywhere in this norm.
        _per_license = []
        _item_names_in_norm: set = set()
        for _row in _norm_rows:
            _item_data, _totals = _license_pivot_data(_row)
            _per_license.append((_row, _item_data, _totals))
            _item_names_in_norm.update(_item_data.keys())

        _ordered_item_names = sorted(
            (n for n in _item_names_in_norm if n != _UNASSIGNED), key=lambda s: s.casefold(),
        )
        if _UNASSIGNED in _item_names_in_norm:
            _ordered_item_names.append(_UNASSIGNED)

        _n_cols = 9 + 3 * len(_ordered_item_names)
        _max_matrix_cols = max(_max_matrix_cols, _n_cols)

        # ── 1. Section header ────────────────────────────────────────────
        _sw.merge_cells(start_row=_sr, start_column=1, end_row=_sr, end_column=_n_cols)
        _nh = _sw.cell(row=_sr, column=1, value=f"SION NORM : {_norm_label}")
        _nh.fill = NORM_BANNER_FILL
        _nh.font = Font(bold=True, color="FFFFFF", size=12)
        _nh.alignment = Alignment(horizontal='center', vertical='center')
        _sr += 1

        # ── 2. Planning Matrix — 2-row header + one row per license ─────
        _hdr_row1, _hdr_row2 = _sr, _sr + 1
        for _col, _label in enumerate([
            'License No', 'Issue Date', 'Expiry Date', 'Exporter', 'SION Norm',
            'Total CIF ($)', 'Debited CIF ($)', 'Allotted CIF ($)', 'Balance CIF ($)',
        ], 1):
            _sw.merge_cells(start_row=_hdr_row1, start_column=_col, end_row=_hdr_row2, end_column=_col)
            _hdr(_sw, _hdr_row1, _col, _label)
        for _i, _name in enumerate(_ordered_item_names):
            _base_col = 10 + _i * 3
            _sw.merge_cells(start_row=_hdr_row1, start_column=_base_col, end_row=_hdr_row1, end_column=_base_col + 2)
            _ic = _sw.cell(row=_hdr_row1, column=_base_col, value=_name)
            _ic.fill = ITEM_HDR_FILL; _ic.font = ITEM_HDR_FONT
            _ic.border = THIN_BORDER
            _ic.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for _sub_col, _sub_label in enumerate(['Available Qty', 'Planned Qty', 'Planned CIF ($)']):
                _hdr(_sw, _hdr_row2, _base_col + _sub_col, _sub_label)
        _sr += 2

        _norm_item_totals = {
            _n: {'available': 0.0, 'planned_qty': 0.0, 'planned_cif': 0.0} for _n in _ordered_item_names
        }
        _norm_totals = {'available_quantity': 0.0, 'planned_quantity': 0.0, 'planned_cif': 0.0}
        _norm_cif_totals = {'total_cif': 0.0, 'debited_cif': 0.0, 'allotted_cif': 0.0, 'balance_cif': 0.0}

        for _idx, (_row, _item_data, _totals) in enumerate(_per_license):
            _rf = None if _idx % 2 == 0 else ALT_FILL
            _ld = _row.get('license_date')
            _ed = _row.get('license_expiry_date')
            _lic_obj = _row['license_obj']

            # Total CIF / Debited CIF — sourced directly from the centralized
            # LicenseBalanceCalculator (the same single source of truth
            # `get_balance_cif`/`calculate_balance` themselves compose from),
            # via the batched `_credit_by_license`/`_debit_by_license` maps
            # computed once above — same numbers `calculate_credit`/
            # `calculate_debit` would give per license (Coalesce'd to 0 when
            # a license has none), just without a query per license.
            # Deliberately NOT the ad-hoc `total_cif`/`total_license_cif`
            # computed earlier in `_write_license_sheet` from the raw
            # BOE/allotment summary-rows loop — that older calculation
            # doesn't exclude BOEs linked to trades and ignores
            # `calculate_trade()`.
            _total_cif = float(_credit_by_license.get(_lic_obj.id, 0))
            _debited_cif = float(_debit_by_license.get(_lic_obj.id, 0))
            # Allotted CIF — per product's explicit instruction, this is the
            # Plan Utilization "Planned CIF" figure (Σ this license's group
            # totals, same number feeding Grand Summary by Norm), NOT
            # LicenseBalanceCalculator.calculate_allotment()'s real-allotment
            # figure. The two can legitimately differ when a plan isn't
            # fully executed as real allotments yet.
            _allotted_cif = _totals['planned_cif']
            # Balance CIF — reuse the already-computed/cached value from
            # `_write_license_sheet` (== `license_obj.get_balance_cif`).
            # NOT recomputed as Total − Debited − Allotted: `get_balance_cif`
            # nets against real allotment transactions + trade CIF, neither
            # of which the other three columns expose, so the four columns
            # will not always reconcile arithmetically — that's expected.
            _balance_cif = _row['balance_cif']

            _cell(_sw, _sr, 1, _row['lic_no'], fill=_rf, bold=True)
            _cell(_sw, _sr, 2, _ld.strftime('%d-%m-%Y') if _ld else '-', fill=_rf, align='center')
            _cell(_sw, _sr, 3, _ed.strftime('%d-%m-%Y') if _ed else '-', fill=_rf, align='center')
            _cell(_sw, _sr, 4, _row.get('exporter_name') or '-', fill=_rf)
            _cell(_sw, _sr, 5, _norm_label, fill=_rf)
            _cell(_sw, _sr, 6, _total_cif, fill=_rf, align='right', num_fmt='#,##0.00')
            _cell(_sw, _sr, 7, _debited_cif, fill=_rf, align='right', num_fmt='#,##0.00')
            _cell(_sw, _sr, 8, _allotted_cif, fill=_rf, align='right', num_fmt='#,##0.00')
            _cell(_sw, _sr, 9, _balance_cif, fill=_rf, align='right', num_fmt='#,##0.00')
            _norm_cif_totals['total_cif'] += _total_cif
            _norm_cif_totals['debited_cif'] += _debited_cif
            _norm_cif_totals['allotted_cif'] += _allotted_cif
            _norm_cif_totals['balance_cif'] += _balance_cif
            for _i, _name in enumerate(_ordered_item_names):
                _base_col = 10 + _i * 3
                _vals = _item_data.get(_name) or {'available': 0.0, 'planned_qty': 0.0, 'planned_cif': 0.0}
                _cell(_sw, _sr, _base_col, _vals['available'], fill=_rf, align='right', num_fmt='#,##0.000')
                _cell(_sw, _sr, _base_col + 1, _vals['planned_qty'], fill=_rf, align='right', num_fmt='#,##0.000')
                _cell(_sw, _sr, _base_col + 2, _vals['planned_cif'], fill=_rf, align='right', num_fmt='#,##0.00')
                _norm_item_totals[_name]['available'] += _vals['available']
                _norm_item_totals[_name]['planned_qty'] += _vals['planned_qty']
                _norm_item_totals[_name]['planned_cif'] += _vals['planned_cif']
            for _k in _norm_totals:
                _norm_totals[_k] += _totals.get(_k, 0.0)

            # Expiry traffic-light — overrides this row's fill/font across
            # every column (License No through the last item's Planned CIF)
            # when the license is expired or expiring soon; otherwise the
            # normal alternating-row fill/font above is left untouched.
            _hl_fill, _hl_font = _expiry_highlight(_ed)
            if _hl_fill is not None:
                for _hc in range(1, _n_cols + 1):
                    _hcell = _sw.cell(row=_sr, column=_hc)
                    _hcell.fill = _hl_fill
                    _hcell.font = _hl_font

            _sr += 1

        # ── 3. Norm Total row ─────────────────────────────────────────────
        for _ci in range(1, _n_cols + 1):
            _sw.cell(row=_sr, column=_ci).fill = TOTAL_FILL
            _sw.cell(row=_sr, column=_ci).border = THIN_BORDER
        _cell(_sw, _sr, 1, 'NORM TOTAL', fill=TOTAL_FILL, bold=True)
        _cell(_sw, _sr, 2, '', fill=TOTAL_FILL); _cell(_sw, _sr, 3, '', fill=TOTAL_FILL); _cell(_sw, _sr, 4, '', fill=TOTAL_FILL)
        _cell(_sw, _sr, 5, '', fill=TOTAL_FILL)
        _cell(_sw, _sr, 6, _norm_cif_totals['total_cif'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        _cell(_sw, _sr, 7, _norm_cif_totals['debited_cif'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        _cell(_sw, _sr, 8, _norm_cif_totals['allotted_cif'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        _cell(_sw, _sr, 9, _norm_cif_totals['balance_cif'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        for _i, _name in enumerate(_ordered_item_names):
            _base_col = 10 + _i * 3
            _t = _norm_item_totals[_name]
            _cell(_sw, _sr, _base_col, _t['available'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
            _cell(_sw, _sr, _base_col + 1, _t['planned_qty'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
            _cell(_sw, _sr, _base_col + 2, _t['planned_cif'], fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
        _sr += 2

        # ── 4. Planning Item Summary — same numbers, transposed ─────────
        _sw.merge_cells(start_row=_sr, start_column=1, end_row=_sr, end_column=4)
        _psh = _sw.cell(row=_sr, column=1, value='PLANNING ITEM SUMMARY')
        _psh.fill = HDR_FILL; _psh.font = Font(bold=True, color="FFFFFF", size=10)
        _psh.alignment = Alignment(horizontal='center', vertical='center')
        _sr += 1
        for _ci, _ch in enumerate(['Planning Item', 'Available Qty', 'Planned Qty', 'Planned CIF ($)'], 1):
            _hdr(_sw, _sr, _ci, _ch)
        _sr += 1
        for _i, _name in enumerate(_ordered_item_names):
            _rf = None if _i % 2 == 0 else ALT_FILL
            _t = _norm_item_totals[_name]
            _cell(_sw, _sr, 1, _name, fill=_rf, bold=True)
            _cell(_sw, _sr, 2, _t['available'], fill=_rf, align='right', num_fmt='#,##0.000')
            _cell(_sw, _sr, 3, _t['planned_qty'], fill=_rf, align='right', num_fmt='#,##0.000')
            _cell(_sw, _sr, 4, _t['planned_cif'], fill=_rf, align='right', num_fmt='#,##0.00')
            _sr += 1
        _sr += 2

        _grand_summary_by_norm.append((
            _norm_label, len(_norm_rows),
            _norm_totals['available_quantity'], _norm_totals['planned_quantity'], _norm_totals['planned_cif'],
        ))

    # ── 5. Grand Summary by Norm ──────────────────────────────────────────
    _sw.merge_cells(start_row=_sr, start_column=1, end_row=_sr, end_column=5)
    _gsh = _sw.cell(row=_sr, column=1, value='GRAND SUMMARY BY NORM')
    _gsh.fill = HDR_FILL; _gsh.font = Font(bold=True, color="FFFFFF", size=11)
    _gsh.alignment = Alignment(horizontal='center', vertical='center')
    _sr += 1
    for _ci, _ch in enumerate(['SION Norm', 'Licenses', 'Available Qty', 'Planned Qty', 'Planned CIF ($)'], 1):
        _hdr(_sw, _sr, _ci, _ch)
    _sr += 1

    _grand_total_licenses = 0
    _grand_total_available = 0.0
    _grand_total_planned_qty = 0.0
    _grand_total_planned_cif = 0.0
    for _i, (_norm_label, _lic_count, _avail, _pqty, _pcif) in enumerate(_grand_summary_by_norm):
        _rf = None if _i % 2 == 0 else ALT_FILL
        _cell(_sw, _sr, 1, _norm_label, fill=_rf, bold=True)
        _cell(_sw, _sr, 2, _lic_count, fill=_rf, align='center')
        _cell(_sw, _sr, 3, _avail, fill=_rf, align='right', num_fmt='#,##0.000')
        _cell(_sw, _sr, 4, _pqty, fill=_rf, align='right', num_fmt='#,##0.000')
        _cell(_sw, _sr, 5, _pcif, fill=_rf, align='right', num_fmt='#,##0.00')
        _grand_total_licenses += _lic_count
        _grand_total_available += _avail
        _grand_total_planned_qty += _pqty
        _grand_total_planned_cif += _pcif
        _sr += 1
    _sr += 1

    # ── 6. Grand Total ────────────────────────────────────────────────────
    for _ci in range(1, 6):
        _sw.cell(row=_sr, column=_ci).fill = TOTAL_FILL
        _sw.cell(row=_sr, column=_ci).border = THIN_BORDER
    _cell(_sw, _sr, 1, 'GRAND TOTAL', fill=TOTAL_FILL, bold=True)
    _cell(_sw, _sr, 2, _grand_total_licenses, fill=TOTAL_FILL, bold=True, align='center')
    _cell(_sw, _sr, 3, _grand_total_available, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
    _cell(_sw, _sr, 4, _grand_total_planned_qty, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
    _cell(_sw, _sr, 5, _grand_total_planned_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')

    # ── Column widths ──────────────────────────────────────────────────────
    _sw.column_dimensions['A'].width = 22  # License No
    _sw.column_dimensions['B'].width = 14  # Issue Date
    _sw.column_dimensions['C'].width = 14  # Expiry Date
    _sw.column_dimensions['D'].width = 24  # Exporter
    _sw.column_dimensions['E'].width = 14  # SION Norm
    _sw.column_dimensions['F'].width = 16  # Total CIF ($)
    _sw.column_dimensions['G'].width = 16  # Debited CIF ($)
    _sw.column_dimensions['H'].width = 16  # Allotted CIF ($)
    _sw.column_dimensions['I'].width = 16  # Balance CIF ($)
    for _col in range(10, _max_matrix_cols + 1):
        _sw.column_dimensions[get_column_letter(_col)].width = 14

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bulk_license_summary.xlsx"'
    return response
