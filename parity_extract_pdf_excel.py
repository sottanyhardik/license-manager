#!/usr/bin/env python3
"""
Extract and verify values from generated PDF and Excel files.
"""

import os
import sys
import django
from decimal import Decimal
from io import BytesIO
import re

backend_path = '/Users/drushahardiksottany/Developer/projects/license-manager/backend'
sys.path.insert(0, backend_path)
os.chdir(backend_path)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'lmanagement.settings')

django.setup()

from openpyxl import load_workbook
import PyPDF2
from apps.core.utils.decimal_utils import to_decimal


def extract_from_pdf(pdf_path):
    """Extract financial data from PDF text."""
    try:
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page in reader.pages:
                text += page.extract_text()

        # Look for key financial values
        # Pattern: "Total ... BILL" or "Purchase" or "Sale" or "Profit"
        purchase_pattern = r'(?:Purchase|PURCHASE|Debit|DEBIT).*?(\d+(?:[,\.]\d+)*)'
        sale_pattern = r'(?:Sale|SALE|Credit|CREDIT).*?(\d+(?:[,\.]\d+)*)'
        profit_pattern = r'(?:Profit|Loss|P/L|P\&L).*?(-?\d+(?:[,\.]\d+)*)'

        results = {
            'text_length': len(text),
            'page_count': len(reader.pages),
            'sample_text': text[:1000] if text else "No text extracted",
        }

        return results
    except Exception as e:
        return {'error': f'PDF extraction failed: {str(e)}'}


def extract_from_excel(excel_path, target_license_id):
    """Extract financial data from Excel file."""
    try:
        wb = load_workbook(excel_path)

        results = {
            'sheets': list(wb.sheetnames),
            'sheet_data': {},
        }

        # Check Financial Ledger sheet
        if 'Financial Ledger' in wb.sheetnames:
            ws = wb['Financial Ledger']
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 20:  # First 20 rows
                    rows.append(row)
            results['sheet_data']['Financial Ledger'] = {
                'first_rows': rows,
                'total_rows': ws.max_row,
            }

        # Check for summary rows
        for ws_name in ['Customs Ledger', 'Reconciliation']:
            if ws_name in wb.sheetnames:
                ws = wb[ws_name]
                results['sheet_data'][ws_name] = {
                    'total_rows': ws.max_row,
                    'total_cols': ws.max_column,
                }

        return results
    except Exception as e:
        return {'error': f'Excel extraction failed: {str(e)}'}


def main():
    """Extract from PDF and Excel files."""
    print(f"\n{'='*120}")
    print("DETAILED PDF AND EXCEL EXTRACTION")
    print(f"{'='*120}")

    # Golden Case 1
    print("\n" + "="*120)
    print("GOLDEN CASE 1: License 0310833996")
    print("="*120)

    pdf_file = '/tmp/license_2616_DFIA_parity.pdf'
    excel_file = '/tmp/license_2616_DFIA_parity.xlsx'

    if os.path.exists(pdf_file):
        print(f"\n1. PDF Extraction ({pdf_file}):")
        pdf_data = extract_from_pdf(pdf_file)
        if 'error' not in pdf_data:
            print(f"   Pages: {pdf_data.get('page_count')}")
            print(f"   Text length: {pdf_data.get('text_length')} chars")
            print(f"   Sample text (first 500 chars):")
            print(f"   {pdf_data.get('sample_text', 'No text')[:500]}")
        else:
            print(f"   ERROR: {pdf_data.get('error')}")
    else:
        print(f"   PDF file not found: {pdf_file}")

    if os.path.exists(excel_file):
        print(f"\n2. Excel Extraction ({excel_file}):")
        excel_data = extract_from_excel(excel_file, 2616)
        if 'error' not in excel_data:
            print(f"   Sheets: {', '.join(excel_data.get('sheets', []))}")

            # Print Financial Ledger info
            fl_data = excel_data.get('sheet_data', {}).get('Financial Ledger', {})
            if fl_data:
                print(f"   Financial Ledger rows: {fl_data.get('total_rows')}")
                print(f"   First 10 rows:")
                for i, row in enumerate(fl_data.get('first_rows', [])[:10]):
                    print(f"      {i}: {row}")
        else:
            print(f"   ERROR: {excel_data.get('error')}")
    else:
        print(f"   Excel file not found: {excel_file}")

    # Golden Case 2
    print(f"\n{'='*120}")
    print("GOLDEN CASE 2: License 0311039916 (Loss Case)")
    print(f"{'='*120}")

    pdf_file2 = '/tmp/license_2033_DFIA_parity.pdf'
    excel_file2 = '/tmp/license_2033_DFIA_parity.xlsx'

    if os.path.exists(pdf_file2):
        print(f"\n1. PDF Extraction ({pdf_file2}):")
        pdf_data2 = extract_from_pdf(pdf_file2)
        if 'error' not in pdf_data2:
            print(f"   Pages: {pdf_data2.get('page_count')}")
            print(f"   Text length: {pdf_data2.get('text_length')} chars")
        else:
            print(f"   ERROR: {pdf_data2.get('error')}")
    else:
        print(f"   PDF file not found: {pdf_file2}")

    if os.path.exists(excel_file2):
        print(f"\n2. Excel Extraction ({excel_file2}):")
        excel_data2 = extract_from_excel(excel_file2, 2033)
        if 'error' not in excel_data2:
            print(f"   Sheets: {', '.join(excel_data2.get('sheets', []))}")
        else:
            print(f"   ERROR: {excel_data2.get('error')}")
    else:
        print(f"   Excel file not found: {excel_file2}")

    print(f"\n{'='*120}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*120}")


if __name__ == '__main__':
    main()
