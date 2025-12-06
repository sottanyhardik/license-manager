# license/tasks.py
from celery import shared_task
from datetime import datetime, timedelta
import logging
import os

from license.models import LicenseImportItemsModel

logger = logging.getLogger(__name__)


@shared_task
def update_items():
    """Update balance values for license items (legacy task)"""
    current_date = datetime.now()
    date_90_days_ago = current_date - timedelta(days=90)
    items = LicenseImportItemsModel.objects.filter(
        license__license_expiry_date__gte=date_90_days_ago
    ).order_by('license__license_expiry_date', 'license__license_date')
    for item in items:
        from bill_of_entry.tasks import update_balance_values_task
        update_balance_values_task(item.id)


@shared_task(bind=True)
def update_all_license_balances(self):
    """
    High-priority task to update balance_cif, is_active, is_expired, and restrictions for all licenses.
    Triggered manually from Item Pivot Report for fast, accurate report generation.

    This task:
    1. Updates balance_cif for all licenses using LicenseBalanceCalculator
    2. Updates is_expired based on license_expiry_date
    3. Updates is_null based on balance < $500
    4. Updates is_active based on expiry (mark inactive if expired)
    5. Checks and updates restriction flags on import items

    Returns:
        dict with status, counts, and timing info
    """
    from django.utils import timezone
    from decimal import Decimal
    from license.models import LicenseDetailsModel, LicenseImportItemsModel
    from license.services.balance_calculator import LicenseBalanceCalculator

    logger.info(f"Starting update_all_license_balances task: task_id={self.request.id}")
    start_time = datetime.now()

    try:
        # Get all licenses
        licenses = LicenseDetailsModel.objects.all()
        total_licenses = licenses.count()

        logger.info(f"Processing {total_licenses} licenses")

        # Update task state
        self.update_state(
            state='PROGRESS',
            meta={'current': 0, 'total': total_licenses, 'status': 'Updating license balances...'}
        )

        updated_count = 0
        error_count = 0
        today = timezone.now().date()

        # Process licenses in batches
        batch_size = 50
        for i, license_obj in enumerate(licenses.iterator(chunk_size=batch_size)):
            try:
                # Calculate balance using centralized service
                balance = LicenseBalanceCalculator.calculate_balance(license_obj)

                # Determine flags
                is_expired = license_obj.license_expiry_date < today if license_obj.license_expiry_date else False
                is_null = balance < Decimal('500')
                is_active = not is_expired  # Mark inactive if expired

                # Update license fields
                LicenseDetailsModel.objects.filter(pk=license_obj.pk).update(
                    balance_cif=balance,
                    is_expired=is_expired,
                    is_null=is_null,
                    is_active=is_active
                )

                updated_count += 1

                # Update progress every batch
                if (i + 1) % batch_size == 0:
                    progress = int(((i + 1) / total_licenses) * 90)  # Reserve 10% for restrictions
                    self.update_state(
                        state='PROGRESS',
                        meta={
                            'current': i + 1,
                            'total': total_licenses,
                            'status': f'Updated {updated_count} licenses...'
                        }
                    )

            except Exception as e:
                error_count += 1
                logger.error(f"Error updating license {license_obj.license_number}: {str(e)}")

        # Update restriction flags on import items
        self.update_state(
            state='PROGRESS',
            meta={'current': 90, 'total': 100, 'status': 'Updating restriction flags...'}
        )

        restriction_count = 0
        import_items = LicenseImportItemsModel.objects.prefetch_related('items')

        for import_item in import_items.iterator(chunk_size=100):
            # Check if any linked ItemNameModel has restriction
            has_restriction = any(
                item.restriction_percentage > 0
                for item in import_item.items.all()
            )

            # Update is_restricted flag if needed
            if has_restriction and not import_item.is_restricted:
                import_item.is_restricted = True
                import_item.save(update_fields=['is_restricted'])
                restriction_count += 1
            elif not has_restriction and import_item.is_restricted:
                import_item.is_restricted = False
                import_item.save(update_fields=['is_restricted'])
                restriction_count += 1

        elapsed = (datetime.now() - start_time).total_seconds()

        result = {
            'status': 'success',
            'updated': updated_count,
            'errors': error_count,
            'restrictions_updated': restriction_count,
            'total_licenses': total_licenses,
            'elapsed_seconds': elapsed,
            'timestamp': datetime.now().isoformat()
        }

        logger.info(f"Update completed: {result}")
        return result

    except Exception as e:
        error_msg = f"Failed to update license balances: {str(e)}"
        logger.error(error_msg)
        return {
            'status': 'error',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }


@shared_task
def sync_all_licenses():
    """
    Daily task to sync all licenses: update balance_cif, flags, and import item balances.
    Runs at 12:00 AM IST every day via Celery Beat.
    """
    from django.core.management import call_command
    from io import StringIO
    import sys

    logger.info("Starting daily license sync task...")

    # Capture command output
    output = StringIO()
    try:
        # Run the sync_licenses management command
        call_command(
            'sync_licenses',
            batch_size=100,
            stdout=output,
            stderr=output
        )

        output_str = output.getvalue()
        logger.info(f"License sync completed successfully:\n{output_str}")
        return {
            'status': 'success',
            'output': output_str,
            'timestamp': datetime.now().isoformat()
        }

    except Exception as e:
        error_msg = f"License sync failed: {str(e)}"
        logger.error(error_msg)
        logger.error(f"Output: {output.getvalue()}")
        return {
            'status': 'error',
            'error': str(e),
            'output': output.getvalue(),
            'timestamp': datetime.now().isoformat()
        }


@shared_task(bind=True)
def generate_item_pivot_excel(self, days=30, sion_norm=None, company_ids=None,
                              exclude_company_ids=None, min_balance=200,
                              license_status='active'):
    """
    Generate item pivot report Excel file as a background task.

    Args:
        days: Number of days to look back
        sion_norm: Filter by SION norm (optional)
        company_ids: Comma-separated company IDs (optional)
        exclude_company_ids: Comma-separated company IDs to exclude (optional)
        min_balance: Minimum balance CIF (default: 200)
        license_status: Filter by status (default: 'active')

    Returns:
        dict with file_path and metadata
    """
    from django.conf import settings
    from license.views.item_pivot_report import ItemPivotReportView
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell import WriteOnlyCell

    logger.info(f"Starting item pivot Excel generation: task_id={self.request.id}")

    # Update task state to PROGRESS
    self.update_state(state='PROGRESS', meta={'current': 0, 'total': 100, 'status': 'Generating report data...'})

    try:
        # Generate report data (sion_norm is optional - if not provided, exports ALL norms)
        view = ItemPivotReportView()

        logger.info(f"Generating report with params: sion_norm={sion_norm}, days={days}, min_balance={min_balance}")

        report_data = view.generate_report(
            days=days,
            sion_norm=sion_norm,
            company_ids=company_ids,
            exclude_company_ids=exclude_company_ids,
            min_balance=min_balance,
            license_status=license_status
        )

        # Check if report_data contains an error
        if isinstance(report_data, dict) and 'error' in report_data:
            raise ValueError(report_data['error'])

        self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100, 'status': 'Creating Excel file...'})

        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"item_pivot_report_{timestamp}_{self.request.id}.xlsx"
        file_path = os.path.join(exports_dir, filename)

        # Use write_only mode for memory efficiency
        workbook = openpyxl.Workbook(write_only=True)

        licenses_by_norm_notification = report_data.get('licenses_by_norm_notification', {})

        # Check if there's data to export
        if not licenses_by_norm_notification:
            raise ValueError('No data found matching the filters. Try adjusting the parameters.')

        total_sheets = sum(len(notif_dict) for notif_dict in licenses_by_norm_notification.values())
        current_sheet = 0

        logger.info(f"Generating {total_sheets} sheets for task {self.request.id}")

        # Create a sheet for each norm-notification combination
        for norm_class in sorted(licenses_by_norm_notification.keys()):
            notifications_dict = licenses_by_norm_notification[norm_class]
            for notification, licenses_list in sorted(notifications_dict.items()):
                current_sheet += 1
                progress = 50 + int((current_sheet / total_sheets) * 40)
                self.update_state(
                    state='PROGRESS',
                    meta={
                        'current': progress,
                        'total': 100,
                        'status': f'Creating sheet {current_sheet}/{total_sheets}: {norm_class} - {notification}'
                    }
                )

                # Sanitize sheet name
                sheet_name = f"{norm_class}_{notification}"[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('[', '(').replace(']', ')')
                worksheet = workbook.create_sheet(title=sheet_name)

                # Title row
                title = f"Item Pivot Report - {norm_class} - {notification}"
                title_cell = WriteOnlyCell(worksheet, value=title)
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
                worksheet.append([title_cell] + [None] * 25)
                worksheet.append([])

                # Build headers - only include items that have data in this norm-notification
                base_headers = ['Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt', 'Exporter', 'Total CIF', 'Balance CIF']
                item_headers = []

                # Filter items to only those with data in this norm-notification
                items_with_data = []
                for item in report_data['items']:
                    item_name = item['name']
                    # Check if ANY license in this list has this item
                    has_data = any(
                        license_data['items'].get(item_name, {}).get('quantity', 0) > 0
                        for license_data in licenses_list
                    )
                    if has_data:
                        items_with_data.append(item)

                for item in items_with_data:
                    item_name = item['name']
                    has_restriction = item.get('has_restriction', False)
                    headers = [
                        f"{item_name} HSN Code",
                        f"{item_name} Product Description",
                        f"{item_name} Total QTY",
                        f"{item_name} Allotted QTY",
                        f"{item_name} Debited QTY",
                        f"{item_name} Balance QTY",
                    ]
                    if has_restriction:
                        headers.extend([f"{item_name} Restriction %", f"{item_name} Restriction Value"])
                    item_headers.extend(headers)

                all_headers = base_headers + item_headers

                # Write headers with styling
                header_row = []
                for header in all_headers:
                    cell = WriteOnlyCell(worksheet, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    header_row.append(cell)
                worksheet.append(header_row)

                # Write data rows
                for idx, license_data in enumerate(licenses_list, 1):
                    row_data = [
                        idx,
                        license_data['license_number'],
                        license_data['license_date'],
                        license_data['license_expiry_date'],
                        license_data['exporter'],
                        license_data['total_cif'],
                        license_data['balance_cif']
                    ]

                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        item_data = license_data['items'].get(item_name, {
                            'hs_code': '', 'description': '', 'quantity': 0,
                            'allotted_quantity': 0, 'debited_quantity': 0,
                            'available_quantity': 0, 'restriction': None, 'restriction_value': 0
                        })

                        row_data.extend([
                            item_data.get('hs_code', ''),
                            item_data.get('description', ''),
                            item_data.get('quantity', 0),
                            item_data.get('allotted_quantity', 0),
                            item_data.get('debited_quantity', 0),
                            item_data.get('available_quantity', 0)
                        ])

                        if has_restriction:
                            restriction_val = item_data.get('restriction')
                            row_data.append(restriction_val if restriction_val else '')
                            row_data.append(item_data.get('restriction_value', 0) if item_data.get('restriction_value') else '')

                    worksheet.append(row_data)

                # Add totals row
                totals_row = [WriteOnlyCell(worksheet, value='TOTAL')]
                totals_row[0].font = Font(bold=True)
                totals_row.extend([None, None, None, None])

                total_cif = sum(lic['total_cif'] for lic in licenses_list)
                balance_cif = sum(lic['balance_cif'] for lic in licenses_list)

                total_cif_cell = WriteOnlyCell(worksheet, value=total_cif)
                total_cif_cell.font = Font(bold=True)
                totals_row.append(total_cif_cell)

                balance_cif_cell = WriteOnlyCell(worksheet, value=balance_cif)
                balance_cif_cell.font = Font(bold=True)
                totals_row.append(balance_cif_cell)

                for item in items_with_data:
                    item_name = item['name']
                    has_restriction = item.get('has_restriction', False)

                    totals_row.extend([None, None])  # Skip HSN and Description

                    for qty_type in ['quantity', 'allotted_quantity', 'debited_quantity', 'available_quantity']:
                        total = sum(lic['items'].get(item_name, {}).get(qty_type, 0) for lic in licenses_list)
                        cell = WriteOnlyCell(worksheet, value=total)
                        cell.font = Font(bold=True)
                        totals_row.append(cell)

                    if has_restriction:
                        totals_row.append(None)
                        total_restriction = sum(lic['items'].get(item_name, {}).get('restriction_value', 0) for lic in licenses_list)
                        cell = WriteOnlyCell(worksheet, value=total_restriction)
                        cell.font = Font(bold=True)
                        totals_row.append(cell)

                worksheet.append(totals_row)

        # Save workbook
        self.update_state(state='PROGRESS', meta={'current': 95, 'total': 100, 'status': 'Saving file...'})
        workbook.save(file_path)
        workbook.close()

        # Get file size
        file_size = os.path.getsize(file_path)

        logger.info(f"Item pivot Excel generated successfully: {filename} ({file_size} bytes)")

        # Return file information
        return {
            'status': 'SUCCESS',
            'file_path': file_path,
            'filename': filename,
            'file_size': file_size,
            'download_url': f'/media/exports/{filename}',
            'generated_at': datetime.now().isoformat(),
            'task_id': self.request.id
        }

    except Exception as e:
        logger.error(f"Error generating item pivot Excel: {str(e)}", exc_info=True)
        # Update task state to FAILURE with error message
        self.update_state(
            state='FAILURE',
            meta={
                'error': str(e),
                'exc_type': type(e).__name__
            }
        )
        # Re-raise to mark task as failed
        raise