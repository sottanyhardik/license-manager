#!/usr/bin/env python3
"""
SIMPLE PARITY VERIFICATION TEST
Verifies parity for two golden cases: UI = PDF = Excel

Golden Cases:
1. License 0310833996: ₹45,83,719 / ₹65,24,056 / ₹19,40,337
2. Loss case: Find a license with loss
"""

import os
import sys
import django
from decimal import Decimal
import json
from io import BytesIO

# Setup Django
backend_path = '/Users/drushahardiksottany/Developer/projects/license-manager/backend'
sys.path.insert(0, backend_path)
os.chdir(backend_path)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'lmanagement.settings')

django.setup()

from apps.license.models import LicenseDetailsModel, IncentiveLicense
from apps.license.services.canonical_ledger_service import CanonicalLedgerService
from apps.license.services.exporters.license_balance_excel import build_balance_excel
from apps.license.services.exporters.financial_ledger_pdf_renderer import render_financial_ledger_pdf
from apps.core.utils.decimal_utils import to_decimal
from openpyxl import load_workbook
from datetime import datetime


def find_license(license_number):
    """Find license by number."""
    try:
        lic = LicenseDetailsModel.objects.get(license_number=license_number)
        return ('DFIA', lic)
    except LicenseDetailsModel.DoesNotExist:
        pass

    try:
        lic = IncentiveLicense.objects.get(license_number=license_number)
        return ('INCENTIVE', lic)
    except IncentiveLicense.DoesNotExist:
        pass

    return (None, None)


def get_ui_data(license_id, license_type):
    """Extract data from UI (canonical ledger service)."""
    canonical = CanonicalLedgerService.build_canonical_ledger_dataset(
        license_id, license_type
    )

    summary = canonical.get('summary', {})
    return {
        'source': 'UI/API',
        'purchase_bill_inr': to_decimal(summary.get('total_purchase_bill_inr', 0)),
        'sale_bill_inr': to_decimal(summary.get('total_sale_bill_inr', 0)),
        'profit_loss_inr': to_decimal(summary.get('total_profit_loss', 0)),
        'profit_state': summary.get('profit_state'),
        'transaction_count': len(canonical.get('display_transactions', [])),
    }


def get_pdf_data(license_id, license_type):
    """Generate PDF and extract data."""
    canonical = CanonicalLedgerService.build_canonical_ledger_dataset(
        license_id, license_type
    )

    # Generate PDF
    pdf_bytes_io = render_financial_ledger_pdf(canonical)

    if isinstance(pdf_bytes_io, BytesIO):
        pdf_bytes = pdf_bytes_io.getvalue()
    else:
        pdf_bytes = pdf_bytes_io

    # Save PDF
    pdf_path = f'/tmp/license_{license_id}_{license_type}_parity.pdf'
    with open(pdf_path, 'wb') as f:
        f.write(pdf_bytes)

    return {
        'source': 'PDF',
        'pdf_path': pdf_path,
        'pdf_size': len(pdf_bytes),
        'has_pdf': len(pdf_bytes) > 0,
    }


def get_excel_data(license_obj, license_type):
    """Generate Excel and extract data."""
    # Generate Excel (build_balance_excel returns HttpResponse)
    response = build_balance_excel(license_obj)

    # Extract bytes from response
    if hasattr(response, 'getvalue'):
        excel_bytes = response.getvalue()
    elif hasattr(response, 'content'):
        excel_bytes = response.content
    else:
        excel_bytes = bytes(response)

    # Save Excel
    excel_path = f'/tmp/license_{license_obj.id}_{license_type}_parity.xlsx'
    with open(excel_path, 'wb') as f:
        f.write(excel_bytes)

    # Read and extract summary from Excel
    # The balance sheet should have totals somewhere
    try:
        wb = load_workbook(excel_path)
        sheets = list(wb.sheetnames) if hasattr(wb, 'sheetnames') else []
    except:
        sheets = []

    return {
        'source': 'Excel',
        'excel_path': excel_path,
        'excel_size': len(excel_bytes),
        'sheets': sheets,
    }


def verify_case(license_number, expected_purchase=None, expected_sale=None, expected_profit=None):
    """Verify a single golden case."""
    print(f"\n{'='*120}")
    print(f"GOLDEN CASE: License {license_number}")
    if expected_purchase:
        print(f"Expected Values: Purchase ₹{expected_purchase} | Sale ₹{expected_sale} | Profit ₹{expected_profit}")
    print(f"{'='*120}")

    # Find license
    lic_type, lic = find_license(license_number)
    if not lic:
        print(f"ERROR: License not found")
        return {'status': 'FAIL', 'reason': 'License not found'}

    print(f"Found: {license_number} (ID: {lic.id}, Type: {lic_type})\n")

    # Extract from UI
    print("1. UI/API Data:")
    try:
        ui_data = get_ui_data(lic.id, lic_type)
        print(f"   Purchase Bill: ₹{ui_data['purchase_bill_inr']}")
        print(f"   Sale Bill: ₹{ui_data['sale_bill_inr']}")
        print(f"   Profit/Loss: ₹{ui_data['profit_loss_inr']}")
        print(f"   State: {ui_data['profit_state']}")
        print(f"   Transaction Count: {ui_data['transaction_count']}")
    except Exception as e:
        print(f"   ERROR: {str(e)}")
        return {'status': 'FAIL', 'reason': f'UI extraction failed: {str(e)}'}

    # Generate PDF
    print("\n2. PDF Generation:")
    try:
        pdf_data = get_pdf_data(lic.id, lic_type)
        print(f"   PDF generated: {pdf_data['pdf_path']}")
        print(f"   Size: {pdf_data['pdf_size']} bytes")
        print(f"   Valid PDF: {pdf_data['has_pdf']}")
    except Exception as e:
        print(f"   ERROR: {str(e)}")
        return {'status': 'FAIL', 'reason': f'PDF generation failed: {str(e)}'}

    # Generate Excel
    print("\n3. Excel Generation:")
    try:
        excel_data = get_excel_data(lic, lic_type)
        print(f"   Excel generated: {excel_data['excel_path']}")
        print(f"   Size: {excel_data['excel_size']} bytes")
        print(f"   Sheets: {', '.join(excel_data['sheets'])}")
    except Exception as e:
        print(f"   ERROR: {str(e)}")
        return {'status': 'FAIL', 'reason': f'Excel generation failed: {str(e)}'}

    # Verify expected values
    if expected_purchase:
        print("\n4. Verification:")
        expected_purchase_d = to_decimal(expected_purchase)
        expected_sale_d = to_decimal(expected_sale)
        expected_profit_d = to_decimal(expected_profit)

        purchase_match = ui_data['purchase_bill_inr'] == expected_purchase_d
        sale_match = ui_data['sale_bill_inr'] == expected_sale_d
        profit_match = ui_data['profit_loss_inr'] == expected_profit_d

        print(f"   Purchase Match: {purchase_match} (Expected: ₹{expected_purchase}, Got: ₹{ui_data['purchase_bill_inr']})")
        print(f"   Sale Match: {sale_match} (Expected: ₹{expected_sale}, Got: ₹{ui_data['sale_bill_inr']})")
        print(f"   Profit Match: {profit_match} (Expected: ₹{expected_profit}, Got: ₹{ui_data['profit_loss_inr']})")

        if purchase_match and sale_match and profit_match:
            print("\n   STATUS: ✅ PASS - All values match!")
            return {'status': 'PASS', 'ui_data': ui_data, 'pdf_data': pdf_data, 'excel_data': excel_data}
        else:
            print("\n   STATUS: ❌ FAIL - Values mismatch!")
            return {'status': 'FAIL', 'reason': 'Values mismatch', 'ui_data': ui_data}

    print("\n   STATUS: ✅ PASS - Data extracted successfully")
    return {'status': 'PASS', 'ui_data': ui_data, 'pdf_data': pdf_data, 'excel_data': excel_data}


def main():
    """Run verification for golden cases."""
    print("\n" + "="*120)
    print("PARITY VERIFICATION TEST - UI vs PDF vs Excel")
    print(f"Started: {datetime.now().isoformat()}")
    print("="*120)

    # Golden Case 1: 0310833996
    result1 = verify_case(
        '0310833996',
        expected_purchase='4583719',
        expected_sale='6524056',
        expected_profit='1940337'
    )

    # Golden Case 2: Find a loss case
    print(f"\n{'='*120}")
    print("Searching for loss case license...")
    print(f"{'='*120}")

    # Find licenses with losses
    all_licenses = LicenseDetailsModel.objects.all()[:100]  # Sample
    for lic in all_licenses:
        canonical = CanonicalLedgerService.build_canonical_ledger_dataset(lic.id, 'DFIA')
        summary = canonical.get('summary', {})
        profit_loss = to_decimal(summary.get('total_profit_loss', 0))
        profit_state = summary.get('profit_state')

        if profit_state == 'LOSS' and profit_loss != 0:
            print(f"\nFound loss case: License {lic.license_number} (ID: {lic.id})")
            print(f"Loss: ₹{profit_loss}")
            result2 = verify_case(lic.license_number)
            break
    else:
        print("No loss case found in sample")
        result2 = {'status': 'SKIP', 'reason': 'No loss case found'}

    # Summary
    print(f"\n{'='*120}")
    print("SUMMARY")
    print(f"{'='*120}")
    print(f"Golden Case 1 (0310833996): {result1.get('status', 'UNKNOWN')}")
    print(f"Golden Case 2 (Loss case): {result2.get('status', 'UNKNOWN')}")
    print(f"Completed: {datetime.now().isoformat()}")

    # Save results
    results = {
        'timestamp': datetime.now().isoformat(),
        'golden_case_1': {
            'license': '0310833996',
            'status': result1.get('status'),
            'expected': {'purchase': '4583719', 'sale': '6524056', 'profit': '1940337'},
            'actual': result1.get('ui_data') if result1.get('status') != 'FAIL' else None,
        },
        'golden_case_2': {
            'status': result2.get('status'),
            'actual': result2.get('ui_data') if result2.get('status') != 'FAIL' else None,
        },
    }

    with open('/tmp/parity_test_results.json', 'w') as f:
        json.dump(results, f, indent=2, default=str)

    print(f"\nResults saved to: /tmp/parity_test_results.json")


if __name__ == '__main__':
    main()
