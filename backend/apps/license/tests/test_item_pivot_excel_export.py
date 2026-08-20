"""Canonical XLSX contract tests for the item-pivot report.

The report has one source of truth: ``ItemPivotService.build``. The JSON
endpoint returns its ``groups`` matrix and the XLSX exporter serializes that
same matrix; these tests protect that parity without using the retired
per-norm report layout.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.core.models import (
    CompanyModel,
    HSCodeModel,
    HeadSIONNormsModel,
    ItemNameModel,
    NotificationNumber,
    PurchaseStatus,
    SchemeCode,
    SionNormClassModel,
)
from apps.license.models import (
    LicenseDetailsModel,
    LicenseExportItemModel,
    LicenseImportItemsModel,
    LicenseItemPlan,
)


User = get_user_model()


@pytest.fixture
def superuser_client(db):
    user = User.objects.create_superuser(
        username="item-pivot-excel-tester",
        email="item-pivot-excel-tester@example.com",
        password="P@ssw0rd12345",
    )
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {RefreshToken.for_user(user).access_token}")
    return client


@pytest.fixture
def pivot_masters(db):
    head_norm = HeadSIONNormsModel.objects.create(name="Item Pivot Test Head Norm")
    norm_class = SionNormClassModel.objects.create(
        head_norm=head_norm, norm_class="PIVOTTEST", is_active=True,
    )
    item_name = ItemNameModel.objects.create(
        name="PIVOT TEST ITEM - PIVOTTEST",
        sion_norm_class=norm_class,
        is_active=True,
        display_order=1,
    )
    return {
        "exporter": CompanyModel.objects.create(iec="9990001111", name="Pivot Excel Exporter"),
        "notification": NotificationNumber.objects.create(code="PIVN1", label="Pivot Notification"),
        "scheme": SchemeCode.objects.create(code="PIVDFIA", label="Pivot DFIA"),
        "purchase_status": PurchaseStatus.objects.create(code="GE", label="GE Purchase"),
        "hs_code": HSCodeModel.objects.create(hs_code="99999999", product_description="Pivot Test Product"),
        "norm_class": norm_class,
        "item_name": item_name,
    }


@pytest.fixture
def pivot_license(db, pivot_masters):
    license_obj = LicenseDetailsModel.objects.create(
        license_number="PIVOT-EXCEL-001",
        license_date=date.today() - timedelta(days=30),
        license_expiry_date=date.today() + timedelta(days=30),
        exporter=pivot_masters["exporter"],
        notification_number=pivot_masters["notification"],
        scheme_code=pivot_masters["scheme"],
        purchase_status=pivot_masters["purchase_status"],
        file_number="PIVOT-FILE-001",
    )
    LicenseExportItemModel.objects.create(
        license=license_obj,
        description="Pivot export item",
        norm_class=pivot_masters["norm_class"],
        cif_fc=Decimal("777.00"),
        cif_inr=Decimal("65000.00"),
    )
    import_item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=1,
        description="Pivot import item",
        hs_code=pivot_masters["hs_code"],
        quantity=Decimal("100.000"),
        allotted_quantity=Decimal("10.000"),
        debited_quantity=Decimal("5.000"),
        available_quantity=Decimal("85.000"),
        debited_value=Decimal("50.00"),
        cif_fc=Decimal("1000.00"),
    )
    import_item.items.add(pivot_masters["item_name"])
    LicenseItemPlan.objects.create(
        license=license_obj,
        import_item=import_item,
        item_name=pivot_masters["item_name"],
        planned_quantity=Decimal("40.000"),
        planned_cif_fc=Decimal("400.00"),
    )
    return license_obj


def _report_params(**overrides):
    params = {"min_balance": 200, "license_status": "active"}
    params.update(overrides)
    return params


def _download_excel(client, **overrides):
    response = client.get(reverse("license:item-pivot-report"), _report_params(format="excel", **overrides))
    assert response.status_code == 200, getattr(response, "data", response.content[:300])
    return b"".join(response.streaming_content) if response.streaming else response.content


def _report_json(client, **overrides):
    response = client.get(reverse("license:item-pivot-report"), _report_params(**overrides))
    assert response.status_code == 200, getattr(response, "data", response.content[:300])
    return response.json()


def _number(value):
    return Decimal(str(value))


def _assert_excel_value_matches_json(actual, expected):
    if expected is None:
        assert actual is None
    else:
        assert _number(actual) == _number(expected)


def _fixture_group(report):
    return next(group for group in report["groups"] if group["notification_number"] == "PIVN1")


@pytest.mark.django_db
def test_item_pivot_excel_is_a_valid_workbook_with_canonical_group_sheet(superuser_client, pivot_license):
    report = _report_json(superuser_client)
    assert report["report_version"] == "canonical-item-pivot-v1"
    group = _fixture_group(report)

    workbook = load_workbook(BytesIO(_download_excel(superuser_client)), data_only=True)
    report_sheets = [sheet for sheet in workbook.worksheets if sheet.title != "TOTAL_Summary"]
    assert len(report_sheets) == 1, workbook.sheetnames
    sheet = report_sheets[0]

    metadata = [cell.value for cell in sheet[1]]
    while metadata and metadata[-1] is None:
        metadata.pop()
    assert metadata == [
        "Notification Number: PIVN1", "GE Purchase", "1 Licences",
    ]
    header = [cell.value for cell in sheet[2]]
    assert header[:9] == [
        "SR NO", "DFIA NO", "EXPIRY DT", "EXPORTER", "TOTAL CIF",
        "DEBITED CIF", "ALLOTTED CIF", "PLANNED CIF", "BALANCE CIF",
    ]
    assert len(header) == 19
    assert header[9:] == [
        "HSN CODE", "DESCRIPTION", "TOTAL QTY", "ALLOTTED QTY", "DEBITED QTY",
        "BALANCE QTY", "RESTRICTION %", "RESTRICTION VAL", "PLAN QTY", "PLANNED CIF",
    ]
    assert group["license_count"] == 1


@pytest.mark.django_db
def test_item_pivot_excel_data_and_totals_match_the_canonical_json_matrix(superuser_client, pivot_license):
    group = _fixture_group(_report_json(superuser_client))
    license_row = next(row for row in group["licenses"] if row["license_number"] == pivot_license.license_number)
    item = next(iter(license_row["items"].values()))
    item_total = next(iter(group["totals"]["items"].values()))

    workbook = load_workbook(BytesIO(_download_excel(superuser_client)), data_only=True)
    sheet = next(sheet for sheet in workbook.worksheets if sheet.title != "TOTAL_Summary")
    data_row = [cell.value for cell in sheet[3]]
    total_row = next(
        [cell.value for cell in row]
        for row in sheet.iter_rows()
        if str(row[0].value).startswith("TOTAL —")
    )

    assert data_row[1] == license_row["license_number"]
    for index, field in ((4, "total_cif"), (5, "debited_cif"), (6, "allotted_cif"), (7, "planned_cif"), (8, "balance_cif")):
        assert _number(data_row[index]) == _number(license_row[field])
        assert _number(total_row[index]) == _number(group["totals"][field])

    assert data_row[9] == item["hsn_code"]
    assert data_row[10] == item["description"]
    for index, field in ((11, "total_qty"), (12, "allotted_qty"), (13, "debited_qty"), (14, "balance_qty"), (17, "plan_qty"), (18, "planned_cif")):
        assert _number(data_row[index]) == _number(item[field])
        assert _number(total_row[index]) == _number(item_total[field])
    assert data_row[15] is None
    assert data_row[16] is None
    assert len(total_row) == len([cell.value for cell in sheet[2]])


@pytest.mark.django_db
def test_item_pivot_excel_total_summary_matches_grand_total_projection(superuser_client, pivot_license):
    report = _report_json(superuser_client)
    workbook = load_workbook(BytesIO(_download_excel(superuser_client)), data_only=True)
    sheet = workbook["TOTAL_Summary"]

    assert [cell.value for cell in sheet[1]] == [
        "TOTAL SUMMARY — ALL NOTIFICATIONS", "SION", "LICENCES", "TOTAL QTY",
        "BOE QTY", "ALLOTTED QTY", "AVAILABLE QTY", "PLANNED QTY", "BALANCE QTY",
        "AVAILABLE CIF", "PLANNED CIF", "BALANCE CIF",
    ]
    grand_total = report["grand_total"]
    grand_total_row = [cell.value for cell in sheet[sheet.max_row]]
    assert grand_total_row[0] == "GRAND TOTAL"
    assert grand_total_row[2] == grand_total["license_count"]
    totals = grand_total["item_summary_totals"]
    for index, field in ((3, "total_qty"), (4, "boe_used_qty"), (5, "allotted_qty"), (6, "available_qty"), (7, "planned_qty"), (8, "balance_qty"), (9, "available_cif"), (10, "planned_cif"), (11, "balance_cif")):
        _assert_excel_value_matches_json(grand_total_row[index], totals[field])


@pytest.mark.django_db
def test_item_pivot_excel_company_filter_matches_json_matrix(superuser_client, pivot_license):
    report = _report_json(superuser_client, company_ids=str(pivot_license.exporter_id))
    workbook = load_workbook(
        BytesIO(_download_excel(superuser_client, company_ids=str(pivot_license.exporter_id))),
        data_only=True,
    )

    group = _fixture_group(report)
    sheet = next(sheet for sheet in workbook.worksheets if sheet.title != "TOTAL_Summary")
    exported_numbers = set()
    for row in sheet.iter_rows(min_row=3):
        if str(row[0].value).startswith("TOTAL —"):
            break
        if row[1].value:
            exported_numbers.add(row[1].value)
    assert exported_numbers == {row["license_number"] for row in group["licenses"]}
