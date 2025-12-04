"""
Excel workbook building utilities.
"""

from dataclasses import dataclass
from decimal import Decimal
from typing import List, Optional, Any, Dict

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class WorksheetConfig:
    """Configuration for worksheet formatting."""
    header_font: Font = None
    header_fill: PatternFill = None
    header_alignment: Alignment = None
    cell_alignment: Alignment = None
    border: Border = None

    def __post_init__(self):
        """Set default styles if not provided."""
        if self.header_font is None:
            self.header_font = Font(bold=True, color="FFFFFF")

        if self.header_fill is None:
            self.header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        if self.header_alignment is None:
            self.header_alignment = Alignment(horizontal="center", vertical="center")

        if self.cell_alignment is None:
            self.cell_alignment = Alignment(horizontal="left", vertical="center")

        if self.border is None:
            thin_border = Side(style='thin', color='000000')
            self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)


class ExcelWorkbookBuilder:
    """
    Helper class for building Excel workbooks with common patterns.
    """

    def __init__(self, worksheet: Worksheet, config: Optional[WorksheetConfig] = None):
        """
        Initialize workbook builder.
        
        Args:
            worksheet: Worksheet to build on
            config: Worksheet configuration
        """
        self.worksheet = worksheet
        self.config = config or WorksheetConfig()
        self.current_row = 1

    def add_title(self, title: str, row: Optional[int] = None) -> 'ExcelWorkbookBuilder':
        """
        Add title row.
        
        Args:
            title: Title text
            row: Row number (uses current_row if None)
            
        Returns:
            Self for chaining
        """
        row = row or self.current_row

        cell = self.worksheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        self.current_row = row + 1
        return self

    def add_header_row(
            self,
            headers: List[str],
            row: Optional[int] = None,
            start_col: int = 1
    ) -> 'ExcelWorkbookBuilder':
        """
        Add header row with styling.
        
        Args:
            headers: List of header strings
            row: Row number (uses current_row if None)
            start_col: Starting column
            
        Returns:
            Self for chaining
        """
        row = row or self.current_row

        for idx, header in enumerate(headers):
            col = start_col + idx
            cell = self.worksheet.cell(row=row, column=col, value=header)
            cell.font = self.config.header_font
            cell.fill = self.config.header_fill
            cell.alignment = self.config.header_alignment
            cell.border = self.config.border

        self.current_row = row + 1
        return self

    def add_data_row(
            self,
            values: List[Any],
            row: Optional[int] = None,
            start_col: int = 1,
            number_columns: Optional[List[int]] = None
    ) -> 'ExcelWorkbookBuilder':
        """
        Add data row.
        
        Args:
            values: List of cell values
            row: Row number (uses current_row if None)
            start_col: Starting column
            number_columns: Indices of columns that contain numbers (0-indexed)
            
        Returns:
            Self for chaining
        """
        row = row or self.current_row
        number_columns = number_columns or []

        for idx, value in enumerate(values):
            col = start_col + idx
            cell = self.worksheet.cell(row=row, column=col, value=value)

            # Apply number formatting if column is numeric
            if idx in number_columns:
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = self.config.cell_alignment

            cell.border = self.config.border

        self.current_row = row + 1
        return self

    def add_data_rows(
            self,
            data: List[List[Any]],
            start_row: Optional[int] = None,
            start_col: int = 1,
            number_columns: Optional[List[int]] = None
    ) -> 'ExcelWorkbookBuilder':
        """
        Add multiple data rows.
        
        Args:
            data: List of rows
            start_row: Starting row (uses current_row if None)
            start_col: Starting column
            number_columns: Indices of columns that contain numbers
            
        Returns:
            Self for chaining
        """
        start_row = start_row or self.current_row

        for idx, row_data in enumerate(data):
            self.add_data_row(row_data, start_row + idx, start_col, number_columns)

        return self

    def add_total_row(
            self,
            values: List[Any],
            row: Optional[int] = None,
            start_col: int = 1,
            number_columns: Optional[List[int]] = None
    ) -> 'ExcelWorkbookBuilder':
        """
        Add total/summary row with bold styling.
        
        Args:
            values: List of cell values
            row: Row number (uses current_row if None)
            start_col: Starting column
            number_columns: Indices of columns that contain numbers
            
        Returns:
            Self for chaining
        """
        row = row or self.current_row
        number_columns = number_columns or []

        for idx, value in enumerate(values):
            col = start_col + idx
            cell = self.worksheet.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True)

            # Apply number formatting if column is numeric
            if idx in number_columns and isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = self.config.cell_alignment

            cell.border = self.config.border
            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

        self.current_row = row + 1
        return self

    def add_blank_row(self, count: int = 1) -> 'ExcelWorkbookBuilder':
        """
        Add blank row(s).
        
        Args:
            count: Number of blank rows to add
            
        Returns:
            Self for chaining
        """
        self.current_row += count
        return self

    def merge_cells(self, start_row: int, start_col: int, end_row: int, end_col: int) -> 'ExcelWorkbookBuilder':
        """
        Merge cells in range.
        
        Args:
            start_row: Start row
            start_col: Start column
            end_row: End row
            end_col: End column
            
        Returns:
            Self for chaining
        """
        self.worksheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col
        )
        return self

    def set_column_width(self, column: int, width: float) -> 'ExcelWorkbookBuilder':
        """
        Set column width.
        
        Args:
            column: Column number (1-indexed)
            width: Width in characters
            
        Returns:
            Self for chaining
        """
        column_letter = self.worksheet.cell(row=1, column=column).column_letter
        self.worksheet.column_dimensions[column_letter].width = width
        return self

    def set_column_widths(self, widths: List[float]) -> 'ExcelWorkbookBuilder':
        """
        Set multiple column widths.
        
        Args:
            widths: List of widths (one per column)
            
        Returns:
            Self for chaining
        """
        for idx, width in enumerate(widths, start=1):
            self.set_column_width(idx, width)
        return self

    def apply_alternating_row_colors(
            self,
            start_row: int,
            end_row: int,
            start_col: int = 1,
            end_col: Optional[int] = None,
            color1: str = "FFFFFF",
            color2: str = "F8F9FA"
    ) -> 'ExcelWorkbookBuilder':
        """
        Apply alternating row colors.
        
        Args:
            start_row: Starting row
            end_row: Ending row
            start_col: Starting column
            end_col: Ending column (uses max column if None)
            color1: First color (hex)
            color2: Second color (hex)
            
        Returns:
            Self for chaining
        """
        if end_col is None:
            end_col = self.worksheet.max_column

        for row in range(start_row, end_row + 1):
            fill_color = color1 if (row - start_row) % 2 == 0 else color2
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col in range(start_col, end_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.fill = fill

        return self

    def add_info_section(
            self,
            info_data: Dict[str, Any],
            start_row: Optional[int] = None,
            label_col: int = 1,
            value_col: int = 2
    ) -> 'ExcelWorkbookBuilder':
        """
        Add information section (key-value pairs).
        
        Args:
            info_data: Dictionary of label-value pairs
            start_row: Starting row (uses current_row if None)
            label_col: Column for labels
            value_col: Column for values
            
        Returns:
            Self for chaining
        """
        start_row = start_row or self.current_row

        for idx, (label, value) in enumerate(info_data.items()):
            row = start_row + idx

            # Label cell
            label_cell = self.worksheet.cell(row=row, column=label_col, value=f"{label}:")
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            label_cell.border = self.config.border
            label_cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")

            # Value cell
            value_cell = self.worksheet.cell(row=row, column=value_col, value=value)
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.border = self.config.border

        self.current_row = start_row + len(info_data)
        return self


def create_simple_table(
        worksheet: Worksheet,
        headers: List[str],
        data: List[List[Any]],
        start_row: int = 1,
        number_columns: Optional[List[int]] = None
) -> None:
    """
    Create a simple table with headers and data.
    
    Args:
        worksheet: Worksheet to write to
        headers: Header row
        data: Data rows
        start_row: Starting row
        number_columns: Indices of columns that contain numbers
    """
    builder = ExcelWorkbookBuilder(worksheet)
    builder.current_row = start_row
    builder.add_header_row(headers)
    builder.add_data_rows(data, number_columns=number_columns)
