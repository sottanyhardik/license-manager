"""
Regression coverage for BL-LEDGER-02's stale-balance reader in the
deprecated-but-still-reachable `build_balance_excel_unused` (`apps/license/
services/exporters/license_balance_excel.py`): both "Balance CIF" cells
(the license-header row and each export-item row) used to read
`license_obj.balance_cif` (the cached column) directly; both now read the
LIVE `LicenseBalanceCalculator.calculate_financial_balance()` figure once
and reuse it, matching the active `build_balance_excel` exporter's
already-live `get_balance_cif`.
"""
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.test import TestCase

from apps.license.models import LicenseExportItemModel
from apps.license.models.core import LicenseBalance
from apps.license.services.balance_calculator import LicenseBalanceCalculator
from apps.license.services.exporters.license_balance_excel import build_balance_excel_unused
from apps.license.tests.test_balance_ledger_views import LicenseBalanceLedgerFixtureMixin


class BuildBalanceExcelUnusedLiveBalanceTests(LicenseBalanceLedgerFixtureMixin, TestCase):
    def test_balance_cif_cells_use_live_balance_not_stale_cache(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("5000.00"))
        item = self.make_item(license_obj, 1)
        boe = self.make_boe(company)
        self.make_debit_row(boe, item, cif_fc=Decimal("1000.00"))

        live_balance = LicenseBalanceCalculator.calculate_financial_balance(license_obj)
        self.assertEqual(live_balance, Decimal("4000.00"))

        # Deliberately desynchronize the cache from the live value.
        LicenseBalance.objects.filter(license=license_obj).update(balance_cif=Decimal("0.00"))
        license_obj.refresh_from_db()
        self.assertEqual(license_obj.balance.balance_cif, Decimal("0.00"))

        response = build_balance_excel_unused(license_obj)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        live_str = f"{float(live_balance):.2f}"
        stale_str = f"{0.0:.2f}"

        # Layout (fixed for this "kept for reference" deprecated function):
        # row1 title; row3 header-row1; row4 values-row1; row6 header-row2
        # ("Purchase Status"/"Balance CIF"/"Get Norm Class"/"Latest Transfer");
        # row7 values-row2 (Balance CIF at column B); row9 "Export Items"
        # section header; row10 export-item headers; row11 first export
        # item's data (Balance CIF at column C).
        header_area_balance = ws.cell(row=7, column=2).value
        export_item_balance = ws.cell(row=11, column=3).value

        self.assertEqual(header_area_balance, live_str, "License-header Balance CIF must show the LIVE balance")
        self.assertEqual(export_item_balance, live_str, "Export-item row Balance CIF must show the LIVE balance")
        self.assertNotEqual(header_area_balance, stale_str)
        self.assertNotEqual(export_item_balance, stale_str)
        # The import-item Bal CIF is supplied by available_value_bulk_map,
        # not the stale LicenseBalance cache or a per-item property call.
        self.assertEqual(ws.cell(row=15, column=10).value, live_str)
