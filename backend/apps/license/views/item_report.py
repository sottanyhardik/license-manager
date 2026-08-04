"""
Item Report - List all License Import Items with filters and inline editing support
"""

import logging

from django.db.models import Prefetch
from django.http import JsonResponse, HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from apps.accounts.permissions import ReportPermission
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.models import ItemNameModel
from apps.license.models import LicenseImportItemsModel


def _safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _merge_report_items_by_group(license_items):
    """
    Merge one licence's raw `report_items` rows (one per raw import item,
    ordered by serial_number — see `generate_report`) into one row per
    `group_key` (`plan_grouping.plan_group_key`, attached to each row by
    `generate_report`) — the same grouping `plan_utilization_rows()` uses
    for the licence-detail API and the balance Excel exports.

    First-seen (= lowest serial number) member's fields seed each merged
    row; additive fields (`available_quantity`, `planned_quantity`,
    `planned_cif`) are summed across the group so grand totals are
    unaffected by merging; `item_names`/`planned_splits` are deduped/unioned
    across every member (defensive, in case a legacy plan line ever landed
    on a non-representative member); `serial_number` becomes the
    comma-joined, ascending list of every merged serial (e.g. "3, 13, 23").
    `hs_code`/`product_description`/`condition_type` fall back to the first
    non-empty value across members when the representative's own is blank.

    Only used by `export_to_excel` — `generate_report`'s JSON `items` list
    stays one row per raw import item (the Item Report table edits item
    names per raw row, so it needs the raw granularity).
    """
    groups: dict = {}
    order: list = []

    for item in license_items:
        key = item.get('group_key') or f"ID:{item.get('id')}"
        group = groups.get(key)
        if group is None:
            group = {
                **item,
                'available_quantity': 0.0,
                'planned_quantity': 0.0,
                'planned_cif': 0.0,
                'item_names': [],
                'planned_splits': [],
                '_serials': [],
                '_item_name_ids': set(),
            }
            groups[key] = group
            order.append(key)

        group['_serials'].append(item.get('serial_number'))
        group['available_quantity'] += float(item.get('available_quantity') or 0)
        group['planned_quantity'] += float(item.get('planned_quantity') or 0)
        group['planned_cif'] += float(item.get('planned_cif') or 0)
        if not group.get('hs_code') and item.get('hs_code'):
            group['hs_code'] = item['hs_code']
        if not group.get('product_description') and item.get('product_description'):
            group['product_description'] = item['product_description']
        if not group.get('condition_type') and item.get('condition_type'):
            group['condition_type'] = item['condition_type']
        for name in item.get('item_names') or []:
            if name['id'] not in group['_item_name_ids']:
                group['_item_name_ids'].add(name['id'])
                group['item_names'].append(name)
        group['planned_splits'].extend(item.get('planned_splits') or [])

    merged = []
    for key in order:
        group = groups[key]
        serials = sorted(s for s in group.pop('_serials') if s is not None)
        group.pop('_item_name_ids')
        group['serial_number'] = ', '.join(str(s) for s in serials)
        merged.append(group)
    return merged


class _ExcelPassthroughRenderer(BaseRenderer):
    """
    Dummy renderer that tells DRF 'excel' is an accepted format so that
    ?format=excel (or ?_format=excel) does not fail content negotiation.
    The view returns a plain Django HttpResponse for Excel which DRF
    passes through without calling this renderer at all.
    """
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'excel'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data  # never reached — view returns HttpResponse directly


logger = logging.getLogger(__name__)


class ItemReportView(APIView):
    """
    Report showing all License Import Items with filters.

    GET parameters:
        - item_names: Comma-separated item name IDs for filtering (multiselect)
        - format / _format: 'json' or 'excel' (default: json)
    """
    permission_classes = [ReportPermission]
    # Register the excel "format" so DRF content negotiation accepts
    # ?format=excel (or ?_format=excel) without raising NotAcceptable (406).
    renderer_classes = [JSONRenderer, _ExcelPassthroughRenderer]

    def get(self, request, *args, **kwargs):
        # DRF intercepts ?format= as a content-negotiation override and raises
        # NotAcceptable for unknown formats like 'excel'.  The frontend sends
        # ?_format=excel (underscore prefix) to bypass that interception.
        # Support both forms for backward compatibility.
        output_format = (
            request.GET.get('_format')
            or request.GET.get('format', 'json')
        ).lower()
        item_names = request.GET.get('item_names')  # Comma-separated item name IDs
        company_ids = request.GET.get('company_ids')  # Comma-separated company IDs
        exclude_company_ids = request.GET.get('exclude_company_ids')  # Comma-separated company IDs to exclude
        min_balance = _safe_int(request.GET.get('min_balance'), 200)
        min_avail_qty = float(request.GET.get('min_avail_qty', 0))
        license_status = request.GET.get('license_status', 'active')
        is_restricted = request.GET.get('is_restricted')  # 'true', 'false', or None for all
        purchase_status = request.GET.get('purchase_status')  # Comma-separated purchase status codes
        product_description = request.GET.get('product_description')  # Product description search
        hsn_code = request.GET.get('hsn_code')  # HSN code search
        norms = request.GET.get('norms')  # Comma-separated SION norm classes
        notification_numbers = request.GET.get('notification_numbers')  # Comma-separated notification numbers
        expiry_date_from = request.GET.get('expiry_date_from')  # YYYY-MM-DD
        expiry_date_to = request.GET.get('expiry_date_to')      # YYYY-MM-DD

        if output_format == 'excel':
            try:
                return self.export_to_excel(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status, product_description, hsn_code, norms, notification_numbers, expiry_date_from, expiry_date_to)
            except Exception as e:
                logger.exception("Error exporting item report to Excel")
                return JsonResponse({'error': str(e)}, status=500)

        # For JSON, generate full report
        try:
            report_data = self.generate_report(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status, product_description, hsn_code, norms, notification_numbers, expiry_date_from, expiry_date_to)
        except Exception as e:
            logger.exception("Error generating item report")
            return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(report_data, safe=False)

    def generate_report(self, item_names=None, company_ids=None, exclude_company_ids=None, min_balance=200, min_avail_qty=0, license_status='active', is_restricted=None, purchase_status=None, product_description=None, hsn_code=None, norms=None, notification_numbers=None, expiry_date_from=None, expiry_date_to=None):
        """
        Generate item report with all license import items.

        Args:
            item_names: Comma-separated item name IDs for filtering
            company_ids: Comma-separated company IDs to include (optional)
            exclude_company_ids: Comma-separated company IDs to exclude (optional)
            min_balance: Minimum balance CIF to include (default 200)
            min_avail_qty: Minimum available quantity to include (default 0)
            license_status: Filter by status - 'active', 'expired', 'expiring_soon', 'all' (default 'active')
            is_restricted: Filter by restriction status - 'true', 'false', or None for all
            purchase_status: Comma-separated purchase status codes (e.g., 'GE,MI,SM')
            product_description: Search text for product description (case-insensitive contains search)
            hsn_code: Search text for HSN code (case-insensitive contains search)
            norms: Comma-separated SION norm classes (e.g., '019/2015,098/2009')
            notification_numbers: Comma-separated license notification numbers (e.g., '019/2015,098/2009')

        Returns:
            Dictionary with report data
        """
        from datetime import date, timedelta
        today = date.today()

        # Base query - all import items with licenses
        from apps.license.models import LicenseTransferModel

        # Prefetch only the latest transfer for each license
        latest_transfer_prefetch = Prefetch(
            'license__transfers',
            queryset=LicenseTransferModel.objects.select_related('from_company', 'to_company').order_by('-transfer_date', '-transfer_initiation_date'),
            to_attr='latest_transfers'
        )

        items = LicenseImportItemsModel.objects.select_related(
            'license',
            'license__exporter',
            'license__ownership__current_owner',
            'license__balance',        # fix N+1: balance_cif
            'license__notes',          # fix N+1: balance_report_notes, condition_sheet
            'license__notification_number',
            'license__purchase_status',
            'hs_code'
        ).prefetch_related('items', latest_transfer_prefetch)

        # Apply license status filter
        if license_status == 'active':
            items = items.filter(
                license__flags__is_active=True,
                license__license_expiry_date__gt=today - timedelta(days=30)
            )
        elif license_status == 'expired':
            items = items.filter(license__license_expiry_date__lt=today)
        elif license_status == 'expiring_soon':
            items = items.filter(
                license__flags__is_active=True,
                license__license_expiry_date__gte=today,
                license__license_expiry_date__lte=today + timedelta(days=30)
            )
        # If 'all', no date or is_active filter applied

        # Apply explicit expiry date range filter
        if expiry_date_from:
            from datetime import datetime as _dt
            items = items.filter(license__license_expiry_date__gte=_dt.strptime(expiry_date_from, '%Y-%m-%d').date())
        if expiry_date_to:
            from datetime import datetime as _dt
            items = items.filter(license__license_expiry_date__lte=_dt.strptime(expiry_date_to, '%Y-%m-%d').date())

        # Filter by min_balance using stored available_value field (can be done in query)
        # This pre-filters before iteration for better performance
        items = items.filter(available_value__gte=min_balance)

        # Filter by min_avail_qty (can be done in query)
        if min_avail_qty > 0:
            items = items.filter(available_quantity__gte=min_avail_qty)

        # Filter by company IDs if specified
        if company_ids:
            company_id_list = [int(cid.strip()) for cid in company_ids.split(',') if cid.strip()]
            items = items.filter(license__exporter_id__in=company_id_list)

        # Exclude company IDs if specified
        if exclude_company_ids:
            exclude_id_list = [int(cid.strip()) for cid in exclude_company_ids.split(',') if cid.strip()]
            items = items.exclude(license__exporter_id__in=exclude_id_list)

        # Filter by item names if specified
        if item_names:
            item_name_ids = [int(id.strip()) for id in item_names.split(',') if id.strip()]
            items = items.filter(items__id__in=item_name_ids).distinct()

        # Filter by is_restricted if specified
        if is_restricted is not None:
            if is_restricted == 'true':
                items = items.filter(is_restricted=True)
            elif is_restricted == 'false':
                items = items.filter(is_restricted=False)

        # Filter by purchase_status if specified
        if purchase_status:
            purchase_status_list = [ps.strip() for ps in purchase_status.split(',') if ps.strip()]
            items = items.filter(license__purchase_status__code__in=purchase_status_list)

        # Filter by product description if specified (case-insensitive contains search)
        if product_description:
            items = items.filter(description__icontains=product_description)

        # Filter by HSN code if specified (case-insensitive contains search)
        if hsn_code:
            items = items.filter(hs_code__hs_code__icontains=hsn_code)

        # Filter by norms (SION norm class) if specified
        if norms:
            norms_list = [n.strip() for n in norms.split(',') if n.strip()]
            items = items.filter(items__sion_norm_class__norm_class__in=norms_list).distinct()

        # Filter by notification numbers (license notification) if specified
        if notification_numbers:
            notification_list = [n.strip() for n in notification_numbers.split(',') if n.strip()]
            items = items.filter(license__notification_number__code__in=notification_list).distinct()

        # Sorted by license expiry date (soonest-expiring first) — the
        # business-report ordering requested for the Item Report — with
        # license number / serial number as stable tie-breakers so a
        # license's own item rows always stay contiguous for grouping.
        items = items.order_by('license__license_expiry_date', 'license__license_number', 'serial_number')

        # Materialise the queryset once so the plan pre-fetch can use the IDs
        # without issuing a second DB round-trip.
        item_list = list(items)

        # Utilization plan per item. Per LICENSE we use the manual plan if one
        # exists, otherwise the norm (E1/E5/E132) plan — never both.
        # Pre-compute for ALL unique licenses in one pass rather than calling
        # effective_plan_for_license() inside the loop (was O(N) DB round-trips).
        from apps.license.services.plan_reporting import plan_map_for_import_items
        from apps.license.services.norm_plan import effective_plan_for_license
        manual_splits = plan_map_for_import_items([it.id for it in item_list])

        # Build per-license effective-plan cache from the already-loaded licenses
        # (select_related already pulled them; no extra queries needed here).
        _eff_cache: dict = {}
        seen_license_ids: set = set()
        for it in item_list:
            lid = it.license_id
            if lid not in seen_license_ids:
                seen_license_ids.add(lid)
                _eff_cache[lid] = effective_plan_for_license(it.license)

        # Build report data
        from apps.license.services.plan_grouping import plan_group_key
        report_items = []
        for item in item_list:
            # Get item names
            item_names_list = [{"id": i.id, "name": i.name} for i in item.items.all()]
            _plan_source, _eff = _eff_cache[item.license_id]
            plan = _eff.get(item.id)
            _ms = manual_splits.get(item.id)

            # Use the stored available_value field (updated by balance update task)
            # This field already contains the correct value:
            # - For restricted items: restriction-based calculated value
            # - For non-restricted items: license balance_cif
            # Note: Make sure to run "Update Balance" in Item Pivot Report to refresh these values
            available_balance = float(item.available_value or 0)

            # Unit Price — the item's own originally-transacted per-unit CIF
            # (cif_fc / quantity). Falls back to the live available balance
            # per unit (available_balance / available_quantity) only when
            # cif_fc isn't usable (e.g. zero on older/incomplete records),
            # so the column is never blank when a sensible value exists.
            quantity = float(item.quantity or 0)
            cif_fc = float(item.cif_fc or 0)
            available_quantity = float(item.available_quantity or 0)
            if quantity > 0 and cif_fc > 0:
                unit_price = cif_fc / quantity
            elif available_quantity > 0:
                unit_price = available_balance / available_quantity
            else:
                unit_price = 0

            # Get latest transfer information
            latest_transfer_info = None
            if hasattr(item.license, 'latest_transfers') and item.license.latest_transfers:
                latest_transfer = item.license.latest_transfers[0]
                latest_transfer_info = str(latest_transfer)  # Uses the __str__ method which formats it nicely

            report_items.append({
                'id': item.id,
                'license_id': item.license.id,
                'license_number': item.license.license_number,
                'license_date': item.license.license_date.isoformat() if item.license.license_date else None,
                'license_expiry_date': item.license.license_expiry_date.isoformat() if item.license.license_expiry_date else None,
                'ledger_date': item.license.ledger_date.isoformat() if item.license.ledger_date else None,
                'exporter_name': item.license.exporter.name if item.license.exporter else None,
                'current_owner': item.license.current_owner.name if item.license.current_owner else None,
                'latest_transfer': latest_transfer_info,
                'hs_code': item.hs_code.hs_code if item.hs_code else None,
                'product_description': item.description or '',
                'item_names': item_names_list,
                'quantity': quantity,
                'available_quantity': available_quantity,
                'available_balance': available_balance,
                'unit_price': unit_price,
                'balance_cif': float(item.license.balance_cif or 0),
                'is_restricted': item.is_restricted,
                'condition_type': item.condition_type or '',
                'notes': item.license.balance_report_notes or '',
                'condition_sheet': item.license.condition_sheet or '',
                'unit': item.unit,
                'serial_number': item.serial_number,
                # Effective plan (manual if the license is manually planned,
                # else norm). Splits are shown only for manual plans.
                'planned_quantity': plan['planned_quantity'] if plan else 0,
                'planned_cif': plan['planned_cif'] if plan else 0,
                'plan_source': _plan_source,
                'planned_splits': (_ms['splits'] if (_plan_source == 'manual' and _ms) else []),
                # Additive, non-breaking: the SAME plan_group_key() the Plan
                # tab / plan_utilization_rows() use to merge S.No rows that
                # share a product. Only consumed by export_to_excel below
                # (to merge rows within a licence's block) — the JSON `items`
                # list itself stays one entry per raw import item, since the
                # Item Report table edits item names per raw row.
                'group_key': plan_group_key(item),
            })

        return {
            'report_date': date.today().isoformat(),
            'total_items': len(report_items),
            'items': report_items
        }

    def export_to_excel(self, item_names=None, company_ids=None, exclude_company_ids=None, min_balance=200, min_avail_qty=0, license_status='active', is_restricted=None, purchase_status=None, product_description=None, hsn_code=None, norms=None, notification_numbers=None, expiry_date_from=None, expiry_date_to=None):
        """Export item report to Excel — single sheet, grouped by license, one row per
        license item, laid out identically to the View Report table (same column
        order/labels, same license-level grouping, same totals)."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        from io import BytesIO
        from apps.license.utils.condition_excel import annotate_cell as _annotate_condition_cell

        # Strip control chars that openpyxl rejects (\x00-\x08, \x0b-\x0c, \x0e-\x1f).
        # PDF-sourced fields like condition_sheet sometimes carry these.
        def _safe(v):
            if isinstance(v, str):
                return ILLEGAL_CHARACTERS_RE.sub('', v)
            return v

        # Generate report data — already sorted by license expiry date (see
        # generate_report), so licenses appear in the same order as the View.
        report_data = self.generate_report(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status, product_description, hsn_code, norms, notification_numbers, expiry_date_from, expiry_date_to)
        items = report_data['items']

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        if not items:
            ws = wb.create_sheet(title="No Data")
            ws.cell(row=1, column=1, value="No items found matching the filter criteria")
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=item_report.xlsx'
            return response

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column order matches the View Report table exactly (View/Excel
        # parity): the 16 required business columns, followed by the
        # pre-existing extra columns the View also keeps (Balance CIF,
        # Is Restricted, Notes, Condition Sheet, Transfer Status).
        # "Condition" sits next to "Serial Number" so it's obvious which
        # licence item the AU / N% restriction applies to.
        headers = [
            'Sr No', 'License No', 'License Date', 'License Expiry Date', 'Ledger Date', 'Exporter Name',
            'Serial Number', 'Condition', 'HSN Code', 'Product Description', 'Item Name',
            'Available Quantity', 'Unit Price', 'Available Balance', 'Plan Qty', 'Plan CIF',
            'Balance CIF', 'Is Restricted', 'Notes', 'Condition Sheet', 'Transfer Status',
        ]
        COL = {name: i + 1 for i, name in enumerate(headers)}
        QTY_FMT = '#,##0.000'
        CIF_FMT = '#,##0.00'
        # License-level columns: written once per license (first item's row)
        # and Excel-merged down the whole group, exactly like the View's
        # rowSpan cells.
        GROUPED_COLS = [
            COL['Sr No'], COL['License No'], COL['License Date'], COL['License Expiry Date'],
            COL['Ledger Date'], COL['Exporter Name'], COL['Available Balance'], COL['Balance CIF'],
            COL['Is Restricted'], COL['Notes'], COL['Condition Sheet'], COL['Transfer Status'],
        ]

        ws = wb.create_sheet(title="Item Report")

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Column widths (openpyxl has no true content-measuring auto-fit —
        # these are sized to comfortably fit typical values per column,
        # the same approach the previous version of this export used).
        widths = {
            'Sr No': 8, 'License No': 18, 'License Date': 15, 'License Expiry Date': 18,
            'Ledger Date': 15, 'Exporter Name': 25, 'Serial Number': 12, 'Condition': 11,
            'HSN Code': 12, 'Product Description': 40, 'Item Name': 25, 'Available Quantity': 18,
            'Unit Price': 14, 'Available Balance': 18, 'Plan Qty': 14, 'Plan CIF': 16,
            'Balance CIF': 18, 'Is Restricted': 14, 'Notes': 30, 'Condition Sheet': 30,
            'Transfer Status': 35,
        }
        for name, col_num in COL.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = widths[name]

        # Group items by license, preserving the license expiry-date order
        # generate_report() already sorted them in.
        grouped_items = {}
        for item in items:
            license_id = item['license_id']
            if license_id not in grouped_items:
                grouped_items[license_id] = []
            grouped_items[license_id].append(item)

        # Merge each licence's raw rows (one per S.No) into one row per
        # planning-item group (the same `plan_group_key` grouping
        # `plan_utilization_rows()` uses) — e.g. 3 S.No rows that share a
        # description collapse into 1, with a comma-joined Serial Number
        # cell. Only the Excel export merges; `report_data['items']`
        # (the JSON API) stays one row per raw import item, since the
        # Item Report table edits item names per raw row.
        grouped_items = {
            license_id: _merge_report_items_by_group(rows)
            for license_id, rows in grouped_items.items()
        }

        # Define border style for merged cells
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Planning split sub-row rendering is shared with
        # license_balance_excel.py — see planning_split_rows.py.
        from apps.license.services.exporters.planning_split_rows import (
            rows_for_splits, write_split_sub_rows,
        )

        # Add data rows with merged cells for same license
        current_row = 2
        sr_no = 1

        for license_id, license_items in grouped_items.items():
            # Per-item visible (qty>0 or cif>0) planning splits, sourced
            # from the same plan_map_for_import_items() map
            # generate_report() already built (item['planned_splits'],
            # scoped to manual-plan items there) — no second
            # LicenseItemPlan query here.
            item_split_rows = [
                rows_for_splits(item.get('planned_splits') or [])
                for item in license_items
            ]
            # Row span now covers each item's own row PLUS its split
            # sub-rows, so the merged license-level columns still span
            # the whole license block correctly.
            row_span = len(license_items) + sum(len(s) for s in item_split_rows)
            start_row = current_row

            # Add each item in this license group
            for item_idx, item in enumerate(license_items):
                item_names_str = ', '.join([i['name'] for i in item['item_names']])

                # License-level columns (only for first row, will be merged)
                if item_idx == 0:
                    ws.cell(row=current_row, column=COL['Sr No'], value=sr_no)
                    ws.cell(row=current_row, column=COL['License No'], value=_safe(item['license_number']))
                    ws.cell(row=current_row, column=COL['License Date'], value=_safe(item['license_date']))
                    ws.cell(row=current_row, column=COL['License Expiry Date'], value=_safe(item['license_expiry_date']))
                    ws.cell(row=current_row, column=COL['Ledger Date'], value=_safe(item.get('ledger_date')))
                    ws.cell(row=current_row, column=COL['Exporter Name'], value=_safe(item['exporter_name']))
                    avail_bal_cell = ws.cell(row=current_row, column=COL['Available Balance'], value=item['available_balance'])
                    avail_bal_cell.number_format = CIF_FMT
                    bal_cif_cell = ws.cell(row=current_row, column=COL['Balance CIF'], value=item['balance_cif'])
                    bal_cif_cell.number_format = CIF_FMT
                    ws.cell(row=current_row, column=COL['Is Restricted'], value='Yes' if item['is_restricted'] else 'No')
                    ws.cell(row=current_row, column=COL['Notes'], value=_safe(item['notes']))
                    ws.cell(row=current_row, column=COL['Condition Sheet'], value=_safe(item['condition_sheet']))
                    ws.cell(row=current_row, column=COL['Transfer Status'], value=_safe(item.get('latest_transfer', '')))

                # Item-level columns (for each row)
                sn_cell = ws.cell(row=current_row, column=COL['Serial Number'], value=item['serial_number'])
                cond = (item.get('condition_type') or '')
                cond_cell = ws.cell(row=current_row, column=COL['Condition'], value=cond)
                # Tint both Serial-Number and Condition cells so the row
                # stands out at a glance.
                _annotate_condition_cell(sn_cell, cond)
                _annotate_condition_cell(cond_cell, cond)
                ws.cell(row=current_row, column=COL['HSN Code'], value=_safe(item['hs_code']))
                ws.cell(row=current_row, column=COL['Product Description'], value=_safe(item['product_description']))
                ws.cell(row=current_row, column=COL['Item Name'], value=_safe(item_names_str))
                qty_cell = ws.cell(row=current_row, column=COL['Available Quantity'], value=item['available_quantity'])
                qty_cell.number_format = QTY_FMT
                price_cell = ws.cell(row=current_row, column=COL['Unit Price'], value=item.get('unit_price') or 0)
                price_cell.number_format = CIF_FMT
                plan_qty_cell = ws.cell(row=current_row, column=COL['Plan Qty'], value=item.get('planned_quantity') or 0)
                plan_qty_cell.number_format = QTY_FMT
                plan_cif_cell = ws.cell(row=current_row, column=COL['Plan CIF'], value=item.get('planned_cif') or 0)
                plan_cif_cell.number_format = CIF_FMT

                current_row += 1

                # ── Planning split sub-rows ─────────────────────────────
                # One indented row per manual-plan split — Planning Item
                # Name / Unit Price / Planned Qty / Planned CIF / "Split N"
                # badge — matching license_balance_excel.py's per-item
                # split rows exactly (same source map, same filter).
                # item_name is sanitized like every other string field in
                # this export (control chars openpyxl rejects).
                _raw_splits = item.get('planned_splits') or []
                _sanitized_splits = [
                    {**s, 'item_name': _safe(s.get('item_name'))} for s in _raw_splits
                ]
                current_row += write_split_sub_rows(
                    ws, current_row, _sanitized_splits,
                    name_col=COL['Item Name'],
                    price_col=COL['Product Description'],  # reused for the "@ $X/unit" label
                    badge_col=COL['Condition'],
                    qty_col=COL['Plan Qty'],
                    cif_col=COL['Plan CIF'],
                    other_cols=(
                        COL['Serial Number'], COL['HSN Code'], COL['Available Quantity'],
                        COL['Unit Price'], COL['Available Balance'], COL['Balance CIF'],
                        COL['Is Restricted'], COL['Notes'], COL['Condition Sheet'], COL['Transfer Status'],
                    ),
                )

            # Merge cells for license-level columns
            if row_span > 1:
                end_row = start_row + row_span - 1
                for col in GROUPED_COLS:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

                # Apply vertical center alignment to merged cells
                for col in GROUPED_COLS:
                    cell = ws.cell(row=start_row, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border

            sr_no += 1

        # ── Totals row — only the numeric columns the View totals ──────────
        # Available Balance is a license-level value (repeated per item, same
        # simplification the View's grouped rowSpan cell uses), so it's summed
        # once per license, not once per raw row, to avoid overcounting.
        last_data_row = current_row - 1
        total_available_quantity = sum(item['available_quantity'] for item in items)
        total_planned_quantity = sum(item.get('planned_quantity') or 0 for item in items)
        total_planned_cif = sum(item.get('planned_cif') or 0 for item in items)
        unique_license_balances = {}
        for item in items:
            unique_license_balances.setdefault(item['license_id'], item['available_balance'])
        total_available_balance = sum(unique_license_balances.values())

        totals_row = current_row
        label_cell = ws.cell(row=totals_row, column=1, value='TOTAL')
        label_cell.font = Font(bold=True)
        ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=COL['Item Name'])
        for name, value, fmt in (
            ('Available Quantity', total_available_quantity, QTY_FMT),
            ('Available Balance', total_available_balance, CIF_FMT),
            ('Plan Qty', total_planned_quantity, QTY_FMT),
            ('Plan CIF', total_planned_cif, CIF_FMT),
        ):
            cell = ws.cell(row=totals_row, column=COL[name], value=value)
            cell.font = Font(bold=True)
            cell.number_format = fmt
        for row in ws[totals_row]:
            row.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Freeze the header row and enable AutoFilter over the data range
        # (the totals row sits below the filterable range, same convention
        # every business report in this app uses).
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{last_data_row}"

        # Save to bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=item_report.xlsx'
        return response


class ItemReportViewSet(viewsets.ViewSet):
    """
    ViewSet for Item Report actions like getting available items for filter.
    """
    permission_classes = [ReportPermission]

    @action(detail=False, methods=['get'], url_path='available-items')
    def available_items(self, request):
        """
        Get item names that actually have linked import items with
        available_value > 0.  Only these names will produce results
        in the report, so showing the rest in the multi-select filter
        is misleading.
        Returns: List of {id, name} dicts ordered by name.
        """
        item_names = (
            ItemNameModel.objects
            .filter(
                is_active=True,
                license_import_item__available_value__gt=0,   # has plannable items
            )
            .distinct()
            .order_by('name')
            .values('id', 'name')
        )
        return Response(list(item_names))

    @action(detail=False, methods=['get'], url_path='planned-item-names')
    def planned_item_names(self, request):
        """
        Get item names actually used as a planning-item target on at least
        one `LicenseItemPlan` line (e.g. E132 Auto-Plan's Nuts / Yeast / Palm
        Kernel Oil / RBD Palmolein Oil / Cheese / Aluminium Foil). Powers the
        "Planned Item Name" filter on the Allotment Available License Items
        screen's Plan mode — only these names can ever produce a match there.
        Returns: List of {id, name} dicts ordered by name.
        """
        item_names = (
            ItemNameModel.objects
            .filter(is_active=True, plan_lines__isnull=False)
            .distinct()
            .order_by('name')
            .values('id', 'name')
        )
        return Response(list(item_names))
