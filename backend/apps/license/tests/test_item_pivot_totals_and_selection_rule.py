"""Canonical Item Pivot v1 totals and export parity regressions.

The report is a read-only projection of persisted plans. It never invokes
Auto Plan as an implicit fallback, and its JSON and XLSX representations must
consume the same canonical matrix.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from apps.license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel, LicenseItemPlan
from apps.license.tests.test_item_pivot_excel_export import (
    _download_excel, pivot_license, pivot_masters, superuser_client,
)


def _get_json(client):
    response = client.get(reverse("license:item-pivot-report"), {
        "min_balance": 200, "license_status": "active",
    })
    assert response.status_code == 200, getattr(response, "data", response.content[:300])
    return response.json()


def _fixture_group(data):
    return next(group for group in data["groups"] if group["notification_number"] == "PIVN1")


def _number(value):
    return Decimal(str(value))


@pytest.mark.django_db
def test_persisted_manual_plan_is_the_authoritative_planned_pair(superuser_client, pivot_license):
    """The persisted 40-unit/$400 plan is displayed verbatim, rather than
    recalculated by the report or replaced with a display-price estimate."""
    group = _fixture_group(_get_json(superuser_client))
    license_row = next(row for row in group["licenses"] if row["license_number"] == "PIVOT-EXCEL-001")
    cell = next(iter(license_row["items"].values()))

    assert _number(cell["plan_qty"]) == Decimal("40.000")
    assert _number(cell["planned_cif"]) == Decimal("400.00")
    assert _number(cell["effective_planned_qty"]) == Decimal("40.000")
    assert _number(cell["effective_planned_cif"]) == Decimal("400.00")
    assert _number(license_row["planned_cif"]) == Decimal("400.00")


@pytest.mark.django_db
def test_report_read_does_not_auto_plan_an_unpersisted_import_item(superuser_client, pivot_masters):
    """A report request must not synthesize plans for an import row. This
    preserves the Celery-only planning boundary and makes a missing plan
    auditable instead of silently presenting an unsaved calculation."""
    license_obj = LicenseDetailsModel.objects.create(
        license_number="PIVOT-EXCEL-NOPLAN-001", license_date=date.today() - timedelta(days=30),
        license_expiry_date=date.today() + timedelta(days=30), exporter=pivot_masters["exporter"],
        notification_number=pivot_masters["notification"], scheme_code=pivot_masters["scheme"],
        purchase_status=pivot_masters["purchase_status"], file_number="PIVOT-FILE-NOPLAN-001",
    )
    LicenseExportItemModel.objects.create(license=license_obj, norm_class=pivot_masters["norm_class"], cif_fc=Decimal("500.00"))
    import_item = LicenseImportItemsModel.objects.create(
        license=license_obj, serial_number=1, description="Pivot import item (no plan)",
        hs_code=pivot_masters["hs_code"], quantity=Decimal("50.000"), available_quantity=Decimal("50.000"),
        cif_fc=Decimal("300.00"),
    )
    import_item.items.add(pivot_masters["item_name"])

    plans_before = LicenseItemPlan.objects.filter(license=license_obj).count()
    _get_json(superuser_client)
    assert LicenseItemPlan.objects.filter(license=license_obj).count() == plans_before


@pytest.mark.django_db
def test_group_totals_are_the_exact_sum_of_canonical_cells(superuser_client, pivot_license):
    group = _fixture_group(_get_json(superuser_client))
    license_row = next(row for row in group["licenses"] if row["license_number"] == "PIVOT-EXCEL-001")
    key, cell = next(iter(license_row["items"].items()))
    totals = group["totals"]

    for field in ("total_cif", "debited_cif", "allotted_cif", "planned_cif", "balance_cif"):
        assert _number(totals[field]) == _number(license_row[field])
    for field in ("total_qty", "allotted_qty", "debited_qty", "balance_qty", "plan_qty", "planned_cif"):
        assert _number(totals["items"][key][field]) == _number(cell[field])


@pytest.mark.django_db
def test_excel_total_row_consumes_the_same_canonical_group_totals(superuser_client, pivot_license):
    data = _get_json(superuser_client)
    group = _fixture_group(data)
    workbook = load_workbook(BytesIO(_download_excel(superuser_client)), data_only=True)
    sheet = next(ws for ws in workbook.worksheets if ws.title != "TOTAL_Summary")

    # Canonical exporter writes metadata at row 1, headers at row 2, and a
    # TOTAL row after the group licences. Fixed columns precede each item's
    # ten fields; compare by index so repeated labels cannot select another
    # canonical item column.
    header = [cell.value for cell in sheet[2]]
    total_row = next([cell.value for cell in row] for row in sheet.iter_rows() if str(row[0].value).startswith("TOTAL —"))
    assert len(total_row) == len(header)
    for index, field in ((4, "total_cif"), (5, "debited_cif"), (6, "allotted_cif"), (7, "planned_cif"), (8, "balance_cif")):
        assert _number(total_row[index]) == _number(group["totals"][field])
    item_totals = next(iter(group["totals"]["items"].values()))
    assert _number(total_row[17]) == _number(item_totals["plan_qty"])
    assert _number(total_row[18]) == _number(item_totals["planned_cif"])
