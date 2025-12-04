"""
Inventory Balance Report by SION Norm

This report shows inventory balance for items grouped by SION norm class.
For each item, it displays:
- Item name
- Total quantity across all licenses
- Total debited quantity (used in BOEs)
- Total allotted quantity (allocated but not yet used)
- Available quantity (remaining balance)
- Total CIF value balance
"""

from decimal import Decimal
from typing import Dict, List, Any, Optional

from django.db.models import Sum, Q, F, DecimalField, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from core.constants import DEC_0, DEC_000
from core.models import SionNormClassModel, ItemNameModel
from license.models import LicenseExportItemModel, LicenseImportItemsModel, LicenseDetailsModel
# Excel export handled inline with openpyxl


@method_decorator(csrf_exempt, name='dispatch')
class InventoryBalanceReportView(View):
    """
    API endpoint for Inventory Balance Report by SION Norm.

    GET parameters:
        - sion_norm: SION norm class (e.g., E1, E5) - required
        - format: 'json' or 'excel' (default: json)
        - include_zero: Include items with zero balance (default: false)
    """

    def get(self, request, *args, **kwargs):
        sion_norm = request.GET.get('sion_norm')
        output_format = request.GET.get('format', 'json').lower()
        include_zero = request.GET.get('include_zero', 'false').lower() == 'true'

        if not sion_norm:
            return JsonResponse({
                'error': 'sion_norm parameter is required',
                'example': '?sion_norm=E1&format=json'
            }, status=400)

        # Fetch report data
        try:
            report_data = self.generate_report(sion_norm, include_zero)
        except SionNormClassModel.DoesNotExist:
            return JsonResponse({
                'error': f'SION norm "{sion_norm}" not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)

        # Return in requested format
        if output_format == 'excel':
            return self.export_to_excel(report_data, sion_norm)
        else:
            return JsonResponse(report_data, safe=False)

    def generate_report(self, sion_norm_class: str, include_zero: bool = False) -> Dict[str, Any]:
        """
        Generate inventory balance report for a specific SION norm.

        Args:
            sion_norm_class: SION norm class code (e.g., E1, E5)
            include_zero: Whether to include items with zero balance

        Returns:
            Dictionary containing report data
        """
        # Get SION norm
        norm = SionNormClassModel.objects.get(norm_class=sion_norm_class)

        # Get all licenses with this SION norm in export items
        licenses_with_norm = LicenseDetailsModel.objects.filter(
            export_license__norm_class=norm,
            is_active=True
        ).distinct()

        # Get all import items from these licenses
        # Group by item name and aggregate quantities
        import_items_query = LicenseImportItemsModel.objects.filter(
            license__in=licenses_with_norm
        ).prefetch_related('items')

        # Build item-level aggregation
        items_data = self._aggregate_by_items(import_items_query)

        # Filter out zero balance items if requested
        if not include_zero:
            items_data = [item for item in items_data if item['available_quantity'] > 0]

        # Calculate totals
        totals = {
            'total_quantity': sum(item['total_quantity'] for item in items_data),
            'debited_quantity': sum(item['debited_quantity'] for item in items_data),
            'allotted_quantity': sum(item['allotted_quantity'] for item in items_data),
            'available_quantity': sum(item['available_quantity'] for item in items_data),
            'total_cif_value': sum(item['total_cif_value'] for item in items_data),
            'available_cif_value': sum(item['available_cif_value'] for item in items_data),
        }

        return {
            'sion_norm': {
                'code': norm.norm_class,
                'description': norm.description or '',
                'head_norm': norm.head_norm.name if norm.head_norm else '',
            },
            'summary': {
                'total_licenses': licenses_with_norm.count(),
                'total_items': len(items_data),
                'total_quantity': float(totals['total_quantity']),
                'total_debited': float(totals['debited_quantity']),
                'total_allotted': float(totals['allotted_quantity']),
                'total_available': float(totals['available_quantity']),
                'total_cif_value': float(totals['total_cif_value']),
                'available_cif_value': float(totals['available_cif_value']),
            },
            'items': items_data,
        }

    def _aggregate_by_items(self, import_items_query) -> List[Dict[str, Any]]:
        """
        Aggregate import items by item name.

        Args:
            import_items_query: QuerySet of LicenseImportItemsModel

        Returns:
            List of dictionaries with aggregated data per item
        """
        # Group by item through M2M relationship
        items_map: Dict[int, Dict[str, Any]] = {}

        for import_item in import_items_query.select_related('hs_code', 'license'):
            # Get all linked items (M2M)
            for item in import_item.items.all():
                if item.id not in items_map:
                    items_map[item.id] = {
                        'item_id': item.id,
                        'item_name': item.name,
                        'hs_code': import_item.hs_code.hs_code if import_item.hs_code else '',
                        'unit': import_item.unit,
                        'description': import_item.description or item.name,
                        'total_quantity': Decimal('0.000'),
                        'debited_quantity': Decimal('0.000'),
                        'allotted_quantity': Decimal('0.000'),
                        'available_quantity': Decimal('0.000'),
                        'total_cif_value': Decimal('0.00'),
                        'available_cif_value': Decimal('0.00'),
                        'licenses': set(),
                    }

                # Aggregate quantities
                items_map[item.id]['total_quantity'] += import_item.quantity or DEC_000
                items_map[item.id]['debited_quantity'] += import_item.debited_quantity or DEC_000
                items_map[item.id]['allotted_quantity'] += import_item.allotted_quantity or DEC_000
                items_map[item.id]['available_quantity'] += import_item.available_quantity or DEC_000
                items_map[item.id]['total_cif_value'] += import_item.cif_fc or DEC_0
                items_map[item.id]['available_cif_value'] += import_item.available_value or DEC_0
                items_map[item.id]['licenses'].add(import_item.license.license_number)

        # Convert to list and format for output
        items_list = []
        for item_data in items_map.values():
            items_list.append({
                'item_name': item_data['item_name'],
                'hs_code': item_data['hs_code'],
                'unit': item_data['unit'],
                'description': item_data['description'],
                'total_quantity': float(item_data['total_quantity']),
                'debited_quantity': float(item_data['debited_quantity']),
                'allotted_quantity': float(item_data['allotted_quantity']),
                'available_quantity': float(item_data['available_quantity']),
                'total_cif_value': float(item_data['total_cif_value']),
                'available_cif_value': float(item_data['available_cif_value']),
                'license_count': len(item_data['licenses']),
            })

        # Sort by item name
        items_list.sort(key=lambda x: x['item_name'])

        return items_list

    def export_to_excel(self, report_data: Dict[str, Any], sion_norm: str) -> HttpResponse:
        """
        Export report data to Excel format.

        Args:
            report_data: Report data dictionary
            sion_norm: SION norm class code

        Returns:
            HttpResponse with Excel file
        """
        # Prepare data for Excel
        headers = [
            'Item Name',
            'HS Code',
            'Unit',
            'Description',
            'Total Quantity',
            'Debited Quantity',
            'Allotted Quantity',
            'Available Quantity',
            'Total CIF Value',
            'Available CIF Value',
            'License Count',
        ]

        rows = []
        for item in report_data['items']:
            rows.append([
                item['item_name'],
                item['hs_code'],
                item['unit'],
                item['description'],
                item['total_quantity'],
                item['debited_quantity'],
                item['allotted_quantity'],
                item['available_quantity'],
                item['total_cif_value'],
                item['available_cif_value'],
                item['license_count'],
            ])

        # Add summary row
        summary = report_data['summary']
        rows.append([])  # Empty row
        rows.append(['SUMMARY', '', '', '', '', '', '', '', '', '', ''])
        rows.append([
            'Total',
            '',
            '',
            '',
            summary['total_quantity'],
            summary['total_debited'],
            summary['total_allotted'],
            summary['total_available'],
            summary['total_cif_value'],
            summary['available_cif_value'],
            summary['total_items'],
        ])

        # Build workbook using openpyxl directly
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"SION {sion_norm}"

        current_row = 1

        # Add title
        title = f"Inventory Balance Report - SION Norm: {sion_norm}"
        if report_data['sion_norm']['description']:
            title += f" ({report_data['sion_norm']['description']})"

        worksheet.merge_cells(f'A{current_row}:K{current_row}')
        title_cell = worksheet[f'A{current_row}']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        current_row += 1

        # Add data rows
        for row_data in rows:
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=current_row, column=col_num, value=value)
            current_row += 1

        # Auto-adjust column widths
        from openpyxl.cell.cell import MergedCell

        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)

            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    # Skip merged cells
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventory_balance_{sion_norm}.xlsx"'

        workbook.save(response)
        return response


# Legacy function-based view for backward compatibility
def inventory_balance_report(request):
    """Function-based view wrapper."""
    view = InventoryBalanceReportView.as_view()
    return view(request)
