# allotment/views_export.py
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.decorators import action

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from core.utils.pdf_utils import create_pdf_exporter

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def add_grouped_export_action(viewset_class):
    """
    Decorator to add grouped export functionality to AllotmentViewSet
    """

    @action(detail=False, methods=['get'], url_path='download')
    def download_grouped_export(self, request):
        """
        Export allotments grouped by company and item name.
        Query params:
            - type: 'AT' for allotment (default), 'TL' for transfer letter
            - export: 'pdf' or 'xlsx' (from query params, e.g., _export=pdf)
            - company: filter by company ID
            - item_name: filter by item name
            - is_boe: filter by is_boe status
        """
        export_type = request.query_params.get('type', 'AT')
        export_format = request.query_params.get('_export', 'pdf').lower()

        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Filter only allotments with details and order by company, item, port
        # Use distinct() without parameters to avoid duplication from prefetch_related
        queryset = queryset.filter(
            allotment_details__isnull=False
        ).distinct().order_by(
            'company__name', 'item_name', 'port__code'
        ).prefetch_related(
            'allotment_details__item__license__exporter',
            'allotment_details__item__hs_code',
            'company',
            'port'
        )

        if export_format == 'pdf':
            return self._export_grouped_pdf(queryset, export_type)
        elif export_format == 'xlsx':
            return self._export_grouped_xlsx(queryset, export_type)
        else:
            return HttpResponse("Invalid export format. Use 'pdf' or 'xlsx'.", status=400)

    def _export_grouped_pdf(self, queryset, export_type):
        """Export grouped allotments to PDF"""
        if not REPORTLAB_AVAILABLE:
            return HttpResponse("PDF export not available", status=500)

        # Group data
        grouped_data = self._group_allotments(queryset)

        # Get active exchange rate
        from core.models import ExchangeRateModel
        active_rate = ExchangeRateModel.get_active_rate()

        # Create PDF exporter
        pdf_exporter = create_pdf_exporter(
            title="Allotment Report",
            filename_prefix="Pending_Allotments",
            orientation='landscape'
        )

        # Create PDF response - inline display for new tab
        response = HttpResponse(content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Allotment_Report_{timestamp}.pdf'
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        buffer = BytesIO()
        doc = pdf_exporter.create_document(buffer)

        elements = []

        # Calculate totals
        total_usd = sum(item['value'] for company in grouped_data.values()
                        for ports in company.values()
                        for items in ports.values()
                        for item in items)

        # Create header section with title on left and exchange rate on right
        if active_rate:
            # Exchange rate mini table
            exchange_rate_data = [
                ['Exch. Rate', f"{active_rate.date.strftime('%d-%m-%Y')}"],
                ['USD', f"{active_rate.usd}"],
                ['EUR', f"{active_rate.euro}"],
                ['GBP', f"{active_rate.pound_sterling}"],
                ['CNY', f"{active_rate.chinese_yuan}"]
            ]

            exchange_table = Table(exchange_rate_data, colWidths=[0.8*inch, 0.8*inch])
            exchange_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), pdf_exporter.HEADER_BG),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            # Create title section
            title_para = Paragraph("Allotment Report", pdf_exporter.title_style)
            subtitle_para = Paragraph(f"Total USD $: {pdf_exporter.format_number(total_usd)}", pdf_exporter.subtitle_style)
            timestamp_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            from reportlab.lib.styles import ParagraphStyle
            timestamp_style = ParagraphStyle(
                'Timestamp',
                parent=pdf_exporter.styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_CENTER,
                spaceAfter=16
            )
            timestamp_para = Paragraph(timestamp_text, timestamp_style)

            # Create a layout table with title on left and exchange rate on right
            header_layout = Table(
                [[title_para, exchange_table]],
                colWidths=[7.5*inch, 1.8*inch]
            )
            header_layout.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ]))

            elements.append(header_layout)
            elements.append(subtitle_para)
            elements.append(timestamp_para)
        else:
            # No exchange rate, use standard title
            subtitle = f"Total USD $: {pdf_exporter.format_number(total_usd)}"
            pdf_exporter.add_title(elements, subtitle=subtitle)

        # Process each company
        for company_name, items_dict in grouped_data.items():
            # Company header
            pdf_exporter.add_company_header(elements, company_name)

            # Track company totals
            company_total_qty = 0
            company_total_value = 0
            company_total_inr = 0

            # Process each item within company
            for item_name_key, ports_dict in items_dict.items():
                # Get the actual item name from first allotment for display
                first_allot = next(iter(next(iter(ports_dict.values())))) if ports_dict else None
                display_item_name = first_allot[
                    'item_name'] if first_allot and 'item_name' in first_allot else item_name_key

                # Item subheader (optional, can be styled differently)
                if len(items_dict) > 1:  # Only show if multiple items
                    pdf_exporter.add_section_header(elements, f"Item: {display_item_name}")

                # Process each port within item
                for port_code, allotments in ports_dict.items():
                    # Table header with DFIA columns + Total CIF INR (removed BOE column)
                    table_data = [[
                        'Sr No', 'Allotment\nDate', 'Port', 'Quantity\n(KGS)',
                        'Unit Price ($)', 'Value ($)', 'Total CIF\nINR', 'Item Name', 'Invoice', 'ETA',
                        'DFIA No.', 'DFIA Date', 'DFIA Port', 'Item Sr.', 'DFIA Qty.', 'DFIA $.', 'DFIA CIF'
                    ]]

                    sr_no = 1
                    item_total_qty = 0
                    item_total_value = 0
                    item_total_inr = 0

                    for allot in allotments:
                        allot_cif_inr = allot['value'] * allot.get('exchange_rate', 89.5)

                        # Add rows for each license detail
                        for idx, detail in enumerate(allot['details']):
                            item_name_text = allot.get('item_name', display_item_name)
                            invoice_text = allot['invoice']

                            if idx == 0:
                                # First license row: show allotment data + license data (removed BOE column)
                                table_data.append([
                                    sr_no,
                                    allot['date'],
                                    allot['port'],
                                    pdf_exporter.format_number(allot['quantity'], decimals=0),
                                    pdf_exporter.format_number(allot['unit_price']),
                                    pdf_exporter.format_number(allot['value']),
                                    pdf_exporter.format_number(allot_cif_inr),
                                    item_name_text,
                                    invoice_text,
                                    allot['eta'],
                                    detail['dfia_no'],
                                    detail['dfia_date'],
                                    detail['dfia_port'],
                                    detail['item_sr_no'],
                                    pdf_exporter.format_number(detail['dfia_qty'], decimals=0),
                                    pdf_exporter.format_number(detail['dfia_value']),
                                    pdf_exporter.format_number(detail['dfia_cif'])
                                ])
                                item_total_qty += allot['quantity']
                                item_total_value += allot['value']
                                item_total_inr += allot_cif_inr
                            else:
                                # Subsequent license rows: empty allotment columns + license data (10 empty cols instead of 11)
                                table_data.append([
                                    '', '', '', '', '', '', '', '', '', '',
                                    detail['dfia_no'],
                                    detail['dfia_date'],
                                    detail['dfia_port'],
                                    detail['item_sr_no'],
                                    pdf_exporter.format_number(detail['dfia_qty'], decimals=0),
                                    pdf_exporter.format_number(detail['dfia_value']),
                                    pdf_exporter.format_number(detail['dfia_cif'])
                                ])
                        sr_no += 1

                    # Add to company totals
                    company_total_qty += item_total_qty
                    company_total_value += item_total_value
                    company_total_inr += item_total_inr

                    # Add totals row (17 columns total, removed BOE column)
                    table_data.append([
                        '', '', 'Total\nQuantity',
                        pdf_exporter.format_number(item_total_qty, decimals=0),
                        'Total USD $',
                        pdf_exporter.format_number(item_total_value),
                        pdf_exporter.format_number(item_total_inr),
                        '', '', '', '', '', '', '', '', '', ''
                    ])

                    # Create table with column widths for 17 columns (reduced Item Sr., increased DFIA CIF)
                    col_widths = [0.35 * inch, 0.6 * inch, 0.5 * inch, 0.6 * inch, 0.65 * inch, 0.65 * inch,
                                  0.85 * inch, 1.0 * inch, 0.8 * inch, 0.6 * inch,
                                  0.75 * inch, 0.6 * inch, 0.6 * inch, 0.4 * inch, 0.6 * inch, 0.65 * inch, 0.75 * inch]

                    # Use shared PDF exporter's create_table method
                    table = pdf_exporter.create_table(table_data, col_widths=col_widths, repeating_rows=1, wrap_text=True)

                    # Apply additional number column alignment for columns: 3, 4, 5, 6, 13, 14, 15, 16 (0-indexed, removed BOE)
                    additional_styles = []
                    for col_idx in [3, 4, 5, 6, 13, 14, 15, 16]:
                        additional_styles.append(
                            ('ALIGN', (col_idx, 1), (col_idx, len(table_data) - 1), 'RIGHT')
                        )

                    # Bold the last row (totals)
                    additional_styles.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
                    additional_styles.append(('BACKGROUND', (0, -1), (-1, -1), pdf_exporter.HEADER_BG))

                    # Apply additional styles
                    table.setStyle(TableStyle(additional_styles))

                    elements.append(table)
                    pdf_exporter.add_spacer(elements, 0.15)

            # Add company total after all items
            summary_data = [[
                'Company Total:',
                f"Qty: {pdf_exporter.format_number(company_total_qty, decimals=0)} KGS",
                f"Value (USD): ${pdf_exporter.format_number(company_total_value)}",
                f"Total CIF INR: {pdf_exporter.format_number(company_total_inr)}"
            ]]
            company_total_table = pdf_exporter.create_summary_table(summary_data)
            elements.append(company_total_table)
            pdf_exporter.add_spacer(elements, 0.2)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    def _export_grouped_xlsx(self, queryset, export_type):
        """Export grouped allotments to Excel"""
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("Excel export not available", status=500)

        # Group data
        grouped_data = self._group_allotments(queryset)

        # Get active exchange rate
        from core.models import ExchangeRateModel
        active_rate = ExchangeRateModel.get_active_rate()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Allotment Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Title
        ws.merge_cells(f'A{row}:O{row}')
        cell = ws[f'A{row}']
        cell.value = "Allotment Report"
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Total and date
        total_usd = sum(item['value'] for company in grouped_data.values()
                        for ports in company.values()
                        for items in ports.values()
                        for item in items)
        ws.merge_cells(f'A{row}:Q{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total USD $: {total_usd:,.2f}    {datetime.now().strftime('%d-%m-%Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Exchange rate info
        if active_rate:
            ws.merge_cells(f'A{row}:Q{row}')
            cell = ws[f'A{row}']
            cell.value = f"Exchange Rate (as of {active_rate.date.strftime('%d-%m-%Y')}): USD {active_rate.usd} | EUR {active_rate.euro} | GBP {active_rate.pound_sterling} | CNY {active_rate.chinese_yuan}"
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(italic=True, size=10)
        row += 2

        # Process each company
        for company_name, items_dict in grouped_data.items():
            # Company header
            ws.merge_cells(f'A{row}:Q{row}')
            cell = ws[f'A{row}']
            cell.value = company_name
            cell.font = Font(bold=True, size=14)
            row += 1

            # Track company totals
            company_total_qty = 0
            company_total_value = 0
            company_total_inr = 0

            # Process each item within company
            for item_name_key, ports_dict in items_dict.items():
                # Get the actual item name from first allotment for display
                first_allot = next(iter(next(iter(ports_dict.values())))) if ports_dict else None
                display_item_name = first_allot[
                    'item_name'] if first_allot and 'item_name' in first_allot else item_name_key

                # Item subheader (optional)
                if len(items_dict) > 1:  # Only show if multiple items
                    ws.merge_cells(f'A{row}:Q{row}')
                    cell = ws[f'A{row}']
                    cell.value = f"Item: {display_item_name}"
                    cell.font = Font(bold=True, size=12, italic=True)
                    row += 1

                # Process each port within item
                for port_code, allotments in ports_dict.items():
                    # Table headers - allotment info + license subheader (removed BOE column)
                    main_headers = ['Sr No', 'Allotment Date', 'Port', 'Quantity (KGS)',
                                    'Unit Price ($)', 'Value ($)', 'Total CIF INR', 'Item Name', 'Invoice', 'ETA']
                    license_headers = ['DFIA No.', 'DFIA Date', 'DFIA Port', 'Item Sr. NO.', 'DFIA Qty.', 'DFIA $.',
                                       'DFIA CIF']

                    # Write main headers
                    for col_idx, header in enumerate(main_headers, 1):
                        cell = ws.cell(row=row, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = border

                    # Merge cells for "DFIA Licenses" header
                    start_col = len(main_headers) + 1
                    end_col = start_col + len(license_headers) - 1
                    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                    cell = ws.cell(row=row, column=start_col, value='DFIA Licenses')
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                    row += 1

                    # Write license sub-headers
                    for col_idx, header in enumerate(license_headers, start_col):
                        cell = ws.cell(row=row, column=col_idx, value=header)
                        cell.font = Font(bold=True, size=9)
                        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = border
                    row += 1

                    sr_no = 1
                    item_total_qty = 0
                    item_total_value = 0
                    item_total_inr = 0

                    for allot in allotments:
                        start_row_for_allot = row
                        allot_cif_inr = allot['value'] * allot.get('exchange_rate', 89.5)

                        # Write allotment data (will be merged vertically if multiple licenses) - removed BOE
                        allot_data = [
                            sr_no,
                            allot['date'],
                            allot['port'],
                            int(allot['quantity']),
                            allot['unit_price'],
                            allot['value'],
                            allot_cif_inr,
                            allot.get('item_name', display_item_name),
                            allot['invoice'],
                            allot['eta']
                        ]

                        # Write license details (one row per license)
                        for detail_idx, detail in enumerate(allot['details']):
                            # Write allotment data only in first row
                            if detail_idx == 0:
                                for col_idx, value in enumerate(allot_data, 1):
                                    cell = ws.cell(row=row, column=col_idx, value=value)
                                    cell.border = border
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    if col_idx in [4, 5, 6, 7]:  # Number columns (added 7 for Total CIF INR)
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                        if col_idx in [5, 6]:  # USD columns
                                            cell.number_format = '#,##0.00'
                                        elif col_idx == 7:  # INR column
                                            cell.number_format = '₹#,##0.00'

                                # Track totals only once per allotment
                                item_total_qty += allot['quantity']
                                item_total_value += allot['value']
                                item_total_inr += allot_cif_inr

                            # Write license details
                            license_data = [
                                detail['dfia_no'],
                                detail['dfia_date'],
                                detail['dfia_port'],
                                detail['item_sr_no'],
                                int(detail['dfia_qty']),
                                detail['dfia_value'],
                                detail['dfia_cif']
                            ]

                            for col_idx, value in enumerate(license_data, len(main_headers) + 1):
                                cell = ws.cell(row=row, column=col_idx, value=value)
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                if col_idx in [len(main_headers) + 5, len(main_headers) + 6,
                                               len(main_headers) + 7]:  # Qty, Value, and CIF columns
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                            row += 1

                        # Merge allotment cells vertically if multiple licenses
                        if len(allot['details']) > 1:
                            for col_idx in range(1, len(main_headers) + 1):
                                ws.merge_cells(start_row=start_row_for_allot, start_column=col_idx,
                                               end_row=row - 1, end_column=col_idx)

                        sr_no += 1

                    # Add to company totals
                    company_total_qty += item_total_qty
                    company_total_value += item_total_value
                    company_total_inr += item_total_inr

                    ws.cell(row=row, column=3, value="Total Quantity").font = Font(bold=True)
                    ws.cell(row=row, column=4, value=int(item_total_qty)).font = Font(bold=True)
                    ws.cell(row=row, column=4).number_format = '#,##0'
                    ws.cell(row=row, column=5, value="Total USD $").font = Font(bold=True)
                    ws.cell(row=row, column=6, value=item_total_value).font = Font(bold=True)
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    ws.cell(row=row, column=7, value=item_total_inr).font = Font(bold=True)
                    ws.cell(row=row, column=7).number_format = '₹#,##0.00'
                    row += 2

            # Add company total after all items
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws.cell(row=row, column=1, value="Company Total:")
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=4, value=int(company_total_qty)).font = Font(bold=True, size=12)
            ws.cell(row=row, column=4).fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=5, value="Total USD $:").font = Font(bold=True, size=12)
            ws.cell(row=row, column=5).fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=6, value=company_total_value).font = Font(bold=True, size=12)
            ws.cell(row=row, column=6).fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=6).number_format = '#,##0.00'

            ws.cell(row=row, column=7, value=company_total_inr).font = Font(bold=True, size=12)
            ws.cell(row=row, column=7).fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=7).number_format = '₹#,##0.00'
            row += 3

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="Pending_Allotments_{timestamp}.xlsx"'
        wb.save(response)

        return response

    def _group_allotments(self, queryset):
        """Group allotments by company → item name → port code (case-insensitive for item name)"""
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

        for allotment in queryset:
            company_name = allotment.company.name if allotment.company else "Unknown"
            port_code = allotment.port.code if allotment.port else "Unknown Port"
            # Use uppercase for grouping key to make it case-insensitive, but keep original for display
            item_name = allotment.item_name or "Unknown"
            item_name_key = item_name.upper()  # Case-insensitive grouping key

            allot_data = {
                'date': allotment.created_on.strftime(
                    '%d-%m-%Y') if allotment.created_on else '--',
                'port': port_code,
                'quantity': float(allotment.required_quantity or 0),
                'unit_price': float(allotment.unit_value_per_unit or 0),
                'value': float(allotment.required_value or 0),
                'exchange_rate': float(allotment.exchange_rate or 89.5),
                'invoice': allotment.invoice or '--',
                'eta': allotment.estimated_arrival_date.strftime(
                    '%d-%m-%Y') if allotment.estimated_arrival_date else '--',
                'is_boe': allotment.is_boe,
                'details': []
            }

            # Add allotment details
            for detail in allotment.allotment_details.all():
                license_obj = detail.item.license if detail.item else None
                allot_data['details'].append({
                    'dfia_no': license_obj.license_number if license_obj else '--',
                    'dfia_date': license_obj.license_date.strftime(
                        '%d-%m-%Y') if license_obj and license_obj.license_date else '--',
                    'dfia_port': license_obj.port.code if license_obj and license_obj.port else '--',
                    'item_sr_no': str(detail.item.serial_number) if detail.item else '--',
                    'dfia_qty': float(detail.qty or 0),
                    'dfia_value': float(detail.cif_fc or 0),
                    'dfia_cif': float(detail.cif_inr or 0)
                })

            # Store original item_name for display
            allot_data['item_name'] = item_name
            grouped_data[company_name][item_name_key][port_code].append(allot_data)

        return grouped_data

    # Add methods to the viewset class
    viewset_class.download_grouped_export = download_grouped_export
    viewset_class._export_grouped_pdf = _export_grouped_pdf
    viewset_class._export_grouped_xlsx = _export_grouped_xlsx
    viewset_class._group_allotments = _group_allotments

    return viewset_class
