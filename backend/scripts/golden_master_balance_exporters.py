#!/usr/bin/env python
"""
Golden-master characterization for the four license *balance exporter* endpoints:

    balance_pdf            GET  /api/licenses/{pk}/balance-pdf/
    balance_excel          GET  /api/licenses/{pk}/balance-excel/
    balance_excel_unused   GET  /api/licenses/{pk}/balance-excel-unused/
    bulk_balance_excel     POST /api/licenses/bulk-balance-excel/

WHY THIS EXISTS
    We are refactoring these endpoints (extracting balance math + PDF/Excel
    rendering out of the god-methods in apps/license/views/license.py) and MUST
    NOT change their output. The committed pytest suite (backend/tests/) cannot
    currently be relied on for this — its conftest imports pre-rename model
    names (`License`, `Company`, ...) and fails to collect. So this script
    fingerprints the *real* output of each endpoint against the live dev
    database (read-only), lets us snapshot a baseline BEFORE the refactor, and
    proves the output is byte/value-identical AFTER.

USAGE (from backend/, venv active)
    python scripts/golden_master_balance_exporters.py record   # snapshot baseline
    python scripts/golden_master_balance_exporters.py check     # compare vs baseline (exit 1 on drift)
    # optional: pass explicit license ids -> ... record 1843 1845 2427

This script only READS the database. It never writes. Auth uses DRF
force_authenticate with an existing superuser (no DB mutation).
"""
import os
import sys
import io
import json
import hashlib
import traceback
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lmanagement.settings")
# Ensure backend/ (parent of scripts/) is importable when run directly.
BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

import django  # noqa: E402
django.setup()  # noqa: E402

# APIRequestFactory issues requests as host 'testserver'; balance_pdf builds
# absolute URIs (logos/images) which triggers host validation. Allow it for
# this in-process read-only harness so we exercise the real code path.
from django.conf import settings as _settings  # noqa: E402
for _h in ("testserver", "localhost"):
    if _h not in _settings.ALLOWED_HOSTS:
        _settings.ALLOWED_HOSTS.append(_h)

from django.contrib.auth import get_user_model  # noqa: E402
from rest_framework.test import APIRequestFactory, force_authenticate  # noqa: E402
import openpyxl  # noqa: E402
from pypdf import PdfReader  # noqa: E402

from apps.license.views.license import LicenseDetailsViewSet  # noqa: E402
from apps.license.models import LicenseDetailsModel  # noqa: E402

BASELINE_DIR = Path(__file__).resolve().parent / ".golden_master"
BASELINE_PATH = BASELINE_DIR / "baseline.json"

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --------------------------------------------------------------------------- #
# License selection — deterministic, variety-seeking                          #
# --------------------------------------------------------------------------- #
def select_license_ids(limit=15):
    """Pick a deterministic, data-rich, variety-spanning sample of DFIA licenses.

    Richness = has export items (credit) + import items (debit surface). We rank
    by combined related-item count so the exporters actually render substance,
    then take the top `limit` (ties broken by id for reproducibility).
    """
    from django.db.models import Count

    qs = (
        LicenseDetailsModel.objects.annotate(
            _n_exp=Count("export_license", distinct=True),
            _n_imp=Count("import_license", distinct=True),
        )
        .order_by("-_n_exp", "-_n_imp", "id")
    )
    ids = list(qs.values_list("id", flat=True)[:limit])
    # Fallback: if annotations produced nothing useful, just take first N by id.
    if not ids:
        ids = list(LicenseDetailsModel.objects.order_by("id").values_list("id", flat=True)[:limit])
    return ids


# --------------------------------------------------------------------------- #
# Fingerprinting                                                              #
# --------------------------------------------------------------------------- #
def _read_content(resp):
    """Return raw bytes from a Django/DRF response, rendering if needed."""
    if hasattr(resp, "render") and not getattr(resp, "is_rendered", True):
        resp.render()
    # streaming (FileResponse) vs regular
    if getattr(resp, "streaming", False):
        return b"".join(resp.streaming_content)
    return resp.content


def _fingerprint_xlsx(content):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    sheets = []
    for ws in wb.worksheets:
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = repr(cell.value)
        sheets.append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged": sorted(str(r) for r in ws.merged_cells.ranges),
                "n_cells": len(cells),
                "cells": cells,
            }
        )
    return {"kind": "xlsx", "n_sheets": len(sheets), "sheets": sheets}


def _fingerprint_pdf(content):
    reader = PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception as exc:  # pragma: no cover - defensive
            pages.append(f"<extract-error:{type(exc).__name__}>")
    text = "\n".join(pages)
    return {
        "kind": "pdf",
        "n_pages": len(reader.pages),
        "text_sha256": hashlib.sha256(text.encode("utf-8", "replace")).hexdigest(),
        "text": text.splitlines(),
    }


def fingerprint_response(resp):
    fp = {
        "status": resp.status_code,
        "content_type": resp.get("Content-Type"),
        "content_disposition": resp.get("Content-Disposition"),
    }
    try:
        content = _read_content(resp)
    except Exception as exc:
        fp["read_error"] = f"{type(exc).__name__}: {exc}"
        return fp

    # NOTE: deliberately do NOT record raw byte length. .xlsx is a zip container
    # whose byte length is not stable run-to-run (compression/metadata ordering)
    # even from identical code; the meaningful invariant is the *decoded* content
    # (sheets/cells/merges below) and, for PDF, the extracted-text hash.
    if resp.status_code != 200:
        # characterize the error body too (truncated, decoded best-effort)
        fp["body"] = content[:2000].decode("utf-8", "replace")
        return fp

    ct = (fp["content_type"] or "")
    try:
        if XLSX_CT in ct:
            fp["body"] = _fingerprint_xlsx(content)
        elif "application/pdf" in ct:
            fp["body"] = _fingerprint_pdf(content)
        else:
            fp["body"] = {"kind": "raw", "sha256": hashlib.sha256(content).hexdigest()}
    except Exception:
        fp["parse_error"] = traceback.format_exc()
    return fp


# --------------------------------------------------------------------------- #
# Endpoint invocation                                                         #
# --------------------------------------------------------------------------- #
def _call_detail(user, action, pk):
    factory = APIRequestFactory()
    view = LicenseDetailsViewSet.as_view({"get": action})
    req = factory.get(f"/api/licenses/{pk}/{action}/")
    force_authenticate(req, user=user)
    return view(req, pk=str(pk))


def _call_bulk(user, license_numbers):
    factory = APIRequestFactory()
    view = LicenseDetailsViewSet.as_view({"post": "bulk_balance_excel"})
    req = factory.post(
        "/api/licenses/bulk-balance-excel/",
        data={"license_numbers": license_numbers},
        format="json",
    )
    force_authenticate(req, user=user)
    return view(req)


def build_snapshot(license_ids):
    User = get_user_model()
    user = User.objects.filter(is_superuser=True).first()
    if user is None:
        raise SystemExit("No superuser found in the dev DB to authenticate as.")

    id_to_number = dict(
        LicenseDetailsModel.objects.filter(id__in=license_ids).values_list("id", "license_number")
    )
    snapshot = {"license_ids": license_ids, "endpoints": {}}

    detail_actions = ["balance_pdf", "balance_excel", "balance_excel_unused"]
    for action in detail_actions:
        per_license = {}
        for pk in license_ids:
            try:
                resp = _call_detail(user, action, pk)
                per_license[str(pk)] = fingerprint_response(resp)
            except Exception:
                per_license[str(pk)] = {"invoke_error": traceback.format_exc()}
        snapshot["endpoints"][action] = per_license

    # bulk: happy path across the sample's license numbers, plus the two error
    # branches (empty list -> 400, all-unknown -> 404) that valid licenses never
    # exercise. Those branches are exactly where an extraction can drop a
    # module-global (e.g. rest_framework Response) unnoticed.
    numbers = [id_to_number.get(pk) for pk in license_ids if id_to_number.get(pk)]
    bulk_cases = {
        "happy": numbers,
        "empty": [],
        "unknown": ["__no_such_license_0000__", "__no_such_license_0001__"],
    }
    bulk = {"license_numbers": numbers}
    for case, nums in bulk_cases.items():
        try:
            bulk[case] = fingerprint_response(_call_bulk(user, nums))
        except Exception:
            bulk[case] = {"invoke_error": traceback.format_exc()}
    snapshot["endpoints"]["bulk_balance_excel"] = bulk
    return snapshot


# --------------------------------------------------------------------------- #
# Diffing                                                                     #
# --------------------------------------------------------------------------- #
def diff(old, new, path=""):
    """Yield human-readable difference strings between two JSON-like structures."""
    if type(old) is not type(new) and not (
        isinstance(old, (int, float)) and isinstance(new, (int, float))
    ):
        yield f"{path or '.'}: type {type(old).__name__} -> {type(new).__name__}"
        return
    if isinstance(old, dict):
        for k in old.keys() | new.keys():
            if k not in old:
                yield f"{path}.{k}: added"
            elif k not in new:
                yield f"{path}.{k}: removed"
            else:
                yield from diff(old[k], new[k], f"{path}.{k}")
    elif isinstance(old, list):
        if len(old) != len(new):
            yield f"{path}: list len {len(old)} -> {len(new)}"
        for i, (a, b) in enumerate(zip(old, new)):
            yield from diff(a, b, f"{path}[{i}]")
    else:
        if old != new:
            yield f"{path}: {old!r} -> {new!r}"


# --------------------------------------------------------------------------- #
# Main                                                                        #
# --------------------------------------------------------------------------- #
def main():
    args = sys.argv[1:]
    if not args or args[0] not in ("record", "check"):
        print(__doc__)
        raise SystemExit(2)
    mode = args[0]
    explicit_ids = [int(a) for a in args[1:] if a.isdigit()]

    if explicit_ids:
        license_ids = explicit_ids
    elif mode == "check" and BASELINE_PATH.exists():
        license_ids = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))["license_ids"]
    else:
        license_ids = select_license_ids()

    print(f"[golden-master] mode={mode} licenses={license_ids}")
    snapshot = build_snapshot(license_ids)

    if mode == "record":
        BASELINE_DIR.mkdir(exist_ok=True)
        BASELINE_PATH.write_text(
            json.dumps(snapshot, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        # quick human summary
        for action, data in snapshot["endpoints"].items():
            if action == "bulk_balance_excel":
                statuses = {c: data[c].get("status") for c in ("happy", "empty", "unknown")}
                print(f"  {action}: {statuses}")
            else:
                statuses = [v.get("status") for v in data.values()]
                print(f"  {action}: statuses={statuses}")
        print(f"[golden-master] baseline written -> {BASELINE_PATH}")
        return

    # check
    if not BASELINE_PATH.exists():
        raise SystemExit(f"No baseline at {BASELINE_PATH}; run `record` first.")
    baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))

    diffs = list(diff(baseline["endpoints"], snapshot["endpoints"], ""))
    if not diffs:
        print("[golden-master] PASS — output identical to baseline for all endpoints.")
        return
    print(f"[golden-master] FAIL — {len(diffs)} difference(s):")
    for d in diffs[:200]:
        print("   ", d)
    if len(diffs) > 200:
        print(f"    ... and {len(diffs) - 200} more")
    raise SystemExit(1)


if __name__ == "__main__":
    main()
