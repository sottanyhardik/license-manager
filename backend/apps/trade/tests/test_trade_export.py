# trade/tests/test_trade_export.py
"""
Regression tests for `GET /api/trades/export/` (the generic
`MasterViewSet` export action, `apps.core.views.master_view._export_xlsx`/
`_export_pdf`) — specifically the `From Company Label`/`To Company Label`/
`Incentive License` columns, which are declared on `LicenseTradeSerializer`
(not plain model attributes) and previously exported as blank/wrong values
because the export walked the raw model instance instead of the serializer.
"""
import itertools
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
import openpyxl
from io import BytesIO

from apps.core.models import CompanyModel, PortModel
from apps.license.models import IncentiveLicense
from apps.trade.models import LicenseTrade, IncentiveTradeLine

_iec_counter = itertools.count(1)


def _unique_iec() -> str:
    return f"{next(_iec_counter):010d}"


class TradeExportColumnsTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="export_admin", email="export_admin@example.com", password="pw12345",
        )
        self.client.force_authenticate(user=self.user)

        self.from_company = CompanyModel.objects.create(iec=_unique_iec(), name="Steel Enterprises")
        self.to_company = CompanyModel.objects.create(iec=_unique_iec(), name="Labdhi Global LLP")
        self.port = PortModel.objects.create(name="Test Port", code="TP1")

        self.incentive_license = IncentiveLicense.objects.create(
            license_type="RODTEP",
            license_number="0811099999",
            license_date=date(2026, 1, 1),
            license_expiry_date=date(2028, 1, 1),
            exporter=self.to_company,
            port_code=self.port,
            license_value=Decimal("50000.00"),
        )

        self.trade = LicenseTrade.objects.create(
            direction=LicenseTrade.DIR_PURCHASE,
            license_type="INCENTIVE",
            from_company=self.from_company,
            to_company=self.to_company,
            invoice_number="EXP-TEST-0001",
            invoice_date=date(2026, 1, 15),
        )
        IncentiveTradeLine.objects.create(
            trade=self.trade,
            incentive_license=self.incentive_license,
            license_value=Decimal("50000.00"),
            rate_pct=Decimal("1.5"),
            amount_inr=Decimal("750.00"),
        )

    def test_excel_export_resolves_company_labels_and_incentive_license(self):
        resp = self.client.get("/api/trades/export/?_export=xlsx")
        self.assertEqual(resp.status_code, 200)

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertIn("From Company Label", headers)
        self.assertIn("To Company Label", headers)
        self.assertIn("Incentive License", headers)

        row_by_invoice = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            row_by_invoice[row_dict.get("Invoice Number")] = row_dict

        row = row_by_invoice["EXP-TEST-0001"]
        # Previously blank: `from_company_label`/`to_company_label` only
        # exist as `CharField(source='from_company.name')`-style declared
        # serializer fields, never as raw model attributes.
        self.assertEqual(row["From Company Label"], "Steel Enterprises")
        self.assertEqual(row["To Company Label"], "Labdhi Global LLP")
        # Previously the raw FK's `str(IncentiveLicense(...))` repr (or
        # blank) — now the serializer's own resolved value.
        self.assertEqual(row["Incentive License"], "0811099999")

    def test_pdf_export_succeeds(self):
        resp = self.client.get("/api/trades/export/?_export=pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertGreater(len(resp.content), 0)
