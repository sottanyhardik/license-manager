# license/utils/condition_excel.py
"""
Shared helper for visualising LicenseImportItemsModel.condition_type in Excel
exports (Item Report, Item Pivot Report, etc.).

Ported verbatim from legacy/backend/apps/license/utils/condition_excel.py.
"""
from __future__ import annotations

from openpyxl.styles import PatternFill, Font

# Cell background → font colour, keyed by the persisted condition_type value.
_FILLS = {
    "AU":  ("DBEAFE", "1E3A8A"),  # blue   — Actual User
    "2%":  ("FEE2E2", "7F1D1D"),  # red    — strictest
    "3%":  ("FED7AA", "7C2D12"),  # orange
    "5%":  ("FEF3C7", "78350F"),  # amber
    "10%": ("D1FAE5", "065F46"),  # green
}


def condition_fill(condition_type) -> PatternFill | None:
    if not condition_type:
        return None
    colors = _FILLS.get(condition_type)
    if not colors:
        bg = "E5E7EB"
    else:
        bg = colors[0]
    return PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def condition_font(condition_type) -> Font | None:
    if not condition_type:
        return None
    colors = _FILLS.get(condition_type)
    if not colors:
        return Font(bold=True, color="374151")
    return Font(bold=True, color=colors[1])


def annotate_cell(cell, condition_type) -> None:
    """Apply both fill + font when a condition is present. No-op otherwise."""
    if not condition_type:
        return
    fill = condition_fill(condition_type)
    font = condition_font(condition_type)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
