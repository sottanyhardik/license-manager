"""
Item-wise Pivot Report

Shows licenses with items as column headers, displaying quantities and values per item.
Similar to the GE DFIA report format.
"""

import logging
from collections import defaultdict
from decimal import Decimal
from typing import Dict, List, Any

from django.db.models import Prefetch
from django.http import JsonResponse, HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from apps.accounts.permissions import ReportPermission
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.constants import DEC_0, DEC_000, GE, MI, CO
from apps.core.models import ItemNameModel
from apps.license.models import (
    LicenseDetailsModel, LicenseImportItemsModel,
    LicenseExportItemModel, LicenseTransferModel,
)
from apps.license.services.plan_grouping import (
    merge_planned_import_items as _merge_planned_import_items,
    merge_items_for_classification as _merge_items_for_classification,
)
from apps.license.views.item_report import _ExcelPassthroughRenderer

def _safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _effective_planned_cif(plan_quantity, plan_cif, planned_cif):
    """The single manual-vs-norm-derived planned-CIF selection rule for one
    license x item cell: prefer the user-authored manual plan (`plan_quantity`/
    `plan_cif`, from `LicenseItemPlan`) when one exists, otherwise fall back
    to the norm-derived `planned_cif` (E1/E5/E132 waterfall). Computed once
    here and exposed per cell as `effective_planned_cif` — every consumer
    (JSON, React, Excel) reads that field instead of re-deriving this rule
    (Phase 2B.2A; see docs/architecture/ITEM_PIVOT_DISPLAY_DATASET_DESIGN.md).
    """
    pq = plan_quantity or 0
    pc = plan_cif or 0
    return pc if (pq or pc) else (planned_cif or 0)


def _effective_planned_quantity(plan_quantity, plan_cif, available_quantity):
    """Quantity counterpart to `_effective_planned_cif` — same manual-vs-norm
    branch (see that docstring), applied to quantity instead of CIF. The
    norm-derived branch falls back to `available_quantity` (not a
    `planned_quantity` field) because the E1/E5/E132 waterfall always plans
    against the full available balance; there is no separate norm-derived
    planned-quantity field to select instead. Computed once here and exposed
    per cell as `effective_planned_quantity` (Phase 2B.2B; see
    docs/architecture/ITEM_PIVOT_NOTIFICATION_SUMMARY_DESIGN.md §5).

    Note: the design doc's reference signature carries an extra unused
    `planned_quantity` parameter for symmetry with `_effective_planned_cif`;
    it is omitted here since it can never affect the return value — this is
    a signature-only deviation with zero output difference.
    """
    pq = plan_quantity or 0
    pc = plan_cif or 0
    return pq if (pq or pc) else (available_quantity or 0)


def _build_notification_summary(licenses, items):
    """Translates `ItemPivotReport.tsx:507-621` (`calculateNotificationSummary`)
    verbatim, including its quirks — see
    docs/architecture/ITEM_PIVOT_NOTIFICATION_SUMMARY_DESIGN.md §1 (spec),
    §3 (restriction-pool worked example), §4 (blended unit price), §9
    (quirks, preserved on purpose per the §12 business decision — do NOT
    "improve" this logic without a corresponding product decision).

    `licenses` — already-built license row dicts (as returned by
    `_build_license_row`) for this scope: either one notification group's
    license list, or every license under a norm flattened across its
    notification groups.
    `items` — the report's ordered item catalogue as `(item_id, item_name)`
    tuples, same order as `sorted_items` / the response's top-level `items`
    list.
    """
    # Pass 1 — opening balance. This is the same sum as
    # `notification_totals[norm][notification_key]['balance_cif']`
    # (Phase 2B.2A) for the per-notification scope, and intentionally kept
    # in sync rather than reused via a shared variable — see design doc
    # §9(c)/§12: consolidating *how* the sum is computed doesn't change the
    # displayed value, so no special-casing is needed here.
    opening_balance = sum(float(lic.get('balance_cif', 0) or 0) for lic in licenses)

    # Pass 2 — restriction pool dedup: a license's `restriction_value` is a
    # shared quota against its restriction percentage, not a per-item value.
    # Dedup key is license_number + percentage, NOT license + item — see
    # design doc §3 for the worked example this reproduces exactly.
    restricted_items_by_percentage = {}
    processed_restrictions = set()
    for lic in licenses:
        lic_number = lic.get('license_number')
        for item_id, item_name in items:
            item_data = (lic.get('items') or {}).get(item_name)
            if item_data and item_data.get('restriction') is not None:
                pct = float(item_data.get('restriction') or 0)
                key = f"{lic_number}_{pct}"
                if key not in processed_restrictions:
                    processed_restrictions.add(key)
                    bucket = restricted_items_by_percentage.setdefault(
                        pct, {'shared_restriction_value': 0.0, 'items': {}}
                    )
                    bucket['shared_restriction_value'] += float(
                        item_data.get('restriction_value', 0) or 0
                    )

    # Pass 3 — per-item aggregation across licenses, in report item order.
    regular_items = {}
    total_available = 0.0
    total_planned_cif = 0.0
    total_planned_qty = 0.0

    for item_id, item_name in items:
        item_available = 0.0
        item_planned = 0.0
        item_planned_qty = 0.0
        has_restriction = False
        restriction_percentage = 0.0

        for lic in licenses:
            item_data = (lic.get('items') or {}).get(item_name)
            if item_data:
                available_quantity = float(item_data.get('available_quantity', 0) or 0)
                item_available += available_quantity

                plan_cif = float(item_data.get('plan_cif', 0) or 0)
                plan_quantity = float(item_data.get('plan_quantity', 0) or 0)
                item_has_manual = plan_cif > 0 or plan_quantity > 0
                planned_cif = float(item_data.get('planned_cif', 0) or 0)
                item_planned += plan_cif if item_has_manual else planned_cif
                item_planned_qty += plan_quantity if item_has_manual else available_quantity

                if item_data.get('restriction') is not None:
                    has_restriction = True
                    # Last license wins if licenses in this scope disagree
                    # on the percentage for this item — deliberate quirk,
                    # preserve verbatim, do not dedupe/reconcile. See design
                    # doc §9(a).
                    restriction_percentage = float(item_data.get('restriction') or 0)

        if item_available > 0 or item_planned > 0:
            item_summary = {
                # Split-planned items with no import counterpart have
                # item_available == 0; fall back to planned qty so the row
                # shows the correct balance quantity instead of 0. See
                # design doc §9(b) — the grand total below intentionally
                # does NOT use this fallback-adjusted value.
                'available': item_available if item_available > 0 else item_planned_qty,
                'planned_cif': item_planned,
                'planned_qty': item_planned_qty,
                'unit_price': (
                    round(item_planned / item_planned_qty, 2) if item_planned_qty > 0 else 0.0
                ),
            }

            if has_restriction:
                bucket = restricted_items_by_percentage.setdefault(
                    restriction_percentage, {'shared_restriction_value': 0.0, 'items': {}}
                )
                bucket['items'][item_name] = item_summary
            else:
                regular_items[item_name] = item_summary

            # Quirk (design doc §9(b)): the grand total sums the RAW
            # per-license `item_available` contribution, NOT the
            # fallback-adjusted `item_summary['available']` value set just
            # above — preserved verbatim from the frontend; do not "fix".
            total_available += item_available
            total_planned_cif += item_planned
            total_planned_qty += item_planned_qty

    blended_unit_price = (
        round(total_planned_cif / total_planned_qty, 2) if total_planned_qty > 0 else 0.0
    )

    return {
        'opening_balance': opening_balance,
        'total_available': total_available,
        'total_planned_cif': total_planned_cif,
        'total_planned_qty': total_planned_qty,
        'blended_unit_price': blended_unit_price,
        'regular_items': regular_items,
        'restricted_items_by_percentage': {
            str(pct): bucket for pct, bucket in restricted_items_by_percentage.items()
        },
    }


def _xlsx_safe_row(row):
    """Strip XML-illegal control characters from string cells before writing.

    Condition-sheet / description text extracted from DGFT PDFs can contain
    control chars (e.g. form-feed) that openpyxl rejects with IllegalCharacterError,
    which would 500 the whole Excel export. Cleans plain string values and the
    value of WriteOnlyCell objects in place; passes numbers/None through.
    """
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    cleaned = []
    for cell in row:
        val = getattr(cell, "value", cell)  # WriteOnlyCell -> its value; else the cell itself
        if isinstance(val, str):
            safe = ILLEGAL_CHARACTERS_RE.sub("", val)
            if val is not cell:  # it's a WriteOnlyCell — clean in place, keep styling
                cell.value = safe
                cleaned.append(cell)
            else:
                cleaned.append(safe)
        else:
            cleaned.append(cell)
    return cleaned


def _planning_split_sheet_rows(licenses_by_norm_notification, item_names):
    """Flat (license, pivot-item-column, split) rows for a "Planning Splits"
    sheet — one row per *visible* LicenseItemPlan split, for every license /
    item-name-column combination that actually has one.

    The main pivot grid is one row per license with a wide, fixed set of
    item-name columns (some of it write-only/append-only), so it can't host
    indented child rows the way license_balance_excel.py / item_report.py do.
    This sheet is additive detail alongside that grid, not a replacement for
    its per-item-name Plan Qty / Plan CIF cells.

    Reuses `rows_for_splits()` (the shared filter/label rules) rather than
    re-deriving the visibility filter here.
    """
    from apps.license.services.exporters.planning_split_rows import rows_for_splits

    rows = []
    for norm_class in sorted(licenses_by_norm_notification.keys()):
        notifications_dict = licenses_by_norm_notification[norm_class]
        for notification, licenses_list in sorted(notifications_dict.items()):
            for lic in licenses_list:
                for item_name in item_names:
                    item_data = lic['items'].get(item_name) or {}
                    for split_row in rows_for_splits(item_data.get('splits') or []):
                        rows.append((
                            lic['license_number'],
                            item_name,
                            split_row['item_name_label'],
                            split_row['split_badge'],
                            split_row['unit_price'],
                            split_row['planned_quantity'],
                            split_row['planned_cif_fc'],
                        ))
    return rows



logger = logging.getLogger(__name__)


class ItemPivotReportView(APIView):
    """
    Report showing licenses with items as columns (pivot format).

    GET parameters:
        - format: 'json' or 'excel' (default: json)
        - days: Number of days to look back (default: 30)
    """
    permission_classes = [ReportPermission]
    # Without this, DRF's own content negotiation intercepts ?format=excel
    # before get() ever runs and raises Http404 (rest_framework.negotiation.
    # DefaultContentNegotiation.filter_renderers has no renderer whose
    # `.format` is 'excel', since DEFAULT_RENDERER_CLASSES only registers
    # JSONRenderer) — i.e. the Export button 404s outright. Registering the
    # passthrough renderer (same fix already applied to ItemReportView) makes
    # 'excel' a recognised format so content negotiation succeeds; the view
    # still returns a plain StreamingHttpResponse for excel, so this renderer
    # is never actually invoked.
    renderer_classes = [JSONRenderer, _ExcelPassthroughRenderer]

    def get(self, request, *args, **kwargs):
        output_format = request.GET.get('format', 'json').lower()
        days = _safe_int(request.GET.get('days'), 30)
        sion_norm = request.GET.get('sion_norm')
        company_ids = request.GET.get('company_ids')  # Comma-separated company IDs
        exclude_company_ids = request.GET.get('exclude_company_ids')  # Comma-separated company IDs to exclude
        min_balance = _safe_int(request.GET.get('min_balance'), 200)
        license_status = request.GET.get('license_status', 'active')
        expiry_date_from = request.GET.get('expiry_date_from')  # YYYY-MM-DD
        expiry_date_to = request.GET.get('expiry_date_to')      # YYYY-MM-DD
        purchase_status = request.GET.get('purchase_status')    # Comma-separated codes

        # For Excel export, use streaming approach to avoid timeout
        if output_format == 'excel':
            try:
                return self.export_to_excel_streaming(days, sion_norm, company_ids, exclude_company_ids, min_balance, license_status, expiry_date_from, expiry_date_to, purchase_status)
            except Exception as e:
                logger.exception("Error exporting item pivot report to Excel")
                return JsonResponse({
                    'error': str(e)
                }, status=500)

        # For JSON, generate full report
        try:
            report_data = self.generate_report(days, sion_norm, company_ids, exclude_company_ids, min_balance,
                                               license_status, expiry_date_from, expiry_date_to, purchase_status)
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)

        return JsonResponse(report_data, safe=False)

    def generate_report(self, days: int = 30, sion_norm: str = None,
                        company_ids: str = None, exclude_company_ids: str = None,
                        min_balance: int = 200, license_status: str = 'active',
                        expiry_date_from: str = None, expiry_date_to: str = None,
                        purchase_status: str = None) -> Dict[str, Any]:
        """
        Generate item-wise pivot report.

        Args:
            days: Number of days to look back for active licenses
            sion_norm: Filter by specific SION norm class (optional)
            company_ids: Comma-separated company IDs to include (optional)
            exclude_company_ids: Comma-separated company IDs to exclude (optional)
            min_balance: Minimum balance CIF to include (default 200)
            license_status: Filter by status - 'active', 'expired', 'expiring_soon', 'all' (default 'active')

        Returns:
            Dictionary with report data
        """
        from datetime import date, timedelta
        today = date.today()

        # Base query - licenses with required purchase status.
        # The frontend sends the chosen codes as a comma-separated string;
        # when omitted, fall back to GE / MI / CO (Global Exim, MITC,
        # Conversion) which is the historical default for this report.
        if purchase_status:
            ps_codes = [c.strip() for c in purchase_status.split(',') if c.strip()]
        else:
            ps_codes = [GE, MI, CO]
        licenses = LicenseDetailsModel.objects.filter(
            purchase_status__code__in=ps_codes
        )

        # Apply license status filter
        if license_status == 'active':
            # Active: expiry date > today - 30 days (not expired more than 30 days ago)
            licenses = licenses.filter(
                flags__is_active=True,
                license_expiry_date__gt=today - timedelta(days=30)
            )
        elif license_status == 'expired':
            # Expired: expiry date < today (don't filter by is_active to include all expired licenses)
            licenses = licenses.filter(license_expiry_date__lt=today)
        elif license_status == 'expiring_soon':
            # Expiring soon: expiry within next 30 days
            licenses = licenses.filter(
                flags__is_active=True,
                license_expiry_date__gte=today,
                license_expiry_date__lte=today + timedelta(days=30)
            )
        # If 'all', no date or is_active filter applied - shows everything

        # Apply explicit expiry date range filter (overrides license_status date logic if provided)
        if expiry_date_from:
            from datetime import datetime as _dt
            licenses = licenses.filter(license_expiry_date__gte=_dt.strptime(expiry_date_from, '%Y-%m-%d').date())
        if expiry_date_to:
            from datetime import datetime as _dt
            licenses = licenses.filter(license_expiry_date__lte=_dt.strptime(expiry_date_to, '%Y-%m-%d').date())

        # Filter by SION norm if specified (optional)
        if sion_norm:
            licenses = licenses.filter(export_license__norm_class__norm_class=sion_norm).distinct()

        # Filter by company IDs if specified
        if company_ids:
            company_id_list = [int(cid.strip()) for cid in company_ids.split(',') if cid.strip()]
            licenses = licenses.filter(exporter_id__in=company_id_list)

        # Exclude company IDs if specified
        if exclude_company_ids:
            exclude_id_list = [int(cid.strip()) for cid in exclude_company_ids.split(',') if cid.strip()]
            licenses = licenses.exclude(exporter_id__in=exclude_id_list)

        # Filter by min_balance at database level using stored balance_cif field
        # This dramatically reduces the number of licenses we need to process
        licenses = licenses.filter(balance__balance_cif__gte=min_balance)

        # Build filtered prefetch querysets based on sion_norm
        import_items_qs = LicenseImportItemsModel.objects.select_related('hs_code')
        export_items_qs = LicenseExportItemModel.objects.select_related('norm_class')
        # NOTE: do NOT filter is_active here. Inactive item names are still needed
        # so a manually-planned-but-inactive item (e.g. WALNUT planned on an E132
        # DFIA) gets its quantities into item_quantities and renders. The column
        # set (all_items) still hides inactive items via its own is_active check,
        # except those that are explicitly planned (added back below).
        item_names_qs = ItemNameModel.objects.select_related('sion_norm_class')

        # If sion_norm specified, filter prefetch queries to only that norm.
        if sion_norm:
            item_names_qs = item_names_qs.filter(sion_norm_class__norm_class=sion_norm)
            export_items_qs = export_items_qs.filter(norm_class__norm_class=sion_norm)

        # Optimize with select_related and prefetch_related to reduce queries.
        # balance_cif / balance_report_notes / condition_sheet / current_owner now
        # live on OneToOne sub-tables (LicenseBalance / LicenseNotes / LicenseOwnership);
        # they're accessed via @property shims, so pull the sub-rows in one shot.
        licenses = licenses.select_related(
            'exporter',
            'port',
            'balance',
            'notes',
            'ownership__current_owner',
            'purchase_status',
            'notification_number',   # Fix #6: was lazy-loaded per licence (~2×N extra queries)
        ).prefetch_related(
            Prefetch('import_license',
                     queryset=import_items_qs.prefetch_related(
                         Prefetch('items', queryset=item_names_qs)
                     ).only('id', 'license_id', 'hs_code_id', 'quantity', 'allotted_quantity',
                            'debited_quantity', 'available_quantity', 'debited_value', 'cif_fc', 'description',
                            'condition_type', 'serial_number')),
            Prefetch('export_license',
                     queryset=export_items_qs.only('id', 'license_id', 'norm_class_id', 'cif_fc')),
            'license_documents',
            # Fix #5: ordered Prefetch avoids re-querying with ORDER BY inside _build_license_row
            Prefetch('transfers',
                     queryset=LicenseTransferModel.objects.order_by('-transfer_date', '-id'),
                     to_attr='transfers_ordered'),
        ).order_by('license_expiry_date', 'license_date')

        # Collect all unique items across all licenses
        # Use list() with prefetch_related for optimal performance (iterator breaks prefetch)
        all_items = {}  # Changed to dict to store item object for sorting
        valid_licenses = list(licenses)  # Licenses already filtered by balance_cif at DB level

        # Norm-specific plannable-item allow-sets: only item names the auto-planner
        # can generate are shown as columns.  Items outside the set (e.g.
        # ESSENTIAL OIL - E1, SUGAR - E1) are silently excluded from the column
        # build here.  The _missing_planned block further below re-admits any item
        # that a user actually planned even if it is outside the allow-set, so
        # intentional manual plans are always surfaced.
        from apps.license.services.e1_auto_plan import E1_PLANNABLE_NAMES as _E1_PLANNABLE
        _NORM_ALLOW: dict[str, frozenset[str]] = {
            'E1': _E1_PLANNABLE,
        }

        for license_obj in valid_licenses:
            for import_item in license_obj.import_license.all():
                for item in import_item.items.all():
                    # Only add items with valid names and that are active (is_active=False hides from pivot)
                    if item and item.name and item.is_active:
                        # Per-norm plannable filter: skip items whose names the
                        # auto-planner for their norm never generates.  This is
                        # keyed off the item's own norm code so it works correctly
                        # regardless of the ?sion_norm request parameter.
                        _item_norm = (item.sion_norm_class.norm_class
                                      if item.sion_norm_class else None)
                        _allow = _NORM_ALLOW.get(_item_norm)
                        if _allow is not None and item.name not in _allow:
                            continue
                        # If filtering by norm, only include items matching that norm
                        if sion_norm:
                            if item.sion_norm_class and item.sion_norm_class.norm_class == sion_norm:
                                all_items[item.id] = item
                        else:
                            all_items[item.id] = item

        # ── "As per planning" per-DFIA item map ────────────────────────────
        # When a DFIA carries a manual utilization plan (LicenseItemPlan), the
        # pivot must show that licence's items *as planned* rather than every
        # import item present on the licence: each row shows only the items that
        # DFIA actually planned, blanking the rest (e.g. BORAX is hidden on the
        # A3627 DFIAs that did not plan it, but still shown on those that did).
        #
        # Licences with NO manual plan are left untouched — they show all their
        # import items as before — so norm-driven norms (E1 / E5 / E132) are
        # unaffected. Column headers remain the union across the report; the
        # filtering is per row/cell in _build_license_row().
        from apps.license.models import LicenseItemPlan

        # import_item_id -> first attached item id, mirroring how a plan's
        # totals are attributed to a single item name in _build_license_row().
        first_item_of_import = {}
        for _lo in valid_licenses:
            for _ii in _lo.import_license.all():
                for _it in _ii.items.all():
                    first_item_of_import[_ii.id] = _it.id
                    break

        # Import-reference helpers (use already-prefetched data — no extra queries).
        # item_name_str_by_id: ItemNameModel.id → name str
        # import_qty_by_import_item: LicenseImportItemsModel.id → quantity
        item_name_str_by_id = {}
        import_qty_by_import_item = {}
        # import_item_ledger_by_id: LicenseImportItemsModel.id -> that ONE
        # import item's own HSN/description/ledger quantities. Built from
        # data already prefetched above (no extra queries) — this is the
        # single source of truth a planned cell's HSN/Description/Total/
        # Allotted/Debited/Balance must come from, never a cross-item merge.
        import_item_ledger_by_id = {}
        for _lo in valid_licenses:
            for _ii in _lo.import_license.all():
                import_qty_by_import_item[_ii.id] = _ii.quantity
                import_item_ledger_by_id[_ii.id] = {
                    'hs_code': _ii.hs_code.hs_code if _ii.hs_code else '',
                    'description': _ii.description or '',
                    'quantity': float(_ii.quantity or 0),
                    'allotted_quantity': float(_ii.allotted_quantity or 0),
                    'debited_quantity': float(_ii.debited_quantity or 0),
                    'available_quantity': float(_ii.available_quantity or 0),
                }
                for _it in _ii.items.all():
                    if _it.id not in item_name_str_by_id:
                        item_name_str_by_id[_it.id] = _it.name

        # license_id -> {item_id: {'q': planned qty, 'cif': planned CIF-FC}}.
        # Attributed to the plan LINE's own item_name (not the import item's
        # first attached name) so e.g. a RUTILE plan line on a BORAX+RUTILE
        # import item lands on RUTILE. Untagged split lines fall back to the
        # import item's first attached name. The key set doubles as "which
        # items this DFIA planned" for the per-row filter below.
        plan_totals_by_license = defaultdict(
            lambda: defaultdict(lambda: {'q': Decimal('0.000'), 'cif': Decimal('0.00')})
        )
        planned_item_ids_all = set()
        # Same query already fetches every LicenseItemPlan row for this page —
        # also keep the raw per-line split (item_name/qty/price/cif) so a
        # "Planning Splits" sheet can be rendered later without a second
        # query. item_name is resolved (id -> name string) once all_items is
        # fully populated, below.
        for _pl in (LicenseItemPlan.objects
                    .filter(license_id__in=[_lo.id for _lo in valid_licenses])
                    .values('license_id', 'import_item_id', 'item_name_id',
                            'planned_quantity', 'unit_price', 'planned_cif_fc')):
            _iname = _pl['item_name_id'] or first_item_of_import.get(_pl['import_item_id'])
            if _iname is None:
                continue
            _cell = plan_totals_by_license[_pl['license_id']][_iname]
            _cell['q'] += _pl['planned_quantity'] or Decimal('0')
            _cell['cif'] += _pl['planned_cif_fc'] or Decimal('0')
            _cell.setdefault('splits', []).append({
                # Resolved to a name string in the pass below, once all_items
                # (including manually-planned-but-inactive items) is complete.
                '_item_name_id': _pl['item_name_id'],
                'planned_quantity': float(_pl['planned_quantity'] or 0),
                'unit_price': float(_pl['unit_price'] or 0),
                'planned_cif_fc': float(_pl['planned_cif_fc'] or 0),
            })
            planned_item_ids_all.add(_iname)
            # Track import reference once per planned-item cell (first plan line wins).
            if 'import_item_name' not in _cell:
                _imp_name_id = first_item_of_import.get(_pl['import_item_id'])
                _cell['import_item_name'] = (
                    item_name_str_by_id.get(_imp_name_id, '') if _imp_name_id else ''
                )
                _cell['import_quantity'] = float(
                    import_qty_by_import_item.get(_pl['import_item_id']) or 0
                )

            # Verification data: the EXACT import item(s) this cell's plan
            # lines actually reference, keyed by import_item id so distinct
            # import items sharing this item-name are never merged into one
            # ledger record. Each entry's HSN/Description/ledger quantities
            # come straight from that ONE import item (import_item_ledger_by_id
            # above) — never summed or "first wins" across import items.
            _iid = _pl['import_item_id']
            _planned_items = _cell.setdefault('planned_import_items', {})
            if _iid not in _planned_items:
                _ledger = import_item_ledger_by_id.get(_iid, {})
                _planned_items[_iid] = {
                    'import_item_id': _iid,
                    'hs_code': _ledger.get('hs_code', ''),
                    'description': _ledger.get('description', ''),
                    'quantity': _ledger.get('quantity', 0.0),
                    'allotted_quantity': _ledger.get('allotted_quantity', 0.0),
                    'debited_quantity': _ledger.get('debited_quantity', 0.0),
                    'available_quantity': _ledger.get('available_quantity', 0.0),
                    'planned_quantity': Decimal('0'),
                    'planned_cif_fc': Decimal('0'),
                }
            # Multiple plan lines can reference the same import item (e.g. a
            # milk item's DWP + SWP split) — sum ONLY the planned qty/CIF,
            # never the ledger fields, which belong to the item itself.
            _planned_items[_iid]['planned_quantity'] += _pl['planned_quantity'] or Decimal('0')
            _planned_items[_iid]['planned_cif_fc'] += _pl['planned_cif_fc'] or Decimal('0')

        # A manually-planned item must appear as a column even if it is INACTIVE
        # in the master (is_active=False) — the user explicitly planned it, so it
        # would otherwise vanish (the column builder above skips inactive items).
        # Add any planned item ids missing from all_items, honouring the norm filter.
        _missing_planned = [iid for iid in planned_item_ids_all if iid not in all_items]
        if _missing_planned:
            for _it in ItemNameModel.objects.filter(id__in=_missing_planned).select_related('sion_norm_class'):
                if not _it.name:
                    continue
                if sion_norm and not (_it.sion_norm_class and _it.sion_norm_class.norm_class == sion_norm):
                    continue
                all_items[_it.id] = _it

        # Resolve each split's own item_name tag (id -> name string) now that
        # all_items includes every planned item, even inactive/filtered ones
        # not directly M2M-attached to their import item. No new query — just
        # a dict merge of data already fetched above. Falls back to None
        # (→ "Split N" badge, same as plan_reporting._build_map) for the rare
        # id that still isn't resolvable (e.g. filtered out by sion_norm).
        _split_item_name_lookup = dict(item_name_str_by_id)
        _split_item_name_lookup.update({iid: (it.name or '') for iid, it in all_items.items()})
        for _lic_cells in plan_totals_by_license.values():
            for _cell in _lic_cells.values():
                for _sp in _cell.get('splits', []):
                    _nid = _sp.pop('_item_name_id', None)
                    _sp['item_name'] = _split_item_name_lookup.get(_nid) if _nid is not None else None

                # planned_import_items: dict-of-dicts (keyed by import_item id,
                # for O(1) accumulation above) -> plain list, Decimal -> float.
                # This is the per-cell verification data: each entry is a real
                # import item a plan line referenced. Distinct import items are
                # then consolidated by `merge_planned_import_items` (HSN +
                # normalized description) purely for readability — e.g. three
                # "PP - E1" import items that are really one physical product
                # split across serial numbers render as one row instead of
                # three; import items that are genuinely different products
                # (different HSN or description) are never merged.
                _pit_map = _cell.get('planned_import_items')
                if _pit_map:
                    _cell['planned_import_items'] = _merge_planned_import_items(
                        {
                            **_pit,
                            'planned_quantity': float(_pit['planned_quantity']),
                            'planned_cif_fc': float(_pit['planned_cif_fc']),
                        }
                        for _pit in _pit_map.values()
                    )
                else:
                    _cell['planned_import_items'] = []

        # Sort items by display_order first, then by name for consistent column order
        sorted_items = sorted(
            [(item.id, item.name) for item in all_items.values()],
            key=lambda x: (all_items[x[0]].display_order, x[1] or '')
        )

        # Batch document-type lookups: one query for the whole page instead of two
        # .exists() calls per licence inside _build_license_row (an N+1 over the report).
        from apps.license.models import LicenseDocumentModel
        doc_types_by_license = defaultdict(set)
        for _lid, _dt in (LicenseDocumentModel.objects
                          .filter(license_id__in=[_lo.id for _lo in valid_licenses])
                          .values_list('license_id', 'type')):
            doc_types_by_license[_lid].add(_dt)

        # Batch the condition-pool computation (was ~13 queries PER licence inside
        # _build_license_row — the report's dominant N+1) into a handful of queries.
        from apps.license.services.condition_pool import compute_condition_pools_bulk
        cond_pools_by_license = compute_condition_pools_bulk([_lo.id for _lo in valid_licenses])

        # Fix #2: batch AllotmentItems query — was 1 query PER licence (N+1).
        # One query returns cif_fc sums keyed by license_id. NOTE: this is a
        # DFIA-specific "Alloted CIF" column (`is_allotted=True` is the DFIA
        # allotment flag, see AllotmentModel's help_text) — a narrower,
        # separate metric from the shared Balance CIF's allotment component
        # (`LicenseBalanceCalculator.calculate_allotment`), not a duplicate
        # of it; left as-is.
        from apps.allotment.models import AllotmentItems as _AllotmentItems
        _license_ids = [_lo.id for _lo in valid_licenses]
        alloted_cif_by_license: dict = defaultdict(lambda: Decimal('0'))
        for _row in _AllotmentItems.objects.filter(
            item__license_id__in=_license_ids,
            allotment__is_allotted=True,
            allotment__bill_of_entry__isnull=True,
        ).values('item__license_id', 'cif_fc'):
            if _row['cif_fc'] is not None:
                alloted_cif_by_license[_row['item__license_id']] += Decimal(str(_row['cif_fc']))

        # Balance CIF must be identical everywhere in the app — read LIVE via
        # the shared `LicenseBalanceCalculator` (same batched method the
        # License List view uses), never the denormalized `balance_cif`
        # column directly, which is only refreshed by a background task/
        # manual trigger and can go stale relative to the live calculation
        # (e.g. right after a Balance Engine formula change, or any edit
        # that doesn't happen to fire a recalculation signal).
        from apps.license.services.balance_calculator import LicenseBalanceCalculator
        # Financial Ledger formula -- see `LicenseDetailsModel.
        # get_balance_cif`'s docstring; must match every other "Balance
        # CIF" in the app.
        live_balance_by_license = LicenseBalanceCalculator.calculate_financial_balance_for_licenses(_license_ids)

        # Build license data with item columns, grouped by norm first, then notification
        # (defaultdict is imported at module level).
        licenses_by_norm_notification = defaultdict(lambda: defaultdict(list))

        for license_obj in valid_licenses:
            license_row = self._build_license_row(
                license_obj, sorted_items,
                item_plan_totals=plan_totals_by_license.get(license_obj.id),
                document_types=doc_types_by_license.get(license_obj.id, frozenset()),
                condition_pools=cond_pools_by_license.get(license_obj.id, {}),
                alloted_cif=alloted_cif_by_license.get(license_obj.id, Decimal('0')),
                balance_cif=live_balance_by_license.get(license_obj.id, Decimal('0')),
            )

            if license_row:
                # Handle blank/empty notification numbers
                notification = (license_obj.notification_number.code if license_obj.notification_number_id else '').strip()
                if not notification:
                    notification = 'Unknown'

                # Fix #4: use prefetch cache — .exists()/.first() bypass it and re-query
                norm_class = 'Unknown'
                _exports = list(license_obj.export_license.all())
                if _exports and _exports[0].norm_class:
                    norm_class = _exports[0].norm_class.norm_class or 'Unknown'

                # Define conversion norms
                conversion_norms = ['E1', 'E5', 'E126', 'E132']
                is_conversion = license_obj.purchase_status and license_obj.purchase_status.code == CO

                # Get exporter name for split sheet logic
                exporter_name = (license_obj.exporter.name or '') if license_obj.exporter else ''
                exporter_name_upper = exporter_name.upper()

                # Determine exporter category for split sheets
                exporter_category = None
                if 'PARLE' in exporter_name_upper:
                    exporter_category = 'Parle'
                elif 'HALDIRAM SNACKS' in exporter_name_upper:
                    exporter_category = 'Haldiram Snacks'
                elif 'HALDIRAM FOODS' in exporter_name_upper:
                    exporter_category = 'Haldiram Foods'
                elif 'HARIOMKAR FOOD' in exporter_name_upper:
                    exporter_category = 'Hariomkar Food'

                # Build notification key based on norm class and purchase status
                if norm_class in conversion_norms and is_conversion:
                    # For conversion licenses in E1, E5, E126, E132
                    if norm_class in ['E5', 'E132']:
                        # E5 and E132 Conversion: split by exporter category
                        if exporter_category:
                            notification_key = f"{notification} - Conversion - {exporter_category}"
                        else:
                            notification_key = f"{notification} - Conversion"
                    else:
                        # E1, E126 Conversion
                        notification_key = f"{notification} - Conversion"

                elif norm_class in ['E5', 'E132']:
                    # E5 and E132 non-conversion: split by exporter category
                    if exporter_category:
                        notification_key = f"{notification} - {exporter_category}"
                    else:
                        notification_key = f"{notification} - Others"

                else:
                    # Regular grouping by notification for other norms
                    notification_key = notification

                # Split every pivot table by PURCHASE STATUS: prefix the group
                # key with the licence's purchase-status label so each rendered
                # table (and its summary / totals / Excel sheet, which all key off
                # this group) contains a single purchase status. The " — " (em
                # dash) delimiter is distinct from the " - " used inside
                # notification_key, so the frontend can split it back apart.
                ps_label = (license_row.get('purchase_status_label')
                            or license_row.get('purchase_status_code') or 'Unknown')
                notification_key = f"{ps_label} — {notification_key}"

                licenses_by_norm_notification[norm_class][notification_key].append(license_row)

        # Determine which items have restrictions
        items_with_restrictions = set()
        for norm_dict in licenses_by_norm_notification.values():
            for licenses_list in norm_dict.values():
                for license_row in licenses_list:
                    for item_id, item_name in sorted_items:
                        item_data = license_row.get('items', {}).get(item_name, {})
                        if item_data.get('restriction') is not None:
                            items_with_restrictions.add(item_id)

        # Convert nested defaultdict to regular dict
        result_dict = {}
        for norm, notification_dict in licenses_by_norm_notification.items():
            result_dict[norm] = dict(notification_dict)

        # Grand totals per (norm, notification) group — the single backend-
        # owned computation of what used to be independently re-summed by
        # both the React page and the Excel exporter's totals row (Phase
        # 2B.2A; see docs/architecture/ITEM_PIVOT_DISPLAY_DATASET_DESIGN.md).
        # Additive top-level key — `licenses_by_norm_notification`'s existing
        # shape is unchanged, no API field renamed.
        # Notification/Norm Summary — the backend-owned translation of the
        # frontend's `calculateNotificationSummary` (Phase 2B.2B; see
        # docs/architecture/ITEM_PIVOT_NOTIFICATION_SUMMARY_DESIGN.md).
        # Built in the same pass over the same `licenses_list`s as
        # `notification_totals` above — no second full iteration needed.
        # Purely additive: nothing in `ItemPivotReport.tsx` reads these keys
        # yet (that is a later, separate phase per the design doc's §7).
        notification_totals = {}
        notification_summary = {}
        norm_summary = {}
        for norm, notification_dict in result_dict.items():
            notification_totals[norm] = {}
            notification_summary[norm] = {}
            for notification_key, licenses_list in notification_dict.items():
                item_totals = {}
                for item_id, item_name in sorted_items:
                    item_totals[item_name] = {
                        'quantity': sum(
                            lic['items'].get(item_name, {}).get('quantity', 0) or 0
                            for lic in licenses_list
                        ),
                        'allotted_quantity': sum(
                            lic['items'].get(item_name, {}).get('allotted_quantity', 0) or 0
                            for lic in licenses_list
                        ),
                        'debited_quantity': sum(
                            lic['items'].get(item_name, {}).get('debited_quantity', 0) or 0
                            for lic in licenses_list
                        ),
                        'available_quantity': sum(
                            lic['items'].get(item_name, {}).get('available_quantity', 0) or 0
                            for lic in licenses_list
                        ),
                        'restriction_value': sum(
                            lic['items'].get(item_name, {}).get('restriction_value', 0) or 0
                            for lic in licenses_list
                        ),
                        # Literal sum of manual plan_quantity only — matches
                        # the on-screen/Excel `totalPlanQty` convention: rows
                        # with no manual plan (norm-derived, shown as a
                        # unit-price rate per-row) contribute 0 here rather
                        # than being folded into a blended rate.
                        'plan_quantity': sum(
                            lic['items'].get(item_name, {}).get('plan_quantity', 0) or 0
                            for lic in licenses_list
                        ),
                        # Manual-vs-norm selection already resolved per cell
                        # by _effective_planned_cif — just sum it here.
                        'effective_planned_cif': sum(
                            lic['items'].get(item_name, {}).get('effective_planned_cif', 0) or 0
                            for lic in licenses_list
                        ),
                    }
                notification_totals[norm][notification_key] = {
                    'total_cif': sum(lic['total_cif'] for lic in licenses_list),
                    'debited_cif': sum(lic.get('debited_cif', 0) for lic in licenses_list),
                    'alloted_cif': sum(lic['alloted_cif'] for lic in licenses_list),
                    'balance_cif': sum(lic['balance_cif'] for lic in licenses_list),
                    # Grand Planned CIF across every item column — equals
                    # summing `total_effective_planned_cif` across licenses,
                    # or summing `effective_planned_cif` across items; kept
                    # as its own field so consumers never need to reduce
                    # either axis themselves.
                    'total_effective_planned_cif': sum(
                        lic.get('total_effective_planned_cif', 0) for lic in licenses_list
                    ),
                    'items': item_totals,
                }

                notification_summary[norm][notification_key] = _build_notification_summary(
                    licenses_list, sorted_items,
                )

            # Norm-level summary: same builder, flattened across every
            # notification group under this norm — mirrors the frontend's
            # `:1531` call site (`Object.values(...).flat()`).
            norm_summary[norm] = _build_notification_summary(
                [lic for licenses_list in notification_dict.values() for lic in licenses_list],
                sorted_items,
            )

        # Fetch notes and conditions for all norms in a single query
        from apps.core.models import SionNormClassModel
        norm_classes_list = list(result_dict.keys())
        sion_norms = SionNormClassModel.objects.filter(
            norm_class__in=norm_classes_list
        ).prefetch_related('notes', 'conditions')

        # Build dict from fetched norms
        norm_notes_conditions = {}
        sion_norms_dict = {sn.norm_class: sn for sn in sion_norms}

        for norm_class in norm_classes_list:
            if norm_class in sion_norms_dict:
                sion_norm = sion_norms_dict[norm_class]
                norm_notes_conditions[norm_class] = {
                    'notes': [
                        {'note_text': note.note_text, 'display_order': note.display_order}
                        for note in sion_norm.notes.all()
                    ],
                    'conditions': [
                        {'condition_text': cond.condition_text, 'display_order': cond.display_order}
                        for cond in sion_norm.conditions.all()
                    ]
                }
            else:
                norm_notes_conditions[norm_class] = {'notes': [], 'conditions': []}

        return {
            'items': [
                {
                    'id': item_id,
                    'name': item_name,
                    'has_restriction': item_id in items_with_restrictions
                }
                for item_id, item_name in sorted_items
            ],
            'licenses_by_norm_notification': result_dict,
            'notification_totals': notification_totals,
            'notification_summary': notification_summary,
            'norm_summary': norm_summary,
            'norm_notes_conditions': norm_notes_conditions,
            'report_date': today.isoformat(),
        }

    def _build_license_row(self, license_obj: LicenseDetailsModel, all_items: List[tuple],
                           item_plan_totals=None, document_types=None,
                           condition_pools=None, alloted_cif=None, balance_cif=None) -> Dict[str, Any]:
        """
        Build a single license row with item columns.

        Args:
            license_obj: LicenseDetailsModel instance
            all_items: List of (item_id, item_name) tuples
            item_plan_totals: When the DFIA is manually planned, a map
                {item_id: {'q': planned qty, 'cif': planned CIF-FC}} of the items
                it actually planned. Drives the per-cell Planned QTY / Planned
                CIF and the "as per planning" filter: items outside this map are
                emitted as empty cells. None => not manually planned, so every
                import item is shown as before.

        Returns:
            Dictionary with license data and item quantities
        """
        # Calculate total CIF from export license items (already prefetched)
        # Start with Decimal('0') to ensure result is always Decimal type
        total_cif = Decimal('0')
        for item in license_obj.export_license.all():
            # Convert to Decimal to handle cases where database returns float
            cif_value = Decimal(str(item.cif_fc)) if item.cif_fc is not None else Decimal('0')
            total_cif += cif_value

        # Fix #2: alloted_cif pre-computed by generate_report in a single batch query.
        # Standalone callers (e.g. Excel export called directly) fall back to the
        # per-licence query so behaviour is unchanged outside the main report path.
        if alloted_cif is None:
            from apps.allotment.models import AllotmentItems as _AI
            _alloted_cif = Decimal('0')
            for _ai in _AI.objects.filter(
                item__license=license_obj,
                allotment__is_allotted=True,
                allotment__bill_of_entry__isnull=True,
            ).values_list('cif_fc', flat=True):
                if _ai is not None:
                    _alloted_cif += Decimal(str(_ai))
            alloted_cif = _alloted_cif

        # Debited CIF = CIF already debited (via BOE) across this licence's import
        # items — the same `debited_value` field the restriction pools treat as
        # debited_cif below. import_license is prefetched, so no extra query.
        debited_cif = Decimal('0')
        for import_item in license_obj.import_license.all():
            debited_cif += Decimal(str(import_item.debited_value)) if import_item.debited_value is not None else DEC_0

        # Aggregate quantities by item (sum across all serial numbers)
        item_quantities = defaultdict(lambda: {
            'quantity': Decimal('0.000'),
            'allotted_quantity': Decimal('0.000'),
            'debited_quantity': Decimal('0.000'),
            'available_quantity': Decimal('0.000'),
            'debited_value': Decimal('0.00'),
            'cif_value': Decimal('0.00'),
            'hs_code': '',
            'description': '',
            'sion_norm_class': None,
            'restriction_percentage': None,
            'condition_type': '',
            # User-authored utilization plan (summed across an item's splits).
            'plan_quantity': Decimal('0.000'),
            'plan_cif': Decimal('0.00'),
        })

        # Fix #1: plan_map_for_license was called per-licence here (N extra queries).
        # generate_report already batches all LicenseItemPlan rows and passes the
        # per-licence totals via item_plan_totals, so this extra call is redundant.
        # Standalone callers that pass item_plan_totals=None get _plan_map=None (same
        # as before — plan_source falls back to 'norm', which is correct).

        # Per condition_type pool — new restriction model. Each "N%" pool is
        # shared by every import item on this licence with that condition_type,
        # and provides the per-cell `restriction_value` shown in the pivot.
        # Use the report's batched pools when provided; else compute per-licence
        # (standalone callers).
        if condition_pools is None:
            from apps.license.services.condition_pool import compute_condition_pools
            condition_pools = compute_condition_pools(license_obj)
        # condition_pools = {"2%": Decimal(...), "3%": Decimal(...), ...}

        for import_item in license_obj.import_license.all():
            # Plan totals are now sourced per plan-line item_name via
            # `item_plan_totals` (see the item-columns loop below); the old
            # first-attached-name attribution is gone as it mis-assigned plans
            # on multi-name import items (e.g. RUTILE on a BORAX+RUTILE item).
            for item in import_item.items.all():
                # Convert all numeric fields to Decimal to handle potential float values from database
                item_quantities[item.id]['quantity'] += Decimal(str(import_item.quantity)) if import_item.quantity is not None else DEC_000
                item_quantities[item.id]['allotted_quantity'] += Decimal(str(import_item.allotted_quantity)) if import_item.allotted_quantity is not None else DEC_000
                item_quantities[item.id]['debited_quantity'] += Decimal(str(import_item.debited_quantity)) if import_item.debited_quantity is not None else DEC_000
                item_quantities[item.id]['available_quantity'] += Decimal(str(import_item.available_quantity)) if import_item.available_quantity is not None else DEC_000
                item_quantities[item.id]['debited_value'] += Decimal(str(import_item.debited_value)) if import_item.debited_value is not None else DEC_0
                item_quantities[item.id]['cif_value'] += Decimal(str(import_item.cif_fc)) if import_item.cif_fc is not None else DEC_0

                if import_item.hs_code and not item_quantities[item.id]['hs_code']:
                    item_quantities[item.id]['hs_code'] = import_item.hs_code.hs_code

                if import_item.description and not item_quantities[item.id]['description']:
                    item_quantities[item.id]['description'] = import_item.description

                # Carry the licence-condition badge through to the pivot cell.
                # If multiple import-item rows map to the same item-name, the
                # first non-empty condition wins (typical case: each item-name
                # appears on one serial number per licence).
                if import_item.condition_type and not item_quantities[item.id]['condition_type']:
                    item_quantities[item.id]['condition_type'] = import_item.condition_type

                # Get restriction from item's sion_norm_class and restriction_percentage
                if item and hasattr(item, 'sion_norm_class') and item.sion_norm_class:
                    sion_norm = item.sion_norm_class.norm_class
                    restriction_pct = item.restriction_percentage

                    item_quantities[item.id]['sion_norm_class'] = sion_norm
                    item_quantities[item.id]['restriction_percentage'] = restriction_pct

        # `balance_cif` is pre-computed LIVE by `generate_report` in a single
        # batched `calculate_financial_balance_for_licenses` call (the
        # Financial Ledger formula, same shared Balance Engine used
        # everywhere else — see `LicenseDetailsModel.get_balance_cif`'s
        # docstring) and passed in here — never the denormalized
        # `license_obj.balance_cif` column directly, which is only
        # refreshed by a background task/manual trigger and can go stale.
        # Standalone callers (balance_cif=None) fall back to a live
        # per-licence calculation so behaviour is unchanged outside the main
        # report path — see `alloted_cif`'s identical fallback above.
        if balance_cif is None:
            from apps.license.services.balance_calculator import LicenseBalanceCalculator
            balance_cif = LicenseBalanceCalculator.calculate_financial_balance(license_obj)

        # Build row data
        # Handle blank/empty notification numbers
        notification_display = (license_obj.notification_number.code if license_obj.notification_number_id else '').strip()
        if not notification_display:
            notification_display = 'Unknown'

        # Check for document types — use the batched map when the report supplies it,
        # else fall back to a per-object query (standalone callers).
        if document_types is None:
            document_types = {d.type for d in license_obj.license_documents.all()}
        has_tl = 'TRANSFER LETTER' in document_types
        has_copy = 'LICENSE COPY' in document_types

        # Fix #5: use pre-ordered to_attr list — avoids re-querying with ORDER BY
        latest_transfer_text = ''
        _transfers = getattr(license_obj, 'transfers_ordered', None)
        if _transfers is None:
            # Standalone caller (no Prefetch to_attr) — fall back to queryset
            _transfers = list(license_obj.transfers.order_by('-transfer_date', '-id'))
        if _transfers:
            latest_transfer_text = str(_transfers[0])
        elif license_obj.current_owner:
            latest_transfer_text = f"Current Owner is {license_obj.current_owner.name}"
        else:
            latest_transfer_text = "Data Not Found"

        # Purchase Status — emitted so the frontend can colour-code each row.
        ps_code  = ''
        ps_label = ''
        if license_obj.purchase_status_id:
            ps = license_obj.purchase_status
            ps_code  = ps.code or ''
            ps_label = ps.label or ''

        row_data = {
            'id': license_obj.id,
            'license_number': license_obj.license_number,
            'license_date': license_obj.license_date.isoformat() if license_obj.license_date else None,
            'license_expiry_date': license_obj.license_expiry_date.isoformat(),
            'ledger_date': license_obj.ledger_date.isoformat() if license_obj.ledger_date else None,
            'exporter': str(license_obj.exporter) if license_obj.exporter else '',
            'port': str(license_obj.port) if license_obj.port else '',
            'notification_number': notification_display,
            'purchase_status_code': ps_code,
            'purchase_status_label': ps_label,
            'total_cif': float(total_cif),
            'debited_cif': float(debited_cif),
            'alloted_cif': float(alloted_cif),
            'balance_cif': float(balance_cif),  # Reuse already calculated balance
            'balance_report_notes': license_obj.balance_report_notes or '',
            'condition_sheet': license_obj.condition_sheet or '',
            'latest_transfer': latest_transfer_text,
            'has_tl': has_tl,
            'has_copy': has_copy,
            # Per-license plan source: 'manual' if the license has any manual
            # plan line, else 'norm'. Uses the already-available item_plan_totals
            # so no extra query is needed (see Fix #1 above).
            'plan_source': 'manual' if item_plan_totals else 'norm',
            'items': {}
        }

        # Calculate unit price for RUTILE - A3627 (Balance CIF / Total Balance QTY of RUTILE)
        rutile_unit_price = None
        rutile_total_balance_qty = Decimal('0')

        # Sum up all RUTILE available quantities
        for item_id, item_name in all_items:
            if item_name == 'RUTILE - A3627' and item_id in item_quantities:
                rutile_total_balance_qty += item_quantities[item_id]['available_quantity']

        # Calculate unit price if we have RUTILE balance qty >= 10, otherwise set to 0
        if rutile_total_balance_qty >= 10:
            rutile_unit_price = float(balance_cif / rutile_total_balance_qty)
        elif rutile_total_balance_qty > 0:
            rutile_unit_price = 0.0

        # ── Per-item Unit Price + Planned CIF (E1 / E5 only) ───────────────
        # Run the same waterfall the bulk Balance Excel runs so the per-item
        # rows in the pivot match the per-category planner exactly. For each
        # item we classify it into a category, compute the category's
        # effective rate (planned_cif / util_qty), then allocate this item's
        # share of the category's planned CIF proportionally to its util qty.
        # Fix #4: use prefetch cache for export_license — .exists()/.first() bypass it
        primary_norm = ''
        _exp_list = list(license_obj.export_license.all())
        if _exp_list and _exp_list[0].norm_class:
            primary_norm = _exp_list[0].norm_class.norm_class or ''

        # `item_plan_data[item_name]` → {'planned_cif': float, 'unit_price': float}
        item_plan_data: Dict[str, Dict[str, float]] = {}
        if primary_norm == 'E1':
            from decimal import Decimal as _Decimal

            from apps.license.services.e1_auto_plan import STEP_ITEM_NAME as _E1_STEP_ITEM_NAME
            from apps.license.services.e1_plan import (
                E1Item as _E1Item, classify_e1_item as _classify, plan_e1_items as _plan_e1_items,
            )

            # Fix #3: reuse the already-prefetched import items instead of issuing
            # a second SELECT per E1 licence. Inactive items are included because
            # the prefetch query (import_items_qs) has no is_active filter.
            # Note: use ii.items.all() not .values_list() — .values_list() bypasses
            # the prefetch cache and would re-query; .all() reads from it for free.
            import_items = license_obj.import_license.all()
            item_ledger_by_id: Dict[int, dict] = {}
            for ii in import_items:
                item_ledger_by_id[ii.id] = {
                    'hs_code': ii.hs_code.hs_code if ii.hs_code else '',
                    'description': ii.description or '',
                    'quantity': float(ii.quantity or 0),
                    'allotted_quantity': float(ii.allotted_quantity or 0),
                    'debited_quantity': float(ii.debited_quantity or 0),
                    'available_quantity': float(_Decimal(str(ii.available_quantity or 0))),
                }

            # Classify and plan ONE merged group per physical product (same
            # HSN + normalized description) instead of once per raw import
            # item — see `merge_items_for_classification`'s docstring: a
            # per-serial classify can split one physical product into two
            # different categories/columns if its serials carry inconsistent
            # master-data tags, which no amount of post-hoc output merging
            # can undo (it only ever sees one category's bucket at a time).
            groups = _merge_items_for_classification(import_items)
            group_by_rep: Dict[int, dict] = {}
            e1_items: list = []
            for g in groups:
                names_text = ', '.join(g['item_names']) if g['item_names'] else (g['description'] or '-')
                cat = _classify(names_text, g['hs_code'], g['description'])
                if not cat:
                    continue
                group_by_rep[g['representative_id']] = g
                e1_items.append(_E1Item(key=g['representative_id'], category=cat, qty=g['available_quantity']))

            # Run the shared per-item engine — the same rules Auto-Plan and
            # norm_plan.py use, so this table (and its Excel export) never
            # drifts from what Auto-Plan would actually commit.
            plan_result = _plan_e1_items(e1_items, _Decimal(str(balance_cif)))

            # Attribute each planned line to the SAME item-name column
            # Auto-Plan would persist it under (STEP_ITEM_NAME) — NOT the
            # import item's own master-data tags (`ii.items.all()`). Several
            # E1 categories (e.g. "FRUIT/COCOA - E1", "EGG ALBUMIN"/"WPC - E1",
            # "PP - E1") are pure planner-output labels that a real import
            # item is rarely pre-tagged with in master data even when it WAS
            # the item the engine selected — attributing by the import item's
            # own tags left those columns' HSN/Description blank despite a
            # real planned CIF. This covers every category the engine can
            # produce (STEP_ITEM_NAME is keyed by all of them), not a
            # per-category patch.
            per_item_util: Dict[str, float] = {}
            per_item_cif: Dict[str, float] = {}
            per_item_planned_items: Dict[str, Dict[int, dict]] = {}
            for line in plan_result.lines:
                nm = _E1_STEP_ITEM_NAME.get(line.step)
                if not nm:
                    continue
                per_item_util[nm] = per_item_util.get(nm, 0.0) + float(line.planned_qty)
                per_item_cif[nm] = per_item_cif.get(nm, 0.0) + float(line.planned_cif)
                _bucket = per_item_planned_items.setdefault(nm, {})
                # `line.key` is the merged group's representative id — fan the
                # group's single planned qty/CIF back across every raw member,
                # proportional to each member's own available_quantity share
                # (the category rate is uniform across the group, so the
                # ratio cancels it out and every member ends up at the same
                # unit rate). This keeps `_merge_planned_import_items` below
                # working unchanged: every member of a physical-product group
                # is now guaranteed to land in this SAME `nm` bucket, so it
                # reliably collapses them into one display row with correct
                # per-member ledger fields.
                group = group_by_rep[line.key]
                members = group['member_ids']
                total_avail = float(group['available_quantity'])
                for _iid in members:
                    _ledger = item_ledger_by_id.get(_iid) or {}
                    share = (
                        _ledger.get('available_quantity', 0.0) / total_avail
                        if total_avail else 1.0 / len(members)
                    )
                    if _iid not in _bucket:
                        _bucket[_iid] = {
                            'import_item_id': _iid,
                            'hs_code': _ledger.get('hs_code', ''),
                            'description': _ledger.get('description', ''),
                            'quantity': _ledger.get('quantity', 0.0),
                            'allotted_quantity': _ledger.get('allotted_quantity', 0.0),
                            'debited_quantity': _ledger.get('debited_quantity', 0.0),
                            'available_quantity': _ledger.get('available_quantity', 0.0),
                            'planned_quantity': 0.0,
                            'planned_cif_fc': 0.0,
                        }
                    _bucket[_iid]['planned_quantity'] += float(line.planned_qty) * share
                    _bucket[_iid]['planned_cif_fc'] += float(line.planned_cif) * share

            for nm, uq in per_item_util.items():
                item_plan = per_item_cif.get(nm, 0.0)
                item_plan_data[nm] = {
                    'unit_price': round(item_plan / uq, 2) if uq else 0.0,
                    'planned_cif': round(item_plan, 2),
                    'planned_import_items': _merge_planned_import_items(
                        per_item_planned_items.get(nm, {}).values(),
                    ),
                }
        elif primary_norm == 'E5':
            from decimal import Decimal as _Decimal

            from apps.license.services.e5_auto_plan import STEP_ITEM_NAME as _E5_STEP_ITEM_NAME
            from apps.license.services.e5_plan import (
                E5Item as _E5Item, classify_e5_item as _classify, plan_e5_items as _plan_e5_items,
            )

            # Fix #3: reuse the already-prefetched import items (see E1 note above).
            import_items = license_obj.import_license.all()
            item_ledger_by_id: Dict[int, dict] = {}
            for ii in import_items:
                item_ledger_by_id[ii.id] = {
                    'hs_code': ii.hs_code.hs_code if ii.hs_code else '',
                    'description': ii.description or '',
                    'quantity': float(ii.quantity or 0),
                    'allotted_quantity': float(ii.allotted_quantity or 0),
                    'debited_quantity': float(ii.debited_quantity or 0),
                    'available_quantity': float(_Decimal(str(ii.available_quantity or 0))),
                }

            # Classify and plan ONE merged group per physical product — see
            # the matching E1 note above for why a per-serial classify can
            # split one product into two categories/columns.
            groups = _merge_items_for_classification(import_items)
            group_by_rep: Dict[int, dict] = {}
            e5_items: list = []
            for g in groups:
                names_text = ', '.join(g['item_names']) if g['item_names'] else (g['description'] or '-')
                cat = _classify(names_text, g['hs_code'], g['description'])
                if not cat:
                    continue
                group_by_rep[g['representative_id']] = g
                e5_items.append(_E5Item(key=g['representative_id'], category=cat, qty=g['available_quantity']))

            # Run the shared per-item engine — the same rules Auto-Plan and
            # norm_plan.py use, so this table (and its Excel export) never
            # drifts from what Auto-Plan would actually commit.
            plan_result = _plan_e5_items(e5_items, _Decimal(str(balance_cif)))

            # Attribute each planned line to the SAME item-name column
            # Auto-Plan would persist it under (STEP_ITEM_NAME) — NOT the
            # import item's own master-data tags. See the matching E1 note
            # above for why this matters for every category, not just the
            # ones already coincidentally pre-tagged in master data.
            per_item_util: Dict[str, float] = {}
            per_item_cif: Dict[str, float] = {}
            per_item_planned_items: Dict[str, Dict[int, dict]] = {}
            for line in plan_result.lines:
                nm = _E5_STEP_ITEM_NAME.get(line.step)
                if not nm:
                    continue
                per_item_util[nm] = per_item_util.get(nm, 0.0) + float(line.planned_qty)
                per_item_cif[nm] = per_item_cif.get(nm, 0.0) + float(line.planned_cif)
                _bucket = per_item_planned_items.setdefault(nm, {})
                # See the matching E1 note above: fan the merged group's one
                # planned line back across every raw member, proportional to
                # each member's available_quantity share.
                group = group_by_rep[line.key]
                members = group['member_ids']
                total_avail = float(group['available_quantity'])
                for _iid in members:
                    _ledger = item_ledger_by_id.get(_iid) or {}
                    share = (
                        _ledger.get('available_quantity', 0.0) / total_avail
                        if total_avail else 1.0 / len(members)
                    )
                    if _iid not in _bucket:
                        _bucket[_iid] = {
                            'import_item_id': _iid,
                            'hs_code': _ledger.get('hs_code', ''),
                            'description': _ledger.get('description', ''),
                            'quantity': _ledger.get('quantity', 0.0),
                            'allotted_quantity': _ledger.get('allotted_quantity', 0.0),
                            'debited_quantity': _ledger.get('debited_quantity', 0.0),
                            'available_quantity': _ledger.get('available_quantity', 0.0),
                            'planned_quantity': 0.0,
                            'planned_cif_fc': 0.0,
                        }
                    _bucket[_iid]['planned_quantity'] += float(line.planned_qty) * share
                    _bucket[_iid]['planned_cif_fc'] += float(line.planned_cif) * share

            for nm, uq in per_item_util.items():
                item_plan = per_item_cif.get(nm, 0.0)
                item_plan_data[nm] = {
                    'unit_price': round(item_plan / uq, 2) if uq else 0.0,
                    'planned_cif': round(item_plan, 2),
                    'planned_import_items': _merge_planned_import_items(
                        per_item_planned_items.get(nm, {}).values(),
                    ),
                }

        # ── Per-item classification plan (E132) ────────────────────────────
        # E132 planning is a deterministic classification (services/e132_plan.py):
        # each item is classified into one planning item and priced at that item's
        # fixed unit price. Unit Price / Planned CIF reuse the E1/E5 columns.
        # Keyed by item name (matching the downstream lookup).
        item_e132_data: Dict[str, Dict[str, Any]] = {}
        if primary_norm == 'E132':
            from apps.license.services.e132_plan import plan_e132_per_item
            _e132_input = []
            for _iid, _inm in all_items:
                if _iid in item_quantities:
                    _d132 = item_quantities[_iid]
                    _e132_input.append({
                        'record_id': _inm,
                        'quantity': float(_d132['available_quantity'] or 0),
                        'hs_code': _d132['hs_code'] or '',
                        'description': _d132['description'] or '',
                    })
            item_e132_data = plan_e132_per_item(_e132_input, float(balance_cif))

        # "As per planning" (AUTOMATED): for E132 the classification IS the plan —
        # only items that classified into a planning item are shown; unclassified
        # items are hidden. A manual plan, when present, takes precedence.
        e132_planned_names = None
        if primary_norm == 'E132' and item_plan_totals is None:
            e132_planned_names = set(item_e132_data.keys())

        # Add item columns
        planned_item_ids = set(item_plan_totals) if item_plan_totals is not None else None
        for item_id, item_name in all_items:
            # Per-product visibility — three-way priority:
            #
            # 1. Manual utilization plan (LicenseItemPlan rows) → show ONLY the
            #    explicitly planned items.  Norm-derived plan (E1/E5 category
            #    waterfall) must NOT cause un-planned import items (e.g. "Milk
            #    Powder") to bleed into a manually-planned license's row — that
            #    is the mixing bug this block fixes.
            #    Also: planned items may be ItemNameModel entries not linked to
            #    the import via M2M (e.g. "SWP - E1" planned from a "Milk Powder"
            #    import item) so the `item_id in item_quantities` guard is
            #    intentionally dropped; the cell falls back to zero import
            #    quantities for such items, which is correct.
            #
            # 2. E132 auto-classification (no manual plan) → show items that the
            #    classifier placed into a planning item; require item_quantities
            #    presence because the classifier works off import data.
            #
            # 3. No planning context → show all items that have import data,
            #    OR that the LIVE norm waterfall (item_plan_data, E1/E5) just
            #    planned a real CIF for — a purely planner-output item-name
            #    (e.g. "FRUIT/COCOA - E1") has no M2M-linked import item on
            #    this licence by construction, so `item_id in item_quantities`
            #    alone would hide a genuinely-planned cell entirely.
            _has_manual = planned_item_ids is not None and item_id in planned_item_ids
            _has_e132   = bool(item_e132_data.get(item_name))
            _has_live_plan = bool(item_plan_data.get(item_name))

            if item_plan_totals is not None:
                # Priority 1 — manual plan: show only planned items, no quantity guard.
                show_item = _has_manual
            elif e132_planned_names is not None:
                # Priority 2 — E132 auto-classification: show classified items only.
                show_item = _has_e132 and item_id in item_quantities
            else:
                # Priority 3 — no planning context: show items with import data
                # or a real live-computed plan.
                show_item = item_id in item_quantities or _has_live_plan

            if show_item:
                item_data = item_quantities[item_id]
                # Per-item manual plan totals (empty for norm-driven licences).
                _item_plan = (item_plan_totals or {}).get(item_id) or {}

                # NEW model: restriction is determined by condition_type set
                # on the licence's import item (from the parsed condition
                # sheet), not by ItemNameModel.restriction_percentage.
                cond_type = item_data.get('condition_type') or ''
                restriction_value = None
                available_cif = Decimal('0')
                if cond_type.endswith('%'):
                    try:
                        restriction_value = float(cond_type.rstrip('%'))
                    except ValueError:
                        restriction_value = None
                    if cond_type in condition_pools:
                        available_cif = condition_pools[cond_type]

                # Use pre-calculated unit price for RUTILE; otherwise fall
                # back to the E1/E5 category rate computed above.
                planner = item_plan_data.get(item_name) or {}
                _e132 = item_e132_data.get(item_name) or {}
                if item_name == 'RUTILE - A3627':
                    unit_price = rutile_unit_price
                    planned_cif = planner.get('planned_cif', 0.0)
                elif _e132:
                    # E132 reuses the Unit Price / Planned CIF columns to show the
                    # classified planning item's fixed unit price and planned value
                    # (balance-capped). Unit Price is None and Planned CIF 0 for any
                    # item whose price is still To-Be-Defined.
                    _e132_up = _e132.get('unit_price')
                    _e132_cif = _e132.get('planned_cif')
                    unit_price = float(_e132_up) if _e132_up is not None else None
                    planned_cif = float(_e132_cif) if _e132_cif is not None else 0.0
                else:
                    unit_price = planner.get('unit_price')
                    planned_cif = planner.get('planned_cif', 0.0)

                # Verification data (Item Pivot Report enhancement): when this
                # cell has actual LicenseItemPlan rows behind it, HSN/
                # Description/ledger quantities must come from the EXACT
                # import item(s) those plan lines reference — never the
                # cross-item-name-merged `item_data` aggregate above. Exactly
                # one planned import item (the overwhelming common case,
                # since Auto-Plan/manual plans normally anchor on one
                # representative item) is used directly; several distinct
                # import items are never merged into one ledger record — the
                # scalar columns are left blank and every one of them is
                # exposed, unmerged, via `planned_import_items` instead.
                # Cells with NO plan line (E132 auto-classification, or no
                # planning context at all) keep the existing aggregate
                # behaviour unchanged — this is a verification aid for
                # planned items only, not a report redesign.
                #
                # Persisted `LicenseItemPlan` rows (`_item_plan`) take
                # priority when present (exact DB rows). Licences that were
                # never explicitly Auto-Plan-"saved" have no such rows, but
                # still show a LIVE norm-waterfall recompute for CIF/rate
                # (`planner`, built above from `item_plan_data`) — that path
                # carries its own `planned_import_items` built the same way,
                # so it must be checked too or a never-saved licence's cells
                # fall back to the old cross-item-name-merged aggregate.
                _pit_list = _item_plan.get('planned_import_items') or planner.get('planned_import_items') or []
                if len(_pit_list) == 1:
                    _pit = _pit_list[0]
                    _hs_code = _pit['hs_code']
                    _description = _pit['description']
                    _quantity = _pit['quantity']
                    _allotted_quantity = _pit['allotted_quantity']
                    _debited_quantity = _pit['debited_quantity']
                    _available_quantity = _pit['available_quantity']
                elif len(_pit_list) > 1:
                    # Genuinely ambiguous — distinct import items were planned
                    # under the same item-name. HSN/Description are STRINGS:
                    # there is no correct single value, so they are left
                    # blank rather than merged, glued, or picked arbitrarily;
                    # the frontend renders each entry from
                    # `planned_import_items` instead. Quantities are NOT
                    # ambiguous the same way — summing them across the
                    # distinct import items sharing this column is factually
                    # correct (identical to the unplanned-cell aggregate
                    # below) and keeps this report's own on-screen totals and
                    # the Excel TOTAL row accurate; zeroing them out here
                    # would silently under-count both.
                    _hs_code = ''
                    _description = ''
                    _quantity = sum((p['quantity'] for p in _pit_list), 0.0)
                    _allotted_quantity = sum((p['allotted_quantity'] for p in _pit_list), 0.0)
                    _debited_quantity = sum((p['debited_quantity'] for p in _pit_list), 0.0)
                    _available_quantity = sum((p['available_quantity'] for p in _pit_list), 0.0)
                else:
                    _hs_code = item_data['hs_code']
                    _description = item_data['description']
                    _quantity = float(item_data['quantity'])
                    _allotted_quantity = float(item_data['allotted_quantity'])
                    _debited_quantity = float(item_data['debited_quantity'])
                    _available_quantity = float(item_data['available_quantity'])

                row_data['items'][item_name] = {
                    'hs_code': _hs_code,
                    'description': _description,
                    'quantity': _quantity,
                    'allotted_quantity': _allotted_quantity,
                    'debited_quantity': _debited_quantity,
                    'available_quantity': _available_quantity,
                    # Verification list: the exact import item(s) behind this
                    # cell's plan lines (empty when the cell has no plan).
                    # Always the authoritative per-item source — HSN/
                    # Description/ledger quantities never mixed across items.
                    'planned_import_items': _pit_list,
                    'restriction': restriction_value,
                    'restriction_value': float(available_cif),
                    'unit_price': unit_price,
                    'planned_cif': planned_cif,
                    # User-authored plan (distinct from the norm-derived
                    # planned_cif), sourced per plan-line item_name.
                    'plan_quantity': float(_item_plan.get('q') or 0),
                    'plan_cif': float(_item_plan.get('cif') or 0),
                    # The single manual-vs-norm selection rule — see
                    # _effective_planned_cif's docstring. Every consumer
                    # should read this instead of re-deriving the rule.
                    'effective_planned_cif': _effective_planned_cif(
                        float(_item_plan.get('q') or 0), float(_item_plan.get('cif') or 0), planned_cif,
                    ),
                    # Quantity counterpart to effective_planned_cif — see
                    # _effective_planned_quantity's docstring. First consumer
                    # is _build_notification_summary's Pass 3 (Phase 2B.2B).
                    'effective_planned_quantity': _effective_planned_quantity(
                        float(_item_plan.get('q') or 0), float(_item_plan.get('cif') or 0),
                        _available_quantity,
                    ),
                    # Raw per-line split breakdown (item_name/qty/unit_price/
                    # cif) for the "Planning Splits" sheet — same source as
                    # plan_quantity/plan_cif above, just not summed.
                    'splits': _item_plan.get('splits', []),
                    'condition_type': cond_type,
                    # E132 sequential-debit fields (None for non-E132 norms).
                    'product_code': _e132.get('product_code'),
                    'unit_rate': _e132.get('unit_rate'),
                    'debit_amount': _e132.get('debit_amount'),
                    'previous_balance': _e132.get('previous_balance'),
                    'new_balance': _e132.get('new_balance'),
                    'debit_status': _e132.get('status'),
                    # Import reference — always from the original import item.
                    # For planned licences the import_item_name/quantity come from
                    # the plan cell; for unplanned licences fall back to the
                    # column name and the import quantity already in item_data.
                    'import_item_name': _item_plan.get('import_item_name') or item_name,
                    'import_quantity': (
                        _item_plan['import_quantity']
                        if 'import_quantity' in _item_plan
                        else float(item_data['quantity'])
                    ),
                }
            else:
                row_data['items'][item_name] = {
                    'hs_code': '',
                    'description': '',
                    'quantity': 0,
                    'allotted_quantity': 0,
                    'debited_quantity': 0,
                    'available_quantity': 0,
                    'planned_import_items': [],
                    'restriction': None,
                    'restriction_value': 0,
                    'unit_price': None,
                    'planned_cif': 0,
                    'plan_quantity': 0,
                    'plan_cif': 0,
                    'effective_planned_cif': 0,
                    'effective_planned_quantity': 0,
                    'splits': [],
                    'condition_type': '',
                    'product_code': None,
                    'unit_rate': None,
                    'debit_amount': None,
                    'previous_balance': None,
                    'new_balance': None,
                    'debit_status': None,
                    'import_item_name': '',
                    'import_quantity': 0,
                }

        # This license's Planned CIF row-total — sum of each item's already-
        # resolved `effective_planned_cif` (Phase 2B.2A). Zero-valued items
        # (no data for this license) don't affect the sum, so this equals
        # summing over any narrower "items with data" subset a consumer
        # might otherwise have filtered to first.
        row_data['total_effective_planned_cif'] = sum(
            item.get('effective_planned_cif', 0) or 0 for item in row_data['items'].values()
        )

        return row_data

    def export_to_excel_streaming(self, days=30, sion_norm=None, company_ids=None,
                                  exclude_company_ids=None, min_balance=200, license_status='active',
                                  expiry_date_from=None, expiry_date_to=None, purchase_status=None):
        """
        Export report to Excel - uses existing generate_report for data, then formats as Excel.
        This ensures consistency with JSON output.

        Returns:
            StreamingHttpResponse with Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.cell import WriteOnlyCell
        from django.http import StreamingHttpResponse
        import tempfile
        import os
        from apps.license.utils.condition_excel import annotate_cell as _annotate_condition_cell

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()

        try:
            # Use the working generate_report method
            report_data = self.generate_report(days, sion_norm, company_ids, exclude_company_ids, min_balance, license_status, expiry_date_from, expiry_date_to, purchase_status)

            workbook = openpyxl.Workbook(write_only=True)
            licenses_by_norm_notif = report_data.get('licenses_by_norm_notification', {})

            for norm_class in sorted(licenses_by_norm_notif.keys()):
                notifications_dict = licenses_by_norm_notif[norm_class]
                for notification, licenses_list in sorted(notifications_dict.items()):
                    # Filter items to only those with actual data in THIS norm-notification.
                    # Must check plan_quantity / plan_cif in addition to quantity: for
                    # manually-planned items whose planned name (e.g. "SWP - E1") is not
                    # a direct import item, quantity = 0 but plan_quantity > 0.
                    items_with_data = []
                    for item in report_data['items']:
                        item_name = item['name']
                        has_data = any(
                            lic['items'].get(item_name, {}).get('quantity', 0) > 0
                            or lic['items'].get(item_name, {}).get('available_quantity', 0) > 0
                            or (lic['items'].get(item_name, {}).get('plan_quantity') or 0) > 0
                            or (lic['items'].get(item_name, {}).get('plan_cif') or 0) > 0
                            for lic in licenses_list
                        )
                        if has_data:
                            items_with_data.append(item)

                    # Create sheet
                    sheet_name = f"{norm_class}_{notification}"[:31].replace('/', '-').replace('\\', '-').replace('*', '-')
                    worksheet = workbook.create_sheet(title=sheet_name)

                    # Title row
                    title_cell = WriteOnlyCell(worksheet, value=f"Item Pivot Report - {norm_class} - {notification}")
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center')
                    worksheet.append(_xlsx_safe_row([title_cell] + [None] * 25))
                    worksheet.append([])

                    # Headers
                    base_headers = ['Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt', 'Exporter', 'Total CIF', 'Debited CIF', 'Alloted CIF', 'Balance CIF', 'Notes', 'Condition Sheet']
                    item_headers = []
                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        headers = [
                            f"{item_name} HSN Code",
                            f"{item_name} Product Description",
                            f"{item_name} Total QTY",
                            f"{item_name} Allotted QTY",
                            f"{item_name} Debited QTY",
                            f"{item_name} Balance QTY"
                        ]
                        headers.extend([
                            f"{item_name} Import Item Name",
                            f"{item_name} Import Qty",
                        ])
                        if has_restriction:
                            headers.extend([
                                f"{item_name} Restriction %",
                                f"{item_name} Restriction Value"
                            ])
                        # Two new per-item columns sourced from the
                        # e1_plan / e5_plan waterfall so the Excel matches
                        # the bulk Balance report cell-for-cell.
                        headers.extend([
                            f"{item_name} Plan Qty",
                            f"{item_name} Planned CIF",
                        ])
                        item_headers.extend(headers)

                    all_headers = base_headers + item_headers
                    header_row = []
                    for header in all_headers:
                        cell = WriteOnlyCell(worksheet, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                        header_row.append(cell)
                    worksheet.append(_xlsx_safe_row(header_row))

                    # Data rows
                    for idx, lic in enumerate(licenses_list, 1):
                        row_data = [
                            idx,
                            lic['license_number'],
                            lic['license_date'],
                            lic['license_expiry_date'],
                            lic['exporter'],
                            lic['total_cif'],
                            lic.get('debited_cif', 0),
                            lic['alloted_cif'],
                            lic['balance_cif'],
                            lic.get('balance_report_notes', ''),
                            lic.get('condition_sheet', '')
                        ]

                        for item in items_with_data:
                            item_name = item['name']
                            has_restriction = item.get('has_restriction', False)
                            item_data = lic['items'].get(item_name, {})
                            cond = item_data.get('condition_type') or ''
                            # When this column is backed by more than one
                            # distinct import item, `hs_code`/`description`
                            # are blank (see _build_license_row — a string
                            # can't be merged across items). Match the UI's
                            # own handling instead of leaving the Excel cell
                            # empty: list each import item's own HSN/
                            # Description on its own line within the cell —
                            # never joined without a separator, never one
                            # value picked over another.
                            _pits = item_data.get('planned_import_items') or []
                            if len(_pits) > 1:
                                _hsn_value = '\n'.join(p.get('hs_code') or '-' for p in _pits)
                                _desc_value = '\n'.join(p.get('description') or '-' for p in _pits)
                            else:
                                _hsn_value = item_data.get('hs_code', '')
                                _desc_value = item_data.get('description', '')
                            # Tint the HSN-code cell for this (licence, item)
                            # pair when a condition is set.
                            hsn_cell = WriteOnlyCell(worksheet, value=_hsn_value)
                            _annotate_condition_cell(hsn_cell, cond)
                            if len(_pits) > 1:
                                hsn_cell.alignment = Alignment(wrap_text=True, vertical='top')
                            row_data.append(hsn_cell)
                            desc_cell = WriteOnlyCell(worksheet, value=_desc_value)
                            if len(_pits) > 1:
                                desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
                            row_data.append(desc_cell)
                            row_data.extend([
                                item_data.get('quantity', 0),
                                item_data.get('allotted_quantity', 0),
                                item_data.get('debited_quantity', 0),
                                item_data.get('available_quantity', 0)
                            ])
                            row_data.extend([
                                item_data.get('import_item_name', ''),
                                item_data.get('import_quantity', 0),
                            ])
                            if has_restriction:
                                row_data.extend([
                                    item_data.get('restriction'),
                                    item_data.get('restriction_value', 0)
                                ])
                            # Per-product: manual plan if this product was manually
                            # planned; else norm-derived unit price / planned CIF.
                            # First column is a display convention (quantity when
                            # manually planned, else the unit price rate) — not a
                            # business calculation. Second column is the CIF value
                            # itself, using the single backend selection rule
                            # (_effective_planned_cif / `effective_planned_cif`).
                            _s_plan_q = item_data.get('plan_quantity') or 0
                            _s_plan_c = item_data.get('plan_cif') or 0
                            if _s_plan_q or _s_plan_c:
                                row_data.append(_s_plan_q or 0)
                            else:
                                row_data.append(item_data.get('unit_price') or 0)
                            row_data.append(item_data.get('effective_planned_cif', 0))

                        worksheet.append(_xlsx_safe_row(row_data))

                    # Totals row — reads the backend-computed totals for
                    # this (norm, notification) group; writes only, no
                    # aggregation happens here (Phase 2B.2A; see
                    # docs/architecture/ITEM_PIVOT_DISPLAY_DATASET_DESIGN.md).
                    # base_headers = ['Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt',
                    #   'Exporter', 'Total CIF', 'Debited CIF', 'Alloted CIF',
                    #   'Balance CIF', 'Notes', 'Condition Sheet']
                    # 'TOTAL' lands under col 1 (Sr no); cols 2-5 (DFIA No/DFIA
                    # Dt/Expiry Dt/Exporter) are blank; cols 6-9 are the four
                    # CIF sums; cols 10-11 (Notes/Condition Sheet) are blank.
                    group_totals = report_data.get('notification_totals', {}).get(norm_class, {}).get(notification, {})
                    item_totals = group_totals.get('items', {})

                    totals_row = [WriteOnlyCell(worksheet, value='TOTAL')]
                    totals_row[0].font = Font(bold=True)
                    totals_row.extend([None, None, None, None])  # DFIA No, DFIA Dt, Expiry Dt, Exporter

                    total_cif_cell = WriteOnlyCell(worksheet, value=group_totals.get('total_cif', 0))
                    total_cif_cell.font = Font(bold=True)
                    totals_row.append(total_cif_cell)

                    debited_cif_cell = WriteOnlyCell(worksheet, value=group_totals.get('debited_cif', 0))
                    debited_cif_cell.font = Font(bold=True)
                    totals_row.append(debited_cif_cell)

                    alloted_cif_cell = WriteOnlyCell(worksheet, value=group_totals.get('alloted_cif', 0))
                    alloted_cif_cell.font = Font(bold=True)
                    totals_row.append(alloted_cif_cell)

                    balance_cif_cell = WriteOnlyCell(worksheet, value=group_totals.get('balance_cif', 0))
                    balance_cif_cell.font = Font(bold=True)
                    totals_row.append(balance_cif_cell)

                    totals_row.extend([None, None])  # Notes, Condition Sheet

                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        item_total = item_totals.get(item_name, {})
                        totals_row.extend([None, None])  # HSN, Description
                        for qty_type in ['quantity', 'allotted_quantity', 'debited_quantity', 'available_quantity']:
                            cell = WriteOnlyCell(worksheet, value=item_total.get(qty_type, 0))
                            cell.font = Font(bold=True)
                            totals_row.append(cell)
                        totals_row.extend([None, None])  # Import Item Name, Import Qty
                        if has_restriction:
                            totals_row.append(None)  # Restriction %
                            cell = WriteOnlyCell(worksheet, value=item_total.get('restriction_value', 0))
                            cell.font = Font(bold=True)
                            totals_row.append(cell)
                        # Plan Qty total: a literal sum of plan_quantity only,
                        # matching the on-screen report's `totalPlanQty` total
                        # (frontend/src/pages/reports/ItemPivotReport.tsx). Rows
                        # with no manual plan (norm-derived, shown as a unit-price
                        # rate per-row) contribute 0 here rather than being folded
                        # into a blended rate.
                        cell = WriteOnlyCell(worksheet, value=item_total.get('plan_quantity', 0))
                        cell.font = Font(bold=True)
                        totals_row.append(cell)

                        # Planned CIF total: the single manual-vs-norm selection
                        # rule (_effective_planned_cif), already resolved per
                        # cell and summed by the backend.
                        cell = WriteOnlyCell(worksheet, value=item_total.get('effective_planned_cif', 0))
                        cell.font = Font(bold=True)
                        totals_row.append(cell)

                    worksheet.append(_xlsx_safe_row(totals_row))

            # "Planning Splits" — a separate sheet listing every visible
            # LicenseItemPlan split (flat, one row each) across all licenses
            # in this report. Additive detail; the pivot grid's per-item-name
            # Plan Qty / Plan CIF cells above are unchanged.
            split_rows = _planning_split_sheet_rows(
                licenses_by_norm_notif,
                [item['name'] for item in report_data['items']],
            )
            splits_ws = workbook.create_sheet(title="Planning Splits")
            split_header_row = []
            for header in ('License No', 'Product', 'Item Name', 'Split',
                           'Unit Price', 'Planned Qty', 'Planned CIF'):
                cell = WriteOnlyCell(splits_ws, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                split_header_row.append(cell)
            splits_ws.append(_xlsx_safe_row(split_header_row))
            for split_row in split_rows:
                splits_ws.append(_xlsx_safe_row(list(split_row)))

            # Save workbook
            workbook.save(temp_file.name)
            workbook.close()

            # Stream file
            def file_iterator(file_path, chunk_size=8192):
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
                try:
                    os.unlink(file_path)
                except OSError:
                    pass

            response = StreamingHttpResponse(
                file_iterator(temp_file.name),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="item_pivot_report.xlsx"'
            return response

        except Exception as e:
            try:
                os.unlink(temp_file.name)
            except OSError:
                pass
            raise e

class ItemPivotViewSet(viewsets.ViewSet):
    """
    ViewSet for Item Pivot Report.

    Permissions: ReportPermission.
    """
    permission_classes = [ReportPermission]

    def list(self, request):
        """
        Get item pivot report.

        Query Parameters:
            days: Number of days to look back (default: 30)
        """
        view = ItemPivotReportView()
        return view.get(request)

    @action(detail=False, methods=['get'], url_path='available-norms')
    def available_norms(self, request):
        """
        Get list of all active norm classes with their descriptions.
        Returns only norms that are marked as active (is_active=True) in SionNormClassModel.
        """
        try:
            # Get only active SION norm classes from the database
            from apps.core.models import SionNormClassModel
            active_norms_data = SionNormClassModel.objects.filter(
                is_active=True
            ).values('norm_class', 'description').order_by('norm_class')

            # Build result with norm_class and description
            result = [
                {
                    'norm_class': norm['norm_class'],
                    'description': norm['description'] or ''
                }
                for norm in active_norms_data
            ]

            return Response(result)
        except Exception as e:
            logger.exception("Error generating item pivot report")
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=['get'], url_path='task-status/(?P<task_id>[^/.]+)')
    def task_status(self, request, task_id=None):
        """
        Check the status of an async Excel generation task.

        Returns:
            state: Task state (PENDING, PROGRESS, SUCCESS, FAILURE)
            current: Current progress (0-100)
            total: Total progress (100)
            status: Status message
            result: Result data (if completed)
        """
        from celery.result import AsyncResult

        task = AsyncResult(task_id)

        if task.state == 'PENDING':
            response = {
                'state': task.state,
                'current': 0,
                'total': 100,
                'status': 'Pending...'
            }
        elif task.state == 'PROGRESS':
            response = {
                'state': task.state,
                'current': task.info.get('current', 0),
                'total': task.info.get('total', 100),
                'status': task.info.get('status', '')
            }
        elif task.state == 'SUCCESS':
            response = {
                'state': task.state,
                'current': 100,
                'total': 100,
                'status': 'Completed!',
                'result': task.info
            }
        else:
            # Something went wrong
            response = {
                'state': task.state,
                'current': 100,
                'total': 100,
                'status': str(task.info) if task.info else 'Unknown error'
            }

        return Response(response)

    @action(detail=False, methods=['post'], url_path='update-balance')
    def update_balance(self, request):
        """
        Trigger high-priority task to update balance_cif, is_active, is_expired, and restrictions.

        This task:
        1. Updates balance_cif for all licenses using LicenseBalanceCalculator
        2. Updates is_expired based on license_expiry_date
        3. Updates is_null based on balance < $500
        4. Updates is_active based on expiry (mark inactive if expired)
        5. Checks and updates restriction flags on import items

        Returns:
            task_id: ID to check status using task-status endpoint
        """
        from apps.license.tasks import update_all_license_balances

        # Get license_status parameter from request body
        license_status = request.data.get('license_status', 'all')

        # Start the Celery task with high priority
        task = update_all_license_balances.apply_async(
            args=[license_status],
            priority=9  # High priority (0-9, 9 is highest)
        )

        return Response({
            'task_id': task.id,
            'status': 'PENDING',
            'license_status': license_status,
            'message': f'Balance update started for {license_status} licenses. Use the task_id to check status.'
        }, status=202)
