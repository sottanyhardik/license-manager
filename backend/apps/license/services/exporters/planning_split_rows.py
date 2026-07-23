"""
Shared rendering for `LicenseItemPlan` "split" sub-rows in Excel exports.

A manually-authored utilization plan can split a single import item across
several planning item names — e.g. "20 units of Wheat Flour @ $5.00/unit"
and "10 units of Milk Powder @ $7.50/unit" both splitting one Milk Powder
import item. Multiple Excel exporters render each *visible* split (real
planned quantity or CIF) as an indented sub-row beneath the item's own
summary row:

  - `apps/license/services/exporters/license_balance_excel.py` — the
    single-license "Plan Utilization" sheet and the bulk multi-sheet variant
    (both used to carry their own copy of this filter/loop/style code).
  - `apps/license/views/item_report.py` — the Item Report export.
  - `apps/license/views/item_pivot_report.py` — a dedicated "Planning
    Splits" sheet (its pivot grid is a wide, one-row-per-license layout,
    some of it write-only/append-only, so it can't host indented child rows
    the way the other two reports do; it reuses `rows_for_splits()` only).

This module is the single source of truth for:
  - `rows_for_splits()` — the filter + label/format rules every caller must
    reproduce exactly (pure, no openpyxl dependency).
  - `write_split_sub_rows()` — a writer for standard (non-write-only)
    worksheets, parameterized by column indices so each report's own column
    layout can be expressed without duplicating the loop/style code.

Note the two pre-existing implementations styled the quantity/CIF sub-row
cells differently (license_balance_excel.py used a plain, non-italic font
and an explicit `#,##0.000`/`#,##0.00` number format; item_report.py reused
the italic "split" font and left the number format as General). Both are
preserved exactly via the `value_font`/`qty_num_fmt`/`cif_num_fmt`
parameters below — this module does not silently unify per-report styling.
"""
from __future__ import annotations

from typing import Any, Dict, Iterable, List, Optional, Sequence

# Shared "split sub-row" styling — EBF3FF fill + 2B5EA7 italic font, proven
# across every pre-existing implementation. Callers of `write_split_sub_rows`
# may override any of these via its `fill`/`font`/`badge_font` kwargs.
SPLIT_FILL_COLOR = "EBF3FF"
SPLIT_FONT_COLOR = "2B5EA7"


def _default_split_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=SPLIT_FILL_COLOR, end_color=SPLIT_FILL_COLOR, fill_type="solid")


def _default_split_font():
    from openpyxl.styles import Font
    return Font(size=9, color=SPLIT_FONT_COLOR, italic=True)


def _default_split_badge_font():
    from openpyxl.styles import Font
    return Font(size=8, color=SPLIT_FONT_COLOR, italic=True)


def rows_for_splits(splits: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Filter + format raw `LicenseItemPlan` split dicts into row-ready records.

    `splits` is the per-item `splits` list as built by
    `apps.license.services.plan_reporting._build_map` (each entry a dict
    with `item_name`, `planned_quantity`, `unit_price`, `planned_cif_fc`).

    Only splits with a real planned quantity or CIF are "visible" — this is
    the exact filter every existing implementation applies. Returned dicts
    are in filtered order, `split_number` numbered 1-based *within that
    filtered list* (not the raw list), matching the current
    `enumerate(_valid_splits)` behaviour.

    Each row also carries the **raw** `item_name` (`None`/'' when the split
    isn't tagged with a planning item name) alongside the display-formatted
    `item_name_label` (which falls back to `"Split N"` for an untagged
    split). Callers that need to bucket splits by planning item name (e.g. a
    pivot column) must key off raw `item_name`, not `item_name_label` — the
    "Split N" fallback is a per-split display label, not a shared bucket
    name, so grouping by it would create one spurious column per untagged
    split instead of one shared "Unassigned" column.
    """
    visible = [
        s for s in (splits or [])
        if float(s.get('planned_quantity') or 0) > 0
        or float(s.get('planned_cif_fc') or 0) > 0
    ]
    rows: List[Dict[str, Any]] = []
    for i, sp in enumerate(visible):
        unit_price = float(sp.get('unit_price') or 0)
        item_name_label = sp.get('item_name') or f'Split {i + 1}'
        rows.append({
            'split_number': i + 1,
            'split_badge': f'Split {i + 1}',
            'item_name': sp.get('item_name'),
            'item_name_label': item_name_label,
            'indented_label': f'  └ {item_name_label}',
            'unit_price': unit_price,
            'unit_price_label': f'@ ${unit_price:,.2f}/unit' if unit_price else '',
            'planned_quantity': float(sp.get('planned_quantity') or 0),
            'planned_cif_fc': float(sp.get('planned_cif_fc') or 0),
        })
    return rows


def write_split_sub_rows(
    ws,
    start_row: int,
    splits: Iterable[Dict[str, Any]],
    *,
    name_col: int,
    badge_col: int,
    qty_col: int,
    cif_col: int,
    price_col: "int | None" = None,
    other_cols: Sequence[int] = (),
    fill=None,
    font=None,
    badge_font=None,
    value_font=None,
    border=None,
    qty_num_fmt: "str | None" = None,
    cif_num_fmt: "str | None" = None,
) -> int:
    """Write one row per *visible* split to a standard (non-write-only)
    worksheet, starting at `start_row`. Returns the number of rows written
    (0 if there are no visible splits) — callers use this to advance their
    own row cursor and for `row_span`/merge-cell math.

    Column layout is entirely caller-specified so each report's own columns
    can be reused as-is:
      - `name_col`: indented planning-item-name label.
      - `price_col` (optional): "@ $X.XX/unit" label, blank if price is 0.
      - `badge_col`: "Split N" badge.
      - `qty_col` / `cif_col`: planned quantity / planned CIF-FC values.
      - `other_cols`: any additional columns that should get the split's
        fill/border painted but no value (e.g. filler columns between the
        item's own columns in a fixed-width table row).

    Styling defaults to the shared EBF3FF fill / 2B5EA7 italic font. `font`
    is used for the name/price cells; `value_font` (defaulting to `font`)
    is used for the qty/cif cells — pass a plain `Font` there to reproduce
    license_balance_excel.py's original (non-italic) qty/cif styling.
    `border` defaults to None (no border touched), matching item_report.py's
    original behaviour; pass your own `Border`/`Side` instance to reproduce
    license_balance_excel.py's THIN_BORDER-everywhere behaviour — each file
    defines its own `Border` instance and they are not assumed interchangeable.
    """
    from openpyxl.styles import Alignment

    rows = rows_for_splits(splits)
    if not rows:
        return 0

    fill = fill or _default_split_fill()
    font = font or _default_split_font()
    badge_font = badge_font or _default_split_badge_font()
    value_font = value_font or font

    touched_cols = set(other_cols) | {name_col, badge_col, qty_col, cif_col}
    if price_col is not None:
        touched_cols.add(price_col)

    row = start_row
    for r in rows:
        # Paint fill(+border) across every touched column first, then
        # overwrite the specific value-bearing cells below — matching the
        # "paint the row, then overwrite specific cells" order every
        # pre-existing implementation used.
        for col in touched_cols:
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            if border is not None:
                cell.border = border

        name_cell = ws.cell(row=row, column=name_col, value=r['indented_label'])
        name_cell.fill = fill
        if border is not None:
            name_cell.border = border
        name_cell.font = font
        name_cell.alignment = Alignment(horizontal='left', vertical='center')

        if price_col is not None:
            price_cell = ws.cell(row=row, column=price_col, value=r['unit_price_label'])
            price_cell.fill = fill
            if border is not None:
                price_cell.border = border
            price_cell.font = font
            price_cell.alignment = Alignment(horizontal='left', vertical='center')

        badge_cell = ws.cell(row=row, column=badge_col, value=r['split_badge'])
        badge_cell.fill = fill
        if border is not None:
            badge_cell.border = border
        badge_cell.font = badge_font
        badge_cell.alignment = Alignment(horizontal='center', vertical='center')

        qty_cell = ws.cell(row=row, column=qty_col, value=r['planned_quantity'])
        qty_cell.fill = fill
        if border is not None:
            qty_cell.border = border
        qty_cell.font = value_font
        qty_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        if qty_num_fmt:
            qty_cell.number_format = qty_num_fmt

        cif_cell = ws.cell(row=row, column=cif_col, value=r['planned_cif_fc'])
        cif_cell.fill = fill
        if border is not None:
            cif_cell.border = border
        cif_cell.font = value_font
        cif_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        if cif_num_fmt:
            cif_cell.number_format = cif_num_fmt

        row += 1

    return len(rows)


def render_plan_utilization_section(
    ws,
    r: int,
    license_obj,
    license_balance: float,
    *,
    plan_map: Optional[Dict[int, Dict[str, Any]]] = None,
    totals_out: Optional[Dict[str, Any]] = None,
) -> int:
    """Render one license's "Plan Utilization" section starting at row `r`.

    This is the shared extraction of the ~140-line block (section header, the
    2-row metrics grid, the 9-column table header, one row per
    `plan_utilization_rows()` group + `write_split_sub_rows()` split
    sub-rows, and the per-license totals row) that used to be duplicated
    verbatim between:
      - the single-license `balance_excel` export
        (`apps.license.services.exporters.license_balance_excel.build_balance_excel`)
      - the bulk `bulk_balance_excel` export's per-license sheet builder
        (`..license_balance_excel._write_license_sheet`)
    and is now also reused by the bulk workbook's "Utilization Planning
    Summary" sheet (one call per exported license).

    Columns written (1-indexed, matching every pre-existing caller):
      1 Item Description | 2 HS Code | 3 S.No | 4 Status | 5 Available Qty |
      6 Planned Qty | 7 Remaining Qty | 8 Planned CIF ($) | 9 Remaining CIF ($)

    Does NOT touch column widths / `freeze_panes` — callers that render this
    section on its own sheet (or want different widths for a stacked-section
    sheet) own that, exactly as the two pre-existing call sites already did.

    Args:
      ws: worksheet to write into (any non-write-only worksheet).
      r: first free row.
      license_obj: the `LicenseDetailsModel` instance.
      license_balance: `float(license_obj.get_balance_cif or 0)` — passed in
        rather than recomputed, since every caller already has it.
      plan_map: pre-built `{import_item_id: {...}}` map from
        `apps.license.services.plan_reporting.plan_map_for_license`. Pass the
        one you already fetched to avoid a second query; computed here
        (one query) when omitted.
      totals_out: if a dict is passed, it is updated in place with this
        license's Plan Utilization totals — `available_quantity`,
        `planned_quantity`, `remaining_quantity`, `planned_cif`,
        `remaining_cif`, `plan_entries` — matching exactly what the rendered
        TOTALS row shows. Callers that only need the next row (both
        pre-existing call sites) can omit this.

    Returns:
      The next free row after the section (its own totals row inclusive).
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from apps.license.services.plan_reporting import plan_map_for_license
    from apps.license.services.plan_utilization import plan_utilization_rows

    if plan_map is None:
        plan_map = plan_map_for_license(license_obj.id)

    HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    YEL_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ALT_FILL = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    BOLD = Font(bold=True, size=9)
    NORM = Font(size=9)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    PLAN_GRN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    PLAN_GRN_FONT = Font(bold=True, size=9, color="375623")
    PLAN_GRY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def _hdr(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HDR_FILL
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return c

    def _cell(row, col, value, fill=None, bold=False, align='left', num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill:
            c.fill = fill
        c.font = BOLD if bold else NORM
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt
        return c

    _plan_data_rows = []
    _g_avail = 0.0
    _g_planned_qty = 0.0
    _g_planned_cif = 0.0

    for _grp in plan_utilization_rows(license_obj, plan_map=plan_map):
        _av = float(_grp['available_quantity'] or 0)
        # Group-level planned totals == Σ LicenseItemPlan.planned_quantity/
        # planned_cif_fc across the whole group (plan_status_for's
        # "original"), which is exactly what summing the group's unioned
        # splits would give — same figure the old per-item code read off
        # plan_map's `total_planned_quantity`/`total_planned_cif`, just
        # aggregated across the merged serials instead of one at a time.
        _pq = float(_grp['original_quantity']) if _grp['has_plan'] else 0.0
        _pc = float(_grp['original_cif_fc']) if _grp['has_plan'] else 0.0
        _g_avail += _av
        _g_planned_qty += _pq
        _g_planned_cif += _pc
        _plan_data_rows.append({
            'desc': _grp['description'] or '-',
            'hs': _grp['hs_code'] or '-',
            'sr': ', '.join(str(s) for s in _grp['serials']),
            'avail': _av, 'pqty': _pq, 'pcif': _pc,
            'rem_qty': _av - _pq,
            'planned': _pq > 0 or _pc > 0,
            'splits': _grp['splits'],
        })
    _g_rem_qty = _g_avail - _g_planned_qty
    _g_rem_cif = license_balance - _g_planned_cif

    # ── Section header ────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:I{r}')
    _ph = ws[f'A{r}']
    _ph.value = 'Plan Utilization'
    _ph.fill = HDR_FILL
    _ph.font = Font(bold=True, color="FFFFFF", size=10)
    _ph.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    # ── Summary metrics (2 rows x 4 label-value pairs) ───────────────────
    _metrics = [
        [('Balance CIF $', f"${license_balance:,.2f}", False),
         ('Planned CIF $', f"${_g_planned_cif:,.2f}", False),
         ('Remaining CIF $', f"${_g_rem_cif:,.2f}", _g_rem_cif < 0),
         ('Remaining CIF', f"${max(0, _g_rem_cif):,.2f}", False)],
        [('Available Qty', f"{_g_avail:,.3f}", False),
         ('Planned Qty', f"{_g_planned_qty:,.3f}", False),
         ('Remaining Qty', f"{_g_rem_qty:,.3f}", _g_rem_qty < 0),
         ('Plan Entries', str(sum(1 for p in _plan_data_rows if p['planned'])), False)],
    ]
    for _row_metrics in _metrics:
        for _mi, (_lbl, _val, _warn) in enumerate(_row_metrics):
            _col = _mi * 2 + 1
            _lc = ws.cell(row=r, column=_col, value=_lbl)
            _lc.fill = HDR_FILL
            _lc.font = Font(bold=True, color="FFFFFF", size=8)
            _lc.border = THIN_BORDER
            _lc.alignment = Alignment(horizontal='center', vertical='center')
            _vc = ws.cell(row=r, column=_col + 1, value=_val)
            _vc.fill = YEL_FILL
            _vc.border = THIN_BORDER
            _vc.font = Font(bold=True, size=9, color="C00000" if _warn else "1F4E79")
            _vc.alignment = Alignment(horizontal='right', vertical='center')
        r += 1
    r += 1  # spacer

    # ── Table headers ─────────────────────────────────────────────────────
    for _ci, _ch in enumerate(['Item Description', 'HS Code', 'S.No', 'Status',
                                'Available Qty', 'Planned Qty', 'Remaining Qty',
                                'Planned CIF ($)', 'Remaining CIF ($)'], 1):
        _hdr(r, _ci, _ch)
    r += 1

    # ── Per-group rows (+ split sub-rows) ──────────────────────────────────
    _running_cif = license_balance
    for _idx, _pr in enumerate(_plan_data_rows):
        _rf = None if _idx % 2 == 0 else ALT_FILL
        _running_cif -= _pr['pcif']
        # ── Item row ──────────────────────────────────────────────────────
        _dc = ws.cell(row=r, column=1, value=_pr['desc'])
        _dc.fill = _rf or PatternFill(fill_type=None)
        _dc.border = THIN_BORDER
        _dc.font = NORM
        _dc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        _hsc = ws.cell(row=r, column=2, value=_pr['hs'])
        _hsc.fill = _rf or PatternFill(fill_type=None)
        _hsc.border = THIN_BORDER
        _hsc.font = NORM
        _hsc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        _snc = ws.cell(row=r, column=3, value=_pr['sr'])
        _snc.fill = _rf or PatternFill(fill_type=None)
        _snc.border = THIN_BORDER
        _snc.font = NORM
        _snc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        _sc = ws.cell(row=r, column=4, value='Planned' if _pr['planned'] else 'Not Planned')
        _sc.fill = PLAN_GRN_FILL if _pr['planned'] else PLAN_GRY_FILL
        _sc.border = THIN_BORDER
        _sc.font = PLAN_GRN_FONT if _pr['planned'] else Font(size=9, color="595959")
        _sc.alignment = Alignment(horizontal='center', vertical='center')
        _cell(r, 5, _pr['avail'], fill=_rf, align='right', num_fmt='#,##0.000')
        if _pr['planned']:
            _cell(r, 6, _pr['pqty'], fill=_rf, align='right', num_fmt='#,##0.000')
            _cell(r, 7, _pr['rem_qty'], fill=PLAN_GRN_FILL if _pr['rem_qty'] <= 0 else _rf,
                  align='right', num_fmt='#,##0.000')
            _cell(r, 8, _pr['pcif'], fill=_rf, align='right', num_fmt='#,##0.00')
            _cell(r, 9, max(0.0, _running_cif), fill=_rf, align='right', num_fmt='#,##0.00')
        else:
            _cell(r, 6, '-', fill=_rf, align='center')
            _cell(r, 7, _pr['avail'], fill=_rf, align='right', num_fmt='#,##0.000')
            _cell(r, 8, '-', fill=_rf, align='center')
            _cell(r, 9, '-', fill=_rf, align='center')
        r += 1
        # ── Split sub-rows (union of every merged serial's splits, only for
        # planned groups) ───────────────────────────────────────────────────
        if _pr['planned']:
            r += write_split_sub_rows(
                ws, r, _pr['splits'],
                name_col=1, price_col=2, badge_col=4, qty_col=6, cif_col=8,
                other_cols=(3, 5, 7, 9),
                value_font=NORM, border=THIN_BORDER,
                qty_num_fmt='#,##0.000', cif_num_fmt='#,##0.00',
            )

    # ── Totals row ────────────────────────────────────────────────────────
    for _ci in range(1, 10):
        ws.cell(row=r, column=_ci).fill = TOTAL_FILL
        ws.cell(row=r, column=_ci).border = THIN_BORDER
    _cell(r, 1, 'TOTALS', fill=TOTAL_FILL, bold=True)
    _cell(r, 2, '', fill=TOTAL_FILL)
    _cell(r, 3, '', fill=TOTAL_FILL)
    _cell(r, 4, '', fill=TOTAL_FILL)
    _cell(r, 5, _g_avail, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
    _cell(r, 6, _g_planned_qty, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
    _cell(r, 7, _g_rem_qty, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.000')
    _cell(r, 8, _g_planned_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
    _cell(r, 9, max(0.0, _g_rem_cif), fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
    r += 1

    if totals_out is not None:
        totals_out.update({
            'available_quantity': _g_avail,
            'planned_quantity': _g_planned_qty,
            'remaining_quantity': _g_rem_qty,
            'planned_cif': _g_planned_cif,
            'remaining_cif': _g_rem_cif,
            'plan_entries': sum(1 for p in _plan_data_rows if p['planned']),
        })

    return r
