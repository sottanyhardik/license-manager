"""
API Tests for License Endpoints
"""
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
import pytest
from django.urls import reverse
from rest_framework import status

from apps.license.models import LicenseDetailsModel


@pytest.mark.api
@pytest.mark.database
class TestLicenseAPI:
    """Test License CRUD operations"""
    
    def test_list_licenses(self, authenticated_client, test_license):
        """Test GET /licenses/"""
        url = reverse('license:licenses-list')
        response = authenticated_client.get(url)
        
        assert response.status_code == status.HTTP_200_OK
        assert 'results' in response.data or isinstance(response.data, list)
    
    def test_retrieve_license(self, authenticated_client, test_license):
        """Test GET /licenses/{id}/"""
        url = reverse('license:licenses-detail', kwargs={'pk': test_license.id})
        response = authenticated_client.get(url)
        
        assert response.status_code == status.HTTP_200_OK
        assert response.data['license_number'] == test_license.license_number
        assert 'import_license' in response.data
    
    def test_license_has_items(self, authenticated_client, test_license):
        """Test license includes related import items"""
        url = reverse('license:licenses-detail', kwargs={'pk': test_license.id})
        response = authenticated_client.get(url)
        
        assert response.status_code == status.HTTP_200_OK
        assert len(response.data['import_license']) == 3
    
    def test_filter_licenses_by_scheme(self, authenticated_client, test_license):
        """Test GET /licenses/?scheme_code=DFIA"""
        url = reverse('license:licenses-list')
        response = authenticated_client.get(url, {'scheme_code': 'DFIA'})
        
        assert response.status_code == status.HTTP_200_OK
    
    def test_search_licenses(self, authenticated_client, test_license):
        """Test GET /licenses/?search={license_number}"""
        url = reverse('license:licenses-list')
        response = authenticated_client.get(url, {'search': test_license.license_number[:5]})
        
        assert response.status_code == status.HTTP_200_OK

    def test_fetch_ledger_returns_explicit_not_configured_response(self, authenticated_client, test_license):
        """Remote DGFT ledger fetch is intentionally unsupported."""
        url = reverse('license:license-actions-fetch-ledger', kwargs={'pk': test_license.id})
        response = authenticated_client.post(url)

        assert response.status_code == status.HTTP_501_NOT_IMPLEMENTED
        assert response.data == {
            'success': False,
            'message': 'Remote DGFT ledger fetch is not configured. Use ledger upload instead.',
            'license_number': test_license.license_number,
        }

    def test_balance_excel_returns_expected_workbook_shape(self, authenticated_client, test_license):
        """Single-license balance Excel keeps its public workbook contract."""
        url = reverse('license:licenses-balance-excel', kwargs={'pk': test_license.id})
        response = authenticated_client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert (
            f'filename="{test_license.license_number}-summary.xlsx"'
            in response['Content-Disposition']
        )

        workbook = load_workbook(BytesIO(response.content), data_only=False)

        assert workbook.sheetnames == ['Summary']
        sheet = workbook['Summary']
        assert sheet['A1'].value.startswith(f'License No: {test_license.license_number}')
        assert sheet['A2'].value == 'Summary (BOE & Allotments)'
        assert any(
            cell.value == 'Summary (Balance Quantity)'
            for row in sheet.iter_rows()
            for cell in row
        )

    def test_bulk_balance_excel_returns_summary_and_license_sheets(
        self,
        authenticated_client,
        test_license,
    ):
        """Bulk balance Excel keeps the summary sheet plus per-license sheet."""
        url = reverse('license:licenses-bulk-balance-excel')
        response = authenticated_client.post(
            url,
            {'license_numbers': [test_license.license_number]},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert 'filename="bulk_license_summary.xlsx"' in response['Content-Disposition']

        workbook = load_workbook(BytesIO(response.content), data_only=False)

        assert workbook.sheetnames[0] == 'Utilization Planning Summary'
        assert test_license.license_number[:31] in workbook.sheetnames
        summary = workbook['Utilization Planning Summary']
        # test_license carries no export items -> norm-grouped into
        # "Unclassified", whose section banner is the sheet's first content.
        assert summary['A1'].value == 'SION NORM : Unclassified'

    def test_bulk_balance_excel_rejects_non_list_license_numbers(self, authenticated_client, test_license):
        """Bulk balance Excel validates request shape before querying licenses."""
        url = reverse('license:licenses-bulk-balance-excel')
        response = authenticated_client.post(
            url,
            {'license_numbers': test_license.license_number},
            format='json',
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert response.data == {
            'error': 'license_numbers must be a non-empty list of strings.',
        }

    def test_bulk_balance_excel_summary_shows_each_license_once_when_sheet_titles_collide(
        self,
        authenticated_client,
        test_company,
        test_port,
    ):
        """Summary sheet identifies each license by its own License No text
        (baked in at build time from `_util_return`), not by the
        openpyxl-deduplicated per-license sheet title — so duplicate-prefix
        license numbers each still get their own header block + Plan
        Utilization section, independent of how openpyxl renamed their
        sheets."""
        prefix = 'LIC-DUPLICATE-SHEET-NAME-12345'
        first = LicenseDetailsModel.objects.create(
            license_number=f'{prefix}A',
            exporter=test_company,
            port=test_port,
        )
        second = LicenseDetailsModel.objects.create(
            license_number=f'{prefix}B',
            exporter=test_company,
            port=test_port,
        )

        url = reverse('license:licenses-bulk-balance-excel')
        response = authenticated_client.post(
            url,
            {'license_numbers': [first.license_number, second.license_number]},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet_titles = workbook.sheetnames[1:]
        assert len(sheet_titles) == 2
        assert len(set(sheet_titles)) == 2

        summary_text = ' '.join(
            str(cell.value)
            for row in workbook['Utilization Planning Summary'].iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        )
        assert first.license_number in summary_text
        assert second.license_number in summary_text

    def test_bulk_balance_excel_summary_is_norm_grouped_planning_matrix(
        self,
        authenticated_client,
        test_license,
        test_company,
        test_port,
    ):
        """Utilization Planning Summary (3rd design): one section per
        distinct SION norm, each with a Planning Matrix pivoted by Planning
        Item Name, a Norm Total row, and a Planning Item Summary -- followed
        by a Grand Summary by Norm and a Grand Total.

        Covers a fabricated norm ("E999", never special-cased anywhere in
        the code) getting its own section for free, the two attribution
        rules (split-exclusive Planned qty/CIF vs. doubled-up Available qty
        across every item-name column with a visible split in that group --
        `item_pivot_report.py`'s `_build_license_row` convention), the
        untagged-split "Unassigned" fallback, and both cross-cutting
        invariants from the two-attribution-rules design."""
        from apps.core.models import HeadSIONNormsModel, ItemNameModel, SionNormClassModel
        from apps.license.models import LicenseExportItemModel, LicenseImportItemsModel, LicenseItemPlan

        # -- Unclassified norm (test_license, no export items): one import
        # item split across two planning item names -- the "Milk split into
        # SWP + DWP" scenario item_pivot_report.py's convention describes.
        swp = ItemNameModel.objects.create(name='Test SWP')
        dwp = ItemNameModel.objects.create(name='Test DWP')
        split_item = test_license.import_license.get(serial_number=1)
        split_item.items.set([swp, dwp])
        LicenseItemPlan.objects.create(
            import_item=split_item, item_name=swp,
            planned_quantity=Decimal('100.000'), unit_price=Decimal('5.00'),
            planned_cif_fc=Decimal('500.00'),
        )
        LicenseItemPlan.objects.create(
            import_item=split_item, item_name=dwp,
            planned_quantity=Decimal('50.000'), unit_price=Decimal('8.00'),
            planned_cif_fc=Decimal('400.00'),
        )

        # -- Fabricated "E999" norm license with an *untagged* split -> the
        # shared "Unassigned" bucket (must not be silently dropped).
        head = HeadSIONNormsModel.objects.create(name='Test Head Norm')
        fabricated_norm = SionNormClassModel.objects.create(head_norm=head, norm_class='E999')
        other_license = LicenseDetailsModel.objects.create(
            license_number='LIC-E999-FABRICATED-NORM-0001',
            exporter=test_company, port=test_port,
        )
        LicenseExportItemModel.objects.create(license=other_license, norm_class=fabricated_norm)
        other_item = LicenseImportItemsModel.objects.create(
            license=other_license, serial_number=1, description='Fabricated item',
            quantity=Decimal('300.000'), available_quantity=Decimal('300.000'),
            cif_fc=Decimal('3000.00'), cif_inr=Decimal('253500.00'),
        )
        LicenseItemPlan.objects.create(
            import_item=other_item, item_name=None,
            planned_quantity=Decimal('30.000'), unit_price=Decimal('5.00'),
            planned_cif_fc=Decimal('150.00'),
        )

        url = reverse('license:licenses-bulk-balance-excel')
        response = authenticated_client.post(
            url,
            {'license_numbers': [test_license.license_number, other_license.license_number]},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        summary = workbook['Utilization Planning Summary']
        all_cells = [cell for row in summary.iter_rows() for cell in row]

        # -- A fabricated norm not equal to E1/E5/E132 gets its own section,
        # with zero code changes (no norm name hardcoded anywhere).
        banner_rows = {
            c.value: c.row for c in all_cells
            if isinstance(c.value, str) and c.value.startswith('SION NORM : ')
        }
        assert set(banner_rows) == {'SION NORM : Unclassified', 'SION NORM : E999'}

        def _norm_total_row(banner_row):
            r = banner_row + 1
            while summary.cell(row=r, column=1).value != 'NORM TOTAL':
                r += 1
                assert r < banner_row + 50  # safety bound, not a real limit
            return r

        def _header_item_cols(banner_row):
            """{item_name: base_col} from the matrix's 2-row header."""
            hdr_row = banner_row + 1
            cols, col = {}, 5
            while summary.cell(row=hdr_row, column=col).value:
                cols[summary.cell(row=hdr_row, column=col).value] = col
                col += 3
            return cols

        # -- Unclassified section: DWP/SWP columns in alphabetical order,
        # doubled-up Available Qty, split-exclusive Planned Qty/CIF.
        uncl_banner = banner_rows['SION NORM : Unclassified']
        uncl_cols = _header_item_cols(uncl_banner)
        assert list(uncl_cols) == ['Test DWP', 'Test SWP']
        uncl_total_row = _norm_total_row(uncl_banner)
        dwp_col, swp_col = uncl_cols['Test DWP'], uncl_cols['Test SWP']
        # Both columns carry the SAME group's full available_quantity (the
        # import item's 1000) -- the intentional double count.
        assert summary.cell(row=uncl_total_row, column=dwp_col).value == pytest.approx(1000.0)
        assert summary.cell(row=uncl_total_row, column=swp_col).value == pytest.approx(1000.0)
        assert summary.cell(row=uncl_total_row, column=dwp_col + 1).value == pytest.approx(50.0)
        assert summary.cell(row=uncl_total_row, column=dwp_col + 2).value == pytest.approx(400.0)
        assert summary.cell(row=uncl_total_row, column=swp_col + 1).value == pytest.approx(100.0)
        assert summary.cell(row=uncl_total_row, column=swp_col + 2).value == pytest.approx(500.0)

        # -- E999 section: untagged split lands in the shared "Unassigned"
        # column, never silently dropped.
        e999_banner = banner_rows['SION NORM : E999']
        e999_cols = _header_item_cols(e999_banner)
        assert list(e999_cols) == ['Unassigned']
        e999_total_row = _norm_total_row(e999_banner)
        unassigned_col = e999_cols['Unassigned']
        assert summary.cell(row=e999_total_row, column=unassigned_col).value == pytest.approx(300.0)
        assert summary.cell(row=e999_total_row, column=unassigned_col + 1).value == pytest.approx(30.0)
        assert summary.cell(row=e999_total_row, column=unassigned_col + 2).value == pytest.approx(150.0)

        # -- Planning Item Summary reuses the SAME Norm Total figures, transposed.
        def _planning_item_summary(total_row):
            r = total_row + 2
            assert summary.cell(row=r, column=1).value == 'PLANNING ITEM SUMMARY'
            r += 2  # skip the section banner + its own header row
            rows = {}
            while summary.cell(row=r, column=1).value:
                rows[summary.cell(row=r, column=1).value] = (
                    summary.cell(row=r, column=2).value,
                    summary.cell(row=r, column=3).value,
                    summary.cell(row=r, column=4).value,
                )
                r += 1
            return rows

        uncl_item_summary = _planning_item_summary(uncl_total_row)
        assert uncl_item_summary['Test DWP'] == pytest.approx((1000.0, 50.0, 400.0))
        assert uncl_item_summary['Test SWP'] == pytest.approx((1000.0, 100.0, 500.0))
        e999_item_summary = _planning_item_summary(e999_total_row)
        assert e999_item_summary['Unassigned'] == pytest.approx((300.0, 30.0, 150.0))

        # -- Grand Summary by Norm / Grand Total.
        grand_summary_header_row = next(
            c.row for c in all_cells if c.value == 'GRAND SUMMARY BY NORM'
        ) + 1
        assert [summary.cell(row=grand_summary_header_row, column=c).value for c in range(1, 6)] == [
            'SION Norm', 'Licenses', 'Available Qty', 'Planned Qty', 'Planned CIF ($)',
        ]
        norm_summary = {}
        r = grand_summary_header_row + 1
        while summary.cell(row=r, column=1).value != 'GRAND TOTAL':
            v = summary.cell(row=r, column=1).value
            if v:
                norm_summary[v] = (
                    summary.cell(row=r, column=2).value,
                    summary.cell(row=r, column=3).value,
                    summary.cell(row=r, column=4).value,
                    summary.cell(row=r, column=5).value,
                )
            r += 1
            assert r < grand_summary_header_row + 20  # safety bound
        grand_total_row = r

        # Invariant B (coverage): every exported license lands in exactly
        # one norm row -- none silently dropped or double-counted.
        assert set(norm_summary) == {'Unclassified', 'E999'}
        assert norm_summary['Unclassified'][0] == 1
        assert norm_summary['E999'][0] == 1
        assert sum(v[0] for v in norm_summary.values()) == 2

        # Invariant A: Σ(Norm Total's Planned CIF across every item column,
        # incl. Unassigned) == Grand Summary by Norm's Planned CIF for that
        # norm == Σ(that norm's licenses' own per-license Plan Utilization
        # "Planned CIF" TOTALS) -- Grand Summary is a GROUP-level total,
        # never derived from the (intentionally double-counted) pivot
        # columns.
        def _per_license_planned_cif(license_number):
            sheet = workbook[license_number[:31]]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == 'TOTALS':
                        return sheet.cell(row=cell.row, column=8).value
            raise AssertionError(f'no TOTALS row found for {license_number}')

        assert norm_summary['Unclassified'][3] == pytest.approx(400.0 + 500.0)
        assert norm_summary['Unclassified'][3] == pytest.approx(
            _per_license_planned_cif(test_license.license_number)
        )
        assert norm_summary['E999'][3] == pytest.approx(150.0)
        assert norm_summary['E999'][3] == pytest.approx(
            _per_license_planned_cif(other_license.license_number)
        )

        # Invariant B (arithmetic): Σ(Grand Summary by Norm) == Grand Total.
        assert summary.cell(row=grand_total_row, column=1).value == 'GRAND TOTAL'
        assert summary.cell(row=grand_total_row, column=2).value == sum(
            v[0] for v in norm_summary.values()
        )
        assert summary.cell(row=grand_total_row, column=3).value == pytest.approx(
            sum(v[1] for v in norm_summary.values())
        )
        assert summary.cell(row=grand_total_row, column=4).value == pytest.approx(
            sum(v[2] for v in norm_summary.values())
        )
        assert summary.cell(row=grand_total_row, column=5).value == pytest.approx(
            sum(v[3] for v in norm_summary.values())
        )


@pytest.mark.api
@pytest.mark.database
class TestLicenseItemAPI:
    """Test License Import Item operations"""
    
    def test_list_license_items(self, authenticated_client, test_license):
        """Test GET /license-items/"""
        url = reverse('license:license-items-list')
        response = authenticated_client.get(url)
        
        assert response.status_code == status.HTTP_200_OK
    
    def test_filter_items_by_license(self, authenticated_client, test_license):
        """Test GET /license-items/?license={id}"""
        url = reverse('license:license-items-list')
        response = authenticated_client.get(url, {'license': test_license.id})
        
        assert response.status_code == status.HTTP_200_OK
        if 'results' in response.data:
            assert len(response.data['results']) == 3


@pytest.mark.api
@pytest.mark.integration
class TestLicenseLedgerUpload:
    """Test License Ledger CSV Upload"""
    
    def test_upload_ledger_csv(self, authenticated_client, tmp_path):
        """Test POST /upload-ledger/ with CSV file"""
        # Create a fake CSV file
        csv_content = """Regn.No.,Regn.Date,Lic.No.,Lic.Date,IEC,Scheme.Cd.,Port,Notification
1234,01/01/2024,0310831825,01/01/2024,0123456789,DFIA,INMUN1,NOTIFICATION
"""
        csv_file = tmp_path / "test_ledger.csv"
        csv_file.write_text(csv_content)
        
        url = reverse('license:upload-ledger')
        with open(csv_file, 'rb') as f:
            response = authenticated_client.post(url, {'ledger': f}, format='multipart')
        
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED]
        assert 'message' in response.data or 'licenses' in response.data
