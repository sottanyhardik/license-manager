"""
Tests for the License Purchase & Profit Report
(`apps.license.services.purchase_profit_report.build_purchase_profit_report`
and `apps.license.views.license_purchase_profit_report.LicensePurchaseProfitReportView`).

The report is a single, flat License Summary table. Covers:
- License selection: earliest QUALIFYING (external, non-internal-linked)
  `LicenseTrade` PURCHASE determines whether a license appears at all.
- Internal linked purchases (`linked_trade IS NOT NULL` — the auto-created
  mirror of an internal company-to-company transfer) are excluded from both
  the first-purchase-date, Purchase Amount, AND Purchase $.
- Purchase Amount (INR) / Purchase $ (`LicenseTradeLine.cif_fc`): full
  lifecycle sums of qualifying external purchase trades, never
  date-filtered once a license qualifies.
- Exporter: the license's own exporter/company, same as elsewhere in the
  License module.
- Exclude License Number: applied AFTER `license_number`/`norm`/
  `exporter_id` inclusion — always wins over an overlapping inclusion.
- Norm(s): every distinct raw SION norm_class attached to the license,
  sorted numerically and deduplicated — not bucketed into "Others".
- View-level contract: JSON shape, Excel sheet name, param validation,
  permissions — unchanged API contract, new payload shape.
"""
import re
import uuid
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A3, A4, landscape
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.core.constants import DEC_0
from apps.core.models import CompanyModel, HSCodeModel, HeadSIONNormsModel, ItemNameModel, SionNormClassModel
from apps.license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel
from apps.license.services.purchase_profit_report import build_purchase_profit_report
from apps.trade.models import LicenseTrade, LicenseTradeLine
from shared.pdf.builders import make_landscape_doc

User = get_user_model()

REPORT_URL = "/api/reports/license-purchase-profit/"

FROM_DATE = date.today() - timedelta(days=30)
TO_DATE = date.today()
OUT_OF_RANGE_DATE = date.today() - timedelta(days=90)


# ---------------------------------------------------------------------------
# Auth fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def report_viewer_client(db):
    user = User.objects.create_user(
        username="ppr-viewer",
        email="ppr-viewer@example.com",
        password="RoleP@ssw0rd123",
    )
    group, _ = Group.objects.get_or_create(name="REPORT_VIEWER")
    user.groups.add(group)
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


@pytest.fixture
def no_role_client(db):
    user = User.objects.create_user(
        username="ppr-norole",
        email="ppr-norole@example.com",
        password="RoleP@ssw0rd123",
    )
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


# ---------------------------------------------------------------------------
# Master-data / fixture builders
# ---------------------------------------------------------------------------

@pytest.fixture
def ppr_masters(db):
    head_norm = HeadSIONNormsModel.objects.create(name="PPR Test Head Norm")
    return {
        "exporter": CompanyModel.objects.create(iec="7770001111", name="PPR Exporter"),
        "supplier_early": CompanyModel.objects.create(iec="7770002222", name="PPR Supplier Early"),
        "supplier_late": CompanyModel.objects.create(iec="7770003333", name="PPR Supplier Late"),
        "hs_code": HSCodeModel.objects.create(hs_code="88888888", product_description="PPR Test Product"),
        "item_a": ItemNameModel.objects.create(name="PPR Item A"),
        "item_b": ItemNameModel.objects.create(name="PPR Item B"),
        "e1_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E1", is_active=True),
        "e5_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E5", is_active=True),
        "e132_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E132", is_active=True),
        "other_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="PPROTHER", is_active=True),
    }


def _make_license(number, exporter):
    return LicenseDetailsModel.objects.create(
        license_number=number,
        license_date=date.today() - timedelta(days=180),
        license_expiry_date=date.today() + timedelta(days=180),
        exporter=exporter,
    )


def _make_export_item(license_obj, norm_class, cif_fc):
    return LicenseExportItemModel.objects.create(
        license=license_obj,
        description=f"Export item for {license_obj.license_number}",
        norm_class=norm_class,
        cif_fc=cif_fc,
        cif_inr=cif_fc * Decimal("84.5"),
    )


def _make_import_item(license_obj, hs_code, serial, item_names=None, description=None, cif_fc=None):
    item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=serial,
        description=description or f"Import item {serial}",
        hs_code=hs_code,
        quantity=Decimal("1000.000"),
        available_quantity=Decimal("1000.000"),
        cif_fc=cif_fc if cif_fc is not None else DEC_0,
    )
    if item_names:
        item.items.set(item_names)
    return item


def _make_purchase_trade(item, amount_inr, invoice_date=None, cif_fc=None, from_company=None):
    """Genuine EXTERNAL purchase: a `LicenseTrade(direction=PURCHASE)` +
    one billed `LicenseTradeLine` against `item`, with no `linked_trade` —
    `amount_inr` is passed straight through as the line's exact INR amount
    (MODE_QTY with qty=1). `cif_fc` is the line's foreign-currency/USD CIF
    value (Purchase $'s source) — defaults to 0 when not exercising that
    specific field. `from_company` is the supplier (Purchase From's
    source) — left unset (`None`) when not exercising that field."""
    trade = LicenseTrade.objects.create(
        direction=LicenseTrade.DIR_PURCHASE,
        invoice_date=invoice_date or date.today(),
        invoice_number=f"PUR-{uuid.uuid4().hex[:8]}",
        from_company=from_company,
    )
    LicenseTradeLine.objects.create(
        trade=trade,
        sr_number=item,
        mode=LicenseTradeLine.MODE_QTY,
        qty_kg=Decimal("1.000"),
        rate_inr_per_kg=Decimal(amount_inr),
        amount_inr=Decimal(amount_inr),
        cif_fc=Decimal(cif_fc) if cif_fc is not None else Decimal("0.00"),
    )
    return trade


def _make_internal_linked_purchase_trade(item, amount_inr, invoice_date=None, cif_fc=None):
    """Mirrors the real `auto_create_paired` flow (`apps/trade/serializers.py`):
    an internal transfer between the business's own companies creates a
    PURCHASE trade AND its auto-generated SALE counterpart, cross-linked via
    `linked_trade` on BOTH sides. Must be ignored entirely for first-purchase
    -date, Purchase Amount, AND Purchase $ — it's a ledger-balancing mirror,
    not a real external acquisition."""
    purchase = _make_purchase_trade(item, amount_inr, invoice_date=invoice_date, cif_fc=cif_fc)
    mirror = LicenseTrade.objects.create(
        direction=LicenseTrade.DIR_SALE,
        invoice_date=purchase.invoice_date,
        invoice_number=f"SALE-{uuid.uuid4().hex[:8]}",
        linked_trade=purchase,
    )
    LicenseTrade.objects.filter(pk=purchase.pk).update(linked_trade=mirror)
    purchase.refresh_from_db()
    return purchase


def _make_debit_trade(item, amount_inr, invoice_date=None, cif_fc=None, qty_kg=None):
    """Genuine EXTERNAL debit: a `LicenseTrade(direction=SALE)` + one billed
    `LicenseTradeLine` against `item`, with no `linked_trade` — the debit-
    side counterpart of `_make_purchase_trade`. `qty_kg` defaults to 1 (MODE_QTY,
    `amount_inr` passed straight through as the line's exact INR amount)."""
    trade = LicenseTrade.objects.create(
        direction=LicenseTrade.DIR_SALE,
        invoice_date=invoice_date or date.today(),
        invoice_number=f"DEBIT-{uuid.uuid4().hex[:8]}",
    )
    LicenseTradeLine.objects.create(
        trade=trade,
        sr_number=item,
        mode=LicenseTradeLine.MODE_QTY,
        qty_kg=Decimal(qty_kg) if qty_kg is not None else Decimal("1.000"),
        rate_inr_per_kg=Decimal(amount_inr),
        amount_inr=Decimal(amount_inr),
        cif_fc=Decimal(cif_fc) if cif_fc is not None else Decimal("0.00"),
    )
    return trade


def _make_internal_linked_debit_trade(item, amount_inr, invoice_date=None, cif_fc=None, qty_kg=None):
    """Mirrors `_make_internal_linked_purchase_trade`, but for the debit
    (SALE) side: an internal transfer's SALE leg, cross-linked via
    `linked_trade` on BOTH sides with its auto-generated PURCHASE
    counterpart. Must be excluded entirely from the debit aggregation."""
    debit = _make_debit_trade(item, amount_inr, invoice_date=invoice_date, cif_fc=cif_fc, qty_kg=qty_kg)
    mirror = LicenseTrade.objects.create(
        direction=LicenseTrade.DIR_PURCHASE,
        invoice_date=debit.invoice_date,
        invoice_number=f"PUR-{uuid.uuid4().hex[:8]}",
        linked_trade=debit,
    )
    LicenseTrade.objects.filter(pk=debit.pk).update(linked_trade=mirror)
    debit.refresh_from_db()
    return debit


# ---------------------------------------------------------------------------
# Service-level tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_single_license_single_purchase_trade(ppr_masters):
    lic = _make_license("PPR-001", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("95000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(
        item, "60000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="705.50",
    )

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert set(report.keys()) == {"summary", "licenses", "item_matrix"}
    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-001")
    assert lic_row["purchase_amount"] == 60000.00
    assert lic_row["purchase_usd"] == 705.50
    assert lic_row["exporter"] == ppr_masters["exporter"].name
    assert lic_row["license_date"] == lic.license_date.isoformat()
    assert lic_row["expiry_date"] == lic.license_expiry_date.isoformat()
    assert lic_row["norms"] == ["E1"]


@pytest.mark.django_db
def test_multiple_purchase_trades_consolidate_to_one_purchase_amount(ppr_masters):
    """Core acceptance criterion: one license, many qualifying external
    purchase trades -> one Purchase Amount, consolidated to their sum,
    even when some trades are dated OUTSIDE the report's range — only the
    EARLIEST trade's date determines whether the license qualifies at all."""
    lic = _make_license("PPR-002", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "20000.00", invoice_date=FROM_DATE + timedelta(days=1))  # earliest, in range
    _make_purchase_trade(item, "15000.00", invoice_date=FROM_DATE + timedelta(days=10))  # in range
    _make_purchase_trade(item, "5000.00", invoice_date=TO_DATE + timedelta(days=50))  # OUT of range, still counted

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-002")
    assert lic_row["purchase_amount"] == 40000.00  # 20000 + 15000 + 5000


@pytest.mark.django_db
def test_first_purchase_date_outside_range_excludes_license(ppr_masters):
    """A license whose EARLIEST qualifying purchase trade falls outside the
    report's date range must not appear at all."""
    lic = _make_license("PPR-OUT-OF-RANGE", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=OUT_OF_RANGE_DATE - timedelta(days=1))

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    license_numbers = {r["license_number"] for r in report["licenses"]}
    assert "PPR-OUT-OF-RANGE" not in license_numbers


@pytest.mark.django_db
def test_internal_linked_purchase_excluded_from_amount_and_first_date(ppr_masters):
    lic = _make_license("PPR-INTERNAL", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])

    # Internal linked purchase: earlier date, much larger amount — must be
    # entirely ignored for BOTH first-purchase-date and Purchase Amount.
    _make_internal_linked_purchase_trade(
        item, "999999.00", invoice_date=OUT_OF_RANGE_DATE - timedelta(days=500),
    )

    # The only genuine external purchase — this alone should determine
    # both qualification and Purchase Amount.
    _make_purchase_trade(item, "25000.00", invoice_date=FROM_DATE + timedelta(days=1))

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-INTERNAL")
    assert lic_row["purchase_amount"] == 25000.00  # NOT 999999 + 25000


@pytest.mark.django_db
def test_license_with_only_internal_linked_purchase_does_not_qualify(ppr_masters):
    lic = _make_license("PPR-ONLY-INTERNAL", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_internal_linked_purchase_trade(item, "40000.00", invoice_date=FROM_DATE + timedelta(days=1))

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert report["licenses"] == []


@pytest.mark.django_db
def test_purchase_usd_consolidates_across_trades_and_excludes_internal_linked(ppr_masters):
    lic = _make_license("PPR-USD", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])

    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="120.00")
    _make_purchase_trade(item, "5000.00", invoice_date=FROM_DATE + timedelta(days=5), cif_fc="60.00")
    # Internal linked purchase: must NOT contribute to Purchase $ either.
    _make_internal_linked_purchase_trade(
        item, "999999.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="88888.00",
    )

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-USD")
    assert lic_row["purchase_amount"] == 15000.00  # 10000 + 5000, NOT + 999999
    assert lic_row["purchase_usd"] == 180.00  # 120 + 60, NOT + 88888


@pytest.mark.django_db
def test_sale_amount_and_sale_usd_sum_qualifying_sale_lines_excluding_internal_linked(ppr_masters):
    """Sale Amount/Sale $ consolidate across every qualifying (external,
    non-internal-linked) SALE trade line for the license — the disposal-
    side mirror of the existing Purchase Amount/Purchase $ tests above."""
    lic = _make_license("PPR-SALE", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="100.00")

    _make_debit_trade(item, "4000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="50.00", qty_kg="100.000")
    _make_debit_trade(item, "1000.00", invoice_date=FROM_DATE + timedelta(days=3), cif_fc="10.00", qty_kg="20.000")
    # Internal linked SALE trade: must NOT contribute to Sale Amount/Sale $.
    _make_internal_linked_debit_trade(
        item, "999999.00", invoice_date=FROM_DATE + timedelta(days=4), cif_fc="88888.00", qty_kg="500.000",
    )

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-SALE")
    assert lic_row["sale_amount"] == 5000.00  # 4000 + 1000, NOT + 999999
    assert lic_row["sale_usd"] == 60.00  # 50 + 10, NOT + 88888


@pytest.mark.django_db
def test_profit_loss_equals_sale_amount_minus_purchase_amount(ppr_masters):
    lic = _make_license("PPR-PROFIT", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "16000.00", invoice_date=FROM_DATE + timedelta(days=2), qty_kg="100.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-PROFIT")
    assert lic_row["purchase_amount"] == 10000.00
    assert lic_row["sale_amount"] == 16000.00
    assert lic_row["profit_loss"] == 6000.00  # 16000 - 10000


@pytest.mark.django_db
def test_sale_amount_not_doubled_for_multi_item_name_import_item(ppr_masters):
    """Regression test for the double-counting fix: an Import Item with 2
    `ItemNameModel`s attached (which the Item Utilization Matrix's
    `debit_qty`/`debit_cif`/`debit_bill` dicts deliberately duplicate the
    debit across, for THAT table's display purpose) must NOT cause
    Sale Amount/Sale $ to be doubled — they come from a separate,
    per-license-only aggregate, not from summing the item-matrix's
    per-item-name cells."""
    name_x = ItemNameModel.objects.create(name="PPR Sale Multi X")
    name_y = ItemNameModel.objects.create(name="PPR Sale Multi Y")
    lic = _make_license("PPR-SALE-MULTI-NAME", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[name_x, name_y])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    # Exactly ONE genuine SALE trade line against the multi-name item.
    _make_debit_trade(item, "4000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="80.00", qty_kg="200.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    # Sanity check: the item_matrix DOES duplicate this debit across both
    # header columns (the correct, documented DISPLAY convention there).
    lic_row_matrix = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-SALE-MULTI-NAME")
    for name in (name_x.name, name_y.name):
        assert lic_row_matrix["items"][name] == {"qty": 200.0, "cif": 80.0, "bill": 4000.0}

    # But Sale Amount/Sale $ on the License Summary row must equal that ONE
    # trade line's raw amount exactly once, NOT doubled to 8000.0/160.0.
    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-SALE-MULTI-NAME")
    assert lic_row["sale_amount"] == 4000.00
    assert lic_row["sale_usd"] == 80.00


@pytest.mark.django_db
def test_exclude_license_number_overrides_license_number_inclusion(ppr_masters):
    """The exclusion filter takes precedence: first License Number/Norm/
    Exporter determine the included set, then Exclude License Number
    removes any of those explicitly."""
    numbers = ["0311050703", "0311051359", "0311051945"]
    for number in numbers:
        lic = _make_license(number, ppr_masters["exporter"])
        _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
        item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
        _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    report = build_purchase_profit_report(
        FROM_DATE, TO_DATE, norm="All",
        license_number=numbers, exclude_license_number=["0311051359"],
    )

    license_numbers = {r["license_number"] for r in report["licenses"]}
    assert license_numbers == {"0311050703", "0311051945"}
    assert "0311051359" not in license_numbers


@pytest.mark.django_db
def test_norms_column_sorted_numerically_and_deduplicated(ppr_masters):
    lic = _make_license("PPR-NORMS", ppr_masters["exporter"])
    # Deliberately out of order, with a duplicate norm_class across two
    # export items.
    _make_export_item(lic, ppr_masters["e132_norm"], Decimal("30000.00"))
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("20000.00"))
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("10000.00"))  # duplicate norm_class
    _make_export_item(lic, ppr_masters["e5_norm"], Decimal("15000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-NORMS")
    assert lic_row["norms"] == ["E1", "E5", "E132"]  # numeric order, deduplicated


@pytest.mark.django_db
def test_norms_column_shows_raw_codes_outside_conversion_norms(ppr_masters):
    """A norm_class outside CONVERSION_NORMS still appears in the Norm(s)
    column under its own raw code — the display column is never bucketed
    into "Others" the way the `norm` FILTER parameter is."""
    lic = _make_license("PPR-RAW-NORM", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["other_norm"], Decimal("30000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-RAW-NORM")
    assert lic_row["norms"] == ["PPROTHER"]


@pytest.mark.django_db
def test_norm_filter_still_narrows_licenses_no_output_grouping(ppr_masters):
    """`norm` remains a valid filter parameter, but the report's output is a
    flat list — never grouped/bucketed by norm."""
    lic_e1 = _make_license("PPR-NORM-E1", ppr_masters["exporter"])
    _make_export_item(lic_e1, ppr_masters["e1_norm"], Decimal("50000.00"))
    item_e1 = _make_import_item(lic_e1, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item_e1, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    lic_other = _make_license("PPR-NORM-OTHER", ppr_masters["exporter"])
    _make_export_item(lic_other, ppr_masters["other_norm"], Decimal("50000.00"))
    item_other = _make_import_item(lic_other, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_b"]])
    _make_purchase_trade(item_other, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    # No norm_class at all -> out of scope, excluded entirely.
    lic_no_norm = _make_license("PPR-NORM-NONE", ppr_masters["exporter"])
    item_no_norm = _make_import_item(lic_no_norm, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item_no_norm, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    all_report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")
    assert set(all_report.keys()) == {"summary", "licenses", "item_matrix"}
    all_license_numbers = {r["license_number"] for r in all_report["licenses"]}
    assert "PPR-NORM-E1" in all_license_numbers
    assert "PPR-NORM-OTHER" in all_license_numbers
    assert "PPR-NORM-NONE" not in all_license_numbers

    e1_only = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="E1")
    assert {r["license_number"] for r in e1_only["licenses"]} == {"PPR-NORM-E1"}

    others_only = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="Others")
    assert {r["license_number"] for r in others_only["licenses"]} == {"PPR-NORM-OTHER"}


@pytest.mark.django_db
def test_purchase_from_is_earliest_qualifying_trades_supplier(ppr_masters):
    """When a license has 2+ qualifying purchases from different
    suppliers, `purchase_from` is the supplier of the EARLIEST one by
    date — verified here by making the earliest trade's supplier
    unambiguous even though it is created/saved AFTER the later trade."""
    lic = _make_license("PPR-FROM", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])

    # Later trade (by invoice_date) created FIRST, to prove ordering is by
    # date, not by insertion/creation order.
    _make_purchase_trade(
        item, "15000.00", invoice_date=FROM_DATE + timedelta(days=10),
        from_company=ppr_masters["supplier_late"],
    )
    _make_purchase_trade(
        item, "20000.00", invoice_date=FROM_DATE + timedelta(days=1),
        from_company=ppr_masters["supplier_early"],
    )

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-FROM")
    assert lic_row["purchase_from"] == ppr_masters["supplier_early"].name
    assert lic_row["purchase_amount"] == 35000.00  # both still consolidated


@pytest.mark.django_db
def test_balance_cif_equals_purchase_usd_minus_sale_usd(ppr_masters):
    """Balance CIF ($) = Purchase $ − Sale $ (Original/acquired CIF $ minus
    Debited/utilized CIF $) — this report's OWN already-computed figures,
    never the broader `LicenseBalanceCalculator` engine (which also
    factors in BOE debits/allotments/an opening-balance anchor this report
    doesn't track or display anywhere else, and would silently diverge
    from the Purchase $/Sale $ columns shown right next to it)."""
    lic = _make_license("PPR-BALANCE", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="500.00")
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="180.00", qty_kg="50.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["licenses"] if r["license_number"] == "PPR-BALANCE")
    assert lic_row["purchase_usd"] == 500.00
    assert lic_row["sale_usd"] == 180.00
    assert lic_row["balance_cif"] == 320.00  # 500.00 - 180.00


@pytest.mark.django_db
def test_summary_totals_match_sum_of_license_rows(ppr_masters):
    """The `summary` block's totals must match manually-computed sums
    across the returned `licenses` rows for a multi-license scenario."""
    lic_1 = _make_license("PPR-SUM-001", ppr_masters["exporter"])
    _make_export_item(lic_1, ppr_masters["e1_norm"], Decimal("50000.00"))
    item_1 = _make_import_item(lic_1, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item_1, "10000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="120.00")

    lic_2 = _make_license("PPR-SUM-002", ppr_masters["exporter"])
    _make_export_item(lic_2, ppr_masters["e5_norm"], Decimal("30000.00"))
    item_2 = _make_import_item(lic_2, ppr_masters["hs_code"], 2, item_names=[ppr_masters["item_b"]])
    _make_purchase_trade(item_2, "25000.50", invoice_date=FROM_DATE + timedelta(days=3), cif_fc="310.25")
    _make_debit_trade(item_1, "5000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="60.00", qty_kg="10.000")
    _make_debit_trade(item_2, "9000.00", invoice_date=FROM_DATE + timedelta(days=4), cif_fc="95.00", qty_kg="15.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    rows = [
        r for r in report["licenses"] if r["license_number"] in {"PPR-SUM-001", "PPR-SUM-002"}
    ]
    assert len(rows) == 2
    assert report["summary"]["total_licenses"] == len(report["licenses"])
    assert report["summary"]["purchase_amount"] == round(sum(r["purchase_amount"] for r in report["licenses"]), 2)
    assert report["summary"]["purchase_usd"] == round(sum(r["purchase_usd"] for r in report["licenses"]), 2)
    assert report["summary"]["balance_cif"] == round(sum(r["balance_cif"] for r in report["licenses"]), 2)
    assert report["summary"]["total_sale_amount"] == round(sum(r["sale_amount"] for r in report["licenses"]), 2)
    assert report["summary"]["total_sale_usd"] == round(sum(r["sale_usd"] for r in report["licenses"]), 2)


@pytest.mark.django_db
def test_total_profit_loss_from_raw_sums_diverges_from_naive_item_matrix_sum(ppr_masters):
    """Grand `total_profit_loss` must equal `total_sale_amount -
    purchase_amount` computed from the RAW per-license ledger sums (the
    `sale_agg` aggregate) — never by re-deriving a "Sale Amount" from the
    Item Utilization Matrix's per-header totals (which duplicate a
    multi-item-name Import Item's debit across every header it maps to).
    This fixture deliberately includes such a multi-name item so the two
    approaches would visibly diverge if the wrong (item-matrix-derived) one
    were used instead of the correct one."""
    name_x = ItemNameModel.objects.create(name="PPR Grand Multi X")
    name_y = ItemNameModel.objects.create(name="PPR Grand Multi Y")

    # License 1: a multi-item-name Import Item with ONE genuine SALE line —
    # the item_matrix duplicates this debit across BOTH its headers.
    lic1 = _make_license("PPR-GRAND-001", ppr_masters["exporter"])
    _make_export_item(lic1, ppr_masters["e1_norm"], Decimal("50000.00"))
    item1 = _make_import_item(lic1, ppr_masters["hs_code"], 1, item_names=[name_x, name_y])
    _make_purchase_trade(item1, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item1, "4000.00", invoice_date=FROM_DATE + timedelta(days=2), qty_kg="200.000")

    # License 2: a plain single-name item, purchase + sale.
    lic2 = _make_license("PPR-GRAND-002", ppr_masters["exporter"])
    _make_export_item(lic2, ppr_masters["e5_norm"], Decimal("30000.00"))
    item2 = _make_import_item(lic2, ppr_masters["hs_code"], 2, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item2, "8000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item2, "5000.00", invoice_date=FROM_DATE + timedelta(days=2), qty_kg="50.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    rows = [r for r in report["licenses"] if r["license_number"] in {"PPR-GRAND-001", "PPR-GRAND-002"}]
    assert len(rows) == 2

    # Correct raw sums, from the actual trade-ledger fixture values.
    expected_total_purchase = Decimal("10000.00") + Decimal("8000.00")
    expected_total_sale = Decimal("4000.00") + Decimal("5000.00")  # NOT 8000 + 5000 = 13000
    expected_profit_loss = float(expected_total_sale - expected_total_purchase)

    assert report["summary"]["purchase_amount"] == float(expected_total_purchase)
    assert report["summary"]["total_sale_amount"] == float(expected_total_sale)
    assert report["summary"]["total_profit_loss"] == expected_profit_loss

    # Demonstrate the divergence: naively summing the item_matrix's own
    # per-header "bill" totals (the wrong source for a grand total) DOES
    # double-count license 1's debit and would NOT match the correct
    # total_sale_amount above.
    naive_sale_from_matrix = sum(
        h["bill"] for h in report["item_matrix"]["totals"].values()
    )
    assert naive_sale_from_matrix != report["summary"]["total_sale_amount"]
    assert naive_sale_from_matrix == 13000.0  # 4000 (x2 headers) + 5000


# ---------------------------------------------------------------------------
# Dynamic Import Item Utilization Matrix — service-level tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_item_matrix_headers_sorted_case_insensitive_and_deduplicated(ppr_masters):
    """Headers come from Import Item names, sorted case-insensitively and
    deduplicated across licenses/import items, regardless of creation or
    trade order."""
    apple = ItemNameModel.objects.create(name="apple")
    banana = ItemNameModel.objects.create(name="Banana")

    lic1 = _make_license("PPR-HDR-001", ppr_masters["exporter"])
    _make_export_item(lic1, ppr_masters["e1_norm"], Decimal("50000.00"))
    item1 = _make_import_item(lic1, ppr_masters["hs_code"], 1, item_names=[banana])
    _make_purchase_trade(item1, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    # A non-zero debit is required so Dynamic Column Optimization doesn't
    # drop this header (an all-zero-Grand-Total header would be removed —
    # see the dedicated `test_item_matrix_*_optimization*` tests below).
    _make_debit_trade(item1, "1000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="15.00", qty_kg="10.000")

    lic2 = _make_license("PPR-HDR-002", ppr_masters["exporter"])
    _make_export_item(lic2, ppr_masters["e1_norm"], Decimal("50000.00"))
    # Two import items on the SAME license reuse the same "apple" item name
    # — must collapse to one header, not two.
    item2a = _make_import_item(lic2, ppr_masters["hs_code"], 1, item_names=[apple])
    item2b = _make_import_item(lic2, ppr_masters["hs_code"], 2, item_names=[apple])
    _make_purchase_trade(item2a, "5000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_purchase_trade(item2b, "5000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item2a, "500.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="8.00", qty_kg="5.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert report["item_matrix"]["headers"] == ["apple", "Banana"]


@pytest.mark.django_db
def test_item_matrix_aggregates_multiple_sale_lines_on_same_import_item(ppr_masters):
    """Multiple SALE trade lines against the same import item aggregate into
    one qty/cif/bill cell."""
    lic = _make_license("PPR-DEBIT-AGG", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")
    _make_debit_trade(item, "2000.00", invoice_date=FROM_DATE + timedelta(days=3), cif_fc="10.00", qty_kg="50.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-DEBIT-AGG")
    cell = lic_row["items"][ppr_masters["item_a"].name]
    assert cell == {"qty": 150.0, "cif": 50.0, "bill": 5000.0}


@pytest.mark.django_db
def test_item_matrix_zero_fills_header_with_no_debit(ppr_masters):
    """A license with an import item but zero SALE trades shows the
    zero-filled dict for that header — the key must be present, not
    omitted. A SECOND license shares the same item name WITH a debit, so
    the header's Grand Total is non-zero and Dynamic Column Optimization
    keeps the column — this test is about per-row zero-fill, not the
    all-zero-Grand-Total column removal covered separately below."""
    lic = _make_license("PPR-ZERO-DEBIT", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    # Deliberately no SALE/debit trade against `item`.

    lic_other = _make_license("PPR-ZERO-DEBIT-PEER", ppr_masters["exporter"])
    _make_export_item(lic_other, ppr_masters["e1_norm"], Decimal("50000.00"))
    item_other = _make_import_item(lic_other, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item_other, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item_other, "1000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="15.00", qty_kg="10.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert ppr_masters["item_a"].name in report["item_matrix"]["headers"]
    lic_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-ZERO-DEBIT")
    assert ppr_masters["item_a"].name in lic_row["items"]
    assert lic_row["items"][ppr_masters["item_a"].name] == {"qty": 0, "cif": 0.0, "bill": 0.0}


@pytest.mark.django_db
def test_item_matrix_multi_name_import_item_contributes_full_debit_to_both(ppr_masters):
    """An import item with 2 `ItemNameModel`s attached contributes its FULL
    debit amount to BOTH header columns — never split."""
    item_x = ItemNameModel.objects.create(name="PPR Multi X")
    item_y = ItemNameModel.objects.create(name="PPR Multi Y")
    lic = _make_license("PPR-MULTI-NAME", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[item_x, item_y])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "4000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="80.00", qty_kg="200.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-MULTI-NAME")
    for name in (item_x.name, item_y.name):
        assert lic_row["items"][name] == {"qty": 200.0, "cif": 80.0, "bill": 4000.0}


@pytest.mark.django_db
def test_item_matrix_totals_match_manual_sums(ppr_masters):
    """`item_matrix.totals` must equal manually-computed sums across the
    test fixtures. (The top-level `summary.total_sale_usd`/
    `total_sale_amount` are NOT derived from these per-header totals — see
    the dedicated Sale Amount/Sale $ tests below — so they are not
    asserted against `item_matrix` sums here.)"""
    item_a_name = ppr_masters["item_a"].name

    lic1 = _make_license("PPR-MTOT-001", ppr_masters["exporter"])
    _make_export_item(lic1, ppr_masters["e1_norm"], Decimal("50000.00"))
    item1 = _make_import_item(lic1, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item1, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item1, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    lic2 = _make_license("PPR-MTOT-002", ppr_masters["exporter"])
    _make_export_item(lic2, ppr_masters["e5_norm"], Decimal("30000.00"))
    item2 = _make_import_item(lic2, ppr_masters["hs_code"], 2, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item2, "8000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item2, "1500.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="20.00", qty_kg="50.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    rows = [
        r for r in report["item_matrix"]["rows"] if r["license_number"] in {"PPR-MTOT-001", "PPR-MTOT-002"}
    ]
    assert len(rows) == 2
    manual_cif = sum(r["items"][item_a_name]["cif"] for r in rows)
    manual_bill = sum(r["items"][item_a_name]["bill"] for r in rows)
    assert report["item_matrix"]["totals"][item_a_name]["cif"] == round(manual_cif, 2)
    assert report["item_matrix"]["totals"][item_a_name]["bill"] == round(manual_bill, 2)


@pytest.mark.django_db
def test_item_matrix_excludes_internal_linked_sale_trades(ppr_masters):
    """Internal-linked SALE trades (`trade.linked_trade` set) are excluded
    from the debit aggregation — same pattern as the existing
    internal-linked PURCHASE exclusion test."""
    lic = _make_license("PPR-INTERNAL-DEBIT", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    # Internal linked debit: much larger amount — must be entirely ignored.
    _make_internal_linked_debit_trade(
        item, "999999.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="88888.00", qty_kg="500.000",
    )
    # The only genuine external debit.
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=3), cif_fc="40.00", qty_kg="100.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    lic_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-INTERNAL-DEBIT")
    cell = lic_row["items"][ppr_masters["item_a"].name]
    assert cell == {"qty": 100.0, "cif": 40.0, "bill": 3000.0}  # NOT +500/+88888/+999999


@pytest.mark.django_db
def test_item_matrix_optimization_removes_all_zero_grand_total_header(ppr_masters):
    """Dynamic Column Optimization: an Import Item column group whose Grand
    Total Qty/CIF $/Bill are ALL zero across every qualifying license is
    dropped entirely — from `headers`, from every row's `items`, and from
    `totals` — while a header with a genuine debit anywhere is kept."""
    keep_name = ItemNameModel.objects.create(name="PPR Opt Keep")
    drop_name = ItemNameModel.objects.create(name="PPR Opt Drop")

    lic_keep = _make_license("PPR-OPT-KEEP", ppr_masters["exporter"])
    _make_export_item(lic_keep, ppr_masters["e1_norm"], Decimal("50000.00"))
    item_keep = _make_import_item(lic_keep, ppr_masters["hs_code"], 1, item_names=[keep_name])
    _make_purchase_trade(item_keep, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item_keep, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    lic_drop = _make_license("PPR-OPT-DROP", ppr_masters["exporter"])
    _make_export_item(lic_drop, ppr_masters["e5_norm"], Decimal("30000.00"))
    item_drop = _make_import_item(lic_drop, ppr_masters["hs_code"], 2, item_names=[drop_name])
    _make_purchase_trade(item_drop, "8000.00", invoice_date=FROM_DATE + timedelta(days=1))
    # Deliberately no SALE/debit trade against `item_drop` anywhere in the
    # report — its Grand Total Qty/CIF/Bill are all zero.

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert keep_name.name in report["item_matrix"]["headers"]
    assert drop_name.name not in report["item_matrix"]["headers"]
    assert drop_name.name not in report["item_matrix"]["totals"]
    for row in report["item_matrix"]["rows"]:
        assert drop_name.name not in row["items"]


@pytest.mark.django_db
def test_item_matrix_optimization_keeps_header_when_only_one_license_is_zero(ppr_masters):
    """The removal decision is based on the GRAND TOTAL, never on an
    individual license's row — a header must NOT be dropped just because
    one license (among several sharing it) has zero values."""
    shared_name = ppr_masters["item_a"]

    lic_zero = _make_license("PPR-OPT-ONE-ZERO", ppr_masters["exporter"])
    _make_export_item(lic_zero, ppr_masters["e1_norm"], Decimal("50000.00"))
    item_zero = _make_import_item(lic_zero, ppr_masters["hs_code"], 1, item_names=[shared_name])
    _make_purchase_trade(item_zero, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    # No debit for this license.

    lic_nonzero = _make_license("PPR-OPT-ONE-NONZERO", ppr_masters["exporter"])
    _make_export_item(lic_nonzero, ppr_masters["e5_norm"], Decimal("30000.00"))
    item_nonzero = _make_import_item(lic_nonzero, ppr_masters["hs_code"], 2, item_names=[shared_name])
    _make_purchase_trade(item_nonzero, "8000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item_nonzero, "1500.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="20.00", qty_kg="50.000")

    report = build_purchase_profit_report(FROM_DATE, TO_DATE, norm="All")

    assert shared_name.name in report["item_matrix"]["headers"]
    zero_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-OPT-ONE-ZERO")
    assert zero_row["items"][shared_name.name] == {"qty": 0, "cif": 0.0, "bill": 0.0}
    nonzero_row = next(r for r in report["item_matrix"]["rows"] if r["license_number"] == "PPR-OPT-ONE-NONZERO")
    assert nonzero_row["items"][shared_name.name] == {"qty": 50.0, "cif": 20.0, "bill": 1500.0}


# ---------------------------------------------------------------------------
# View-level tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_view_json_response_has_expected_top_level_keys(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-VIEW-001", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code == 200
    data = response.json()
    assert set(data.keys()) == {"summary", "licenses", "item_matrix"}
    assert isinstance(data["licenses"], list)
    assert set(data["summary"].keys()) == {
        "total_licenses", "purchase_amount", "purchase_usd", "balance_cif",
        "total_sale_usd", "total_sale_amount", "total_profit_loss",
    }
    row = data["licenses"][0]
    assert set(row.keys()) >= {
        "license_number", "license_date", "expiry_date", "exporter", "norms",
        "purchase_from", "purchase_amount", "purchase_usd",
        "sale_amount", "sale_usd", "profit_loss", "balance_cif",
    }
    assert "license_id" not in row

    assert set(data["item_matrix"].keys()) == {"headers", "rows", "totals"}
    assert isinstance(data["item_matrix"]["headers"], list)
    assert isinstance(data["item_matrix"]["rows"], list)
    assert isinstance(data["item_matrix"]["totals"], dict)


@pytest.mark.django_db
def test_view_exclude_license_number_param_is_comma_separated(report_viewer_client, ppr_masters):
    numbers = ["PPR-KEEP", "PPR-DROP"]
    for number in numbers:
        lic = _make_license(number, ppr_masters["exporter"])
        _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
        item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
        _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {
            "from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(),
            "exclude_license_number": "PPR-DROP, PPR-NONEXISTENT",
        },
    )
    assert response.status_code == 200
    license_numbers = {r["license_number"] for r in response.json()["licenses"]}
    assert license_numbers == {"PPR-KEEP"}


@pytest.mark.django_db
def test_view_excel_export_returns_valid_workbook(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-VIEW-002", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["License Summary", "Item Utilization Matrix"]


@pytest.mark.django_db
def test_view_excel_item_matrix_grand_total_row_shows_static_column_totals(report_viewer_client, ppr_masters):
    """The Item Utilization Matrix sheet's GRAND TOTAL row must show real
    values for Purchase Amount/Purchase $/Sale Amount/Sale $/Profit-Loss/
    Balance CIF ($) (columns 7-12) — not blanks left over from a label that
    used to span all 12 static columns. Cross-checked against the same
    request's JSON `summary` so a column-offset regression (e.g. `n_static`
    vs. the new `n_label` getting confused) would fail this test."""
    lic = _make_license("PPR-XL-TOTALS", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    params = {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    json_response = report_viewer_client.get(REPORT_URL, params)
    summary = json_response.json()["summary"]

    excel_response = report_viewer_client.get(REPORT_URL, {**params, "format": "excel"})
    workbook = load_workbook(BytesIO(excel_response.content))
    matrix_ws = workbook["Item Utilization Matrix"]

    # hdr_row1=1, hdr_row2=2, one data row=3, GRAND TOTAL=4.
    total_row = 4
    assert matrix_ws.cell(row=total_row, column=1).value == "GRAND TOTAL"
    assert matrix_ws.cell(row=total_row, column=7).value == summary["purchase_amount"]
    assert matrix_ws.cell(row=total_row, column=8).value == summary["purchase_usd"]
    assert matrix_ws.cell(row=total_row, column=9).value == summary["total_sale_amount"]
    assert matrix_ws.cell(row=total_row, column=10).value == summary["total_sale_usd"]
    assert matrix_ws.cell(row=total_row, column=11).value == summary["total_profit_loss"]
    assert matrix_ws.cell(row=total_row, column=12).value == summary["balance_cif"]
    # Sanity: Profit/Loss is a real, non-zero (here: negative) figure —
    # confirms this isn't accidentally reading an empty/zeroed cell.
    assert summary["total_profit_loss"] == -7000.0


@pytest.mark.django_db
def test_view_pdf_export_returns_pdf(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-VIEW-003", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"


@pytest.mark.django_db
def test_view_missing_date_params_returns_400(report_viewer_client):
    response = report_viewer_client.get(REPORT_URL)
    assert response.status_code == 400

    response = report_viewer_client.get(REPORT_URL, {"from_date": FROM_DATE.isoformat()})
    assert response.status_code == 400


@pytest.mark.django_db
def test_view_invalid_date_format_returns_400(report_viewer_client):
    response = report_viewer_client.get(
        REPORT_URL, {"from_date": "01-01-2026", "to_date": "31-01-2026"}
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_view_requires_report_permission(no_role_client):
    response = no_role_client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code == 403


@pytest.mark.django_db
def test_view_unauthenticated_is_rejected():
    client = APIClient()
    response = client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code in (401, 403)


# ---------------------------------------------------------------------------
# shared/pdf/builders.py — make_landscape_doc `pagesize` kwarg (additive)
# ---------------------------------------------------------------------------

def test_make_landscape_doc_defaults_to_a4():
    """No `pagesize` arg -> today's exact behavior, `landscape(A4)`, for
    every existing caller that doesn't (and won't) pass the new kwarg."""
    doc = make_landscape_doc(BytesIO())
    assert doc.pagesize == landscape(A4)


def test_make_landscape_doc_honors_explicit_pagesize():
    doc = make_landscape_doc(BytesIO(), pagesize=A3)
    assert doc.pagesize == landscape(A3)
    assert doc.pagesize != landscape(A4)


# ---------------------------------------------------------------------------
# Excel export — Generated On/By, Filters, formatting pass
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_view_excel_header_shows_generated_by_and_on(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-XL-GEN", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["License Summary"]

    header_area_text = " ".join(
        str(worksheet.cell(row=r, column=1).value or "") for r in range(1, 6)
    )
    assert "Generated On" in header_area_text
    assert "Generated By" in header_area_text
    # The viewer-role test user has no first/last name -> falls back to
    # `username`, never a blank string or a 500.
    assert "ppr-viewer" in header_area_text
    assert "Filters" in header_area_text
    assert "Norm : All" in header_area_text
    assert "Exporter : All" in header_area_text


@pytest.mark.django_db
def test_view_excel_summary_metric_table_matches_json_summary(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-XL-SUMTBL", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1), cif_fc="120.00")
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    params = {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    summary = report_viewer_client.get(REPORT_URL, params).json()["summary"]

    response = report_viewer_client.get(REPORT_URL, {**params, "format": "excel"})
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["License Summary"]

    # Find the "Metric" / "Value" header row and read the 7 rows under it.
    metric_header_row = next(
        r for r in range(1, 15)
        if worksheet.cell(row=r, column=1).value == "Metric"
        and worksheet.cell(row=r, column=2).value == "Value"
    )
    values_by_metric = {}
    for offset in range(1, 8):
        row = metric_header_row + offset
        metric = worksheet.cell(row=row, column=1).value
        values_by_metric[metric] = worksheet.cell(row=row, column=2).value

    assert values_by_metric["Total Licenses"] == summary["total_licenses"]
    assert values_by_metric["Purchase Amount"] == summary["purchase_amount"]
    assert values_by_metric["Purchase $"] == summary["purchase_usd"]
    assert values_by_metric["Sale Amount"] == summary["total_sale_amount"]
    assert values_by_metric["Sale $"] == summary["total_sale_usd"]
    assert values_by_metric["Profit / Loss"] == summary["total_profit_loss"]
    assert values_by_metric["Balance CIF ($)"] == summary["balance_cif"]

    # The Value column must be real numeric cells, not text baked into a
    # concatenated string — `number_format` can only apply to real numbers.
    value_cell = worksheet.cell(row=metric_header_row + 2, column=2)  # "Purchase Amount" row
    assert value_cell.number_format == '#,##0.00'


@pytest.mark.django_db
def test_view_excel_freeze_panes_and_auto_filter_set_on_both_sheets(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-XL-FREEZE", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["License Summary"]
    matrix_ws = workbook["Item Utilization Matrix"]

    assert worksheet.freeze_panes is not None
    assert worksheet.auto_filter.ref is not None

    # Item Utilization Matrix keeps its unchanged 2-row header at rows 1-2
    # (no Generated-On/Filters preamble on this sheet) -> freeze below row 2.
    assert matrix_ws.freeze_panes == "A3"
    assert matrix_ws.auto_filter.ref is not None


@pytest.mark.django_db
def test_view_excel_money_cell_has_number_format(report_viewer_client, ppr_masters):
    lic = _make_license("PPR-XL-FMT", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["License Summary"]

    # Locate the main table's header row (the one with 'License No.'), then
    # check the 'Purchase Amount' data cell (column 7) directly below it.
    header_row = next(
        r for r in range(1, 20) if worksheet.cell(row=r, column=1).value == "License No."
    )
    money_cell = worksheet.cell(row=header_row + 1, column=7)
    assert money_cell.number_format == '#,##0.00'
    assert money_cell.number_format != 'General'
    assert money_cell.alignment.horizontal == 'right'

    text_cell = worksheet.cell(row=header_row + 1, column=1)
    assert text_cell.alignment.horizontal == 'left'


# ---------------------------------------------------------------------------
# PDF export — Metric|Value summary table, page size A3/A2 switch
# ---------------------------------------------------------------------------

def _pdf_mediabox_width(pdf_bytes: bytes) -> float:
    """
    Parse the first `/MediaBox [x0 y0 x1 y1]` out of raw PDF bytes to get
    the page width in points — a black-box way to tell A3-landscape
    (~1190pt wide) apart from A2-landscape (~1684pt wide) apart from
    A4-landscape (~842pt wide) without needing access to the
    `SimpleDocTemplate` instance built inside the view.
    """
    match = re.search(rb"/MediaBox\s*\[\s*[\d.]+\s+[\d.]+\s+([\d.]+)\s+[\d.]+\s*\]", pdf_bytes)
    assert match, "No /MediaBox found in PDF"
    return float(match.group(1))


@pytest.mark.django_db
def test_view_pdf_export_uses_a3_for_few_dynamic_items(report_viewer_client, ppr_masters):
    """12 static columns + 3*1 dynamic-item columns = 15 <= 20 -> A3."""
    lic = _make_license("PPR-PDF-A3", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))
    _make_debit_trade(item, "3000.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="40.00", qty_kg="100.000")

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    width = _pdf_mediabox_width(response.content)
    assert abs(width - landscape(A3)[0]) < 1.0


@pytest.mark.django_db
def test_view_pdf_export_uses_a2_for_many_dynamic_items(report_viewer_client, ppr_masters):
    """12 static columns + 3*6 dynamic-item columns = 30 > 20 -> A2."""
    lic = _make_license("PPR-PDF-A2", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    for serial in range(1, 7):
        item_name = ItemNameModel.objects.create(name=f"PPR PDF A2 Item {serial}")
        item = _make_import_item(lic, ppr_masters["hs_code"], serial, item_names=[item_name])
        _make_purchase_trade(item, "1000.00", invoice_date=FROM_DATE + timedelta(days=1))
        # A non-zero debit per item so Dynamic Column Optimization doesn't
        # drop any of the 6 headers before the column count is computed.
        _make_debit_trade(
            item, "500.00", invoice_date=FROM_DATE + timedelta(days=2), cif_fc="10.00", qty_kg="5.000",
        )

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    width = _pdf_mediabox_width(response.content)
    assert width > landscape(A3)[0]  # strictly wider than A3 landscape -> escalated to A2


@pytest.mark.django_db
def test_view_pdf_export_with_generated_by_still_returns_valid_pdf(report_viewer_client, ppr_masters):
    """
    Smoke test for the `request.user` -> `draw_page_footer`/Generated-By
    wiring: the `report_viewer_client` user has no first/last name set
    (`_get_generated_by` must fall back to `username` without raising), and
    the PDF must still build successfully via the new
    `onFirstPage`/`onLaterPages` canvas-callback footer (replacing
    `append_generated_footer` for this report only). PDF content streams
    are flate-encoded by default, so "Generated By"/"Page N" text isn't
    reliably greppable in the raw bytes here — this only asserts the
    export doesn't 500 and still returns a well-formed PDF.
    """
    lic = _make_license("PPR-PDF-GEN", ppr_masters["exporter"])
    _make_export_item(lic, ppr_masters["e1_norm"], Decimal("50000.00"))
    item = _make_import_item(lic, ppr_masters["hs_code"], 1, item_names=[ppr_masters["item_a"]])
    _make_purchase_trade(item, "10000.00", invoice_date=FROM_DATE + timedelta(days=1))

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response.content[:5] == b"%PDF-"
    assert b"%%EOF" in response.content[-64:]
