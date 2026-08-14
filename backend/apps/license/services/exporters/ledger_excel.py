"""Excel renderers for the License Ledger reports.

Parallel to the PDF exporters (ledger_pdf.py), these generate Excel workbooks
for all three ledger views:
1. Summary (all licenses)
2. Detailed (with transactions)
3. Company-scoped

All use canonical service data to ensure consistency with PDF/API outputs.
"""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from datetime import datetime
import logging

from shared.pdf.builders import format_indian_number

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY EXCEL: All licenses in table format (parallel to generate_all_licenses_pdf)
# ─────────────────────────────────────────────────────────────────────────────

def generate_ledger_summary_excel(licenses_data, query_params):
    """
    Generate Excel containing all licenses in tabular format with profit/loss.
    Matches the PDF summary structure exactly.

    Returns: (content_bytes, filename_str)
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "License Summary"

        # ── Styles ────────────────────────────────────────────────────────────
        HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        DATA_FILL = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        NO_PURCHASE_FILL = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
        NEGATIVE_BALANCE_FILL = PatternFill(start_color="d32f2f", end_color="d32f2f", fill_type="solid")
        NEGATIVE_BALANCE_FONT = Font(color="FFFFFF", bold=True, size=9)
        TOTAL_FILL = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        TOTAL_FONT = Font(bold=True, size=10)
        PROFIT_FONT = Font(color="2e7d32", bold=True, size=9)
        LOSS_FONT = Font(color="d32f2f", bold=True, size=9)

        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ── Title & Info ──────────────────────────────────────────────────────
        current_row = 1
        ws.merge_cells(f'A{current_row}:L{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "LICENSE LEDGER - SUMMARY"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = HEADER_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Filter info
        license_type = query_params.get('license_type', 'ALL')
        active_only = query_params.get('active_only', 'true').lower() == 'true'
        status_text = 'Active Only' if active_only else 'All'

        ws.merge_cells(f'A{current_row}:L{current_row}')
        filter_cell = ws[f'A{current_row}']
        filter_cell.value = f"Filter: License Type = {license_type} | Status = {status_text} | Total = {len(licenses_data)} licenses"
        filter_cell.font = Font(italic=True, size=9)
        current_row += 1

        # No Purchase warning
        no_purchase_count = sum(1 for lic in licenses_data if not lic.get('purchase_amount') or lic.get('purchase_amount') == 0)
        if no_purchase_count > 0:
            ws.merge_cells(f'A{current_row}:L{current_row}')
            warn_cell = ws[f'A{current_row}']
            warn_cell.value = f"⚠ WARNING: {no_purchase_count} license(s) with no purchase transactions"
            warn_cell.font = Font(bold=True, size=10, color="7D6608")
            warn_cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            warn_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[current_row].height = 18
        current_row += 2

        # ── Headers ───────────────────────────────────────────────────────────
        headers = [
            'License No.', 'Type', 'Exporter', 'License Date', 'Expiry Date',
            'Purchase ($)', 'Sold ($)', 'Balance ($)',
            'Purchase Amt (INR)', 'Sale Amt (INR)', 'P/L (INR)', 'Status'
        ]
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1

        # ── Data Rows ─────────────────────────────────────────────────────────
        total_purchase_usd = 0.0
        total_sold_usd = 0.0
        total_balance_usd = 0.0
        total_purchase_amt = 0.0
        total_sale_amt = 0.0
        total_pl = 0.0

        for idx, lic in enumerate(licenses_data):
            lic_date = lic.get('license_date')
            exp_date = lic.get('license_expiry_date')
            lic_date_str = lic_date.strftime('%d-%b-%y') if lic_date else '-'
            exp_date_str = exp_date.strftime('%d-%b-%y') if exp_date else '-'

            currency = lic.get('currency', 'USD')
            purchase_usd = float(lic.get('total_value', 0) or 0)
            sold_usd = float(lic.get('sold_value', 0) or 0)
            balance_usd = float(lic.get('balance_value', 0) or 0)
            purchase_amt = float(lic.get('purchase_amount', 0) or 0)
            sale_amt = float(lic.get('sale_amount', 0) or 0)
            profit_loss = float(lic.get('total_profit_loss', 0) or 0)
            status = 'Active' if lic.get('is_active', False) else 'Expired'

            total_purchase_usd += purchase_usd
            total_sold_usd += sold_usd
            total_balance_usd += balance_usd
            total_purchase_amt += purchase_amt
            total_sale_amt += sale_amt
            total_pl += profit_loss

            # Determine row styling
            has_no_purchase = purchase_amt == 0
            has_negative_balance = balance_usd < 0

            # Build row
            values = [
                lic.get('license_number', '-'),
                lic.get('license_type', '-'),
                lic.get('exporter_name', '-') or '-',
                lic_date_str,
                exp_date_str,
                f"${format_indian_number(purchase_usd, 2)}" if currency == 'USD' else f"INR {format_indian_number(purchase_usd, 2)}",
                f"${format_indian_number(sold_usd, 2)}" if currency == 'USD' else f"INR {format_indian_number(sold_usd, 2)}",
                f"${format_indian_number(balance_usd, 2)}" if currency == 'USD' else f"INR {format_indian_number(balance_usd, 2)}",
                format_indian_number(purchase_amt, 2),
                format_indian_number(sale_amt, 2),
                format_indian_number(profit_loss, 2),
                status
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='right' if col_num >= 6 and col_num <= 11 else 'left', vertical='center')

                # Apply row background
                if has_no_purchase:
                    cell.fill = NO_PURCHASE_FILL

                # Highlight negative balance column
                if col_num == 8 and has_negative_balance:
                    cell.fill = NEGATIVE_BALANCE_FILL
                    cell.font = NEGATIVE_BALANCE_FONT

                # Color P/L column
                if col_num == 11:
                    cell.font = PROFIT_FONT if profit_loss >= 0 else LOSS_FONT

            current_row += 1

        # ── Totals Row ────────────────────────────────────────────────────────
        totals_data = [
            'TOTAL', '', '', '', '',
            f"${format_indian_number(total_purchase_usd, 2)}",
            f"${format_indian_number(total_sold_usd, 2)}",
            f"${format_indian_number(total_balance_usd, 2)}",
            format_indian_number(total_purchase_amt, 2),
            format_indian_number(total_sale_amt, 2),
            format_indian_number(total_pl, 2),
            ''
        ]

        for col_num, value in enumerate(totals_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right' if col_num >= 6 and col_num <= 11 else 'left', vertical='center')

            # Bold P/L in totals
            if col_num == 11:
                cell.font = Font(bold=True, size=10, color="2e7d32" if total_pl >= 0 else "d32f2f")

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 15  # License No
        ws.column_dimensions['B'].width = 12  # Type
        ws.column_dimensions['C'].width = 20  # Exporter
        ws.column_dimensions['D'].width = 14  # License Date
        ws.column_dimensions['E'].width = 14  # Expiry
        ws.column_dimensions['F'].width = 14  # Purchase ($)
        ws.column_dimensions['G'].width = 14  # Sold ($)
        ws.column_dimensions['H'].width = 14  # Balance ($)
        ws.column_dimensions['I'].width = 16  # Purchase Amt
        ws.column_dimensions['J'].width = 16  # Sale Amt
        ws.column_dimensions['K'].width = 14  # P/L
        ws.column_dimensions['L'].width = 12  # Status

        ws.freeze_panes = f'A{header_row + 1}'

        # Save
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ledger_summary_{timestamp}.xlsx"

        return excel_file.read(), filename

    except Exception as e:
        logger.exception(f"Failed to generate ledger summary Excel: {e}")
        raise


# ─────────────────────────────────────────────────────────────────────────────
# DETAILED EXCEL: With transaction details (parallel to generate_detailed_licenses_pdf)
# ─────────────────────────────────────────────────────────────────────────────

def generate_ledger_detailed_excel(licenses_data, query_params):
    """
    Generate Excel with detailed transactions for each license.
    One sheet per license with transaction history and P/L.

    Returns: (content_bytes, filename_str)
    """
    from apps.license.services.exporters.ledger_pdf import get_license_transactions

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # ── Styles ────────────────────────────────────────────────────────────
        HEADER_FILL = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        SECTION_FILL = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
        DATA_FILL = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        PROFIT_FONT = Font(color="2e7d32", bold=True, size=9)
        LOSS_FONT = Font(color="d32f2f", bold=True, size=9)

        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        company_id = query_params.get('company')

        # ── Create sheet per license ──────────────────────────────────────────
        for lic_data in licenses_data:
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = (lic_data.get('license_number', 'License') or 'License')[:28]
            ws = wb.create_sheet(title=sheet_name)

            current_row = 1

            # Title
            ws.merge_cells(f'A{current_row}:J{current_row}')
            title_cell = ws[f'A{current_row}']
            lic_number = lic_data.get('license_number', 'N/A')
            title_cell.value = f"LICENSE LEDGER - {lic_number}"
            title_cell.font = Font(bold=True, size=12, color="FFFFFF")
            title_cell.fill = HEADER_FILL
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # License info
            ws.merge_cells(f'A{current_row}:J{current_row}')
            exporter = lic_data.get('exporter_name', 'N/A')
            lic_type = lic_data.get('license_type', 'N/A')
            info_cell = ws[f'A{current_row}']
            info_cell.value = f"Exporter: {exporter} | Type: {lic_type}"
            info_cell.font = Font(italic=True, size=10)
            current_row += 1

            # License dates
            lic_date = lic_data.get('license_date')
            exp_date = lic_data.get('license_expiry_date')
            lic_date_str = lic_date.strftime('%d-%b-%Y') if lic_date else '-'
            exp_date_str = exp_date.strftime('%d-%b-%Y') if exp_date else '-'

            ws.merge_cells(f'A{current_row}:J{current_row}')
            date_cell = ws[f'A{current_row}']
            date_cell.value = f"License Date: {lic_date_str} | Expiry Date: {exp_date_str}"
            current_row += 1
            current_row += 1

            # Transaction table
            transactions = get_license_transactions(lic_data, company_id=company_id)

            if transactions:
                # Headers (NEW ORDER: Date, Particulars, Type, Items, Debit($), Credit($), Sale Bill(₹), Purchase Bill(₹), Balance($), P/L(₹), Purchase Bill, SION)
                txn_headers = [
                    'Date', 'Particulars', 'Type', 'Items',
                    'Debit ($)', 'Credit ($)', 'Sale Bill (₹)', 'Purchase Bill (₹)',
                    'Balance ($)', 'P/L (₹)', 'Purchase Bill', 'SION'
                ]
                header_row = current_row
                for col_num, header in enumerate(txn_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                current_row += 1

                # Transaction rows
                for txn in transactions:
                    date_str = txn['date'].strftime('%d-%b-%y') if txn.get('date') else '-'
                    particular = txn.get('particular', '-')
                    txn_type = txn.get('type', '-')
                    item_names = txn.get('item_names', '-')

                    # Handle debit/credit columns: show "-" for None/0 values
                    debit_cif = txn.get('debit_cif', 0)
                    credit_cif = txn.get('credit_cif', 0)
                    debit_amount = txn.get('debit_amount', 0)
                    credit_amount = txn.get('credit_amount', 0)
                    balance = txn.get('balance')
                    total_pl = txn.get('total_profit_loss', 0)

                    # Canonical fields: purchase bill status and SION norms
                    has_purchase_bill = txn.get('has_purchase_bill', False)
                    is_sion_norm_empty = txn.get('is_sion_norm_empty', True)
                    sion_norm = txn.get('sion_norm', '')

                    # Helper to convert numeric to float or "-"
                    def fmt_numeric(val):
                        try:
                            num_val = float(val) if val else 0
                            return num_val if num_val > 0 else '-'
                        except (ValueError, TypeError):
                            return '-'

                    # Format purchase bill status
                    purchase_bill_status = 'WITH_PURCHASE_BILL' if has_purchase_bill else 'NO_PURCHASE_BILL'

                    # Format SION: show "N/A" if empty, otherwise show the value
                    sion_display = 'N/A' if is_sion_norm_empty else sion_norm

                    values = [
                        date_str,
                        particular,
                        txn_type,
                        item_names,
                        fmt_numeric(debit_cif),
                        fmt_numeric(credit_cif),
                        fmt_numeric(debit_amount),
                        fmt_numeric(credit_amount),
                        float(balance) if balance is not None else '-',
                        float(total_pl) if total_pl and float(total_pl) != 0 else '-',
                        purchase_bill_status,
                        sion_display,
                    ]

                    for col_num, value in enumerate(values, 1):
                        # Convert "-" to None for Excel display, but keep numeric values
                        cell_value = None if value == '-' else value
                        cell = ws.cell(row=current_row, column=col_num, value=cell_value)
                        cell.border = THIN_BORDER

                        # Display "-" for empty cells
                        if value == '-':
                            cell.value = '-'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Format numeric columns (5-10 are numeric: Debit($), Credit($), Sale Bill, Purchase Bill, Balance, P/L)
                        elif col_num in [5, 6, 7, 8, 9, 10] and value != '-' and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        # Columns 11-12 are text (Purchase Bill, SION) - center-aligned
                        elif col_num in [11, 12]:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Color P/L (column 10)
                        if col_num == 10 and value != '-' and isinstance(value, (int, float)) and value != 0:
                            cell.font = PROFIT_FONT if value >= 0 else LOSS_FONT

                    current_row += 1

                ws.freeze_panes = f'A{header_row + 1}'
            else:
                ws.merge_cells(f'A{current_row}:J{current_row}')
                no_txn_cell = ws[f'A{current_row}']
                no_txn_cell.value = "No transactions found for this license"
                no_txn_cell.font = Font(italic=True)
                current_row += 1

            # Column widths (12 columns now: added Purchase Bill and SION)
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 24  # Particulars
            ws.column_dimensions['C'].width = 14  # Type
            ws.column_dimensions['D'].width = 30  # Items
            ws.column_dimensions['E'].width = 14  # Debit ($)
            ws.column_dimensions['F'].width = 14  # Credit ($)
            ws.column_dimensions['G'].width = 14  # Sale Bill (₹)
            ws.column_dimensions['H'].width = 14  # Purchase Bill (₹)
            ws.column_dimensions['I'].width = 14  # Balance ($)
            ws.column_dimensions['J'].width = 12  # P/L (₹)
            ws.column_dimensions['K'].width = 18  # Purchase Bill
            ws.column_dimensions['L'].width = 14  # SION

        # Save
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ledger_detailed_{timestamp}.xlsx"

        return excel_file.read(), filename

    except Exception as e:
        logger.exception(f"Failed to generate ledger detailed Excel: {e}")
        raise


# ─────────────────────────────────────────────────────────────────────────────
# COMPANY EXCEL: Company-scoped ledger (parallel to generate_company_ledger_pdf)
# ─────────────────────────────────────────────────────────────────────────────

def generate_ledger_company_excel(licenses_data, company_name, query_params):
    """
    Generate Excel for company-specific ledger showing profit/loss per license.
    Parallel to generate_company_ledger_pdf.

    Returns: (content_bytes, filename_str)
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Company Ledger"

        # ── Styles ────────────────────────────────────────────────────────────
        HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        DATA_FILL = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        TOTAL_FONT = Font(bold=True, size=10)
        PROFIT_FONT = Font(color="2e7d32", bold=True, size=9)
        LOSS_FONT = Font(color="d32f2f", bold=True, size=9)

        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ── Title ─────────────────────────────────────────────────────────────
        current_row = 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"COMPANY LEDGER - {company_name.upper()}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = HEADER_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Filter info
        license_type = query_params.get('license_type', 'ALL')
        active_only = query_params.get('active_only', 'true').lower() == 'true'
        status_text = 'Active Only' if active_only else 'All'

        ws.merge_cells(f'A{current_row}:H{current_row}')
        filter_cell = ws[f'A{current_row}']
        filter_cell.value = f"Filter: License Type = {license_type} | Status = {status_text} | Total = {len(licenses_data)} licenses"
        filter_cell.font = Font(italic=True, size=9)
        current_row += 2

        # ── Headers ───────────────────────────────────────────────────────────
        headers = [
            'License No.', 'Type', 'Exporter', 'License Date', 'Expiry Date',
            'Total Value', 'Balance', 'P/L (INR)'
        ]
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1

        # ── Data Rows ─────────────────────────────────────────────────────────
        total_pl = 0.0

        for idx, lic in enumerate(licenses_data):
            lic_date = lic.get('license_date')
            exp_date = lic.get('license_expiry_date')
            lic_date_str = lic_date.strftime('%d-%b-%y') if lic_date else '-'
            exp_date_str = exp_date.strftime('%d-%b-%y') if exp_date else '-'

            currency = 'USD' if lic.get('license_type') == 'DFIA' else 'INR'
            total_value = float(lic.get('total_value', 0) or 0)
            balance = float(lic.get('available_balance', 0) or 0)
            profit_loss = float(lic.get('total_profit_loss', 0) or 0)

            total_pl += profit_loss

            # Build row
            values = [
                lic.get('license_number', '-'),
                lic.get('license_type', '-'),
                lic.get('exporter_name', '-') or '-',
                lic_date_str,
                exp_date_str,
                total_value,
                balance,
                profit_loss,
            ]

            # Apply background color
            bg_color = DATA_FILL if (idx % 2 == 0) else None

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = THIN_BORDER
                if bg_color:
                    cell.fill = bg_color

                # Format currency/numeric columns
                if col_num in [6, 7, 8]:
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                        # Color P/L column
                        if col_num == 8:
                            cell.font = PROFIT_FONT if value >= 0 else LOSS_FONT
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # Prefix currency
                if col_num == 6:
                    if isinstance(value, float):
                        currency_prefix = "$" if currency == "USD" else "INR "
                        cell.value = value
                elif col_num == 7:
                    if isinstance(value, float):
                        currency_prefix = "$" if currency == "USD" else "INR "
                        cell.value = value

            current_row += 1

        # ── Totals Row ────────────────────────────────────────────────────────
        totals_data = [
            'TOTAL', '', '', '', '',
            '',
            '',
            total_pl,
        ]

        for col_num, value in enumerate(totals_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = THIN_BORDER

            if col_num == 8 and isinstance(value, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.font = Font(bold=True, size=10, color="2e7d32" if value >= 0 else "d32f2f")
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 15  # License No
        ws.column_dimensions['B'].width = 12  # Type
        ws.column_dimensions['C'].width = 20  # Exporter
        ws.column_dimensions['D'].width = 14  # License Date
        ws.column_dimensions['E'].width = 14  # Expiry
        ws.column_dimensions['F'].width = 16  # Total Value
        ws.column_dimensions['G'].width = 16  # Balance
        ws.column_dimensions['H'].width = 14  # P/L

        ws.freeze_panes = f'A{header_row + 1}'

        # Save
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Sanitize company name for filename
        safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_')).strip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"company_ledger_{safe_company_name}_{timestamp}.xlsx"

        return excel_file.read(), filename

    except Exception as e:
        logger.exception(f"Failed to generate company ledger Excel: {e}")
        raise
