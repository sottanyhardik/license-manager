"""Presentation-only Excel renderer for canonical Financial Ledger data.

The module deliberately imports no model or business service.  Amounts, group
totals, balances and profit/loss are copied from the canonical DTO unchanged.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
PALE_BLUE = "D9EAF7"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
THIN_GREY = Side(style="thin", color="B7C9D6")
USD_FORMAT = '$#,##0.00;[Red]-$#,##0.00;–'
INR_FORMAT = '[$₹-en-IN]#,##0.00;[Red]-[$₹-en-IN]#,##0.00;–'
DATE_FORMAT = "dd-mmm-yyyy"


def render_financial_ledger_excel(canonical_data: dict) -> BytesIO:
    """Render an already-filtered canonical collection without querying/calculating."""
    workbook = openpyxl.Workbook()
    list_sheet = workbook.active
    list_sheet.title = "License Summary"
    licenses = list(canonical_data.get("licenses") or [])
    _render_license_list(list_sheet, canonical_data)

    # SION is an added summary layer. Every export also retains the complete
    # canonical per-license transaction ledger; no transaction is collapsed
    # into or replaced by a SION total.
    if licenses:
        _render_detail(workbook.create_sheet("Financial Trade Ledger"), licenses)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _render_license_list(sheet, canonical_data: dict) -> None:
    headers = (
        "Company", "SION Norm", "License Number", "Type", "Date", "1st Purchase Date",
        "Balance ($)", "Purchase (₹)", "Sale (₹)", "P/L (₹)",
    )
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "LICENSE LEDGER"
    _style_title(sheet["A1"])
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet[3])

    # Company membership and totals are canonical DTO fields.  The workbook
    # does not infer ownership from transactions or sum any money column.
    for group in canonical_data.get("company_groups") or []:
        company_name = _display(group.get("company_name"))
        company_row = sheet.max_row + 1
        sheet.merge_cells(start_row=company_row, start_column=1, end_row=company_row, end_column=10)
        company_cell = sheet.cell(company_row, 1, company_name)
        company_cell.font = Font(bold=True, color=NAVY)
        company_cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        sion_groups = group.get("sion_groups")
        if sion_groups is None:  # Compatibility with canonical payloads made before SION grouping.
            sion_groups = [{"sion_label": None, "licenses": group.get("licenses") or []}]
        for sion_group in sion_groups:
            if sion_group.get("sion_label") is not None:
                row = sheet.max_row + 1
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                cell = sheet.cell(row, 1, f"SION: {sion_group['sion_label']}")
                cell.font = Font(bold=True, italic=True, color=NAVY)
                cell.fill = PatternFill("solid", fgColor="EAF2F8")
                cell.alignment = Alignment(indent=1)
            for data in sion_group.get("licenses") or []:
                _append_license_row(sheet, data, company_name, sion_group.get("sion_label"))
            if sion_group.get("sion_label") is not None:
                _append_total_row(sheet, f"SION Total — {sion_group['sion_label']}", sion_group)
        _append_total_row(sheet, f"Total — {company_name}", group)

    grand_total = canonical_data.get("grand_total")
    if grand_total:
        _append_total_row(sheet, "GRAND TOTAL", grand_total, fill=NAVY, font_color=WHITE)

    last_row = max(sheet.max_row, 3)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:J{last_row}"
    sheet.print_title_rows = "1:3"
    sheet.print_area = f"A1:J{last_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.oddFooter.center.text = "License Ledger | Page &P of &N"
    _set_widths(sheet, (30, 20, 22, 14, 15, 19, 18, 20, 20, 20))


def _render_detail(sheet, licenses: list[dict]) -> None:
    headers = (
        "Company", "SION", "License Number", "Date", "Particulars", "Invoice Number", "Type", "Items",
        "Credit ($)", "Debit ($)", "Purchase (₹)", "Sale (₹)", "Balance ($)", "P/L (₹)",
    )
    sheet.merge_cells("A1:N1")
    sheet["A1"] = "FINANCIAL TRADE LEDGER — INDIVIDUAL LICENSES"
    _style_title(sheet["A1"])
    sheet.append(headers)
    _style_header(sheet[2])

    for data in licenses:
        summary = data.get("summary") or {}
        transactions = list(data.get("display_transactions") or [])
        companies = data.get("license_wise_companies") or []
        company_names = {company.get("company_id"): company.get("company_name") for company in companies}
        for transaction in transactions:
            document = transaction.get("invoice_document") or {}
            row = sheet.max_row + 1
            sheet.append([
                _display(company_names.get(transaction.get("company_id")) or transaction.get("company_name")),
                _display(data.get("sion_norms")), _display(data.get("license_number")),
                transaction.get("date"), _display(transaction.get("party_name")),
                _display(document.get("invoice_number")), _display(transaction.get("type")),
                _display(", ".join(transaction.get("item_names") or [])),
                transaction.get("purchase_amount"), transaction.get("sale_amount"),
                transaction.get("purchase_bill_amount"), transaction.get("sale_bill_amount"),
                transaction.get("license_running_balance"), transaction.get("profit_loss_inr"),
            ])
            if document.get("document_exists") and document.get("secure_url"):
                invoice_cell = sheet.cell(row, 6)
                invoice_cell.hyperlink = document["secure_url"]
                invoice_cell.style = "Hyperlink"
                invoice_cell.comment = openpyxl.comments.Comment(
                    "SIGNED" if document.get("signed") else "UNSIGNED", "License Manager",
                )
            elif document.get("status") == "COPY_UNAVAILABLE":
                sheet.cell(row, 6).comment = openpyxl.comments.Comment("Copy unavailable", "License Manager")
            _format_flat_detail_row(sheet, row)
        row = sheet.max_row + 1
        sheet.append([
            None, _display(data.get("sion_norms")), _display(data.get("license_number")),
            None, "LICENSE TOTAL", None, None, None, None, None,
            summary.get("total_purchase_bill_inr"), summary.get("total_sale_bill_inr"),
            summary.get("current_balance"), summary.get("total_profit_loss"),
        ])
        for cell in sheet[row]:
            cell.font = Font(bold=True, color=NAVY)
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        _format_flat_detail_row(sheet, row)

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:N{max(sheet.max_row, 2)}"
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:N{max(sheet.max_row, 1)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.oddFooter.left.text = "Financial Trade Ledger"
    sheet.oddFooter.center.text = "Page &P of &N"
    _set_widths(sheet, (28, 18, 22, 15, 31, 24, 20, 34, 17, 17, 19, 19, 18, 19))


def _format_flat_detail_row(sheet, row: int) -> None:
    if isinstance(sheet.cell(row, 4).value, (date, datetime)):
        sheet.cell(row, 4).number_format = DATE_FORMAT
    for column in (9, 10, 13):
        sheet.cell(row, column).number_format = USD_FORMAT
    for column in (11, 12, 14):
        sheet.cell(row, column).number_format = INR_FORMAT
    for cell in sheet[row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def _append_license_row(sheet, data: dict, company_name, sion_label) -> None:
    row = sheet.max_row + 1
    sheet.append([
        _display(company_name), _display(sion_label), _display(data.get("license_number")),
        _display(data.get("license_type")), data.get("license_date"), data.get("first_purchase_date"),
        data.get("current_balance"),
        data.get("purchase_bill_inr") if data.get("has_purchase_bill") else None,
        data.get("sale_bill_inr"), data.get("profit_loss_inr"),
    ])
    if not data.get("has_purchase_bill"):
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor=PALE_RED)
        sheet.cell(row, 8).value = "-"
        sheet.cell(row, 8).comment = openpyxl.comments.Comment("NO PURCHASE BILL", "License Manager")
    _format_list_row(sheet, row)


def _append_total_row(sheet, label: str, totals: dict, *, fill=PALE_BLUE, font_color=NAVY) -> None:
    row = sheet.max_row + 1
    sheet.append([
        label, None, None, None, None, None, totals.get("total_balance"),
        totals.get("total_purchase_bill_inr"), totals.get("total_sale_bill_inr"),
        totals.get("total_profit_loss_inr"),
    ])
    for cell in sheet[row]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = PatternFill("solid", fgColor=fill)
    _format_list_row(sheet, row)


def _display(value: Any) -> Any:
    """Approved presentation placeholder; never manufacture a financial value."""
    return value if value not in (None, "", []) else "-"


def _style_title(cell) -> None:
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_header(cells) -> None:
    for cell in cells:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def _format_list_row(sheet, row: int) -> None:
    for column in (5, 6):
        if isinstance(sheet.cell(row, column).value, (date, datetime)):
            sheet.cell(row, column).number_format = DATE_FORMAT
    sheet.cell(row, 7).number_format = USD_FORMAT
    for column in (8, 9, 10):
        sheet.cell(row, column).number_format = INR_FORMAT
    for cell in sheet[row]:
        cell.border = Border(bottom=THIN_GREY)


def _set_widths(sheet, widths: Iterable[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
