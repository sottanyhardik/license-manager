"""
Tests for the License Trading Register & Profit Report
(`apps.license.services.license_trading_register_report.
build_license_trading_register_report` and
`apps.license.views.license_trading_register_report.LicenseTradingRegisterReportView`).

Covers:
- Purchase + Sale consolidation into a single per-license chronological
  register with running profit and Open/Closed status (boundary:
  purchase == sale -> Closed).
- COMMISSION_PURCHASE/COMMISSION_SALE trades excluded entirely (scope
  decision 2) — never affect Purchase/Sale/Profit or appear in the register.
- Date-window filtering scopes the whole register (both sides).
- Multi-item `LicenseTradeLine` labeled as one joined-name row (scope
  decision 6), preserving exact partition identity between the register
  and the item summary.
- Reconciliation at every level: license item summary == license summary,
  norm summary == sum of its licenses, norm item summary == norm summary,
  grand == sum of norms, grand item summary == grand summary.
- Item/Norm filters narrow the Transaction Register itself (scope
  decision 5), not just the candidate license set.
- View-level contract: JSON/Excel/PDF response codes + content-type, param
  validation, permissions.
"""
import uuid
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import connection
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.core.models import CompanyModel, HSCodeModel, HeadSIONNormsModel, ItemNameModel, SionNormClassModel
from apps.license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel
from apps.license.services.license_trading_register_report import build_license_trading_register_report
from apps.trade.models import LicenseTrade, LicenseTradeLine

User = get_user_model()

REPORT_URL = "/api/reports/license-trading-register/"

FROM_DATE = date.today() - timedelta(days=60)
TO_DATE = date.today()
OUT_OF_RANGE_DATE = date.today() - timedelta(days=120)


# ---------------------------------------------------------------------------
# Auth fixtures (mirrors test_license_purchase_profit_report.py).
# ---------------------------------------------------------------------------

@pytest.fixture
def report_viewer_client(db):
    user = User.objects.create_user(
        username="tr-viewer",
        email="tr-viewer@example.com",
        password="RoleP@ssw0rd123",
    )
    group, _ = Group.objects.get_or_create(name="REPORT_VIEWER")
    user.groups.add(group)
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


@pytest.fixture
def no_role_client(db):
    user = User.objects.create_user(
        username="tr-norole",
        email="tr-norole@example.com",
        password="RoleP@ssw0rd123",
    )
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


# ---------------------------------------------------------------------------
# Master-data / fixture builders
# ---------------------------------------------------------------------------

@pytest.fixture
def tr_masters(db):
    head_norm = HeadSIONNormsModel.objects.create(name="TR Test Head Norm")
    return {
        "exporter": CompanyModel.objects.create(iec="8880001111", name="TR Exporter"),
        "supplier": CompanyModel.objects.create(iec="8880002222", name="TR Supplier"),
        "customer": CompanyModel.objects.create(iec="8880003333", name="TR Customer"),
        "other_customer": CompanyModel.objects.create(iec="8880004444", name="TR Other Customer"),
        "hs_code": HSCodeModel.objects.create(hs_code="77777777", product_description="TR Test Product"),
        "item_a": ItemNameModel.objects.create(name="TR Item A"),
        "item_b": ItemNameModel.objects.create(name="TR Item B"),
        "e126_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E126", is_active=True),
        "e132_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E132", is_active=True),
    }


def _make_license(number, exporter):
    return LicenseDetailsModel.objects.create(
        license_number=number,
        license_date=date.today() - timedelta(days=180),
        license_expiry_date=date.today() + timedelta(days=180),
        exporter=exporter,
    )


def _make_export_item(license_obj, norm_class, cif_fc=Decimal("100000.00")):
    return LicenseExportItemModel.objects.create(
        license=license_obj,
        description=f"Export item for {license_obj.license_number}",
        norm_class=norm_class,
        cif_fc=cif_fc,
        cif_inr=cif_fc * Decimal("84.5"),
    )


def _make_import_item(license_obj, hs_code, serial, item_names=None, description=None):
    item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=serial,
        description=description or f"Import item {serial}",
        hs_code=hs_code,
        quantity=Decimal("1000.000"),
        available_quantity=Decimal("1000.000"),
    )
    if item_names:
        item.items.set(item_names)
    return item


def _make_trade(direction, from_company=None, to_company=None, invoice_date=None, invoice_number=None):
    return LicenseTrade.objects.create(
        direction=direction,
        license_type=LicenseTrade.LICENSE_TYPE_DFIA,
        from_company=from_company,
        to_company=to_company,
        invoice_number=invoice_number or f"INV-{direction}-{uuid.uuid4().hex[:12]}",
        invoice_date=invoice_date or date.today(),
    )


def _make_line(trade, sr_number, amount, qty=Decimal("1.0000")):
    """Uses MODE_QTY (qty x rate=amount) so `amount_inr` == amount exactly,
    the same trick `test_license_purchase_profit_report.py::_make_purchase`
    uses for `LicensePurchase`."""
    return LicenseTradeLine.objects.create(
        trade=trade,
        sr_number=sr_number,
        mode=LicenseTradeLine.MODE_QTY,
        qty_kg=qty,
        rate_inr_per_kg=(Decimal(str(amount)) / qty).quantize(Decimal("0.01")),
    )


def _find_license(report, license_number):
    for norm_block in report["norms"]:
        for lic in norm_block["licenses"]:
            if lic["license_number"] == license_number:
                return lic
    return None


def _find_norm(report, norm):
    return next((n for n in report["norms"] if n["norm"] == norm), None)


# ---------------------------------------------------------------------------
# Service-level tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_empty_input_returns_empty_report():
    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    assert report["norms"] == []
    assert report["grand_item_summary"] == []
    assert report["dashboard"]["total_licenses"] == 0
    assert report["grand_summary"]["licenses_count"] == 0


@pytest.mark.django_db
def test_single_license_single_purchase_and_sale(tr_masters):
    lic = _make_license("TR-001", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    purchase_trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(purchase_trade, item, "60000.00")

    sale_trade = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(sale_trade, item, "90000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-001")
    assert lic_row is not None
    assert lic_row["summary"]["purchase"] == 60000.00
    assert lic_row["summary"]["sale"] == 90000.00
    assert lic_row["summary"]["profit"] == 30000.00
    assert lic_row["summary"]["margin_pct"] == pytest.approx(33.33, abs=0.01)  # profit/sale
    assert lic_row["summary"]["status"] == "Closed"
    assert len(lic_row["transactions"]) == 2
    assert lic_row["transactions"][0]["direction"] == LicenseTrade.DIR_PURCHASE
    assert lic_row["transactions"][1]["direction"] == LicenseTrade.DIR_SALE
    assert lic_row["transactions"][0]["running_profit"] == -60000.00
    assert lic_row["transactions"][1]["running_profit"] == 30000.00


@pytest.mark.django_db
def test_purchase_only_is_open(tr_masters):
    lic = _make_license("TR-002", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    purchase_trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(purchase_trade, item, "40000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-002")
    assert lic_row["summary"]["purchase"] == 40000.00
    assert lic_row["summary"]["sale"] == 0.0
    assert lic_row["summary"]["status"] == "Open"


@pytest.mark.django_db
def test_status_boundary_purchase_equals_sale_is_closed(tr_masters):
    lic = _make_license("TR-BOUNDARY", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    purchase_trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(purchase_trade, item, "50000.00")
    sale_trade = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(sale_trade, item, "50000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-BOUNDARY")
    assert lic_row["summary"]["purchase"] == lic_row["summary"]["sale"] == 50000.00
    assert lic_row["summary"]["status"] == "Closed"


@pytest.mark.django_db
def test_running_profit_crossover(tr_masters):
    lic = _make_license("TR-004", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    p1 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p1, item, "50000.00")
    s1 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(s1, item, "20000.00")
    s2 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=10),
    )
    _make_line(s2, item, "40000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    txns = _find_license(report, "TR-004")["transactions"]
    assert len(txns) == 3
    assert txns[0]["running_profit"] == -50000.00
    assert txns[1]["running_profit"] == -30000.00
    assert txns[2]["running_profit"] == 10000.00


@pytest.mark.django_db
def test_commission_trades_excluded_entirely(tr_masters):
    """Scope decision 2: COMMISSION_PURCHASE/COMMISSION_SALE never affect
    Purchase/Sale/Profit and never appear in the register."""
    lic = _make_license("TR-005", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    purchase_trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(purchase_trade, item, "50000.00")

    commission_purchase = _make_trade(
        LicenseTrade.DIR_COMMISSION_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=2),
    )
    _make_line(commission_purchase, item, "1000.00")

    sale_trade = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(sale_trade, item, "80000.00")

    commission_sale = _make_trade(
        LicenseTrade.DIR_COMMISSION_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=6),
    )
    _make_line(commission_sale, item, "2000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-005")
    assert lic_row["summary"]["purchase"] == 50000.00  # commission NOT included
    assert lic_row["summary"]["sale"] == 80000.00  # commission NOT included
    assert len(lic_row["transactions"]) == 2  # only the genuine purchase + sale rows


@pytest.mark.django_db
def test_date_window_filters_both_purchase_and_sale_sides(tr_masters):
    """Scope decision 3: the date filter scopes the WHOLE register, both
    sides — unlike the sibling Purchase & Profit report."""
    lic = _make_license("TR-DATE", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    out_of_range_purchase = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=OUT_OF_RANGE_DATE,
    )
    _make_line(out_of_range_purchase, item, "10000.00")

    in_range_purchase = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(in_range_purchase, item, "20000.00")

    out_of_range_sale = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=TO_DATE + timedelta(days=30),
    )
    _make_line(out_of_range_sale, item, "50000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-DATE")
    assert lic_row["summary"]["purchase"] == 20000.00  # out-of-range purchase excluded
    assert lic_row["summary"]["sale"] == 0.0  # out-of-range sale excluded
    assert len(lic_row["transactions"]) == 1


@pytest.mark.django_db
def test_multi_item_line_labeled_as_one_joined_row(tr_masters):
    """Scope decision 6: an SR with multiple `items` (M2M) is ONE
    transaction row labeled with the joined item names, not split."""
    lic = _make_license("TR-MULTIITEM", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(
        lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"], tr_masters["item_b"]],
    )
    trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade, item, "10000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    lic_row = _find_license(report, "TR-MULTIITEM")
    assert len(lic_row["transactions"]) == 1
    assert lic_row["transactions"][0]["item"] == "TR Item A, TR Item B"
    assert len(lic_row["item_summary"]) == 1
    assert lic_row["item_summary"][0]["item"] == "TR Item A, TR Item B"


@pytest.mark.django_db
def test_multi_license_multi_norm(tr_masters):
    lic1 = _make_license("TR-MN-001", tr_masters["exporter"])
    _make_export_item(lic1, tr_masters["e126_norm"])
    item1 = _make_import_item(lic1, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    p1 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p1, item1, "10000.00")

    lic2 = _make_license("TR-MN-002", tr_masters["exporter"])
    _make_export_item(lic2, tr_masters["e132_norm"])
    item2 = _make_import_item(lic2, tr_masters["hs_code"], 1, item_names=[tr_masters["item_b"]])
    p2 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p2, item2, "5000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    norms_present = {n["norm"] for n in report["norms"]}
    assert norms_present == {"E126", "E132"}
    assert _find_license(report, "TR-MN-001") is not None
    assert _find_license(report, "TR-MN-002") is not None


@pytest.mark.django_db
def test_trade_spanning_two_licenses_split_correctly(tr_masters):
    lic1 = _make_license("TR-SPAN-001", tr_masters["exporter"])
    _make_export_item(lic1, tr_masters["e126_norm"])
    item1 = _make_import_item(lic1, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    lic2 = _make_license("TR-SPAN-002", tr_masters["exporter"])
    _make_export_item(lic2, tr_masters["e126_norm"])
    item2 = _make_import_item(lic2, tr_masters["hs_code"], 1, item_names=[tr_masters["item_b"]])

    combined_trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(combined_trade, item1, "30000.00")
    _make_line(combined_trade, item2, "45000.00")
    combined_trade.refresh_from_db()
    assert combined_trade.subtotal_amount == Decimal("75000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    row1 = _find_license(report, "TR-SPAN-001")
    row2 = _find_license(report, "TR-SPAN-002")
    assert row1["summary"]["purchase"] == 30000.00
    assert row2["summary"]["purchase"] == 45000.00
    assert len(row1["transactions"]) == 1
    assert len(row2["transactions"]) == 1


@pytest.mark.django_db
def test_item_filter_narrows_the_register_itself(tr_masters):
    """Scope decision 5: item/norm filters narrow the Transaction Register
    itself, not just the candidate license set — a license with SOME
    matching and SOME non-matching lines only shows the matching ones."""
    lic = _make_license("TR-ITEMFILTER", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item_a = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    item_b = _make_import_item(lic, tr_masters["hs_code"], 2, item_names=[tr_masters["item_b"]])

    trade_a = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade_a, item_a, "10000.00")
    trade_b = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=2),
    )
    _make_line(trade_b, item_b, "5000.00")

    lic_other = _make_license("TR-ITEMFILTER-OTHER", tr_masters["exporter"])
    _make_export_item(lic_other, tr_masters["e126_norm"])
    item_other = _make_import_item(lic_other, tr_masters["hs_code"], 1, item_names=[tr_masters["item_b"]])
    trade_other = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade_other, item_other, "9000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All", item_id=tr_masters["item_a"].id)

    # A license with zero matching lines is excluded entirely.
    assert _find_license(report, "TR-ITEMFILTER-OTHER") is None

    lic_row = _find_license(report, "TR-ITEMFILTER")
    assert lic_row is not None
    # Only the item_a line survives -- the item_b line on the SAME license
    # is dropped from the register, not merely from the license-selection.
    assert len(lic_row["transactions"]) == 1
    assert lic_row["transactions"][0]["item"] == "TR Item A"
    assert lic_row["summary"]["purchase"] == 10000.00
    items_shown = {it["item"] for it in lic_row["item_summary"]}
    assert items_shown == {"TR Item A"}


@pytest.mark.django_db
def test_norm_filter_narrows_the_register_and_buckets_correctly(tr_masters):
    lic = _make_license("TR-NORMFILTER", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item_e126 = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    item_e132 = _make_import_item(lic, tr_masters["hs_code"], 2, item_names=[tr_masters["item_b"]])

    trade_1 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade_1, item_e126, "10000.00")
    trade_2 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=2),
    )
    _make_line(trade_2, item_e132, "5000.00")

    e126_report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="E126")
    e126_lic = _find_license(e126_report, "TR-NORMFILTER")
    assert e126_lic is not None
    assert len(e126_lic["transactions"]) == 1
    assert e126_lic["summary"]["purchase"] == 10000.00
    assert {n["norm"] for n in e126_report["norms"]} == {"E126"}


@pytest.mark.django_db
def test_license_type_filter_non_dfia_returns_empty_report(tr_masters):
    lic = _make_license("TR-LT-001", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade, item, "10000.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All", license_type="INCENTIVE")
    assert report["norms"] == []
    assert report["dashboard"]["total_licenses"] == 0


@pytest.mark.django_db
def test_customer_filter_narrows_sale_side_only(tr_masters):
    lic = _make_license("TR-CUST", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])

    purchase = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(purchase, item, "10000.00")

    sale_matching = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(sale_matching, item, "8000.00")

    sale_other = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["other_customer"],
        invoice_date=FROM_DATE + timedelta(days=6),
    )
    _make_line(sale_other, item, "7000.00")

    report = build_license_trading_register_report(
        FROM_DATE, TO_DATE, norm="All", customer_id=tr_masters["customer"].id,
    )
    lic_row = _find_license(report, "TR-CUST")
    assert lic_row is not None
    # Purchase side untouched by the customer filter.
    assert lic_row["summary"]["purchase"] == 10000.00
    # Only the matching-customer sale counts.
    assert lic_row["summary"]["sale"] == 8000.00
    assert len(lic_row["transactions"]) == 2


@pytest.mark.django_db
def test_reconciliation_chain_license_norm_grand(tr_masters):
    lic1 = _make_license("TR-REC-001", tr_masters["exporter"])
    _make_export_item(lic1, tr_masters["e126_norm"])
    item1a = _make_import_item(lic1, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    item1b = _make_import_item(lic1, tr_masters["hs_code"], 2, item_names=[tr_masters["item_b"]])

    p1 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p1, item1a, "30000.00")
    _make_line(p1, item1b, "17777.00")  # awkward number to exercise rounding

    s1 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(s1, item1a, "45000.00")
    _make_line(s1, item1b, "25000.00")

    lic2 = _make_license("TR-REC-002", tr_masters["exporter"])
    _make_export_item(lic2, tr_masters["e126_norm"])
    item2a = _make_import_item(lic2, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    p2 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=2),
    )
    _make_line(p2, item2a, "12345.00")
    s2 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=8),
    )
    _make_line(s2, item2a, "22222.00")

    # A second norm, to exercise the grand == sum-of-norms leg too.
    lic3 = _make_license("TR-REC-003", tr_masters["exporter"])
    _make_export_item(lic3, tr_masters["e132_norm"])
    item3 = _make_import_item(lic3, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    p3 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p3, item3, "8888.00")
    s3 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=3),
    )
    _make_line(s3, item3, "9999.00")

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")

    # --- License item summary == license summary --------------------------
    for number in ("TR-REC-001", "TR-REC-002", "TR-REC-003"):
        lic_row = _find_license(report, number)
        item_purchase_sum = sum(Decimal(str(it["purchase_value"])) for it in lic_row["item_summary"])
        item_sale_sum = sum(Decimal(str(it["sale_value"])) for it in lic_row["item_summary"])
        item_profit_sum = sum(Decimal(str(it["profit"])) for it in lic_row["item_summary"])
        assert item_purchase_sum == Decimal(str(lic_row["summary"]["purchase"]))
        assert item_sale_sum == Decimal(str(lic_row["summary"]["sale"]))
        assert item_profit_sum == Decimal(str(lic_row["summary"]["profit"]))

    # --- Norm summary == sum of its licenses; norm item summary == norm summary
    e126 = _find_norm(report, "E126")
    e126_licenses = e126["licenses"]
    assert {l["license_number"] for l in e126_licenses} == {"TR-REC-001", "TR-REC-002"}
    lic_purchase_sum = sum(Decimal(str(l["summary"]["purchase"])) for l in e126_licenses)
    lic_sale_sum = sum(Decimal(str(l["summary"]["sale"])) for l in e126_licenses)
    lic_profit_sum = sum(Decimal(str(l["summary"]["profit"])) for l in e126_licenses)
    assert lic_purchase_sum == Decimal(str(e126["summary"]["purchase"]))
    assert lic_sale_sum == Decimal(str(e126["summary"]["sale"]))
    assert lic_profit_sum == Decimal(str(e126["summary"]["profit"]))
    assert e126["summary"]["licenses_count"] == 2

    norm_item_purchase_sum = sum(Decimal(str(it["purchase_value"])) for it in e126["item_summary"])
    norm_item_sale_sum = sum(Decimal(str(it["sale_value"])) for it in e126["item_summary"])
    norm_item_profit_sum = sum(Decimal(str(it["profit"])) for it in e126["item_summary"])
    assert norm_item_purchase_sum == Decimal(str(e126["summary"]["purchase"]))
    assert norm_item_sale_sum == Decimal(str(e126["summary"]["sale"]))
    assert norm_item_profit_sum == Decimal(str(e126["summary"]["profit"]))

    # --- Grand == sum of norms ----------------------------------------------
    norm_purchase_sum = sum(Decimal(str(n["summary"]["purchase"])) for n in report["norms"])
    norm_sale_sum = sum(Decimal(str(n["summary"]["sale"])) for n in report["norms"])
    norm_profit_sum = sum(Decimal(str(n["summary"]["profit"])) for n in report["norms"])
    grand = report["grand_summary"]
    assert norm_purchase_sum == Decimal(str(grand["purchase"]))
    assert norm_sale_sum == Decimal(str(grand["sale"]))
    assert norm_profit_sum == Decimal(str(grand["profit"]))
    assert grand["licenses_count"] == 3

    # --- Grand item summary == grand summary --------------------------------
    grand_item_purchase_sum = sum(Decimal(str(r["purchase_value"])) for r in report["grand_item_summary"])
    grand_item_sale_sum = sum(Decimal(str(r["sale_value"])) for r in report["grand_item_summary"])
    grand_item_profit_sum = sum(Decimal(str(r["profit"])) for r in report["grand_item_summary"])
    assert grand_item_purchase_sum == Decimal(str(grand["purchase"]))
    assert grand_item_sale_sum == Decimal(str(grand["sale"]))
    assert grand_item_profit_sum == Decimal(str(grand["profit"]))

    # Also true independently for the dashboard rollup.
    dash = report["dashboard"]
    assert Decimal(str(dash["total_purchase"])) == Decimal(str(grand["purchase"]))
    assert Decimal(str(dash["total_sale"])) == Decimal(str(grand["sale"]))
    assert Decimal(str(dash["total_profit"])) == Decimal(str(grand["profit"]))
    assert dash["total_licenses"] == 3


@pytest.mark.django_db
def test_dashboard_open_closed_counts(tr_masters):
    lic1 = _make_license("TR-DASH-001", tr_masters["exporter"])
    _make_export_item(lic1, tr_masters["e126_norm"])
    item1 = _make_import_item(lic1, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    p1 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p1, item1, "10000.00")
    s1 = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=2),
    )
    _make_line(s1, item1, "15000.00")  # Closed

    lic2 = _make_license("TR-DASH-002", tr_masters["exporter"])
    _make_export_item(lic2, tr_masters["e132_norm"])
    item2 = _make_import_item(lic2, tr_masters["hs_code"], 1, item_names=[tr_masters["item_b"]])
    p2 = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(p2, item2, "20000.00")  # Open (no sale)

    report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    dash = report["dashboard"]
    assert dash["total_licenses"] == 2
    assert dash["open_licenses"] == 1
    assert dash["closed_licenses"] == 1


@pytest.mark.django_db
def test_query_count_stays_constant_regardless_of_fixture_size(tr_masters):
    def _build_fixture(n_licenses, trades_per_license, prefix):
        for i in range(n_licenses):
            lic = _make_license(f"TR-PERF-{prefix}-{i}", tr_masters["exporter"])
            _make_export_item(lic, tr_masters["e126_norm"])
            item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
            for j in range(trades_per_license):
                pt = _make_trade(
                    LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
                    invoice_date=FROM_DATE + timedelta(days=j + 1), invoice_number=f"P-{prefix}-{i}-{j}",
                )
                _make_line(pt, item, "1000.00")
                st = _make_trade(
                    LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
                    invoice_date=FROM_DATE + timedelta(days=j + 30), invoice_number=f"S-{prefix}-{i}-{j}",
                )
                _make_line(st, item, "1500.00")

    _build_fixture(2, 2, "A")
    with CaptureQueriesContext(connection) as small_ctx:
        small_report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    assert small_report["dashboard"]["total_licenses"] == 2

    _build_fixture(6, 5, "B")
    with CaptureQueriesContext(connection) as large_ctx:
        large_report = build_license_trading_register_report(FROM_DATE, TO_DATE, norm="All")
    assert large_report["dashboard"]["total_licenses"] == 8

    assert len(large_ctx.captured_queries) == len(small_ctx.captured_queries)


# ---------------------------------------------------------------------------
# View-level tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_view_json_response_has_expected_top_level_keys(report_viewer_client, tr_masters):
    lic = _make_license("TR-VIEW-001", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade, item, "10000.00")

    response = report_viewer_client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code == 200
    assert response["Content-Type"].startswith("application/json")
    data = response.json()
    assert "dashboard" in data
    assert "norms" in data
    assert "grand_summary" in data
    assert "grand_item_summary" in data
    assert "licenses_count" in data["grand_summary"]


@pytest.mark.django_db
def test_view_excel_export_returns_valid_workbook(report_viewer_client, tr_masters):
    lic = _make_license("TR-VIEW-002", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade, item, "10000.00")
    sale = _make_trade(
        LicenseTrade.DIR_SALE, from_company=tr_masters["exporter"], to_company=tr_masters["customer"],
        invoice_date=FROM_DATE + timedelta(days=5),
    )
    _make_line(sale, item, "15000.00")

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.content))
    assert "Trading Register" in workbook.sheetnames


@pytest.mark.django_db
def test_view_pdf_export_returns_pdf_content_type(report_viewer_client, tr_masters):
    lic = _make_license("TR-VIEW-003", tr_masters["exporter"])
    _make_export_item(lic, tr_masters["e126_norm"])
    item = _make_import_item(lic, tr_masters["hs_code"], 1, item_names=[tr_masters["item_a"]])
    trade = _make_trade(
        LicenseTrade.DIR_PURCHASE, from_company=tr_masters["supplier"], to_company=tr_masters["exporter"],
        invoice_date=FROM_DATE + timedelta(days=1),
    )
    _make_line(trade, item, "10000.00")

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


@pytest.mark.django_db
def test_view_empty_report_excel_and_pdf_still_render(report_viewer_client):
    """No matching data at all -- exports must not blow up on empty lists."""
    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "excel"},
    )
    assert response.status_code == 200

    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "format": "pdf"},
    )
    assert response.status_code == 200


@pytest.mark.django_db
def test_view_missing_date_params_returns_400(report_viewer_client):
    response = report_viewer_client.get(REPORT_URL)
    assert response.status_code == 400

    response = report_viewer_client.get(REPORT_URL, {"from_date": FROM_DATE.isoformat()})
    assert response.status_code == 400


@pytest.mark.django_db
def test_view_invalid_date_format_returns_400(report_viewer_client):
    response = report_viewer_client.get(
        REPORT_URL, {"from_date": "01-01-2026", "to_date": "31-01-2026"}
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_view_invalid_int_param_returns_400(report_viewer_client):
    response = report_viewer_client.get(
        REPORT_URL,
        {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat(), "exporter_id": "abc"},
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_view_requires_report_permission(no_role_client):
    response = no_role_client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code == 403


@pytest.mark.django_db
def test_view_unauthenticated_is_rejected():
    client = APIClient()
    response = client.get(
        REPORT_URL, {"from_date": FROM_DATE.isoformat(), "to_date": TO_DATE.isoformat()}
    )
    assert response.status_code in (401, 403)
