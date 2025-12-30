# bill_of_entry/views_export.py
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.decorators import action

from core.utils.pdf_utils import create_pdf_exporter

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def add_grouped_export_action(viewset_class):
    """
    Decorator to add grouped export functionality to BillOfEntryViewSet
    """

    @action(detail=False, methods=['get'], url_path='export')
    def export_bill_of_entries(self, request):
        """
        Export bill of entries grouped by company.
        URL: /api/bill-of-entries/export/?_export=pdf

        Query params:
            - _export: 'pdf' or 'xlsx' (default: pdf)
            - company: filter by company ID
            - bill_of_entry_date_after: filter by date
            - bill_of_entry_date_before: filter by date
        """
        export_format = request.query_params.get('_export', 'pdf').lower()

        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Filter only BOEs with item details and order by date first, then company, item, port
        queryset = queryset.filter(
            item_details__isnull=False
        ).distinct().order_by(
            'bill_of_entry_date', 'company__name', 'product_name', 'port__code'
        ).prefetch_related(
            'item_details__sr_number__license',
            'item_details__sr_number__hs_code',
            'company',
            'port'
        )

        if export_format == 'pdf':
            return self._export_grouped_pdf(queryset)
        elif export_format == 'xlsx':
            return self._export_grouped_xlsx(queryset)
        else:
            return HttpResponse("Invalid export format. Use 'pdf' or 'xlsx'.", status=400)

    def _export_grouped_pdf(self, queryset):
        """Export grouped bill of entries to PDF with business-level presentation"""
        pdf_exporter = create_pdf_exporter(
            title="Bill of Entry Report",
            filename_prefix="Bill_of_Entries",
            orientation='landscape'
        )

        if not pdf_exporter:
            return HttpResponse("PDF export not available", status=500)

        # Group data
        grouped_data = self._group_boe(queryset)

        # Calculate grand totals
        total_inr = sum(boe['total_inr']
                        for company in grouped_data.values()
                        for license_dict in company.values()
                        for product in license_dict.values()
                        for port in product.values()
                        for boe in port)
        total_fc = sum(boe['total_fc']
                       for company in grouped_data.values()
                       for license_dict in company.values()
                       for product in license_dict.values()
                       for port in product.values()
                       for boe in port)

        # Create PDF response - inline display for new tab
        response = HttpResponse(content_type='application/pdf')
        today = datetime.now().strftime('%d-%m-%Y')
        filename = f'BOE Report - {today}.pdf'
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        buffer = BytesIO()
        doc = pdf_exporter.create_document(buffer)

        elements = []

        # Add title with totals
        subtitle = (f"Total CIF (FC): ${pdf_exporter.format_number(total_fc)} | "
                    f"Total CIF (INR): {pdf_exporter.format_number(total_inr)}")
        pdf_exporter.add_title(elements, subtitle=subtitle)

        # Process each company (sorted alphabetically, None values last)
        for company_name, license_dict in sorted(grouped_data.items(), key=lambda x: (x[0] is None, x[0] or '')):
            # Company header
            pdf_exporter.add_company_header(elements, company_name)

            # Track company totals
            company_total_qty = 0
            company_total_inr = 0
            company_total_fc = 0

            # Flatten all BOEs for this company into a single list grouped by port
            company_boes_by_port = defaultdict(list)
            for license_serial, products_dict in license_dict.items():
                for product_name, ports_dict in products_dict.items():
                    for port_code, boe_list in ports_dict.items():
                        for boe in boe_list:
                            company_boes_by_port[port_code].append(boe)

            # Create single consolidated table for entire company
            sr_no = 1

            # Table header - 18 columns (added Exporter and Exchange Rate)
            table_data = [[
                'Sr\nNo', 'BOE\nNumber', 'BOE\nDate', 'Port', 'Quantity\n(KGS)',
                'Unit\nPrice ($)', 'Value\n($)', 'Exchange\nRate', 'Total CIF\nINR', 'Item\nName', 'Invoice',
                'Exporter', 'License\nNo.', 'License\nDate', 'License\nPort', 'Item\nSr.',
                'BOE\nQty.', 'BOE\n$.', 'BOE\nCIF'
            ]]

            # Process each port and its BOEs
            for port_code in sorted(company_boes_by_port.keys()):
                boe_list = company_boes_by_port[port_code]
                port_total_qty = 0
                port_total_value = 0
                port_total_inr = 0

                for boe in boe_list:
                    # Main BOE row (first license detail)
                    if boe['license_details']:
                        first_detail = boe['license_details'][0]

                        table_data.append([
                            str(sr_no),
                            boe['boe_number'],
                            boe['boe_date'],
                            port_code,
                            pdf_exporter.format_number(boe['total_quantity'], decimals=0),
                            pdf_exporter.format_number(
                                boe['total_fc'] / boe['total_quantity'] if boe['total_quantity'] > 0 else 0
                            ),
                            pdf_exporter.format_number(boe['total_fc']),
                            pdf_exporter.format_number(boe['exchange_rate']),
                            pdf_exporter.format_number(boe['total_inr']),
                            boe['product_name'],  # Use the product_name from boe data
                            boe['invoice_no'],
                            first_detail['exporter_name'],  # Exporter name
                            first_detail['license_no'],     # License number only
                            first_detail['license_date'],
                            first_detail['license_port'],
                            first_detail['item_sr_no'],
                            pdf_exporter.format_number(first_detail['qty'], decimals=0),
                            pdf_exporter.format_number(first_detail['cif_fc']),
                            pdf_exporter.format_number(first_detail['cif_inr'])
                        ])

                        # Additional license detail rows (if multiple licenses)
                        for detail in boe['license_details'][1:]:
                            table_data.append([
                                '', '', '', '', '', '', '', '', '',  # Empty main BOE info columns
                                '', '',  # Item Name, Invoice
                                detail['exporter_name'],  # Exporter name
                                detail['license_no'],     # License number only
                                detail['license_date'],
                                detail['license_port'],
                                detail['item_sr_no'],
                                pdf_exporter.format_number(detail['qty'], decimals=0),
                                pdf_exporter.format_number(detail['cif_fc']),
                                pdf_exporter.format_number(detail['cif_inr'])
                            ])

                    sr_no += 1
                    port_total_qty += boe['total_quantity']
                    port_total_value += boe['total_fc']
                    port_total_inr += boe['total_inr']
                    company_total_qty += boe['total_quantity']
                    company_total_inr += boe['total_inr']
                    company_total_fc += boe['total_fc']

                # Add port totals row (18 columns now) - properly aligned
                table_data.append([
                    '',  # Sr No
                    '',  # BOE Number
                    '',  # BOE Date
                    f'Port Total ({port_code})',  # Port
                    pdf_exporter.format_number(port_total_qty, decimals=0),  # Quantity (KGS)
                    '',  # Unit Price ($)
                    pdf_exporter.format_number(port_total_value),  # Value ($)
                    '',  # Exchange Rate
                    pdf_exporter.format_number(port_total_inr),  # Total CIF INR
                    '',  # Item Name
                    '',  # Invoice
                    '',  # Exporter
                    '',  # License No.
                    '',  # License Date
                    '',  # License Port
                    '',  # Item Sr.
                    '',  # BOE Qty.
                    '',  # BOE $.
                    ''   # BOE CIF
                ])

            # Create table with column widths (18 columns) - optimized to fit A4 landscape with small margins
            from reportlab.lib.units import inch
            col_widths = [
                0.28 * inch,  # Sr No
                0.6 * inch,   # BOE Number
                0.52 * inch,  # BOE Date
                0.48 * inch,  # Port
                0.52 * inch,  # Quantity (KGS)
                0.48 * inch,  # Unit Price ($)
                0.6 * inch,   # Value ($)
                0.48 * inch,  # Exchange Rate
                0.72 * inch,  # Total CIF INR
                0.7 * inch,   # Item Name
                0.52 * inch,  # Invoice
                0.75 * inch,  # Exporter
                0.68 * inch,  # License No.
                0.52 * inch,  # License Date
                0.48 * inch,  # License Port
                0.35 * inch,  # Item Sr.
                0.48 * inch,  # BOE Qty.
                0.52 * inch,  # BOE $.
                0.6 * inch    # BOE CIF
            ]  # Total: ~10.78 inches

            table = pdf_exporter.create_table(table_data, col_widths=col_widths, repeating_rows=1)

            # Apply number column alignment for columns: 4, 5, 6, 7, 8, 15, 16, 17, 18 (0-indexed)
            from reportlab.platypus import TableStyle
            additional_styles = []
            for col_idx in [4, 5, 6, 7, 8, 15, 16, 17]:
                additional_styles.append(
                    ('ALIGN', (col_idx, 1), (col_idx, len(table_data) - 1), 'RIGHT')
                )

            # Find rows where Port Total appears and make them bold
            for row_idx, row in enumerate(table_data):
                if row[3] and 'Port Total' in str(row[3]):
                    additional_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                    additional_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), pdf_exporter.HEADER_BG))

            # Apply additional styles
            table.setStyle(TableStyle(additional_styles))

            elements.append(table)
            pdf_exporter.add_spacer(elements, 0.2)

            # Add company total summary
            summary_data = [[
                'Grand Total:',
                f"Qty: {pdf_exporter.format_number(company_total_qty, decimals=0)} KGS",
                f"CIF (FC): ${pdf_exporter.format_number(company_total_fc)}",
                f"CIF (INR): {pdf_exporter.format_number(company_total_inr)}"
            ]]
            summary_table = pdf_exporter.create_summary_table(summary_data)
            elements.append(summary_table)
            pdf_exporter.add_spacer(elements, 0.3)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    def _export_grouped_xlsx(self, queryset):
        """Export grouped bill of entries to Excel"""
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("Excel export not available", status=500)

        # Group data
        grouped_data = self._group_boe(queryset)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bill of Entries"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Title
        ws.merge_cells(f'A{row}:R{row}')
        cell = ws[f'A{row}']
        cell.value = "Bill of Entry Report"
        cell.font = Font(bold=True, size=16, color="1e3a8a")
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Total and date
        total_inr = sum(boe['total_inr']
                        for company in grouped_data.values()
                        for license_dict in company.values()
                        for product in license_dict.values()
                        for port in product.values()
                        for boe in port)
        total_fc = sum(boe['total_fc']
                       for company in grouped_data.values()
                       for license_dict in company.values()
                       for product in license_dict.values()
                       for port in product.values()
                       for boe in port)
        ws.merge_cells(f'A{row}:R{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total INR: ₹{total_inr:,.2f}    Total FC: ${total_fc:,.2f}    {datetime.now().strftime('%d-%m-%Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Process each company (sorted alphabetically, None values last)
        for company_name, license_dict in sorted(grouped_data.items(), key=lambda x: (x[0] is None, x[0] or '')):
            # Company header
            ws.merge_cells(f'A{row}:R{row}')
            cell = ws[f'A{row}']
            cell.value = company_name
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

            # Track company totals
            company_total_qty = 0
            company_total_inr = 0
            company_total_fc = 0

            # Process each license serial within company
            for license_serial, products_dict in license_dict.items():
                # Process each product (Item) within license
                for product_name, ports_dict in products_dict.items():
                    # Product subheader
                    ws.merge_cells(f'A{row}:R{row}')
                    cell = ws[f'A{row}']
                    cell.value = f"Item: {product_name}"
                    cell.font = Font(bold=True, size=11, color="3b82f6")
                    row += 1

                # Table headers (18 columns - added Exporter and Exchange Rate)
                headers = ['Sr No', 'BOE Number', 'BOE Date', 'Port', 'Quantity (KGS)', 'Unit Price ($)', 'Value ($)',
                           'Exchange Rate', 'Total CIF INR', 'Item Name', 'Invoice', 'Exporter', 'License No.', 'License Date',
                           'License Port', 'Item Sr.', 'BOE Qty.', 'BOE $.', 'BOE CIF']

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                    cell.border = border
                row += 1

                sr_no = 1
                product_total_qty = 0
                product_total_value = 0
                product_total_inr = 0

                # Process each port within product
                for port_code, boe_list in ports_dict.items():
                    for boe in boe_list:
                        if boe['license_details']:
                            first_detail = boe['license_details'][0]
                            unit_price = boe['total_fc'] / boe['total_quantity'] if boe['total_quantity'] > 0 else 0

                            # Main BOE row (18 columns - added Exporter)
                            data = [
                                sr_no,
                                boe['boe_number'],
                                boe['boe_date'],
                                port_code,
                                int(boe['total_quantity']),
                                round(unit_price, 2),
                                round(boe['total_fc'], 2),
                                round(boe['exchange_rate'], 2),  # Exchange Rate
                                round(boe['total_inr'], 2),  # Total CIF INR
                                product_name,
                                boe['invoice_no'],
                                first_detail['exporter_name'],  # Exporter
                                first_detail['license_no'],     # License No.
                                first_detail['license_date'],
                                first_detail['license_port'],
                                first_detail['item_sr_no'],
                                int(first_detail['qty']),
                                round(first_detail['cif_fc'], 2),
                                round(first_detail['cif_inr'], 2)
                            ]

                            for col_idx, value in enumerate(data, 1):
                                cell = ws.cell(row=row, column=col_idx, value=value)
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                if col_idx in [5, 6, 7, 8, 9, 17, 18, 19]:  # Number columns (updated indices)
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                    if col_idx in [6, 7, 18]:  # USD columns
                                        cell.number_format = '#,##0.00'
                                    elif col_idx in [9, 19]:  # INR columns
                                        cell.number_format = '₹#,##0.00'
                                    elif col_idx == 8:  # Exchange Rate column
                                        cell.number_format = '#,##0.00'
                            row += 1

                            # Additional license detail rows
                            for detail in boe['license_details'][1:]:
                                data = [
                                    '', '', '', '', '', '', '', '', '',  # Empty main BOE columns
                                    '', '',  # Item Name, Invoice
                                    detail['exporter_name'],  # Exporter
                                    detail['license_no'],     # License No.
                                    detail['license_date'],
                                    detail['license_port'],
                                    detail['item_sr_no'],
                                    int(detail['qty']),
                                    round(detail['cif_fc'], 2),
                                    round(detail['cif_inr'], 2)
                                ]

                                for col_idx, value in enumerate(data, 1):
                                    cell = ws.cell(row=row, column=col_idx, value=value)
                                    cell.border = border
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    if col_idx in [17, 18, 19]:  # Updated column indices
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                        if col_idx in [18, 19]:  # Updated column indices
                                            cell.number_format = '#,##0.00'
                                row += 1

                        sr_no += 1
                        product_total_qty += boe['total_quantity']
                        product_total_value += boe['total_fc']
                        product_total_inr += boe['total_inr']
                        company_total_qty += boe['total_quantity']
                        company_total_inr += boe['total_inr']
                        company_total_fc += boe['total_fc']

                # Product totals
                ws.cell(row=row, column=3, value="Total").font = Font(bold=True)
                ws.cell(row=row, column=4, value=int(product_total_qty)).font = Font(bold=True)
                ws.cell(row=row, column=4).number_format = '#,##0'
                ws.cell(row=row, column=6, value=round(product_total_value, 2)).font = Font(bold=True)
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=7, value=round(product_total_inr, 2)).font = Font(bold=True)
                ws.cell(row=row, column=7).number_format = '₹#,##0.00'
                row += 2

            # Add grand total after all products
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws.cell(row=row, column=1, value="Grand Total:")
            cell.font = Font(bold=True, size=12, color="1e3a8a")
            cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=4, value=int(company_total_qty)).font = Font(bold=True, size=11)
            ws.cell(row=row, column=4).fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=4).number_format = '#,##0'

            ws.cell(row=row, column=6, value=round(company_total_fc, 2)).font = Font(bold=True, size=11)
            ws.cell(row=row, column=6).fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=6).number_format = '#,##0.00'

            ws.merge_cells(f'H{row}:I{row}')
            cell = ws.cell(row=row, column=8, value=f"CIF INR: ₹{company_total_inr:,.2f}")
            cell.font = Font(bold=True, size=11, color="1e3a8a")
            cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 3

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        today = datetime.now().strftime('%d-%m-%Y')
        response['Content-Disposition'] = f'attachment; filename="BOE Report - {today}.xlsx"'
        wb.save(response)

        return response

    def _group_boe(self, queryset):
        """Group bill of entries by company → license_serial_number → product_name → port"""
        # Structure: {company: {license_serial: {product_name: {port: [boe_list]}}}}
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))

        for boe in queryset:
            company_name = boe.company.name if boe.company else "Unknown"
            port_code = boe.port.code if boe.port else "Unknown Port"
            # Normalize product name: uppercase, strip spaces for grouping
            raw_product_name = boe.product_name or "Unknown Product"
            product_name = raw_product_name.strip().upper()

            # Get first license serial number from item details
            first_item = boe.item_details.first()
            license_serial = "Unknown License"
            if first_item and first_item.sr_number:
                license_obj = first_item.sr_number.license if first_item.sr_number else None
                if license_obj:
                    license_serial = f"{license_obj.license_number} (Sr. {first_item.sr_number.serial_number})"

            # Collect license details for this BOE
            license_details = []
            for detail in boe.item_details.all():
                license_obj = detail.sr_number.license if detail.sr_number else None

                # Store exporter name and license number separately
                if license_obj:
                    exporter_name = license_obj.exporter.name if license_obj.exporter else '--'
                    license_number = license_obj.license_number
                else:
                    exporter_name = '--'
                    license_number = '--'

                license_details.append({
                    'exporter_name': exporter_name,
                    'license_no': license_number,
                    'license_port': license_obj.port.code if (license_obj and license_obj.port) else '--',
                    'license_date': license_obj.license_date.strftime('%d-%m-%Y') if license_obj else '--',
                    'item_sr_no': str(detail.sr_number.serial_number) if detail.sr_number else '--',
                    'qty': float(detail.qty or 0),
                    'cif_fc': float(detail.cif_fc or 0),
                    'cif_inr': float(detail.cif_inr or 0)
                })

            # Calculate exchange rate: use boe.exchange_rate if exists, otherwise calculate from total_inr / total_fc
            total_fc = float(boe.get_total_fc or 0)
            total_inr = float(boe.get_total_inr or 0)

            if boe.exchange_rate and float(boe.exchange_rate) > 0:
                exchange_rate = float(boe.exchange_rate)
            elif total_fc > 0:
                exchange_rate = total_inr / total_fc
            else:
                exchange_rate = 0

            boe_data = {
                'boe_number': boe.bill_of_entry_number or '--',
                'boe_date': boe.bill_of_entry_date.strftime('%d-%m-%Y') if boe.bill_of_entry_date else '--',
                'port': port_code,
                'product_name': product_name,
                'invoice_no': boe.invoice_no or '--',
                'total_quantity': float(boe.get_total_quantity or 0),
                'total_fc': total_fc,
                'total_inr': total_inr,
                'exchange_rate': exchange_rate,  # Add exchange rate
                'license_details': license_details,  # List of license entries
            }

            grouped_data[company_name][license_serial][product_name][port_code].append(boe_data)

        return grouped_data

    # Add methods to the viewset class
    viewset_class.export_bill_of_entries = export_bill_of_entries
    viewset_class._export_grouped_pdf = _export_grouped_pdf
    viewset_class._export_grouped_xlsx = _export_grouped_xlsx
    viewset_class._group_boe = _group_boe

    return viewset_class
