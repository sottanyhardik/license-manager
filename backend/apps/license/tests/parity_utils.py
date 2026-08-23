"""
Parity Testing Utilities

Shared utilities for comparing UI (API), PDF, and Excel outputs.
Used by test_ui_pdf_excel_parity_golden.py and other parity tests.
"""

from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Any


class APIDataExtractor:
    """Extract structured data from API (ledger) responses."""

    @staticmethod
    def extract_summary(api_response: Dict) -> Dict[str, Decimal]:
        """
        Extract financial summary from API response.

        Returns:
        {
            'purchase_bill': Decimal,
            'sale_bill': Decimal,
            'profit': Decimal,
            'profit_state': str,
        }
        """
        summary = api_response.get('summary', {})
        return {
            'purchase_bill': Decimal(str(summary.get('total_purchase_bill_inr', 0))).quantize(Decimal('0.01')),
            'sale_bill': Decimal(str(summary.get('total_sale_bill_inr', 0))).quantize(Decimal('0.01')),
            'profit': Decimal(str(summary.get('total_profit_loss', 0))).quantize(Decimal('0.01')),
            'profit_state': summary.get('profit_state', 'UNAVAILABLE'),
        }

    @staticmethod
    def extract_transactions(api_response: Dict) -> List[Dict]:
        """
        Extract transaction list from API response.

        Each transaction has:
        - id: transaction ID
        - type: PURCHASE, SALE, COMMISSION, OPENING
        - amount: license value (USD for DFIA)
        - bill_amount: invoice amount (INR, None for OPENING)
        - license_running_balance: running balance after this transaction
        """
        return api_response.get('transactions', [])

    @staticmethod
    def extract_company_utilizations(api_response: Dict) -> Dict[str, Decimal]:
        """
        Extract per-company balance breakdown from API response.

        Returns:
        {
            'company_id:company_name': Decimal(balance),
            ...
        }
        """
        utilizations = {}
        for company_id, util_data in api_response.get('company_utilizations', {}).items():
            company_name = util_data.get('company_name', 'Unknown')
            balance = Decimal(str(util_data.get('utilization_balance', 0))).quantize(Decimal('0.01'))
            utilizations[f"{company_id}:{company_name}"] = balance
        return utilizations

    @staticmethod
    def transaction_count(api_response: Dict) -> int:
        """Get count of all transactions in API response."""
        return len(api_response.get('transactions', []))

    @staticmethod
    def display_transaction_count(api_response: Dict) -> int:
        """
        Get count of display transactions (respects display rule).

        Display rule: PURCHASE + SALE only, OPENING shown only when no PURCHASE.
        """
        display = api_response.get('display_transactions', [])
        opening = api_response.get('opening_display')

        count = len(display)
        if opening is not None:
            count += 1

        return count


class PDFDataExtractor:
    """Extract structured data from PDF exports."""

    @staticmethod
    def extract_text(pdf_bytes: bytes) -> str:
        """Extract all text from PDF as normalized whitespace-joined string."""
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(pdf_bytes))
        full_text = " ".join(
            " ".join(page.extract_text().split())
            for page in reader.pages
        )
        return full_text

    @staticmethod
    def normalize_number_for_search(decimal_value: Decimal) -> List[str]:
        """
        Generate all reasonable text representations of a number for PDF search.

        PDF might render as:
        - "4583719.00" (plain, no separators)
        - "45,83,719.00" (Indian format: X,XX,XX,XXX)
        - "45 83 719.00" (Indian with spaces)
        - "4583719" (no decimal)

        Returns all formats to try when searching PDF text.
        """
        integer_part = int(decimal_value)
        plain_str = f"{decimal_value:.2f}"

        formats = [
            plain_str,  # "4583719.00"
            str(integer_part),  # "4583719"
            plain_str.replace('.', ''),  # "458371900"
        ]

        # Indian format: X,XX,XX,XXX
        if integer_part >= 100000:
            s = str(integer_part)
            if len(s) > 5:
                indian = s[:-5] + ',' + s[-5:-2] + ',' + s[-2:]
                formats.append(f"{indian}.00")  # "45,83,719.00"
                formats.append(indian)  # "45,83,719"
                # Variant with spaces
                formats.append(s[:-5] + ' ' + s[-5:-2] + ' ' + s[-2:] + '.00')

        return formats

    @staticmethod
    def contains_value(pdf_text: str, decimal_value: Decimal) -> bool:
        """Check if PDF text contains this number in any reasonable format."""
        search_formats = PDFDataExtractor.normalize_number_for_search(decimal_value)
        return any(fmt in pdf_text for fmt in search_formats)

    @staticmethod
    def get_page_count(pdf_bytes: bytes) -> int:
        """Get number of pages in PDF."""
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(pdf_bytes))
        return len(reader.pages)


class ExcelDataExtractor:
    """Extract structured data from Excel exports."""

    @staticmethod
    def load_workbook(excel_bytes: bytes):
        """Load Excel workbook from bytes."""
        from openpyxl import load_workbook
        return load_workbook(BytesIO(excel_bytes))

    @staticmethod
    def find_value_in_worksheet(
        worksheet,
        target_value: Decimal,
        search_cols: Optional[range] = None,
        max_rows: int = 50,
    ) -> Optional[Tuple[int, int]]:
        """
        Search for a specific Decimal value in worksheet cells.

        Returns:
        - (row, col) tuple if found (1-indexed)
        - None if not found
        """
        if search_cols is None:
            search_cols = range(1, worksheet.max_column + 1)

        target_decimal = Decimal(str(target_value)).quantize(Decimal('0.01'))

        for row in range(1, min(max_rows, worksheet.max_row + 1)):
            for col in search_cols:
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    try:
                        cell_decimal = Decimal(str(cell_value)).quantize(Decimal('0.01'))
                        if cell_decimal == target_decimal:
                            return (row, col)
                    except (ValueError, TypeError):
                        pass

        return None

    @staticmethod
    def find_sheet_by_name_pattern(
        workbook,
        pattern: str,
    ) -> Optional[str]:
        """
        Find first sheet name matching pattern (case-insensitive substring).

        Returns first matching sheet name, or None.
        """
        pattern_lower = pattern.lower()
        for name in workbook.sheetnames:
            if pattern_lower in name.lower():
                return name
        return None

    @staticmethod
    def get_cell_value(worksheet, row: int, col: int) -> Any:
        """Get value from specific cell (1-indexed)."""
        return worksheet.cell(row=row, column=col).value

    @staticmethod
    def get_row_values(worksheet, row: int) -> List[Any]:
        """Get all values from a row (1-indexed)."""
        return [worksheet.cell(row=row, column=col).value
                for col in range(1, worksheet.max_column + 1)]


class ParityAssertions:
    """Common assertion helpers for parity tests."""

    @staticmethod
    def assert_values_equal(
        actual: Decimal,
        expected: Decimal,
        label: str,
        tolerance: Decimal = Decimal('0.01'),
    ) -> None:
        """Assert two decimal values are equal within tolerance."""
        actual_dec = Decimal(str(actual)).quantize(Decimal('0.01'))
        expected_dec = Decimal(str(expected)).quantize(Decimal('0.01'))

        if abs(actual_dec - expected_dec) > tolerance:
            raise AssertionError(
                f"{label}: expected {expected_dec}, got {actual_dec} "
                f"(difference: {abs(actual_dec - expected_dec)})"
            )

    @staticmethod
    def assert_precision_2_places(value: Decimal, label: str) -> None:
        """Assert value uses exactly 2 decimal places."""
        value_dec = Decimal(str(value))
        quantized = value_dec.quantize(Decimal('0.01'))

        if value_dec != quantized:
            raise AssertionError(
                f"{label}: {value_dec} does not have 2 decimal places "
                f"(normalized to {quantized})"
            )

    @staticmethod
    def assert_all_values_present(
        actual_dict: Dict[str, Any],
        required_keys: List[str],
        context: str = "response",
    ) -> None:
        """Assert all required keys present in dictionary."""
        missing = [k for k in required_keys if k not in actual_dict]

        if missing:
            raise AssertionError(
                f"{context}: missing required keys {missing}. "
                f"Present: {list(actual_dict.keys())}"
            )


class ParityDataComparison:
    """Compare financial data across outputs."""

    @staticmethod
    def compare_summary_blocks(
        api_summary: Dict[str, Decimal],
        expected_summary: Dict[str, Decimal],
    ) -> Dict[str, Any]:
        """
        Compare API summary to expected values.

        Returns:
        {
            'purchase_bill': {'expected': X, 'actual': Y, 'match': bool},
            'sale_bill': {...},
            'profit': {...},
            'profit_state': {...},
        }
        """
        comparison = {}

        for key in ['purchase_bill', 'sale_bill', 'profit']:
            expected = expected_summary.get(key)
            actual = api_summary.get(key)

            comparison[key] = {
                'expected': expected,
                'actual': actual,
                'match': expected == actual if expected and actual else None,
            }

        # Profit state (string comparison)
        comparison['profit_state'] = {
            'expected': expected_summary.get('profit_state'),
            'actual': api_summary.get('profit_state'),
            'match': expected_summary.get('profit_state') == api_summary.get('profit_state'),
        }

        return comparison

    @staticmethod
    def all_match(comparison: Dict[str, Any]) -> bool:
        """Check if all values in comparison matched."""
        return all(v.get('match') for v in comparison.values())

    @staticmethod
    def mismatch_report(comparison: Dict[str, Any]) -> str:
        """Generate human-readable mismatch report."""
        lines = []
        for key, result in comparison.items():
            if not result.get('match'):
                lines.append(
                    f"  {key}: expected {result['expected']}, got {result['actual']}"
                )
        return "Mismatches:\n" + "\n".join(lines) if lines else "All values match"
