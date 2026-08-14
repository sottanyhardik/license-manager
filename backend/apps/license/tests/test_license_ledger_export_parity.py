"""Golden parity: UI API, PDF and Excel consume one canonical ledger dataset."""
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
from pypdf import PdfReader
import pytest
from django.urls import reverse

from apps.license.services.exporters.financial_ledger_excel_renderer import render_financial_ledger_excel
from apps.license.services.exporters.financial_ledger_pdf_renderer import render_financial_ledger_pdf
from apps.license.services.canonical_ledger_service import CanonicalLedgerService


def canonical_golden_data():
    purchase = {
        "date": date(2026, 4, 1), "id": 1, "type": "PURCHASE",
        "company_id": 766, "company_name": "Plan 766",
        "party_name": "Supplier", "item_names": ["Item 766"],
        "purchase_amount": Decimal("1250.50"), "sale_amount": None,
        "purchase_bill_amount": Decimal("100000.25"), "sale_bill_amount": None,
        "invoice_document": {
            "invoice_number": "PUR-001", "document_exists": True, "signed": True,
            "status": "SIGNED", "secure_url": "/api/invoice-documents/view/purchase-token/",
        },
    }
    sale = {
        "date": date(2026, 4, 2), "id": 2, "type": "SALE",
        "company_id": 766, "company_name": "Plan 766",
        "party_name": "Buyer", "item_names": ["Item 766"],
        "purchase_amount": None, "sale_amount": Decimal("250.25"),
        "purchase_bill_amount": None, "sale_bill_amount": Decimal("30000.75"),
        "invoice_document": {
            "invoice_number": "SAL-001", "document_exists": True, "signed": False,
            "status": "UNSIGNED", "secure_url": "/api/invoice-documents/view/sale-token/",
        },
    }
    return {
        "license_id": 2436, "license_number": "2436", "license_type": "DFIA",
        "license_date": date(2026, 4, 1), "expiry_date": date(2027, 4, 1),
        "first_purchase_date": date(2026, 4, 1), "sion_norms": "SION-A",
        "company_name": "Plan 766", "has_purchase_bill": True,
        "exporter_name": "Golden Exporter", "opening_display": None,
        "display_transactions": [purchase, sale], "transactions": [purchase, sale],
        "company_utilizations": {766: {"company_name": "Plan 766", "utilization_balance": Decimal("1000.25")}},
        "license_wise_companies": [{
            "company_id": 766, "company_name": "Plan 766",
            "purchase_total": Decimal("100000.25"), "sale_total": Decimal("30000.75"),
            "current_balance": Decimal("1000.25"), "profit_loss": Decimal("-69999.50"),
            "profit_state": "LOSS",
        }],
        "summary": {
            "total_purchase": Decimal("1250.50"), "total_sale": Decimal("250.25"),
            "total_purchase_bill_inr": Decimal("100000.25"), "total_sale_bill_inr": Decimal("30000.75"),
            "current_balance": Decimal("1000.25"), "total_profit_loss": Decimal("-69999.50"),
            "balance_currency": "USD", "bill_currency": "INR", "profit_currency": "INR",
        },
    }


def test_pdf_and_excel_render_the_exact_canonical_golden_values():
    canonical = canonical_golden_data()  # same object serialized to the frontend API
    canonical_data = {
        "scope": "detail",
        "licenses": [canonical],
        "summary": CanonicalLedgerService.build_collection_summary([canonical]),
        "company_groups": CanonicalLedgerService.build_collection_company_groups([canonical]),
    }

    pdf_reader = PdfReader(render_financial_ledger_pdf(canonical_data))
    pdf_text = "\n".join(page.extract_text() or "" for page in pdf_reader.pages)
    workbook = load_workbook(render_financial_ledger_excel(canonical_data), data_only=True)
    values = [cell.value for sheet in workbook for row in sheet.iter_rows() for cell in row]

    for value in ("1,250.50", "250.25", "1,000.25", "1,00,000.25", "30,000.75", "-69,999.50"):
        assert value in pdf_text
    for value in (
        canonical["summary"]["total_purchase"], canonical["summary"]["total_sale"],
        canonical["summary"]["current_balance"], canonical["summary"]["total_profit_loss"],
        canonical["company_utilizations"][766]["utilization_balance"],
    ):
        assert value in values
    assert canonical["display_transactions"][0]["purchase_bill_amount"] in values
    assert canonical["display_transactions"][1]["sale_bill_amount"] in values
    assert "PUR-001" in pdf_text and "SIGNED" in pdf_text
    assert "SAL-001" in pdf_text and "UNSIGNED" in pdf_text
    pdf_links = {
        annotation.get_object()["/A"]["/URI"]
        for page in pdf_reader.pages
        for annotation in (page.get("/Annots") or [])
        if annotation.get_object().get("/A")
    }
    assert pdf_links == {
        "/api/invoice-documents/view/purchase-token/",
        "/api/invoice-documents/view/sale-token/",
    }


def test_excel_list_and_detail_reconcile_verbatim_with_canonical_dto():
    """Workbook cells are a projection of DTO values, never exporter arithmetic."""
    canonical = canonical_golden_data()
    workbook = load_workbook(render_financial_ledger_excel({
        "scope": "detail", "licenses": [canonical], "summary": {},
        "company_groups": CanonicalLedgerService.build_collection_company_groups([canonical]),
    }), data_only=True)

    list_sheet = workbook["License Summary"]
    assert tuple(cell.value for cell in list_sheet[3]) == (
        "Company", "SION Norm", "License Number", "Type", "Date", "1st Purchase Date",
        "Balance ($)", "Purchase (₹)", "Sale (₹)", "P/L (₹)",
    )
    list_row = next(row for row in list_sheet.iter_rows(values_only=True) if row[2] == "2436")
    assert list_row == (
        "Plan 766", "SION-A", "2436", "DFIA", datetime(2026, 4, 1), datetime(2026, 4, 1),
        Decimal("1000.25"), Decimal("100000.25"), Decimal("30000.75"), Decimal("-69999.50"),
    )

    detail = workbook["Financial Trade Ledger"]
    assert any(tuple(cell.value for cell in row) == (
        "Company", "SION", "License Number", "Date", "Particulars", "Invoice Number", "Type", "Items",
        "Credit ($)", "Debit ($)", "Purchase (₹)", "Sale (₹)", "Balance ($)", "P/L (₹)",
    ) for row in detail.iter_rows())
    purchase_row = next(row for row in detail.iter_rows(values_only=True) if row[3] == datetime(2026, 4, 1))
    sale_row = next(row for row in detail.iter_rows(values_only=True) if row[3] == datetime(2026, 4, 2))
    assert purchase_row[1:12] == (
        "SION-A", "2436", datetime(2026, 4, 1), "Supplier", "PUR-001", "PURCHASE", "Item 766",
        Decimal("1250.50"), None, Decimal("100000.25"), None,
    )
    assert sale_row[1:12] == (
        "SION-A", "2436", datetime(2026, 4, 2), "Buyer", "SAL-001", "SALE", "Item 766",
        None, Decimal("250.25"), None, Decimal("30000.75"),
    )
    total_row = next(row for row in detail.iter_rows(values_only=True) if row[4] == "LICENSE TOTAL")
    assert total_row[10:14] == (
        Decimal("100000.25"), Decimal("30000.75"), Decimal("1000.25"), Decimal("-69999.50"),
    )

    purchase_invoice_cell = next(row[5] for row in detail.iter_rows() if row[3].value == datetime(2026, 4, 1))
    sale_invoice_cell = next(row[5] for row in detail.iter_rows() if row[3].value == datetime(2026, 4, 2))
    assert purchase_invoice_cell.hyperlink.target == "/api/invoice-documents/view/purchase-token/"
    assert purchase_invoice_cell.comment.text == "SIGNED"
    assert sale_invoice_cell.hyperlink.target == "/api/invoice-documents/view/sale-token/"
    assert sale_invoice_cell.comment.text == "UNSIGNED"

    assert list_sheet.freeze_panes == "A4"
    assert list_sheet.auto_filter.ref
    assert list_sheet.print_area
    assert list_sheet.print_title_rows == "$1:$3"
    all_cells = [cell for sheet in workbook for row in sheet.iter_rows() for cell in row]
    assert not any(isinstance(cell.value, str) and cell.value.startswith("=") for cell in all_cells)
    assert "N/A" not in {cell.value for cell in all_cells}


def test_excel_company_sion_hierarchy_handles_composite_and_empty_norm_groups():
    """Excel consumes canonical SION groups; it neither splits nor duplicates licences."""
    def row(number, norms, balance, purchase, sale):
        return {
            "license_number": number, "license_type": "DFIA",
            "license_date": date(2026, 1, 1), "first_purchase_date": date(2026, 1, 2),
            "sion_norms": norms, "current_balance": Decimal(balance),
            "purchase_bill_inr": Decimal(purchase), "sale_bill_inr": Decimal(sale),
            "profit_loss_inr": Decimal(sale) - Decimal(purchase), "has_purchase_bill": True,
        }

    composite = row("MULTI-1", "E1, E5", "70", "100", "140")
    empty = row("EMPTY-1", "", "20", "50", "30")
    company = {
        "company_name": "LABDHI MERCANTILE LLP", "licenses": [composite, empty],
        "sion_groups": [
            {
                "sion_norm": "E1, E5", "sion_label": "E1, E5", "licenses": [composite],
                "license_count": 1, "total_balance": Decimal("70"),
                "total_purchase_bill_inr": Decimal("100"), "total_sale_bill_inr": Decimal("140"),
                "total_profit_loss_inr": Decimal("40"),
            },
            {
                "sion_norm": "", "sion_label": "N/A / EMPTY", "licenses": [empty],
                "license_count": 1, "total_balance": Decimal("20"),
                "total_purchase_bill_inr": Decimal("50"), "total_sale_bill_inr": Decimal("30"),
                "total_profit_loss_inr": Decimal("-20"),
            },
        ],
        "total_balance": Decimal("90"), "total_purchase_bill_inr": Decimal("150"),
        "total_sale_bill_inr": Decimal("170"), "total_profit_loss_inr": Decimal("20"),
    }
    grand = {
        "license_count": 2, "total_balance": Decimal("90"),
        "total_purchase_bill_inr": Decimal("150"), "total_sale_bill_inr": Decimal("170"),
        "total_profit_loss_inr": Decimal("20"),
    }
    workbook = load_workbook(render_financial_ledger_excel({
        "scope": "list", "licenses": [], "summary": {},
        "company_groups": [company], "grand_total": grand,
    }), data_only=True)
    rows = list(workbook["License Summary"].iter_rows(values_only=True))
    first_column = [entry[0] for entry in rows]
    license_numbers = [entry[2] for entry in rows]

    assert license_numbers.count("MULTI-1") == 1
    assert license_numbers.count("EMPTY-1") == 1
    assert first_column.index("SION: E1, E5") < license_numbers.index("MULTI-1")
    assert first_column.index("SION: N/A / EMPTY") < license_numbers.index("EMPTY-1")
    assert next(entry for entry in rows if entry[0] == "SION Total — E1, E5")[6:10] == (70, 100, 140, 40)
    assert next(entry for entry in rows if entry[0] == "SION Total — N/A / EMPTY")[6:10] == (20, 50, 30, -20)
    assert next(entry for entry in rows if entry[0] == "GRAND TOTAL")[6:10] == (90, 150, 170, 20)


def test_renderers_contain_no_query_or_business_service_dependency():
    import apps.license.services.exporters.financial_ledger_excel_renderer as excel
    import apps.license.services.exporters.financial_ledger_pdf_renderer as pdf

    for module in (pdf, excel):
        source_names = set(module.__dict__)
        assert "CanonicalLedgerService" not in source_names
        assert "LicenseTrade" not in source_names


@pytest.mark.django_db
def test_excel_renderer_executes_zero_database_queries(django_assert_num_queries):
    """Excel is a pure projection of a materialized canonical DTO."""
    canonical = canonical_golden_data()
    dto = {
        "scope": "detail",
        "licenses": [canonical],
        "summary": CanonicalLedgerService.build_collection_summary([canonical]),
        "company_groups": CanonicalLedgerService.build_collection_company_groups([canonical]),
    }

    with django_assert_num_queries(0):
        output = render_financial_ledger_excel(dto)

    assert output.getbuffer().nbytes > 0


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("file_format", "content_type", "disposition"),
    [
        ("pdf", "application/pdf", "inline"),
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "attachment"),
    ],
)
def test_authenticated_export_endpoint_uses_one_route_for_both_formats(
    authenticated_client, test_license, file_format, content_type, disposition,
):
    response = authenticated_client.get(
        reverse("license:license-ledger-export"),
        {"file_format": file_format, "license_id": test_license.id, "license_type": "DFIA", "item_id": 766},
    )
    assert response.status_code == 200
    assert response["Content-Type"] == content_type
    assert response["Content-Disposition"].startswith(disposition)
    assert f"license-ledger-{test_license.id}-766.{file_format}" in response["Content-Disposition"]


@pytest.mark.django_db
def test_retained_collection_endpoints_work_without_filter_parameters(authenticated_client):
    summary = authenticated_client.get(reverse("license:license-ledger-summary"))
    license_wise = authenticated_client.get(reverse("license:license-ledger-license-wise"))

    assert summary.status_code == 200
    assert set(summary.json()) == {"dfia", "incentive"}
    assert license_wise.status_code == 200
    payload = license_wise.json()
    assert set(payload) == {"licenses", "company_groups", "grand_total"}
    assert isinstance(payload["company_groups"], list)
    assert "total_profit_loss_inr" in payload["grand_total"]
