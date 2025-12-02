"""
Expiring Licenses Report

Shows licenses expiring within the next month (today's date to today + 30 days)
with detailed item-level balance information for each license.
"""

from datetime import date, timedelta
from decimal import Decimal
from typing import Dict, List, Any

from django.db.models import Sum, Q, F, DecimalField, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny

from core.constants import DEC_0, DEC_000, GE, MI, IP, SM
from license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel


class ExpiringLicensesReportView(View):
    """
    Report showing licenses expiring in the next month with item balances.

    GET parameters:
        - days: Number of days from today (default: 30)
        - format: 'json' or 'excel' (default: json)
        - sion_norm: Filter by SION norm (optional)
    """

    def get(self, request, *args, **kwargs):
        days = int(request.GET.get('days', 30))
        output_format = request.GET.get('format', 'json').lower()
        sion_norm = request.GET.get('sion_norm', None)

        try:
            report_data = self.generate_report(days, sion_norm)
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)

        if output_format == 'excel':
            return self.export_to_excel(report_data, days)
        else:
            return JsonResponse(report_data, safe=False)

    def generate_report(self, days: int = 30, sion_norm: str = None) -> Dict[str, Any]:
        """
        Generate expiring licenses report.

        Args:
            days: Number of days from today to check expiry
            sion_norm: Optional SION norm filter

        Returns:
            Dictionary with report data
        """
        today = date.today()
        expiry_date = today + timedelta(days=days)

        # Get licenses expiring between today and expiry_date
        # Filter by specific purchase statuses: GE, MI (NP), IP, SM
        licenses_query = LicenseDetailsModel.objects.filter(
            license_expiry_date__gte=today,
            license_expiry_date__lte=expiry_date,
            is_active=True,
            purchase_status__in=[GE, MI, IP, SM]
        ).select_related('exporter', 'port').prefetch_related(
            'export_license__norm_class',
            'import_license__items',
            'import_license__hs_code'
        )

        # Filter by SION norm if provided
        if sion_norm:
            licenses_query = licenses_query.filter(
                export_license__norm_class__norm_class=sion_norm
            ).distinct()

        licenses = licenses_query.order_by('license_expiry_date', 'license_date')

        # Build report data
        licenses_data = []
        total_licenses = 0
        total_balance_cif = Decimal('0.00')
        total_items = 0

        for license_obj in licenses:
            # Skip licenses with balance less than 100
            balance = license_obj.get_balance_cif
            if balance < Decimal('100.00'):
                continue

            license_data = self._build_license_data(license_obj)
            licenses_data.append(license_data)
            total_licenses += 1
            total_balance_cif += Decimal(str(license_data['balance_cif']))
            total_items += len(license_data['items'])

        return {
            'report_period': {
                'from_date': today.isoformat(),
                'to_date': expiry_date.isoformat(),
                'days': days,
            },
            'summary': {
                'total_licenses': total_licenses,
                'total_items': total_items,
                'total_balance_cif': float(total_balance_cif),
            },
            'licenses': licenses_data,
        }

    def _build_license_data(self, license_obj: LicenseDetailsModel) -> Dict[str, Any]:
        """
        Build detailed data for a single license.

        Args:
            license_obj: LicenseDetailsModel instance

        Returns:
            Dictionary with license and item data
        """
        # Calculate days until expiry
        days_to_expiry = (license_obj.license_expiry_date - date.today()).days

        # Get SION norms
        sion_norms = list(
            license_obj.export_license.filter(
                norm_class__isnull=False
            ).values_list('norm_class__norm_class', flat=True).distinct()
        )

        # Get export items summary
        export_items = license_obj.export_license.aggregate(
            total_quantity=Coalesce(Sum('net_quantity'), Value(DEC_0), output_field=DecimalField()),
            total_cif_fc=Coalesce(Sum('cif_fc'), Value(DEC_0), output_field=DecimalField()),
            total_fob_fc=Coalesce(Sum('fob_fc'), Value(DEC_0), output_field=DecimalField()),
        )

        # Get import items with balances
        items_data = self._get_items_with_balances(license_obj)

        # Calculate totals
        total_quantity = sum(item['quantity'] for item in items_data)
        total_debited = sum(item['debited_quantity'] for item in items_data)
        total_allotted = sum(item['allotted_quantity'] for item in items_data)
        total_available = sum(item['available_quantity'] for item in items_data)

        return {
            'license_number': license_obj.license_number,
            'notification_number': license_obj.notification_number or '',
            'license_date': license_obj.license_date.isoformat() if license_obj.license_date else None,
            'license_expiry_date': license_obj.license_expiry_date.isoformat(),
            'days_to_expiry': days_to_expiry,
            'exporter': str(license_obj.exporter) if license_obj.exporter else '',
            'port': str(license_obj.port) if license_obj.port else '',
            'sion_norms': sion_norms,
            'condition_sheet': license_obj.condition_sheet or '',
            'export_summary': {
                'total_quantity': float(export_items['total_quantity']),
                'total_cif_fc': float(export_items['total_cif_fc']),
                'total_fob_fc': float(export_items['total_fob_fc']),
            },
            'balance_cif': float(license_obj.get_balance_cif),
            'import_summary': {
                'total_quantity': float(total_quantity),
                'debited_quantity': float(total_debited),
                'allotted_quantity': float(total_allotted),
                'available_quantity': float(total_available),
            },
            'items': items_data,
        }

    def _get_items_with_balances(self, license_obj: LicenseDetailsModel) -> List[Dict[str, Any]]:
        """
        Get all import items with their balance information.
        Merges items with the same FK item and aggregates quantities.

        Args:
            license_obj: LicenseDetailsModel instance

        Returns:
            List of item dictionaries with merged balance data
        """
        import_items = license_obj.import_license.all()

        # Group items by FK item (items field)
        items_map = {}

        for import_item in import_items:
            # Get all linked items (M2M) with their IDs
            linked_items = list(import_item.items.values('id', 'name'))

            # If no FK items, use description as fallback
            if not linked_items:
                key = f"no_item_{import_item.description}"
                item_name = import_item.description or ''
                item_id = None
            else:
                # Use first item as primary key (or combine all)
                item_id = linked_items[0]['id']
                item_name = ', '.join([item['name'] for item in linked_items])
                key = f"item_{item_id}"

            if key not in items_map:
                items_map[key] = {
                    'item_id': item_id,
                    'item_name': item_name,
                    'description': import_item.description or '',
                    'hs_code': import_item.hs_code.hs_code if import_item.hs_code else '',
                    'unit': import_item.unit,
                    'quantity': Decimal('0.000'),
                    'debited_quantity': Decimal('0.000'),
                    'allotted_quantity': Decimal('0.000'),
                    'available_quantity': Decimal('0.000'),
                    'cif_fc': Decimal('0.00'),
                    'available_value': Decimal('0.00'),
                    'serial_numbers': [],
                    'conditions': [],
                }

            # Aggregate quantities
            items_map[key]['quantity'] += import_item.quantity or DEC_000
            items_map[key]['debited_quantity'] += import_item.debited_quantity or DEC_000
            items_map[key]['allotted_quantity'] += import_item.allotted_quantity or DEC_000
            items_map[key]['available_quantity'] += import_item.available_quantity or DEC_000
            items_map[key]['cif_fc'] += import_item.cif_fc or DEC_0
            items_map[key]['available_value'] += import_item.available_value or DEC_0

            # Collect serial numbers
            items_map[key]['serial_numbers'].append(import_item.serial_number)

            # Collect conditions if available
            if import_item.comment and import_item.comment.strip():
                items_map[key]['conditions'].append(f"Sr.{import_item.serial_number}: {import_item.comment.strip()}")

        # Convert to list and format
        items_list = []
        for item_data in items_map.values():
            # Format serial numbers as "1, 10, 20"
            serial_numbers_str = ', '.join(map(str, sorted(item_data['serial_numbers'])))

            # Format conditions
            conditions_str = '\n'.join(item_data['conditions']) if item_data['conditions'] else ''

            items_list.append({
                'item_id': item_data['item_id'],
                'serial_numbers': serial_numbers_str,
                'item_name': item_data['item_name'],
                'description': item_data['description'],
                'hs_code': item_data['hs_code'],
                'unit': item_data['unit'],
                'quantity': float(item_data['quantity']),
                'debited_quantity': float(item_data['debited_quantity']),
                'allotted_quantity': float(item_data['allotted_quantity']),
                'available_quantity': float(item_data['available_quantity']),
                'cif_fc': float(item_data['cif_fc']),
                'available_value': float(item_data['available_value']),
                'conditions': conditions_str,
            })

        # Sort by the first serial number in the list (lowest serial number)
        items_list.sort(key=lambda x: int(x['serial_numbers'].split(',')[0].strip()) if x['serial_numbers'] else 0)

        return items_list

    def export_to_excel(self, report_data: Dict[str, Any], days: int) -> HttpResponse:
        """
        Export report to Excel format with separate sheets for each SION norm.

        Args:
            report_data: Report data dictionary
            days: Number of days for report period

        Returns:
            HttpResponse with Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Group licenses by SION norms
        licenses_by_norm = {}
        for license_data in report_data['licenses']:
            norms = license_data['sion_norms']
            if not norms:
                norm_key = 'No Norm'
            else:
                # Use first norm as primary grouping
                norm_key = norms[0]

            if norm_key not in licenses_by_norm:
                licenses_by_norm[norm_key] = []
            licenses_by_norm[norm_key].append(license_data)

        # If no licenses found, create a single sheet with message
        if not licenses_by_norm:
            worksheet = workbook.create_sheet(title="No Data")
            worksheet['A1'] = "No expiring licenses found for the specified criteria"
            worksheet['A1'].font = Font(bold=True)
            worksheet.column_dimensions['A'].width = 60

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="expiring_licenses_{days}_days.xlsx"'
            workbook.save(response)
            return response

        # Create a sheet for each norm
        for norm_name, licenses_list in sorted(licenses_by_norm.items()):
            # Create worksheet with norm name (sanitize for Excel sheet naming)
            sheet_name = norm_name[:31]  # Excel sheet name limit
            worksheet = workbook.create_sheet(title=sheet_name)

            current_row = 1

            # Title
            title = f"Licenses Expiring in Next {days} Days - {norm_name}"
            worksheet.merge_cells(f'A{current_row}:L{current_row}')
            title_cell = worksheet[f'A{current_row}']
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Report period
            period_text = f"Period: {report_data['report_period']['from_date']} to {report_data['report_period']['to_date']}"
            worksheet.merge_cells(f'A{current_row}:L{current_row}')
            period_cell = worksheet[f'A{current_row}']
            period_cell.value = period_text
            period_cell.alignment = Alignment(horizontal='center')
            current_row += 2

            # Process each license in this norm
            for license_data in licenses_list:
                # License header
                license_header = f"License: {license_data['license_number']} | Expiry: {license_data['license_expiry_date']} | Days Left: {license_data['days_to_expiry']}"
                worksheet.merge_cells(f'A{current_row}:L{current_row}')
                header_cell = worksheet[f'A{current_row}']
                header_cell.value = license_header
                header_cell.font = Font(bold=True, size=11)
                header_cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                current_row += 1

                # License details
                details = [
                    ['Notification Number:', license_data['notification_number'], 'License Date:', license_data['license_date']],
                    ['Exporter:', license_data['exporter'], 'Port:', license_data['port']],
                    ['SION Norms:', ', '.join(license_data['sion_norms']), 'Balance CIF:', f"${license_data['balance_cif']:.2f}"],
                ]
                for detail_row in details:
                    for col_num, value in enumerate(detail_row, 1):
                        worksheet.cell(row=current_row, column=col_num, value=value)
                    current_row += 1

                current_row += 1

                # Items table header
                item_headers = [
                    'Sr. No.', 'Item Name', 'HS Code', 'Unit',
                    'Quantity', 'Debited', 'Allotted', 'Available',
                    'CIF Value', 'Available Value'
                ]
                for col_num, header in enumerate(item_headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1

                # Items data
                for item in license_data['items']:
                    row_data = [
                        item['serial_numbers'],
                        item['item_name'],
                        item['hs_code'],
                        item['unit'],
                        item['quantity'],
                        item['debited_quantity'],
                        item['allotted_quantity'],
                        item['available_quantity'],
                        item['cif_fc'],
                        item['available_value'],
                    ]
                    for col_num, value in enumerate(row_data, 1):
                        worksheet.cell(row=current_row, column=col_num, value=value)
                    current_row += 1

                    # Add conditions as notes below item if available
                    if item.get('conditions') and item['conditions'].strip():
                        worksheet.merge_cells(f'B{current_row}:J{current_row}')
                        notes_cell = worksheet.cell(row=current_row, column=2)
                        notes_cell.value = f"Conditions: {item['conditions']}"
                        notes_cell.font = Font(italic=True, size=9)
                        notes_cell.fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                        notes_cell.alignment = Alignment(wrap_text=True, vertical='top')
                        current_row += 1

                # License summary
                summary = license_data['import_summary']
                current_row += 1
                summary_data = [
                    'TOTAL', '', '', '',
                    summary['total_quantity'],
                    summary['debited_quantity'],
                    summary['allotted_quantity'],
                    summary['available_quantity'],
                    '', ''
                ]
                for col_num, value in enumerate(summary_data, 1):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.font = Font(bold=True)
                current_row += 3

            # Norm summary (for this sheet)
            worksheet.merge_cells(f'A{current_row}:L{current_row}')
            norm_summary_cell = worksheet[f'A{current_row}']
            norm_summary_cell.value = f"SUMMARY FOR {norm_name}"
            norm_summary_cell.font = Font(bold=True, size=12)
            norm_summary_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            norm_summary_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Calculate norm-specific summary
            norm_total_licenses = len(licenses_list)
            norm_total_items = sum(len(lic['items']) for lic in licenses_list)
            norm_total_balance = sum(lic['balance_cif'] for lic in licenses_list)

            # Calculate total quantities across all items in this norm
            norm_total_quantity = Decimal('0.000')
            norm_total_debited = Decimal('0.000')
            norm_total_allotted = Decimal('0.000')
            norm_total_available = Decimal('0.000')
            norm_total_cif = Decimal('0.00')
            norm_total_available_value = Decimal('0.00')

            for lic in licenses_list:
                for item in lic['items']:
                    norm_total_quantity += Decimal(str(item['quantity']))
                    norm_total_debited += Decimal(str(item['debited_quantity']))
                    norm_total_allotted += Decimal(str(item['allotted_quantity']))
                    norm_total_available += Decimal(str(item['available_quantity']))
                    norm_total_cif += Decimal(str(item['cif_fc']))
                    norm_total_available_value += Decimal(str(item['available_value']))

            norm_summary_rows = [
                ['Total Licenses:', norm_total_licenses],
                ['Total Items:', norm_total_items],
                ['Total Balance CIF:', f"${norm_total_balance:.2f}"],
            ]
            for summary_row in norm_summary_rows:
                for col_num, value in enumerate(summary_row, 1):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.value = value
                    if col_num == 1:
                        cell.font = Font(bold=True)
                current_row += 1

            # Add item-wise summary table
            current_row += 1
            worksheet.merge_cells(f'A{current_row}:L{current_row}')
            items_summary_cell = worksheet[f'A{current_row}']
            items_summary_cell.value = "ITEM-WISE SUMMARY"
            items_summary_cell.font = Font(bold=True, size=11)
            items_summary_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            items_summary_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Build item-wise aggregation
            itemwise_summary = {}
            for lic in licenses_list:
                for item in lic['items']:
                    item_id = item.get('item_id')
                    item_name = item['item_name']

                    # Use item_id as key, fallback to item_name if no ID
                    key = f"item_{item_id}" if item_id else f"no_item_{item_name}"

                    if key not in itemwise_summary:
                        itemwise_summary[key] = {
                            'item_name': item_name,
                            'unit': item['unit'],
                            'quantity': Decimal('0.000'),
                            'debited_quantity': Decimal('0.000'),
                            'allotted_quantity': Decimal('0.000'),
                            'available_quantity': Decimal('0.000'),
                            'cif_fc': Decimal('0.00'),
                            'available_value': Decimal('0.00'),
                        }

                    itemwise_summary[key]['quantity'] += Decimal(str(item['quantity']))
                    itemwise_summary[key]['debited_quantity'] += Decimal(str(item['debited_quantity']))
                    itemwise_summary[key]['allotted_quantity'] += Decimal(str(item['allotted_quantity']))
                    itemwise_summary[key]['available_quantity'] += Decimal(str(item['available_quantity']))
                    itemwise_summary[key]['cif_fc'] += Decimal(str(item['cif_fc']))
                    itemwise_summary[key]['available_value'] += Decimal(str(item['available_value']))

            # Item-wise summary headers
            itemwise_headers = [
                'Item Name', 'Unit', 'Quantity', 'Debited', 'Allotted',
                'Available', 'CIF Value', 'Available Value'
            ]
            for col_num, header in enumerate(itemwise_headers, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Item-wise summary data rows
            for item_data in sorted(itemwise_summary.values(), key=lambda x: x['item_name']):
                row_data = [
                    item_data['item_name'],
                    item_data['unit'],
                    float(item_data['quantity']),
                    float(item_data['debited_quantity']),
                    float(item_data['allotted_quantity']),
                    float(item_data['available_quantity']),
                    float(item_data['cif_fc']),
                    float(item_data['available_value']),
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                current_row += 1

            # Grand total row
            current_row += 1
            grand_total_data = [
                'GRAND TOTAL', '',
                float(norm_total_quantity),
                float(norm_total_debited),
                float(norm_total_allotted),
                float(norm_total_available),
                float(norm_total_cif),
                float(norm_total_available_value)
            ]
            for col_num, value in enumerate(grand_total_data, 1):
                cell = worksheet.cell(row=current_row, column=col_num)
                cell.value = value
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
            current_row += 1

            # Auto-adjust column widths for this sheet
            from openpyxl.cell.cell import MergedCell

            for col_idx in range(1, 13):  # Columns A to L
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)

                for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        # Skip merged cells
                        if isinstance(cell, MergedCell):
                            continue
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="expiring_licenses_{days}_days.xlsx"'

        workbook.save(response)
        return response


class ExpiringLicensesViewSet(viewsets.ViewSet):
    """
    ViewSet for Expiring Licenses Report.

    Permissions: AllowAny - accessible to all users
    """
    permission_classes = [AllowAny]

    def list(self, request):
        """
        Get expiring licenses report.

        Query Parameters:
            days: Number of days from today (default: 30)
            sion_norm: Filter by SION norm (optional)

        Returns:
            Report with licenses expiring in specified period
        """
        days = int(request.query_params.get('days', 30))
        sion_norm = request.query_params.get('sion_norm', None)

        report_view = ExpiringLicensesReportView()

        try:
            report_data = report_view.generate_report(days, sion_norm)
            return Response(report_data)
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export expiring licenses report to Excel.

        Query Parameters:
            days: Number of days from today (default: 30)
            sion_norm: Filter by SION norm (optional)

        Returns:
            Excel file download
        """
        days = int(request.query_params.get('days', 30))
        sion_norm = request.query_params.get('sion_norm', None)

        report_view = ExpiringLicensesReportView()

        try:
            report_data = report_view.generate_report(days, sion_norm)
            return report_view.export_to_excel(report_data, days)
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Get summary of expiring licenses by period.

        Query Parameters:
            days: Number of days from today (default: 30)

        Returns:
            Summary statistics
        """
        days = int(request.query_params.get('days', 30))

        report_view = ExpiringLicensesReportView()
        report_data = report_view.generate_report(days, None)

        return Response({
            'report_period': report_data['report_period'],
            'summary': report_data['summary'],
        })
