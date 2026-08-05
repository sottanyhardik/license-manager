"""
License Trading Register & Profit Report — view layer.

Consolidates every DFIA `LicenseTrade`/`LicenseTradeLine` purchase/sale row
into Dashboard -> Norm -> License -> Transaction Register -> License
Summary -> License Item Summary -> Norm Summary -> Norm Item Summary ->
Grand Summary -> Grand Item Summary. See
`apps.license.services.license_trading_register_report.
build_license_trading_register_report` for the full business-rule
rationale and scope decisions.

GET parameters:
    - from_date, to_date: required, 'YYYY-MM-DD'
    - norm: 'All' (default) or a specific SION norm code (e.g. 'E1')
    - license_type: 'DFIA' (default); any other value returns an empty report
    - license_number: optional, case-insensitive contains search
    - exporter_id: optional
    - item_id: optional (ItemNameModel id)
    - customer_id: optional (CompanyModel id, matched against SALE rows)
    - supplier_id: optional (CompanyModel id, matched against PURCHASE rows)
    - format / _format: 'json' (default) / 'excel' / 'pdf'
"""
import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict

from django.http import HttpResponse, JsonResponse
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.views import APIView

from apps.accounts.permissions import ReportPermission
from apps.license.services.license_trading_register_report import build_license_trading_register_report
from apps.license.views.item_report import _ExcelPassthroughRenderer

logger = logging.getLogger(__name__)


class _PdfPassthroughRenderer(BaseRenderer):
    """
    Same trick as `_ExcelPassthroughRenderer` (`item_report.py`), for 'pdf':
    tells DRF content negotiation that ?format=pdf is an accepted format so
    it doesn't raise NotAcceptable/404 before `get()` ever runs. The view
    always returns a plain Django `HttpResponse` for PDF, so this renderer's
    `render()` is never actually invoked.
    """
    media_type = 'application/pdf'
    format = 'pdf'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


def _parse_int_param(request, name):
    """Returns (value, error_response). error_response is None on success."""
    raw = request.GET.get(name)
    if not raw:
        return None, None
    try:
        return int(raw), None
    except ValueError:
        return None, JsonResponse({'error': f'{name} must be an integer'}, status=400)


class LicenseTradingRegisterReportView(APIView):
    permission_classes = [ReportPermission]
    # Without this, DRF content negotiation raises NotAcceptable/404 for
    # ?format=excel / ?format=pdf before get() ever runs — same fix as
    # ItemPivotReportView / ItemReportView / LicensePurchaseProfitReportView.
    renderer_classes = [JSONRenderer, _ExcelPassthroughRenderer, _PdfPassthroughRenderer]

    def get(self, request, *args, **kwargs):
        output_format = (
            request.GET.get('_format') or request.GET.get('format', 'json')
        ).lower()

        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')
        if not from_date_str or not to_date_str:
            return JsonResponse(
                {'error': 'from_date and to_date parameters are required',
                 'example': '?from_date=2026-01-01&to_date=2026-01-31'},
                status=400,
            )
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse(
                {'error': 'from_date and to_date must be in YYYY-MM-DD format'},
                status=400,
            )

        norm = request.GET.get('norm', 'All')
        license_type = request.GET.get('license_type', 'DFIA')
        license_number = request.GET.get('license_number') or None

        exporter_id, err = _parse_int_param(request, 'exporter_id')
        if err:
            return err
        item_id, err = _parse_int_param(request, 'item_id')
        if err:
            return err
        customer_id, err = _parse_int_param(request, 'customer_id')
        if err:
            return err
        supplier_id, err = _parse_int_param(request, 'supplier_id')
        if err:
            return err

        try:
            report_data = build_license_trading_register_report(
                from_date, to_date,
                norm=norm,
                license_type=license_type,
                license_number=license_number,
                exporter_id=exporter_id,
                item_id=item_id,
                customer_id=customer_id,
                supplier_id=supplier_id,
            )
        except Exception as e:
            logger.exception("Error generating License Trading Register & Profit Report")
            return JsonResponse({'error': str(e)}, status=500)

        if output_format == 'excel':
            try:
                return self.export_to_excel(report_data, from_date, to_date)
            except Exception as e:
                logger.exception("Error exporting License Trading Register & Profit Report to Excel")
                return JsonResponse({'error': str(e)}, status=500)

        if output_format == 'pdf':
            try:
                return self.export_to_pdf(report_data, from_date, to_date)
            except Exception as e:
                logger.exception("Error exporting License Trading Register & Profit Report to PDF")
                return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(report_data, safe=False)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def export_to_excel(self, report_data: Dict[str, Any], from_date, to_date) -> HttpResponse:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Alignment, Font, PatternFill

        MAX_COLS = 9  # widest table is the Transaction Register (9 columns)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading Register"
        # Collapse control sits on the License Summary row (ABOVE its
        # Transaction Register detail rows) — same convention as
        # license_balance_excel.py's Financial Ledger sheet
        # (row_dimensions[r].outline_level = 1 on the child/detail rows).
        ws.sheet_properties.outlinePr.summaryBelow = False

        TITLE_FONT = Font(bold=True, size=14)
        BANNER_FONT = Font(bold=True, size=12, color='FFFFFF')
        BANNER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        SUB_BANNER_FONT = Font(bold=True, size=11, color='1A1A1A')
        SUB_BANNER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        HEADER_FILL = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        SUMMARY_FONT = Font(bold=True)
        SUMMARY_FILL = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')
        TOTAL_FONT = Font(bold=True, color='FFFFFF')
        TOTAL_FILL = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
        CENTER = Alignment(horizontal='center')
        MONEY_FMT = '#,##0.00'
        QTY_FMT = '#,##0.000'

        current_row = 1

        def merge_banner(text, font, fill, height=None):
            nonlocal current_row
            cell = ws.cell(row=current_row, column=1, value=text)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=MAX_COLS)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if height:
                ws.row_dimensions[current_row].height = height
            current_row += 1

        def write_header_row(headers):
            nonlocal current_row
            header_row_num = current_row
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=current_row, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
            current_row += 1
            return header_row_num

        def write_data_row(values, *, outline=0, bold=False, fill=None, money_cols=(), qty_cols=()):
            nonlocal current_row
            for col, val in enumerate(values, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                if bold:
                    c.font = SUMMARY_FONT
                if fill:
                    c.fill = fill
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if col in qty_cols:
                        c.number_format = QTY_FMT
                        c.alignment = Alignment(horizontal='right')
                    elif col in money_cols:
                        c.number_format = MONEY_FMT
                        c.alignment = Alignment(horizontal='right')
            if outline:
                ws.row_dimensions[current_row].outline_level = outline
            current_row += 1

        # --- Title ------------------------------------------------------
        merge_banner("License Trading Register & Profit Report", TITLE_FONT, PatternFill())
        merge_banner(f"Period: {from_date.isoformat()} to {to_date.isoformat()}", Font(italic=True), PatternFill())
        current_row += 1

        # --- Dashboard ----------------------------------------------------
        merge_banner("Dashboard", BANNER_FONT, BANNER_FILL)
        dash = report_data['dashboard']
        write_header_row(['Total Licenses', 'Open', 'Closed', 'Total Purchase', 'Total Sale', 'Total Profit', 'Overall Margin %'])
        # Freeze right below the Dashboard header row so it stays visible
        # while scrolling through the Norm/License sections below.
        dashboard_header_row = current_row - 1
        write_data_row(
            [dash['total_licenses'], dash['open_licenses'], dash['closed_licenses'],
             dash['total_purchase'], dash['total_sale'], dash['total_profit'], dash['overall_margin_pct']],
            money_cols=(4, 5, 6),
        )
        ws.freeze_panes = f'A{dashboard_header_row + 1}'
        current_row += 1

        # --- Per-Norm sections --------------------------------------------
        for norm_block in report_data['norms']:
            merge_banner(f"Norm: {norm_block['norm']}", BANNER_FONT, BANNER_FILL)

            for lic in norm_block['licenses']:
                merge_banner(
                    f"License: {lic['license_number']}  |  Exporter: {lic['exporter']}",
                    SUB_BANNER_FONT, SUB_BANNER_FILL,
                )

                summary = lic['summary']
                write_header_row(['Purchase', 'Sale', 'Profit', 'Margin %', 'Status'])
                # License Summary row is the OUTLINE PARENT (level 0) for
                # the Transaction Register rows written immediately below —
                # collapsing it in Excel hides the whole register.
                write_data_row(
                    [summary['purchase'], summary['sale'], summary['profit'], summary['margin_pct'], summary['status']],
                    bold=True, fill=SUMMARY_FILL, money_cols=(1, 2, 3),
                )

                write_header_row(['Date', 'Direction', 'Invoice Number', 'From Company', 'To Company', 'Item', 'Purchase', 'Sale', 'Running Profit'])
                for txn in lic['transactions']:
                    write_data_row(
                        [txn['date'], txn['direction'], txn['invoice_number'], txn['from_company'],
                         txn['to_company'], txn['item'], txn['purchase'], txn['sale'], txn['running_profit']],
                        outline=1, money_cols=(7, 8, 9),
                    )

                write_header_row(['Item', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %'])
                for it in lic['item_summary']:
                    write_data_row(
                        [it['item'], it['purchase_qty'], it['sale_qty'], it['purchase_value'], it['sale_value'], it['profit'], it['margin_pct']],
                        money_cols=(4, 5, 6), qty_cols=(2, 3),
                    )
                current_row += 1

            n_summary = norm_block['summary']
            merge_banner(f"Norm Summary: {norm_block['norm']}", SUB_BANNER_FONT, SUB_BANNER_FILL)
            write_header_row(['Licenses', 'Purchase', 'Sale', 'Profit', 'Margin %'])
            write_data_row(
                [n_summary['licenses_count'], n_summary['purchase'], n_summary['sale'], n_summary['profit'], n_summary['margin_pct']],
                bold=True, fill=SUMMARY_FILL, money_cols=(2, 3, 4),
            )
            current_row += 1

            write_header_row(['Item', 'Licenses', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %'])
            for it in norm_block['item_summary']:
                write_data_row(
                    [it['item'], it['licenses_count'], it['purchase_qty'], it['sale_qty'],
                     it['purchase_value'], it['sale_value'], it['profit'], it['margin_pct']],
                    money_cols=(5, 6, 7), qty_cols=(3, 4),
                )
            current_row += 2

        # --- Grand Summary --------------------------------------------------
        merge_banner("Grand Summary", BANNER_FONT, BANNER_FILL)
        grand = report_data['grand_summary']
        write_header_row(['Licenses', 'Purchase', 'Sale', 'Profit', 'Margin %'])
        write_data_row(
            [grand['licenses_count'], grand['purchase'], grand['sale'], grand['profit'], grand['margin_pct']],
            bold=True, fill=TOTAL_FILL, money_cols=(2, 3, 4),
        )
        current_row += 2

        # --- Grand Item Summary ----------------------------------------------
        merge_banner("Grand Item Summary", BANNER_FONT, BANNER_FILL)
        write_header_row(['Norm', 'Item', 'Licenses', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %'])
        for row in report_data['grand_item_summary']:
            write_data_row(
                [row['norm'], row['item'], row['licenses_count'], row['purchase_qty'], row['sale_qty'],
                 row['purchase_value'], row['sale_value'], row['profit'], row['margin_pct']],
                money_cols=(6, 7, 8), qty_cols=(4, 5),
            )

        # --- Auto-fit column widths -----------------------------------------
        for col_idx in range(1, MAX_COLS + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="license_trading_register_report.xlsx"'
        wb.save(response)
        return response

    # ------------------------------------------------------------------
    # PDF export
    # ------------------------------------------------------------------
    def export_to_pdf(self, report_data: Dict[str, Any], from_date, to_date) -> HttpResponse:
        from reportlab.lib.units import inch
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import PageBreak, Paragraph, Spacer, Table, TableStyle

        from shared.pdf.builders import (
            append_generated_footer,
            format_indian_number,
            make_data_grid_commands,
            make_header_table_style_commands,
            make_landscape_doc,
            make_section_title_style,
            make_subtitle_style,
            make_title_style,
            pl_paragraph,
        )

        buffer = BytesIO()
        doc = make_landscape_doc(buffer)
        styles = getSampleStyleSheet()
        wrap_style = styles["Normal"]
        elements = []

        def simple_table(headers, rows, *, bold_last=False):
            data = [headers] + rows
            table = Table(data, repeatRows=1)
            commands = make_header_table_style_commands() + make_data_grid_commands()
            if bold_last:
                commands += [('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')]
            table.setStyle(TableStyle(commands))
            return table

        # --- Title + Dashboard -------------------------------------------
        elements.append(Paragraph("License Trading Register & Profit Report", make_title_style(styles)))
        elements.append(
            Paragraph(f"Period: {from_date.isoformat()} to {to_date.isoformat()}", make_subtitle_style(styles))
        )
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph("Dashboard", make_section_title_style(styles)))
        dash = report_data['dashboard']
        elements.append(simple_table(
            ['Total Licenses', 'Open', 'Closed', 'Total Purchase', 'Total Sale', 'Total Profit', 'Overall Margin %'],
            [[
                dash['total_licenses'], dash['open_licenses'], dash['closed_licenses'],
                format_indian_number(dash['total_purchase']), format_indian_number(dash['total_sale']),
                pl_paragraph(dash['total_profit'], wrap_style), f"{dash['overall_margin_pct']:.2f}",
            ]],
        ))

        # Page break per Norm (ledger_pdf.py::generate_detailed_licenses_pdf
        # convention: break before each new group, not before the first).
        for idx, norm_block in enumerate(report_data['norms']):
            elements.append(PageBreak())

            elements.append(Paragraph(f"Norm: {norm_block['norm']}", make_section_title_style(styles)))

            for lic in norm_block['licenses']:
                elements.append(
                    Paragraph(
                        f"License: {lic['license_number']} | Exporter: {lic['exporter']} | Status: {lic['summary']['status']}",
                        styles["Heading3"],
                    )
                )
                summary = lic['summary']
                elements.append(simple_table(
                    ['Purchase', 'Sale', 'Profit', 'Margin %', 'Status'],
                    [[
                        format_indian_number(summary['purchase']), format_indian_number(summary['sale']),
                        pl_paragraph(summary['profit'], wrap_style), f"{summary['margin_pct']:.2f}", summary['status'],
                    ]],
                ))
                elements.append(Spacer(1, 0.1 * inch))

                txn_header = ['Date', 'Direction', 'Invoice No.', 'From Company', 'To Company', 'Item', 'Purchase', 'Sale', 'Running Profit']
                txn_rows = [
                    [
                        txn['date'] or '-', txn['direction'], txn['invoice_number'] or '-',
                        txn['from_company'] or '-', txn['to_company'] or '-', txn['item'],
                        format_indian_number(txn['purchase']), format_indian_number(txn['sale']),
                        pl_paragraph(txn['running_profit'], wrap_style),
                    ]
                    for txn in lic['transactions']
                ]
                elements.append(simple_table(txn_header, txn_rows))
                elements.append(Spacer(1, 0.1 * inch))

                item_header = ['Item', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %']
                item_rows = [
                    [
                        it['item'], format_indian_number(it['purchase_qty'], 3), format_indian_number(it['sale_qty'], 3),
                        format_indian_number(it['purchase_value']), format_indian_number(it['sale_value']),
                        pl_paragraph(it['profit'], wrap_style), f"{it['margin_pct']:.2f}",
                    ]
                    for it in lic['item_summary']
                ]
                elements.append(simple_table(item_header, item_rows))
                elements.append(Spacer(1, 0.2 * inch))

            n_summary = norm_block['summary']
            elements.append(Paragraph(f"Norm Summary: {norm_block['norm']}", styles["Heading3"]))
            elements.append(simple_table(
                ['Licenses', 'Purchase', 'Sale', 'Profit', 'Margin %'],
                [[
                    n_summary['licenses_count'], format_indian_number(n_summary['purchase']),
                    format_indian_number(n_summary['sale']),
                    pl_paragraph(n_summary['profit'], wrap_style, bold=True), f"{n_summary['margin_pct']:.2f}",
                ]],
            ))
            elements.append(Spacer(1, 0.1 * inch))

            n_item_header = ['Item', 'Licenses', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %']
            n_item_rows = [
                [
                    it['item'], it['licenses_count'], format_indian_number(it['purchase_qty'], 3),
                    format_indian_number(it['sale_qty'], 3),
                    format_indian_number(it['purchase_value']), format_indian_number(it['sale_value']),
                    pl_paragraph(it['profit'], wrap_style), f"{it['margin_pct']:.2f}",
                ]
                for it in norm_block['item_summary']
            ]
            elements.append(simple_table(n_item_header, n_item_rows))

        # --- Grand Summary + Grand Item Summary: final page -----------------
        elements.append(PageBreak())
        elements.append(Paragraph("Grand Summary", make_section_title_style(styles)))
        grand = report_data['grand_summary']
        elements.append(simple_table(
            ['Licenses', 'Purchase', 'Sale', 'Profit', 'Margin %'],
            [[
                grand['licenses_count'], format_indian_number(grand['purchase']), format_indian_number(grand['sale']),
                pl_paragraph(grand['profit'], wrap_style, bold=True), f"{grand['margin_pct']:.2f}",
            ]],
        ))
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("Grand Item Summary", make_section_title_style(styles)))
        gitem_header = ['Norm', 'Item', 'Licenses', 'Purchase Qty', 'Sale Qty', 'Purchase Value', 'Sale Value', 'Profit', 'Margin %']
        gitem_rows = [
            [
                row['norm'], row['item'], row['licenses_count'],
                format_indian_number(row['purchase_qty'], 3), format_indian_number(row['sale_qty'], 3),
                format_indian_number(row['purchase_value']), format_indian_number(row['sale_value']),
                pl_paragraph(row['profit'], wrap_style), f"{row['margin_pct']:.2f}",
            ]
            for row in report_data['grand_item_summary']
        ]
        elements.append(simple_table(gitem_header, gitem_rows))

        append_generated_footer(elements, styles)

        # Page-numbering canvasmaker hook — `make_landscape_doc` has no
        # canvasmaker param today; this is the standard reportlab
        # onFirstPage/onLaterPages pattern, self-contained here so no
        # shared helper needs touching for other reports to keep working.
        def _draw_page_number(canvas, pdf_doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(pdf_doc.pagesize[0] - 30, 20, f"Page {pdf_doc.page}")
            canvas.restoreState()

        doc.build(elements, onFirstPage=_draw_page_number, onLaterPages=_draw_page_number)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="license_trading_register_report.pdf"'
        return response
