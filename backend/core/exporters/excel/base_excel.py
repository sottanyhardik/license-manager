"""
Base Excel exporter class with common Excel generation functionality.
"""

from dataclasses import dataclass
from io import BytesIO
from typing import Optional, Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..base import BaseExporter


@dataclass
class ExcelConfig:
    """Configuration for Excel generation."""
    sheet_name: str = "Sheet1"
    title: str = ""
    author: str = ""
    freeze_panes: Optional[str] = None  # e.g., "A2" to freeze first row


class BaseExcelExporter(BaseExporter):
    """
    Base class for Excel exporters using openpyxl.
    
    Provides common functionality for creating Excel workbooks.
    """

    def __init__(self, config: Optional[ExcelConfig] = None, **kwargs):
        """
        Initialize Excel exporter.
        
        Args:
            config: Excel configuration
            **kwargs: Additional configuration passed to BaseExporter
        """
        super().__init__(**kwargs)
        self.config = config or ExcelConfig()
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None

    def get_content_type(self) -> str:
        """Get Excel MIME type."""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def get_file_extension(self) -> str:
        """Get Excel file extension."""
        return 'xlsx'

    def _create_workbook(self) -> Workbook:
        """
        Create Excel workbook with configured settings.
        
        Returns:
            Workbook instance
        """
        workbook = openpyxl.Workbook()

        # Set document properties
        if self.config.title:
            workbook.properties.title = self.config.title
        if self.config.author:
            workbook.properties.creator = self.config.author

        # Get or create active worksheet
        worksheet = workbook.active
        worksheet.title = self.config.sheet_name

        self.worksheet = worksheet

        return workbook

    def add_sheet(self, sheet_name: str) -> Worksheet:
        """
        Add a new worksheet to workbook.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Worksheet instance
        """
        if not self.workbook:
            raise ValueError("Workbook not initialized")

        return self.workbook.create_sheet(title=sheet_name)

    def set_active_sheet(self, sheet_name: str) -> None:
        """
        Set active worksheet.
        
        Args:
            sheet_name: Name of the sheet
        """
        if not self.workbook:
            raise ValueError("Workbook not initialized")

        self.worksheet = self.workbook[sheet_name]

    def write_cell(self, row: int, col: int, value: Any) -> None:
        """
        Write value to cell.
        
        Args:
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            value: Value to write
        """
        if not self.worksheet:
            raise ValueError("Worksheet not initialized")

        self.worksheet.cell(row=row, column=col, value=value)

    def write_row(self, row: int, values: list, start_col: int = 1) -> None:
        """
        Write values to a row.
        
        Args:
            row: Row number (1-indexed)
            values: List of values
            start_col: Starting column (1-indexed)
        """
        for idx, value in enumerate(values):
            self.write_cell(row, start_col + idx, value)

    def write_rows(self, data: list, start_row: int = 1, start_col: int = 1) -> None:
        """
        Write multiple rows of data.
        
        Args:
            data: List of rows (each row is a list of values)
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
        """
        for idx, row_data in enumerate(data):
            self.write_row(start_row + idx, row_data, start_col)

    def apply_freeze_panes(self, cell: Optional[str] = None) -> None:
        """
        Apply freeze panes to worksheet.
        
        Args:
            cell: Cell reference (e.g., "A2" to freeze first row)
                  Uses config value if not specified
        """
        if not self.worksheet:
            raise ValueError("Worksheet not initialized")

        freeze_cell = cell or self.config.freeze_panes
        if freeze_cell:
            self.worksheet.freeze_panes = freeze_cell

    def auto_size_columns(self, min_width: int = 10, max_width: int = 50) -> None:
        """
        Auto-size columns based on content.
        
        Args:
            min_width: Minimum column width
            max_width: Maximum column width
        """
        if not self.worksheet:
            raise ValueError("Worksheet not initialized")

        for column in self.worksheet.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                if column_letter is None:
                    column_letter = cell.column_letter

                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            if column_letter:
                self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_workbook(self) -> None:
        """Save workbook to buffer."""
        if not self.workbook:
            raise ValueError("Workbook not initialized")

        self.workbook.save(self.buffer)

    def generate(self, data: Any) -> BytesIO:
        """
        Generate Excel document.
        
        This method should be overridden by subclasses to implement
        specific document generation logic.
        
        Args:
            data: Data to export
            
        Returns:
            BytesIO buffer with Excel file
        """
        self.reset_buffer()

        # Create workbook
        self.workbook = self._create_workbook()

        # Subclasses should override this to add content
        self._add_content(data)

        # Apply freeze panes if configured
        self.apply_freeze_panes()

        # Auto-size columns
        self.auto_size_columns()

        # Save to buffer
        self.save_workbook()

        return self.buffer

    def _add_content(self, data: Any) -> None:
        """
        Add content to workbook.
        
        This method should be overridden by subclasses.
        
        Args:
            data: Data to add to workbook
        """
        raise NotImplementedError("Subclasses must implement _add_content()")
