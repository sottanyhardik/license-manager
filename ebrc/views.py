from datetime import datetime
from io import BytesIO

import xlsxwriter
from django.http import HttpResponseRedirect
from django.shortcuts import render
# Create your views here.
from django.urls import reverse
from django.views import View
from django.views.generic import FormView, ListView
from django_tables2 import SingleTableView
from django_tables2.export import ExportMixin
from openpyxl import load_workbook

from ebrc import forms
from ebrc.tables import ShippingTable

try:
    from StringIO import StringIO
except ImportError:
    from io import StringIO

from django.http import HttpResponse

from ebrc.forms import EbrcCaptcha
from ebrc.models import EbrcDetails, FileUploadDetails, ShippingDetails
from ebrc.scripts.ebrc import get_cookies, get_captcha
from lmanagement.tasks import fetch_file


class EBRCMain(FormView):
    template_name = 'ebrc/forms.html'
    form_class = forms.EbrcForm

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            iec = form.cleaned_data.get('iec')
            ifsc = form.cleaned_data.get('ifsc')
            excel_file = request.FILES.get("file")
            now = datetime.now()
            if excel_file:
                file, bool = FileUploadDetails.objects.get_or_create(iec=iec, ifsc=ifsc, user=request.user,
                                                                     file_name=excel_file._name, upload_time=str(now))
                wb = load_workbook(excel_file)
                # getting a particular sheet by name out of many sheets
                worksheet = wb.active
                excel_data = list()
                # iterating over the rows and
                # getting value from each cell in row
                first = True
                list1 = []
                for index, row in enumerate(worksheet.iter_rows()):
                    if not first:
                        if row[0].value:
                            list1.append(
                                ShippingDetails(
                                    shipping_bill=str(row[0].value).replace('\xa0', ''),
                                    file=file
                                ))
                    else:
                        first = False
                try:
                    ShippingDetails.objects.bulk_create(list1)
                except:
                    for data in list1:
                        try:
                            data.save()
                        except:
                            pass
            else:
                file, bool = FileUploadDetails.objects.get_or_create(iec=iec, ifsc=ifsc, user=request.user,
                                                                     file_name=str(now), upload_time=str(now))
            return HttpResponseRedirect(reverse('ebrc_list'))
        return render(request, self.template_name, {'form': form})


class EBRCFileList(ListView):
    template_name = 'ebrc/list.html'
    model_name = FileUploadDetails
    context_object_name = 'files_list'

    def get_queryset(self):
        if self.request.user.is_superuser:
            return self.model_name.objects.all()
        else:
            return self.request.user.uploaded_files.all()


class Ebrcdump(View):
    def get(self, request, data):
        output = BytesIO()
        # Feed a buffer to workbook
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("users")
        ebrc_details = EbrcDetails.objects.filter(file=data)
        shipping_bill_nos = [ebrc_detail.shipping_bill_no for ebrc_detail in ebrc_details]
        bold = workbook.add_format({'bold': True})
        columns = ["Sr No.", "ebrcNumb", "Date of Realisation", "IEC", "BRC Date", "BRC Status", "Shipping Bill No",
                   "Shipping Bill Port", "Shipping Date", "Realised Value	Currency", "BRC Utilisation Status"]
        # Fill first row with columns
        not_found = ShippingDetails.objects.filter(file_id=data).exclude(shipping_bill__in=shipping_bill_nos)
        row = 0
        for i, elem in enumerate(columns):
            worksheet.write(row, i, elem, bold)
        row += 1
        # Now fill other rows with columns
        for index, ebrc_detail in enumerate(ebrc_details):
            worksheet.write(row, 0, index)
            worksheet.write(row, 1, ebrc_detail.ebrcNumb)
            worksheet.write(row, 2, ebrc_detail.date_of_realisation)
            worksheet.write(row, 3, ebrc_detail.iec)
            worksheet.write(row, 4, ebrc_detail.brc_date)
            worksheet.write(row, 5, ebrc_detail.brc_status)
            worksheet.write(row, 6, ebrc_detail.shipping_bill_no)
            worksheet.write(row, 7, ebrc_detail.shipping_port)
            worksheet.write(row, 8, ebrc_detail.shipping_date)
            worksheet.write(row, 9, ebrc_detail.realised_value)
            worksheet.write(row, 10, ebrc_detail.currency)
            worksheet.write(row, 11, ebrc_detail.brc_utilisation_status)
            row += 1
        for index, ebrc_detail in enumerate(not_found):
            worksheet.write(row, 0, ebrc_detail.sr_no)
            worksheet.write(row, 1, '')
            worksheet.write(row, 2, '')
            worksheet.write(row, 3, '')
            worksheet.write(row, 4, '')
            worksheet.write(row, 5, '')
            worksheet.write(row, 6, ebrc_detail.shipping_bill)
            worksheet.write(row, 7, ebrc_detail.shipping_port)
            worksheet.write(row, 8, ebrc_detail.shipping_date)
            worksheet.write(row, 9, '')
            worksheet.write(row, 10, '')
            worksheet.write(row, 11, '')
            row += 1
        # Close workbook for building file
        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return response


class EBRCFetch(FormView):
    template_name = 'ebrc/fetch.html'
    form_class = EbrcCaptcha

    def get_context_data(self, **kwargs):
        context = super(EBRCFetch, self).get_context_data(**kwargs)
        fetch_cookies = get_cookies()
        context['captcha_url'] = get_captcha(fetch_cookies)
        context['fetch_cookies'] = fetch_cookies
        return context

    def post(self, request, *args, **kwargs):
        captcha = self.request.POST.get('captcha')
        cookies = self.request.POST.get('cookies')
        data = self.kwargs.get('data')
        fetch_file.delay(data, cookies, captcha)
        return HttpResponseRedirect(reverse('ebrc_detail_list', kwargs={'data': data}))


class EBRCList(ExportMixin, SingleTableView):
    template_name = 'ebrc/detail_list.html'
    model_name = ShippingDetails
    context_object_name = 'ebrc_list'
    table_class = ShippingTable
    paginate_by = 500

    def get_queryset(self):
        return ShippingDetails.objects.filter(file_id=self.kwargs['data'])
