# license/views/expiring_licenses_report.py
"""
Expiring Licenses Report.

Ported exactly from:
  legacy/backend/apps/license/views/expiring_licenses_report.py

Shows licenses expiring within the next N days (default 30) with
detailed item-level balance data. Supports JSON and Excel output.

Field mapping (legacy → new model):
  import_item.quantity          → quantity
  import_item.debited_quantity  → debited_quantity
  import_item.allotted_quantity → allotted_quantity
  import_item.available_quantity → available_quantity
  import_item.cif_fc            → cif_fc
  import_item.available_value   → available_value
  import_item.items (M2M)       → items (M2M → ItemNameModel)
  import_item.hs_code           → hs_code (FK → HSCodeModel)
  license_obj.get_balance_cif   → license_obj.balance.balance_cif
  license_obj.condition_sheet   → license_obj.notes.condition_sheet

Purchase status codes used as filters (legacy GE, MI, IP, SM):
  GE = "GE"  (General Entry)
  MI = "MI"  (NP / Market Import)
  IP = "IP"  (In Progress)
  SM = "SM"  (Scheme Matched)
These are the legacy codes stored in PurchaseStatus.code — preserved exactly.
"""
from datetime import date, timedelta
from decimal import Decimal
from typing import Any, Dict, List

from django.db.models import DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.accounts.permissions import ReportPermission
from apps.license.models import LicenseDetailsModel

DEC_0 = Decimal("0")
DEC_000 = Decimal("0.000")

# Legacy purchase status codes preserved verbatim
_ACTIVE_STATUS_CODES = ["GE", "MI", "IP", "SM"]


def _get_balance_cif(license_obj) -> Decimal:
    """Safe balance_cif accessor — returns 0 if balance row missing."""
    try:
        return license_obj.balance.balance_cif or DEC_0
    except Exception:
        return DEC_0


def _get_condition_sheet(license_obj) -> str:
    """Safe notes.condition_sheet accessor."""
    try:
        return license_obj.notes.condition_sheet or ""
    except Exception:
        return ""


def _get_items_with_balances(license_obj: LicenseDetailsModel) -> List[Dict[str, Any]]:
    """
    Get all import items with their balance information.
    Groups items sharing the same ItemNameModel; aggregates quantities.
    Ported verbatim from legacy ExpiringLicensesReportView._get_items_with_balances.
    """
    import_items = license_obj.import_license.all()
    items_map: Dict[str, dict] = {}

    for import_item in import_items:
        linked_items = list(import_item.items.values("id", "name"))

        if not linked_items:
            key = f"no_item_{import_item.description}"
            item_name = import_item.description or ""
            item_id = None
        else:
            item_id = linked_items[0]["id"]
            item_name = ", ".join(i["name"] for i in linked_items)
            key = f"item_{item_id}"

        if key not in items_map:
            items_map[key] = {
                "item_id": item_id,
                "item_name": item_name,
                "description": import_item.description or "",
                "hs_code": import_item.hs_code.hs_code if import_item.hs_code else "",
                "unit": import_item.unit,
                "quantity": DEC_000,
                "debited_quantity": DEC_000,
                "allotted_quantity": DEC_000,
                "available_quantity": DEC_000,
                "cif_fc": DEC_0,
                "available_value": DEC_0,
                "serial_numbers": [],
                "conditions": [],
            }

        items_map[key]["quantity"] += import_item.quantity or DEC_000
        items_map[key]["debited_quantity"] += import_item.debited_quantity or DEC_000
        items_map[key]["allotted_quantity"] += import_item.allotted_quantity or DEC_000
        items_map[key]["available_quantity"] += import_item.available_quantity or DEC_000
        items_map[key]["cif_fc"] += import_item.cif_fc or DEC_0
        items_map[key]["available_value"] += import_item.available_value or DEC_0
        items_map[key]["serial_numbers"].append(import_item.serial_number)

        if import_item.comment and import_item.comment.strip():
            items_map[key]["conditions"].append(
                f"Sr.{import_item.serial_number}: {import_item.comment.strip()}"
            )

    items_list = []
    for item_data in items_map.values():
        serial_numbers_str = ", ".join(
            map(str, sorted(item_data["serial_numbers"]))
        )
        conditions_str = "\n".join(item_data["conditions"]) if item_data["conditions"] else ""

        items_list.append({
            "item_id": item_data["item_id"],
            "serial_numbers": serial_numbers_str,
            "item_name": item_data["item_name"],
            "description": item_data["description"],
            "hs_code": item_data["hs_code"],
            "unit": item_data["unit"],
            "quantity": float(item_data["quantity"]),
            "debited_quantity": float(item_data["debited_quantity"]),
            "allotted_quantity": float(item_data["allotted_quantity"]),
            "available_quantity": float(item_data["available_quantity"]),
            "cif_fc": float(item_data["cif_fc"]),
            "available_value": float(item_data["available_value"]),
            "conditions": conditions_str,
        })

    items_list.sort(
        key=lambda x: (
            int(x["serial_numbers"].split(",")[0].strip())
            if x["serial_numbers"]
            else 0
        )
    )
    return items_list


def _build_license_data(license_obj: LicenseDetailsModel) -> Dict[str, Any]:
    """Build full license dict for one expiring license."""
    days_to_expiry = (
        (license_obj.license_expiry_date - date.today()).days
        if license_obj.license_expiry_date
        else None
    )

    sion_norms = list(
        license_obj.export_license.filter(
            norm_class__isnull=False
        ).values_list("norm_class__norm_class", flat=True).distinct()
    )

    export_items = license_obj.export_license.aggregate(
        total_quantity=Coalesce(
            Sum("net_quantity"), Value(DEC_0), output_field=DecimalField()
        ),
        total_cif_fc=Coalesce(
            Sum("cif_fc"), Value(DEC_0), output_field=DecimalField()
        ),
        total_fob_fc=Coalesce(
            Sum("fob_fc"), Value(DEC_0), output_field=DecimalField()
        ),
    )

    items_data = _get_items_with_balances(license_obj)

    total_quantity = sum(item["quantity"] for item in items_data)
    total_debited = sum(item["debited_quantity"] for item in items_data)
    total_allotted = sum(item["allotted_quantity"] for item in items_data)
    total_available = sum(item["available_quantity"] for item in items_data)

    return {
        "license_number": license_obj.license_number,
        "notification_number": (
            license_obj.notification_number.code
            if license_obj.notification_number_id
            else ""
        ),
        "license_date": (
            license_obj.license_date.isoformat() if license_obj.license_date else None
        ),
        "license_expiry_date": (
            license_obj.license_expiry_date.isoformat()
            if license_obj.license_expiry_date
            else None
        ),
        "ledger_date": (
            license_obj.balance.ledger_date.isoformat()
            if (
                hasattr(license_obj, "balance")
                and license_obj.balance
                and license_obj.balance.ledger_date
            )
            else None
        ),
        "days_to_expiry": days_to_expiry,
        "exporter": str(license_obj.exporter) if license_obj.exporter else "",
        "port": str(license_obj.port) if license_obj.port else "",
        "sion_norms": sion_norms,
        "condition_sheet": _get_condition_sheet(license_obj),
        "export_summary": {
            "total_quantity": float(export_items["total_quantity"]),
            "total_cif_fc": float(export_items["total_cif_fc"]),
            "total_fob_fc": float(export_items["total_fob_fc"]),
        },
        "balance_cif": float(_get_balance_cif(license_obj)),
        "import_summary": {
            "total_quantity": float(total_quantity),
            "debited_quantity": float(total_debited),
            "allotted_quantity": float(total_allotted),
            "available_quantity": float(total_available),
        },
        "items": items_data,
    }


def _generate_expiring_report(days: int = 30, sion_norm: str = None) -> Dict[str, Any]:
    """Core report logic — returns dict suitable for JSON or Excel export."""
    today = date.today()
    expiry_date = today + timedelta(days=days)

    licenses_query = LicenseDetailsModel.objects.filter(
        license_expiry_date__gte=today,
        license_expiry_date__lte=expiry_date,
        flags__is_active=True,
        purchase_status__code__in=_ACTIVE_STATUS_CODES,
    ).select_related(
        "exporter", "port", "notification_number", "balance", "flags", "notes",
        "purchase_status",
    ).prefetch_related(
        "export_license__norm_class",
        "import_license__items",
        "import_license__hs_code",
    )

    if sion_norm:
        licenses_query = licenses_query.filter(
            export_license__norm_class__norm_class=sion_norm
        ).distinct()

    licenses = licenses_query.order_by("license_expiry_date", "license_date")

    licenses_data = []
    total_balance_cif = DEC_0
    total_items_count = 0

    for license_obj in licenses:
        balance = _get_balance_cif(license_obj)
        if balance < Decimal("100.00"):
            continue
        license_data = _build_license_data(license_obj)
        licenses_data.append(license_data)
        total_balance_cif += Decimal(str(license_data["balance_cif"]))
        total_items_count += len(license_data["items"])

    return {
        "report_period": {
            "from_date": today.isoformat(),
            "to_date": expiry_date.isoformat(),
            "days": days,
        },
        "summary": {
            "total_licenses": len(licenses_data),
            "total_items": total_items_count,
            "total_balance_cif": float(total_balance_cif),
        },
        "licenses": licenses_data,
    }


def _export_expiring_to_excel(report_data: Dict[str, Any], days: int) -> HttpResponse:
    """Export expiring licenses report to Excel. Ported from legacy verbatim."""
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    licenses_by_norm: Dict[str, list] = {}
    for license_data in report_data["licenses"]:
        norms = license_data["sion_norms"]
        norm_key = norms[0] if norms else "No Norm"
        licenses_by_norm.setdefault(norm_key, []).append(license_data)

    if not licenses_by_norm:
        ws = workbook.create_sheet(title="No Data")
        ws["A1"] = "No expiring licenses found for the specified criteria"
        ws["A1"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 60
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="expiring_licenses_{days}_days.xlsx"'
        )
        workbook.save(response)
        return response

    for norm_name, licenses_list in sorted(licenses_by_norm.items()):
        sheet_name = norm_name[:31]
        ws = workbook.create_sheet(title=sheet_name)
        cur = 1

        ws.merge_cells(f"A{cur}:L{cur}")
        cell = ws[f"A{cur}"]
        cell.value = f"Licenses Expiring in Next {days} Days - {norm_name}"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cur += 1

        ws.merge_cells(f"A{cur}:L{cur}")
        period_cell = ws[f"A{cur}"]
        period_cell.value = (
            f"Period: {report_data['report_period']['from_date']} "
            f"to {report_data['report_period']['to_date']}"
        )
        period_cell.alignment = Alignment(horizontal="center")
        cur += 2

        for ld in licenses_list:
            ws.merge_cells(f"A{cur}:L{cur}")
            hdr = ws[f"A{cur}"]
            hdr.value = (
                f"License: {ld['license_number']} | "
                f"Expiry: {ld['license_expiry_date']} | "
                f"Days Left: {ld['days_to_expiry']}"
            )
            hdr.font = Font(bold=True, size=11)
            hdr.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            cur += 1

            details = [
                ["Notification Number:", ld["notification_number"], "License Date:", ld["license_date"]],
                ["Exporter:", ld["exporter"], "Port:", ld["port"]],
                ["SION Norms:", ", ".join(ld["sion_norms"]), "Balance CIF:", f"${ld['balance_cif']:.2f}"],
                ["Ledger Date:", ld.get("ledger_date") or "-", "", ""],
            ]
            for detail_row in details:
                for col_num, value in enumerate(detail_row, 1):
                    ws.cell(row=cur, column=col_num, value=value)
                cur += 1
            cur += 1

            item_headers = [
                "Sr. No.", "Item Name", "HS Code", "Unit",
                "Quantity", "Debited", "Allotted", "Available",
                "CIF Value", "Available Value",
            ]
            for col_num, header in enumerate(item_headers, 1):
                c = ws.cell(row=cur, column=col_num)
                c.value = header
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                c.alignment = Alignment(horizontal="center")
            cur += 1

            for item in ld["items"]:
                for col_num, value in enumerate([
                    item["serial_numbers"], item["item_name"], item["hs_code"],
                    item["unit"], item["quantity"], item["debited_quantity"],
                    item["allotted_quantity"], item["available_quantity"],
                    item["cif_fc"], item["available_value"],
                ], 1):
                    ws.cell(row=cur, column=col_num, value=value)
                cur += 1

                if item.get("conditions") and item["conditions"].strip():
                    ws.merge_cells(f"B{cur}:J{cur}")
                    nc = ws.cell(row=cur, column=2)
                    nc.value = f"Conditions: {item['conditions']}"
                    nc.font = Font(italic=True, size=9)
                    nc.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                    nc.alignment = Alignment(wrap_text=True, vertical="top")
                    cur += 1

            summary = ld["import_summary"]
            cur += 1
            for col_num, value in enumerate([
                "TOTAL", "", "", "",
                summary["total_quantity"], summary["debited_quantity"],
                summary["allotted_quantity"], summary["available_quantity"],
                "", "",
            ], 1):
                c = ws.cell(row=cur, column=col_num)
                c.value = value
                c.font = Font(bold=True)
            cur += 3

        # Norm summary section
        ws.merge_cells(f"A{cur}:L{cur}")
        ns = ws[f"A{cur}"]
        ns.value = f"SUMMARY FOR {norm_name}"
        ns.font = Font(bold=True, size=12)
        ns.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ns.alignment = Alignment(horizontal="center")
        cur += 1

        norm_total_balance = sum(lic["balance_cif"] for lic in licenses_list)
        for label, value in [
            ("Total Licenses:", len(licenses_list)),
            ("Total Items:", sum(len(lic["items"]) for lic in licenses_list)),
            ("Total Balance CIF:", f"${norm_total_balance:.2f}"),
        ]:
            ws.cell(row=cur, column=1, value=label).font = Font(bold=True)
            ws.cell(row=cur, column=2, value=value)
            cur += 1

        # Item-wise summary
        cur += 1
        ws.merge_cells(f"A{cur}:L{cur}")
        iws = ws[f"A{cur}"]
        iws.value = "ITEM-WISE SUMMARY"
        iws.font = Font(bold=True, size=11)
        iws.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        iws.alignment = Alignment(horizontal="center")
        cur += 1

        itemwise: Dict[str, dict] = {}
        for lic in licenses_list:
            for item in lic["items"]:
                iid = item.get("item_id")
                iname = item["item_name"]
                ikey = f"item_{iid}" if iid else f"no_item_{iname}"
                if ikey not in itemwise:
                    itemwise[ikey] = {
                        "item_name": iname,
                        "unit": item["unit"],
                        "quantity": DEC_000,
                        "debited_quantity": DEC_000,
                        "allotted_quantity": DEC_000,
                        "available_quantity": DEC_000,
                        "cif_fc": DEC_0,
                        "available_value": DEC_0,
                    }
                for field in ("quantity", "debited_quantity", "allotted_quantity",
                              "available_quantity", "cif_fc", "available_value"):
                    itemwise[ikey][field] += Decimal(str(item[field]))

        iw_headers = [
            "Item Name", "Unit", "Quantity", "Debited", "Allotted",
            "Available", "CIF Value", "Available Value",
        ]
        for col_num, hdr in enumerate(iw_headers, 1):
            c = ws.cell(row=cur, column=col_num)
            c.value = hdr
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            c.alignment = Alignment(horizontal="center")
        cur += 1

        norm_totals = {k: DEC_0 for k in ("quantity", "debited_quantity",
                                           "allotted_quantity", "available_quantity",
                                           "cif_fc", "available_value")}
        for id_data in sorted(itemwise.values(), key=lambda x: x["item_name"]):
            row_data = [
                id_data["item_name"], id_data["unit"],
                float(id_data["quantity"]), float(id_data["debited_quantity"]),
                float(id_data["allotted_quantity"]), float(id_data["available_quantity"]),
                float(id_data["cif_fc"]), float(id_data["available_value"]),
            ]
            for col_num, val in enumerate(row_data, 1):
                c = ws.cell(row=cur, column=col_num)
                c.value = val
                c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cur += 1
            for k in norm_totals:
                norm_totals[k] += id_data[k]

        cur += 1
        grand_row = [
            "GRAND TOTAL", "",
            float(norm_totals["quantity"]), float(norm_totals["debited_quantity"]),
            float(norm_totals["allotted_quantity"]), float(norm_totals["available_quantity"]),
            float(norm_totals["cif_fc"]), float(norm_totals["available_value"]),
        ]
        for col_num, val in enumerate(grand_row, 1):
            c = ws.cell(row=cur, column=col_num)
            c.value = val
            c.font = Font(bold=True, size=11)
            c.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        cur += 1

        # Auto-adjust column widths
        for col_idx in range(1, 13):
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for c in row:
                    if isinstance(c, MergedCell):
                        continue
                    try:
                        if c.value:
                            max_length = max(max_length, len(str(c.value)))
                    except (TypeError, AttributeError):
                        pass
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="expiring_licenses_{days}_days.xlsx"'
    )
    workbook.save(response)
    return response


# ---------------------------------------------------------------------------
# ViewSet — registered as its own router entry (matches legacy pattern)
# ---------------------------------------------------------------------------

class ExpiringLicensesViewSet(viewsets.ViewSet):
    """
    ViewSet for Expiring Licenses Report.
    Route: /api/v1/licenses/expiring-licenses/
    """

    permission_classes = [ReportPermission]

    def list(self, request):
        """
        GET /api/v1/licenses/expiring-licenses/
        Query params: days (default 30), sion_norm, format (json|excel)
        """
        try:
            days = int(request.query_params.get("days", 30))
        except (TypeError, ValueError):
            days = 30
        sion_norm = request.query_params.get("sion_norm")
        output_format = request.query_params.get("format", "json").lower()

        try:
            report_data = _generate_expiring_report(days, sion_norm)
        except Exception as exc:
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        if output_format == "excel":
            return _export_expiring_to_excel(report_data, days)
        return Response(report_data)

    @action(detail=False, methods=["get"])
    def export(self, request):
        """GET /api/v1/licenses/expiring-licenses/export/ — Excel download."""
        try:
            days = int(request.query_params.get("days", 30))
        except (TypeError, ValueError):
            days = 30
        sion_norm = request.query_params.get("sion_norm")

        try:
            report_data = _generate_expiring_report(days, sion_norm)
            return _export_expiring_to_excel(report_data, days)
        except Exception as exc:
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """GET /api/v1/licenses/expiring-licenses/summary/ — summary stats only."""
        try:
            days = int(request.query_params.get("days", 30))
        except (TypeError, ValueError):
            days = 30

        report_data = _generate_expiring_report(days, None)
        return Response({
            "report_period": report_data["report_period"],
            "summary": report_data["summary"],
        })
