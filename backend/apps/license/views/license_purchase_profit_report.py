"""
License Purchase & Profit Report — view layer.

Acquisition-focused report: a single flat License Summary table — License
No. / License Date / Expiry Date / Exporter / Norm(s) / Purchase From /
Purchase Amount / Purchase $ / Sale Amount / Sale $ / Profit / Loss /
Trade Balance ($) — one row per qualifying license, not grouped by norm,
plus a `summary` grand-total block (rendered as a GRAND TOTAL row at the
bottom of the License Summary table/sheet in Excel/PDF); and a second,
pivot-style `item_matrix` block — the Dynamic Import Item Utilization
Matrix — with dynamic columns per Import Item name and SALE-direction
trade-ledger debit (qty/CIF $/Bill ₹) cell values. See
`apps.license.services.purchase_profit_report.build_purchase_profit_report`
for the full business-rule rationale (trade-ledger-sourced Purchase Amount/
Purchase $/Purchase From, license-selection rule, Norm(s) sourcing, why
Trade Balance ($) is its own figure rather than the centralized (license-wide)
Balance CIF engine, the separately-computed Sale Amount/
Sale $/Profit-Loss figures, and the `item_matrix` headers/debit sourcing).
Excel/PDF exports only format the `summary`/`licenses`/`item_matrix` the
builder already computed — never recompute.

GET parameters:
    - from_date, to_date: required, 'YYYY-MM-DD'
    - norm: 'All' (default) / 'E1' / 'E5' / 'E126' / 'E132' / 'Others'
      (filters which licenses qualify; the report is never grouped by norm)
    - license_number: optional, case-insensitive contains search
    - exclude_license_number: optional, comma-separated license numbers to
      exclude entirely — applied AFTER `license_number`/`norm`/`exporter_id`
      inclusion, so it always wins over an overlapping inclusion.
    - exporter_id: optional
    - format / _format: 'json' (default) / 'excel' / 'pdf'
"""
import functools
import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.http import HttpResponse, JsonResponse
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.views import APIView

from apps.accounts.permissions import ReportPermission
from apps.core.reports.export_naming import build_export_filename
from apps.license.services.purchase_profit_report import build_purchase_profit_report
from apps.license.views.item_report import _ExcelPassthroughRenderer

logger = logging.getLogger(__name__)


def _get_generated_by(request) -> str:
    """
    ``request.user.get_full_name() or request.user.username`` — guarded so
    a user object with no callable ``get_full_name`` (or one that raises
    or returns an empty string) never turns an export into a 500. Falls
    back to ``"Unknown"`` if even ``username`` is unavailable.
    """
    user = getattr(request, 'user', None)
    if user is None:
        return 'Unknown'
    get_full_name = getattr(user, 'get_full_name', None)
    full_name = None
    if callable(get_full_name):
        try:
            full_name = get_full_name()
        except Exception:
            full_name = None
    return full_name or getattr(user, 'username', None) or 'Unknown'


def _build_applied_filters(
    from_date, to_date, norm, license_number, exporter_id,
    exclude_license_number: Optional[List[str]],
) -> List[str]:
    """
    The "Filters:" line shared verbatim by both Excel and PDF exports —
    every filter is listed explicitly (From/To/Norm/Exporter/License/
    Exclude), showing the literal string ``"All"`` when a filter is at its
    default rather than omitting the line, e.g. ``"Norm : E1"`` /
    ``"Exporter : All"``.
    """
    return [
        f"From : {from_date.isoformat() if from_date else 'All'}",
        f"To : {to_date.isoformat() if to_date else 'All'}",
        f"Norm : {norm or 'All'}",
        f"Exporter : {exporter_id if exporter_id else 'All'}",
        f"License Number : {license_number or 'All'}",
        "Exclude License Number : " + (
            ', '.join(exclude_license_number) if exclude_license_number else 'All'
        ),
    ]


def _summary_metric_rows(summary: Dict[str, Any]) -> List[tuple]:
    """
    The 7 Metric/Value rows shared verbatim by both exports' summary
    table — sourced entirely from ``report_data['summary']``, never
    recomputed here.
    """
    return [
        ("Total Licenses", summary['total_licenses']),
        ("Purchase Amount", summary['purchase_amount']),
        ("Purchase $", summary['purchase_usd']),
        ("Sale Amount", summary['total_sale_amount']),
        ("Sale $", summary['total_sale_usd']),
        ("Profit / Loss", summary['total_profit_loss']),
        ("Trade Balance ($)", summary['trade_balance_usd']),
    ]


class _PdfPassthroughRenderer(BaseRenderer):
    """
    Dummy renderer that tells DRF 'pdf' is an accepted format so that
    ?format=pdf does not fail content negotiation — same pattern as
    `_ExcelPassthroughRenderer` (`item_report.py`). The view returns a
    plain Django HttpResponse for PDF, which DRF passes through without
    calling this renderer at all. Without this, `?format=pdf` was
    previously rejected by DRF's content negotiation before `get()` ever
    ran, despite this view's own docstring/comments claiming otherwise —
    the PDF export button was silently unreachable.
    """
    media_type = 'application/pdf'
    format = 'pdf'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data  # never reached — view returns HttpResponse directly


class LicensePurchaseProfitReportView(APIView):
    permission_classes = [ReportPermission]
    # Without this, DRF content negotiation raises NotAcceptable/404 for
    # ?format=excel / ?format=pdf before get() ever runs — same fix as
    # ItemPivotReportView / ItemReportView.
    renderer_classes = [JSONRenderer, _ExcelPassthroughRenderer, _PdfPassthroughRenderer]

    def get(self, request, *args, **kwargs):
        output_format = (
            request.GET.get('_format') or request.GET.get('format', 'json')
        ).lower()

        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')
        if not from_date_str or not to_date_str:
            return JsonResponse(
                {'error': 'from_date and to_date parameters are required',
                 'example': '?from_date=2026-01-01&to_date=2026-01-31'},
                status=400,
            )
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse(
                {'error': 'from_date and to_date must be in YYYY-MM-DD format'},
                status=400,
            )

        norm = request.GET.get('norm', 'All')
        license_number = request.GET.get('license_number') or None
        exclude_license_number_raw = request.GET.get('exclude_license_number') or ''
        exclude_license_number = [
            v.strip() for v in exclude_license_number_raw.split(',') if v.strip()
        ] or None
        exporter_id_raw = request.GET.get('exporter_id')
        exporter_id = None
        if exporter_id_raw:
            try:
                exporter_id = int(exporter_id_raw)
            except ValueError:
                return JsonResponse({'error': 'exporter_id must be an integer'}, status=400)

        try:
            report_data = build_purchase_profit_report(
                from_date, to_date, norm=norm, license_number=license_number, exporter_id=exporter_id,
                exclude_license_number=exclude_license_number,
            )
        except Exception as e:
            logger.exception("Error generating License Purchase & Profit Report")
            return JsonResponse({'error': str(e)}, status=500)

        if output_format == 'excel':
            try:
                return self.export_to_excel(
                    request, report_data, from_date, to_date, norm, license_number, exporter_id,
                    exclude_license_number,
                )
            except Exception as e:
                logger.exception("Error exporting License Purchase & Profit Report to Excel")
                return JsonResponse({'error': str(e)}, status=500)

        if output_format == 'pdf':
            try:
                return self.export_to_pdf(
                    request, report_data, from_date, to_date, norm, license_number, exporter_id,
                    exclude_license_number,
                )
            except Exception as e:
                logger.exception("Error exporting License Purchase & Profit Report to PDF")
                return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(report_data, safe=False)

    # ------------------------------------------------------------------
    # Excel export — "License Summary" + "Item Utilization Matrix" sheets
    # ------------------------------------------------------------------
    def export_to_excel(
        self, request, report_data: Dict[str, Any], from_date, to_date, norm, license_number, exporter_id,
        exclude_license_number=None,
    ) -> HttpResponse:
        """
        NOTE: this method's signature was extended beyond `(request,
        report_data)` to also take the filter parameters, so its "Filters:"
        header block (see `_build_applied_filters`) can show the same
        From/To/Norm/Exporter/License/Exclude values the PDF export already
        shows — every value is display-only, sourced from the already-
        validated `get()` params, never recomputed. `report_data` itself
        (summary/licenses/item_matrix) is still rendered completely
        verbatim, with zero recalculation.
        """
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "License Summary"

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        center = Alignment(horizontal='center')
        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        zebra_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        def _apply_grid(ws, *, header_rows, data_rows, text_cols, numeric_cols, max_col, zebra_rows=frozenset()):
            """
            Shared formatting pass for a table's header + data rows: a thin
            border on every cell, right-aligned `#,##0.00` on numeric data
            cells, left-aligned on text data cells, and light zebra-striping
            on `zebra_rows` (callers omit the GRAND TOTAL row from
            `zebra_rows` — it keeps its bold-only styling). Never touches
            `.value`, `.font`, or header `.fill` — those are set by callers.
            """
            for row in header_rows:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = thin_border
            for row in data_rows:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col in numeric_cols:
                        cell.number_format = '#,##0.00'
                        cell.alignment = right_align
                    elif col in text_cols:
                        cell.alignment = left_align
                    if row in zebra_rows:
                        cell.fill = zebra_fill

        generated_by = _get_generated_by(request)
        generated_on = datetime.now().strftime('%d-%b-%Y %I:%M %p')
        applied_filters = _build_applied_filters(
            from_date, to_date, norm, license_number, exporter_id, exclude_license_number,
        )

        headers = [
            'License No.', 'License Date', 'Expiry Date', 'Exporter', 'Norm(s)', 'Purchase From',
            'Purchase Amount', 'Purchase $', 'Sale Amount', 'Sale $', 'Profit / Loss', 'Trade Balance ($)',
        ]
        n_static = 6  # 'License No.' through 'Purchase From' — spanned by the GRAND TOTAL label below.
        max_cols = len(headers)
        current_row = 1

        title_cell = worksheet.cell(row=current_row, column=1, value="License Purchase & Profit Report — License Summary")
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center
        current_row += 1

        generated_cell = worksheet.cell(
            row=current_row, column=1,
            value=f"Generated On: {generated_on}    Generated By: {generated_by}",
        )
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
        generated_cell.font = Font(italic=True, size=9)
        generated_cell.alignment = center
        current_row += 1

        filters_cell = worksheet.cell(
            row=current_row, column=1, value="Filters: " + " | ".join(applied_filters),
        )
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
        filters_cell.font = Font(italic=True, size=9)
        filters_cell.alignment = center
        current_row += 2

        # Summary as a "Metric | Value" table — same 7 rows/values the PDF
        # export's own summary table below shows, read verbatim off
        # `report_data['summary']`.
        summary = report_data['summary']
        summary_header_row = current_row
        for col_num, label in enumerate(("Metric", "Value"), 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        current_row += 1
        summary_first_data_row = current_row
        for metric, value in _summary_metric_rows(summary):
            worksheet.cell(row=current_row, column=1, value=metric)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1
        summary_last_data_row = current_row - 1
        _apply_grid(
            worksheet,
            header_rows=[summary_header_row],
            data_rows=range(summary_first_data_row, summary_last_data_row + 1),
            text_cols={1}, numeric_cols={2}, max_col=2,
            zebra_rows={
                r for i, r in enumerate(range(summary_first_data_row, summary_last_data_row + 1)) if i % 2 == 1
            },
        )
        current_row += 1

        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        current_row += 1
        first_data_row = current_row

        for lic in report_data['licenses']:
            row_data = [
                lic['license_number'], lic['license_date'], lic['expiry_date'], lic['exporter'],
                ', '.join(lic['norms']), lic['purchase_from'],
                lic['purchase_amount'], lic['purchase_usd'],
                lic['sale_amount'], lic['sale_usd'], lic['profit_loss'],
                lic['trade_balance_usd'],
            ]
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=current_row, column=col_num, value=value)
            current_row += 1
        last_data_row = current_row - 1

        # GRAND TOTAL row — sourced from `report_data['summary']`, never a
        # re-sum of the rows above (same convention as the Item Utilization
        # Matrix's own GRAND TOTAL row further below).
        total_row = current_row
        total_label_cell = worksheet.cell(row=current_row, column=1, value='GRAND TOTAL')
        total_label_cell.font = Font(bold=True)
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_static)
        total_row_values = [
            summary['purchase_amount'], summary['purchase_usd'],
            summary['total_sale_amount'], summary['total_sale_usd'], summary['total_profit_loss'],
            summary['trade_balance_usd'],
        ]
        for offset, value in enumerate(total_row_values):
            cell = worksheet.cell(row=current_row, column=n_static + 1 + offset, value=value)
            cell.font = Font(bold=True)
        current_row += 1

        # Formatting pass (number_format/alignment/borders/zebra) covers the
        # header row through the GRAND TOTAL row inclusive — GRAND TOTAL
        # gets number_format/alignment/borders too, just no zebra fill.
        _apply_grid(
            worksheet,
            header_rows=[header_row],
            data_rows=range(first_data_row, total_row + 1),
            text_cols={1, 2, 3, 4, 5, 6}, numeric_cols={7, 8, 9, 10, 11, 12}, max_col=max_cols,
            zebra_rows={r for i, r in enumerate(range(first_data_row, last_data_row + 1)) if i % 2 == 1},
        )
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = (
            f"A{header_row}:{openpyxl.utils.get_column_letter(max_cols)}{header_row}"
        )

        for col_idx in range(1, max_cols + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        # --------------------------------------------------------------
        # Second sheet — Item Utilization Matrix: dynamic per-Import-Item
        # -name pivot with a 2-row merged header (static columns span both
        # header rows; each item name horizontally spans its 3 Qty/CIF $/
        # Bill ₹ sub-columns). Same merge_cells technique as
        # `license_balance_excel.py`'s Planning Matrix section (~line
        # 1739), reimplemented inline here per this view's own
        # Excel/PDF-stays-inline convention — not imported from there.
        # --------------------------------------------------------------
        item_matrix = report_data.get('item_matrix') or {"headers": [], "rows": [], "totals": {}}
        matrix_ws = workbook.create_sheet("Item Utilization Matrix")

        static_headers = [
            'License No.', 'License Date', 'Expiry Date', 'Exporter', 'Norm(s)', 'Purchase From',
            'Purchase Amount', 'Purchase $', 'Sale Amount', 'Sale $', 'Profit / Loss', 'Trade Balance ($)',
        ]
        item_headers = item_matrix['headers']
        n_static = len(static_headers)
        matrix_max_cols = n_static + 3 * len(item_headers)
        hdr_row1, hdr_row2 = 1, 2

        for col_idx, label in enumerate(static_headers, 1):
            matrix_ws.merge_cells(start_row=hdr_row1, start_column=col_idx, end_row=hdr_row2, end_column=col_idx)
            cell = matrix_ws.cell(row=hdr_row1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for i, item_name in enumerate(item_headers):
            base_col = n_static + 1 + i * 3
            matrix_ws.merge_cells(start_row=hdr_row1, start_column=base_col, end_row=hdr_row1, end_column=base_col + 2)
            name_cell = matrix_ws.cell(row=hdr_row1, column=base_col, value=item_name)
            name_cell.font = header_font
            name_cell.fill = header_fill
            name_cell.alignment = center
            for sub_idx, sub_label in enumerate(['Qty', 'CIF $', 'Bill ₹']):
                sub_cell = matrix_ws.cell(row=hdr_row2, column=base_col + sub_idx, value=sub_label)
                sub_cell.font = header_font
                sub_cell.fill = header_fill
                sub_cell.alignment = center

        matrix_first_data_row = hdr_row2 + 1
        data_row = matrix_first_data_row
        for row in item_matrix['rows']:
            row_values = [
                row['license_number'], row['license_date'], row['expiry_date'], row['exporter'],
                ', '.join(row['norms']), row['purchase_from'],
                row['purchase_amount'], row['purchase_usd'],
                row['sale_amount'], row['sale_usd'], row['profit_loss'], row['trade_balance_usd'],
            ]
            for col_idx, value in enumerate(row_values, 1):
                matrix_ws.cell(row=data_row, column=col_idx, value=value)
            for i, item_name in enumerate(item_headers):
                cell_data = row['items'].get(item_name, {"qty": 0, "cif": 0.0, "bill": 0.0})
                base_col = n_static + 1 + i * 3
                matrix_ws.cell(row=data_row, column=base_col, value=cell_data['qty'])
                matrix_ws.cell(row=data_row, column=base_col + 1, value=cell_data['cif'])
                matrix_ws.cell(row=data_row, column=base_col + 2, value=cell_data['bill'])
            data_row += 1
        matrix_last_data_row = data_row - 1

        # GRAND TOTAL row — static-column totals from `report_data['summary']`
        # (the SAME numbers the License Summary sheet's own GRAND TOTAL row
        # shows, above), per-item totals from `item_matrix['totals']`.
        # Neither is recomputed here — both are read verbatim off the DTO.
        matrix_totals = item_matrix['totals']
        n_label = 6  # 'License No.' through 'Purchase From' — spanned by the GRAND TOTAL label.
        total_label_cell = matrix_ws.cell(row=data_row, column=1, value='GRAND TOTAL')
        total_label_cell.font = Font(bold=True)
        matrix_ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=n_label)
        matrix_static_totals = [
            summary['purchase_amount'], summary['purchase_usd'],
            summary['total_sale_amount'], summary['total_sale_usd'], summary['total_profit_loss'],
            summary['trade_balance_usd'],
        ]
        for offset, value in enumerate(matrix_static_totals):
            static_total_cell = matrix_ws.cell(row=data_row, column=n_label + 1 + offset, value=value)
            static_total_cell.font = Font(bold=True)
        for i, item_name in enumerate(item_headers):
            base_col = n_static + 1 + i * 3
            item_totals = matrix_totals.get(item_name, {"qty": 0, "cif": 0.0, "bill": 0.0})
            for offset, key in enumerate(('qty', 'cif', 'bill')):
                total_cell = matrix_ws.cell(row=data_row, column=base_col + offset, value=item_totals[key])
                total_cell.font = Font(bold=True)
        matrix_total_row = data_row

        # Formatting pass — text cols are the 6 leading static columns
        # (License No. through Purchase From); everything from column 7
        # onward (the remaining static financial columns plus every
        # dynamic Qty/CIF $/Bill ₹ column) is numeric. Covers the GRAND
        # TOTAL row too (number_format/alignment/borders, no zebra fill).
        matrix_text_cols = {1, 2, 3, 4, 5, 6}
        matrix_numeric_cols = set(range(7, matrix_max_cols + 1))
        _apply_grid(
            matrix_ws,
            header_rows=[hdr_row1, hdr_row2],
            data_rows=range(matrix_first_data_row, matrix_total_row + 1),
            text_cols=matrix_text_cols, numeric_cols=matrix_numeric_cols, max_col=matrix_max_cols,
            zebra_rows={
                r for i, r in enumerate(range(matrix_first_data_row, matrix_last_data_row + 1)) if i % 2 == 1
            },
        )
        matrix_ws.freeze_panes = "A3"  # below the 2-row header (hdr_row1, hdr_row2).
        matrix_ws.auto_filter.ref = (
            f"A{hdr_row2}:{openpyxl.utils.get_column_letter(matrix_max_cols)}{hdr_row2}"
        )

        for col_idx in range(1, matrix_max_cols + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row_cells in matrix_ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row_cells:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
            matrix_ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = build_export_filename('purchase-profit-report', 'xlsx', from_date=from_date, to_date=to_date)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    # ------------------------------------------------------------------
    # PDF export — Header, Applied Filters, Summary, License Summary
    # ------------------------------------------------------------------
    def export_to_pdf(
        self, request, report_data: Dict[str, Any], from_date, to_date, norm, license_number, exporter_id,
        exclude_license_number=None,
    ) -> HttpResponse:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A2, A3
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

        from shared.pdf.builders import (
            draw_page_footer,
            format_indian_number,
            make_data_grid_commands,
            make_header_table_style_commands,
            make_landscape_doc,
            make_section_title_style,
            make_subtitle_style,
            make_title_style,
        )

        # Page-size decision needs the Item Utilization Matrix's column
        # count up front, before the doc/buffer are created below.
        item_matrix = report_data.get('item_matrix') or {"headers": [], "rows": [], "totals": {}}
        item_headers = item_matrix['headers']
        n_static = 12  # the 12 static License Summary columns, same count used by the Excel matrix sheet.
        n_matrix_cols = n_static + 3 * len(item_headers)

        buffer = BytesIO()
        # A3 landscape is preferred per spec. Escalate to A2 only when the
        # Item Utilization Matrix would exceed ~20 columns (12 static + 3
        # sub-columns per dynamic Import Item, i.e. more than ~2-3 dynamic
        # items) at this report's existing ~7pt font / 4pt cell padding —
        # a deliberate, coarse single-tier heuristic picked from A3's rough
        # practical column capacity at that font/padding, NOT a measured
        # guarantee: reportlab never auto-shrinks a `Table` to fit a page,
        # so beyond this it can still overflow — consistent with this
        # file's already-accepted "column-wise pagination for very many
        # items" limitation on the Item Utilization Matrix table itself.
        page_size = A2 if n_matrix_cols > 20 else A3
        doc = make_landscape_doc(buffer, pagesize=page_size)
        styles = getSampleStyleSheet()
        elements = []

        generated_by = _get_generated_by(request)
        generated_on = datetime.now().strftime('%d-%b-%Y %I:%M %p')

        elements.append(Paragraph("License Purchase & Profit Report", make_title_style(styles)))
        elements.append(
            Paragraph(f"Period: {from_date.isoformat()} to {to_date.isoformat()}", make_subtitle_style(styles))
        )
        elements.append(
            Paragraph(
                f"Generated On: {generated_on}    Generated By: {generated_by}", make_subtitle_style(styles),
            )
        )

        applied_filters = _build_applied_filters(
            from_date, to_date, norm, license_number, exporter_id, exclude_license_number,
        )
        elements.append(Paragraph("Filters: " + " | ".join(applied_filters), styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

        # Summary as a "Metric | Value" table — same 7 rows the Excel
        # export's own summary table above shows, read verbatim off
        # `report_data['summary']`.
        summary = report_data['summary']
        summary_table_data = [["Metric", "Value"]] + [
            [
                metric,
                str(value) if metric == "Total Licenses" else format_indian_number(value),
            ]
            for metric, value in _summary_metric_rows(summary)
        ]
        summary_table = Table(summary_table_data, colWidths=[2.5 * inch, 2.5 * inch])
        summary_table.setStyle(TableStyle(make_header_table_style_commands() + make_data_grid_commands()))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.2 * inch))

        elements.append(Paragraph("License Summary", make_section_title_style(styles)))
        lic_header = [
            'License No.', 'License Date', 'Expiry Date', 'Exporter', 'Norm(s)', 'Purchase From',
            'Purchase Amount', 'Purchase $', 'Sale Amount', 'Sale $', 'Profit / Loss', 'Trade Balance ($)',
        ]
        lic_n_static = 6  # 'License No.' through 'Purchase From' — spanned by the GRAND TOTAL label below.
        lic_data = [lic_header]
        for lic in report_data['licenses']:
            lic_data.append([
                lic['license_number'],
                lic['license_date'] or '-',
                lic['expiry_date'] or '-',
                lic['exporter'] or '-',
                ', '.join(lic['norms']) or '-',
                lic['purchase_from'] or '-',
                format_indian_number(lic['purchase_amount']),
                format_indian_number(lic['purchase_usd']),
                format_indian_number(lic['sale_amount']),
                format_indian_number(lic['sale_usd']),
                format_indian_number(lic['profit_loss']),
                format_indian_number(lic['trade_balance_usd']),
            ])

        # GRAND TOTAL row — sourced from `report_data['summary']`, never a
        # re-sum of the rows above (same convention as the Item Utilization
        # Matrix's own GRAND TOTAL row below).
        lic_data.append(
            ['GRAND TOTAL'] + [''] * (lic_n_static - 1) + [
                format_indian_number(summary['purchase_amount']),
                format_indian_number(summary['purchase_usd']),
                format_indian_number(summary['total_sale_amount']),
                format_indian_number(summary['total_sale_usd']),
                format_indian_number(summary['total_profit_loss']),
                format_indian_number(summary['trade_balance_usd']),
            ]
        )
        lic_total_row_idx = len(lic_data) - 1

        lic_table = Table(lic_data, repeatRows=1)
        lic_table.setStyle(TableStyle(
            make_header_table_style_commands() + make_data_grid_commands() + [
                ("FONTNAME", (0, lic_total_row_idx), (-1, lic_total_row_idx), "Helvetica-Bold"),
                ("SPAN", (0, lic_total_row_idx), (lic_n_static - 1, lic_total_row_idx)),
            ]
        ))
        elements.append(lic_table)

        # ------------------------------------------------------------------
        # Item Utilization Matrix — second table, 2-row grouped header built
        # via reportlab `SPAN` commands (no existing 2-row-grouped-header
        # PDF table elsewhere in this codebase to reuse — built inline
        # here): horizontal SPANs merge each item name across its 3 Qty/
        # CIF $/Bill ₹ sub-columns in header row 0; vertical SPANs merge
        # each static column's header across both header rows (0 and 1).
        # `repeatRows=2` repeats both header rows on page breaks.
        # Column-wise pagination for very many dynamic item columns is an
        # accepted limitation, not solved here.
        # ------------------------------------------------------------------
        item_matrix = report_data.get('item_matrix') or {"headers": [], "rows": [], "totals": {}}
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Item Utilization Matrix", make_section_title_style(styles)))

        matrix_static_headers = [
            'License No.', 'License Date', 'Expiry Date', 'Exporter', 'Norm(s)', 'Purchase From',
            'Purchase Amount', 'Purchase $', 'Sale Amount', 'Sale $', 'Profit / Loss', 'Trade Balance ($)',
        ]
        item_headers = item_matrix['headers']
        n_static = len(matrix_static_headers)

        header_row_0 = list(matrix_static_headers) + [name for name in item_headers for _ in range(3)]
        header_row_1 = [''] * n_static + ['Qty', 'CIF $', 'Bill ₹'] * len(item_headers)
        matrix_data = [header_row_0, header_row_1]

        for row in item_matrix['rows']:
            data_row = [
                row['license_number'],
                row['license_date'] or '-',
                row['expiry_date'] or '-',
                row['exporter'] or '-',
                ', '.join(row['norms']) or '-',
                row['purchase_from'] or '-',
                format_indian_number(row['purchase_amount']),
                format_indian_number(row['purchase_usd']),
                format_indian_number(row['sale_amount']),
                format_indian_number(row['sale_usd']),
                format_indian_number(row['profit_loss']),
                format_indian_number(row['trade_balance_usd']),
            ]
            for item_name in item_headers:
                cell_data = row['items'].get(item_name, {"qty": 0, "cif": 0.0, "bill": 0.0})
                data_row.extend([
                    format_indian_number(cell_data['qty']),
                    format_indian_number(cell_data['cif']),
                    format_indian_number(cell_data['bill']),
                ])
            matrix_data.append(data_row)

        matrix_totals = item_matrix['totals']
        matrix_n_label = 6  # 'License No.' through 'Purchase From' — spanned by the GRAND TOTAL label.
        total_row = ['GRAND TOTAL'] + [''] * (matrix_n_label - 1) + [
            format_indian_number(summary['purchase_amount']),
            format_indian_number(summary['purchase_usd']),
            format_indian_number(summary['total_sale_amount']),
            format_indian_number(summary['total_sale_usd']),
            format_indian_number(summary['total_profit_loss']),
            format_indian_number(summary['trade_balance_usd']),
        ]
        for item_name in item_headers:
            item_totals = matrix_totals.get(item_name, {"qty": 0, "cif": 0.0, "bill": 0.0})
            total_row.extend([
                format_indian_number(item_totals['qty']),
                format_indian_number(item_totals['cif']),
                format_indian_number(item_totals['bill']),
            ])
        matrix_data.append(total_row)
        total_row_idx = len(matrix_data) - 1

        matrix_style_commands = make_data_grid_commands() + [
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 8),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("VALIGN", (0, 0), (-1, 1), "MIDDLE"),
            # Vertical merge: each static column's header spans both header rows.
            *[("SPAN", (col, 0), (col, 1)) for col in range(n_static)],
            # Bolded GRAND TOTAL row, its label spanning only the leading
            # non-numeric columns — the numeric static columns (Purchase
            # Amount through Trade Balance ($)) show real totals, not blanks.
            ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
            ("SPAN", (0, total_row_idx), (matrix_n_label - 1, total_row_idx)),
        ]
        # Horizontal merge: each item name spans its 3 Qty/CIF $/Bill ₹ sub-columns.
        for i in range(len(item_headers)):
            col1 = n_static + i * 3
            col2 = col1 + 2
            matrix_style_commands.append(("SPAN", (col1, 0), (col2, 0)))

        matrix_table = Table(matrix_data, repeatRows=2)
        matrix_table.setStyle(TableStyle(matrix_style_commands))
        elements.append(matrix_table)

        # Canvas-callback footer (page numbers must be drawn per-page — a
        # flowable `Paragraph`, like `append_generated_footer` appends, is
        # only drawn once) — replaces `append_generated_footer` for THIS
        # report only; other PDF exports keep using it unchanged.
        footer_cb = functools.partial(draw_page_footer, generated_by=generated_by)
        doc.build(elements, onFirstPage=footer_cb, onLaterPages=footer_cb)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = build_export_filename('purchase-profit-report', 'pdf', from_date=from_date, to_date=to_date)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
