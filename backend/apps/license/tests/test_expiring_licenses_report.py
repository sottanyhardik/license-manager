"""
Tests for the Expiring Licenses Report (backend/apps/license/views/expiring_licenses_report.py).

Phase 2A regression coverage — mirrors test_active_licenses_report.py.
Includes the permanent regression test for a real bug found while writing
these tests: this view had no `renderer_classes` override, so DRF's content
negotiation raised `Http404` for `?format=excel` BEFORE `get()` ever ran
(see `rest_framework.negotiation.DefaultContentNegotiation.filter_renderers`
— it raises `Http404`, not `NotAcceptable`, when no renderer's `.format`
matches the query param). Fixed by registering
`apps.core.reports.renderers.ExcelPassthroughRenderer`, the same pattern
`ItemReportView`/`LicensePurchaseProfitReportView` already used.
`test_excel_export_returns_valid_workbook_matching_json` below is the
permanent regression test for that fix — it exercises the exact HTTP path
(`?format=excel`) the frontend's `LicenseExportPanel` calls.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.core.models import CompanyModel, HSCodeModel, HeadSIONNormsModel, ItemNameModel, PurchaseStatus, SionNormClassModel
from apps.license.models import LicenseDetailsModel, LicenseExportItemModel, LicenseImportItemsModel

User = get_user_model()

REPORT_URL = "/api/reports/expiring-licenses/"


@pytest.fixture
def report_viewer_client(db):
    user = User.objects.create_user(
        username="expiring-licenses-viewer",
        email="expiring-licenses-viewer@example.com",
        password="RoleP@ssw0rd123",
    )
    group, _ = Group.objects.get_or_create(name="REPORT_VIEWER")
    user.groups.add(group)
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


@pytest.fixture
def el_masters(db):
    head_norm = HeadSIONNormsModel.objects.create(name="Expiring Licenses Test Head Norm")
    return {
        "exporter": CompanyModel.objects.create(iec="4440002222", name="Expiring Licenses Exporter"),
        "hs_code": HSCodeModel.objects.create(hs_code="55559999", product_description="Expiring Licenses Test Product"),
        "item_a": ItemNameModel.objects.create(name="Expiring Licenses Item A"),
        "e1_norm": SionNormClassModel.objects.create(head_norm=head_norm, norm_class="E1", is_active=True),
        "ge_status": PurchaseStatus.objects.create(code="GE", label="Genuine"),
    }


def _make_license(number, exporter, purchase_status=None, *, expiry_days=15):
    return LicenseDetailsModel.objects.create(
        license_number=number,
        license_date=date.today() - timedelta(days=60),
        license_expiry_date=date.today() + timedelta(days=expiry_days),
        exporter=exporter,
        purchase_status=purchase_status,
    )


def _make_export_item(license_obj, norm_class, cif_fc=Decimal("50000.00")):
    return LicenseExportItemModel.objects.create(
        license=license_obj,
        description=f"Export item for {license_obj.license_number}",
        norm_class=norm_class,
        cif_fc=cif_fc,
        cif_inr=cif_fc * Decimal("84.5"),
    )


def _make_import_item(license_obj, hs_code, item_names, *, serial=1, quantity=Decimal("1000.000")):
    item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=serial,
        description=f"Import item {serial}",
        hs_code=hs_code,
        quantity=quantity,
        available_quantity=quantity,
    )
    item.items.set(item_names)
    return item


@pytest.mark.django_db
def test_json_response_matches_envelope_shape(report_viewer_client, el_masters):
    lic = _make_license("EXP-LIC-001", el_masters["exporter"], el_masters["ge_status"])
    _make_export_item(lic, el_masters["e1_norm"])
    _make_import_item(lic, el_masters["hs_code"], [el_masters["item_a"]])

    response = report_viewer_client.get(REPORT_URL, {"days": 30})
    assert response.status_code == 200
    data = response.json()

    assert set(data.keys()) == {"report_period", "summary", "licenses"}
    assert data["summary"]["total_licenses"] == 1
    assert len(data["licenses"]) == 1
    assert data["licenses"][0]["license_number"] == "EXP-LIC-001"


@pytest.mark.django_db
def test_excel_export_returns_valid_workbook_matching_json(report_viewer_client, el_masters):
    """Permanent regression test for the `?format=excel` Http404 bug (see
    module docstring) — a real HTTP request through DRF's dispatch/content-
    negotiation cycle, not a direct method call, so it actually exercises
    the code path that was broken."""
    lic = _make_license("EXP-LIC-010", el_masters["exporter"], el_masters["ge_status"])
    _make_export_item(lic, el_masters["e1_norm"])
    _make_import_item(lic, el_masters["hs_code"], [el_masters["item_a"]])

    json_response = report_viewer_client.get(REPORT_URL, {"days": 30})
    license_data = json_response.json()["licenses"][0]

    excel_response = report_viewer_client.get(REPORT_URL, {"days": 30, "format": "excel"})

    assert excel_response.status_code == 200
    assert excel_response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(excel_response.content))
    assert "E1" in workbook.sheetnames
    sheet = workbook["E1"]

    header_row = next(
        r for r in range(1, 20)
        if isinstance(sheet.cell(row=r, column=1).value, str)
        and sheet.cell(row=r, column=1).value.startswith("License: EXP-LIC-010")
    )
    # Layout: header row, then 4 detail rows (Notification/License Date,
    # Exporter/Port, SION Norms/Balance CIF, Ledger Date) — see
    # ExpiringLicensesReportView.export_to_excel's `details` list.
    balance_row = header_row + 3
    assert sheet.cell(row=balance_row, column=3).value == "Balance CIF:"
    excel_balance_cif = float(str(sheet.cell(row=balance_row, column=4).value).lstrip("$"))
    assert excel_balance_cif == license_data["balance_cif"]


@pytest.mark.django_db
def test_export_filename_follows_standard_convention(report_viewer_client, el_masters):
    lic = _make_license("EXP-LIC-020", el_masters["exporter"], el_masters["ge_status"])
    _make_export_item(lic, el_masters["e1_norm"])
    _make_import_item(lic, el_masters["hs_code"], [el_masters["item_a"]])

    response = report_viewer_client.get(REPORT_URL, {"days": 30, "format": "excel"})
    disposition = response["Content-Disposition"]
    assert disposition.startswith('attachment; filename="expiring-licenses_')
    assert disposition.endswith('.xlsx"')
