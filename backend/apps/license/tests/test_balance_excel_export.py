"""
Tests for the Licence Balance Workspace Excel export
(`apps/license/services/exporters/license_balance_excel.py`,
`build_balance_excel`), restructured into 5 worksheets — "Financial
Ledger", "Customs Ledger", "Timeline", "Reconciliation", "Audit Log" — all
rendered from `LicenseBalanceLedgerBuilder` so they can never independently
drift from the JSON API workspace or the PDF report.

Uses the same DB-backed fixture helpers as
`apps/license/tests/test_balance_ledger_views.py` /
`apps/reconciliation/tests/test_reconciliation.py`.
"""
import io
from decimal import Decimal

from django.test import TestCase

import openpyxl

from apps.license.models import LicenseExportItemModel
from apps.license.services.exporters.license_balance_excel import build_balance_excel
from apps.license.services.license_balance_ledger_builder import LicenseBalanceLedgerBuilder
from apps.reconciliation.models import ReconciliationLog
from apps.reconciliation.services.allocation_service import create_invoice_boe_allocation
from apps.reconciliation.tests.test_reconciliation import ReconciliationFixtureMixin

from .test_balance_ledger_views import LicenseBalanceLedgerFixtureMixin

# A minimal license (no BOEs, allotments, or trades) has NO Purchase/Sale
# activity, so the Financial Ledger sheet is omitted entirely — see
# `has_trading_activity` in `build_financial_ledger`'s docstring.
EXPECTED_SHEET_NAMES_NO_TRADING = ["Customs Ledger", "Timeline", "Reconciliation", "Audit Log"]


class BalanceExcelStructureTests(LicenseBalanceLedgerFixtureMixin, TestCase):
    """Baseline structural checks against a minimal license (no BOEs,
    allotments, or trades) — the empty-state path for Timeline/Audit Log."""

    def _load_workbook(self, license_obj):
        response = build_balance_excel(license_obj)
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def test_five_sheets_present_in_correct_order(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("10000.00"))

        wb = self._load_workbook(license_obj)

        self.assertEqual(wb.sheetnames, EXPECTED_SHEET_NAMES_NO_TRADING)

    def test_each_sheet_has_real_rows(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("10000.00"))

        wb = self._load_workbook(license_obj)

        # Financial Ledger sheet is entirely absent (no trading activity).
        self.assertNotIn("Financial Ledger", wb.sheetnames)
        # Customs Ledger: Customs Summary block + header + at least opening/final rows.
        self.assertGreaterEqual(wb["Customs Ledger"].max_row, 10)
        # Reconciliation: license info row + reconciliation block + BOE/Allotment summary + plan utilization.
        self.assertGreaterEqual(wb["Reconciliation"].max_row, 5)

    def test_timeline_empty_state_writes_single_explanatory_row(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("10000.00"))

        wb = self._load_workbook(license_obj)
        tl = wb["Timeline"]

        # Title row + one explanatory row, nothing else.
        self.assertEqual(tl.max_row, 2)
        self.assertEqual(tl["A2"].value, "No timeline events recorded for this licence yet.")

    def test_audit_log_empty_state_writes_explanatory_rows_for_both_sections(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("10000.00"))

        wb = self._load_workbook(license_obj)
        al = wb["Audit Log"]

        values = [al.cell(row=r, column=1).value for r in range(1, al.max_row + 1)]
        self.assertIn("No reconciliation actions recorded for this licence yet.", values)
        self.assertIn("No ignored/restored warnings recorded for this licence yet.", values)


class BalanceExcelFinancialLedgerHierarchyTests(
    LicenseBalanceLedgerFixtureMixin, ReconciliationFixtureMixin, TestCase
):
    """Children of a consolidated 'BOE Allocation' row must render on the
    Financial Ledger sheet with Excel-native outline grouping (collapsible
    in Excel itself), immediately below their parent."""

    def _build_license_with_two_fully_allocated_boes(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("222360.00"))

        item = self.make_item(license_obj, 1)
        boe1 = self.make_boe(company, number="7650222")
        boe2 = self.make_boe(company, number="7650224")
        row1 = self.make_debit_row(boe1, item, cif_fc=Decimal("87120.00"), qty=Decimal("99000.000"))
        row2 = self.make_debit_row(boe2, item, cif_fc=Decimal("87120.00"), qty=Decimal("99000.000"))

        trade = self.make_sale_trade(company, invoice_number="LML/2025-26/0125")
        trade_line = self.make_trade_line(trade, item, cif_fc=Decimal("174240.00"), qty_kg=Decimal("198000.0000"))

        create_invoice_boe_allocation(
            trade_line, row1, qty=row1.qty, cif_fc=row1.cif_fc, cif_inr=row1.cif_inr, user=None,
        )
        create_invoice_boe_allocation(
            trade_line, row2, qty=row2.qty, cif_fc=row2.cif_fc, cif_inr=row2.cif_inr, user=None,
        )
        return license_obj

    def test_fully_matched_invoice_produces_one_trade_row_with_children_in_financial_ledger_sheet(self):
        """
        `build_financial_ledger()`'s BOE rows are skipped once fully
        allocated (their `contributed` is 0) — the two BOEs in this fixture
        are both fully allocated to the SAME invoice, so neither BOE
        produces its own row. Their combined 174,240 debit is instead
        carried by the SALE trade line's own "Licence Trade (Sold)" row
        (see that method's docstring), with the two underlying allocations
        rendered as Excel-outlined child rows immediately below it.
        """
        license_obj = self._build_license_with_two_fully_allocated_boes()

        response = build_balance_excel(license_obj)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Financial Ledger"]

        doc_numbers = [ws.cell(row=r, column=4).value for r in range(3, ws.max_row + 1)]
        self.assertIn("LML/2025-26/0125", doc_numbers)

        # No SEPARATE row for either individual BOE (both fully allocated
        # into the one trade row above) — the trade row's own BOE Number
        # column is the joined display string of both.
        boe_numbers = [ws.cell(row=r, column=5).value for r in range(3, ws.max_row + 1)]
        self.assertIn("7650222, 7650224", boe_numbers)

        # summaryBelow=False -> the collapse control sits with the parent
        # (which is ABOVE its children here), not below the detail block.
        self.assertFalse(ws.sheet_properties.outlinePr.summaryBelow)

    def test_audit_log_includes_reconciliation_log_row(self):
        license_obj = self._build_license_with_two_fully_allocated_boes()

        # Two allocations were created above -> two ACTION_ALLOCATE
        # ReconciliationLog rows exist for this licence's import item.
        self.assertTrue(
            ReconciliationLog.objects.filter(
                license_item__license=license_obj, action=ReconciliationLog.ACTION_ALLOCATE,
            ).exists()
        )

        response = build_balance_excel(license_obj)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        al = wb["Audit Log"]

        actions = [al.cell(row=r, column=3).value for r in range(1, al.max_row + 1)]
        self.assertIn("Allocate", actions)


class BalanceExcelCustomsLedgerTests(LicenseBalanceLedgerFixtureMixin, TestCase):
    """The Customs Ledger sheet's summary block must carry exactly the
    numbers `build_customs_ledger()` computes — no independent calculation
    in the exporter."""

    def test_customs_summary_matches_builder(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("50000.00"))

        item = self.make_item(license_obj, 1)
        boe = self.make_boe(company, number="7650999")
        self.make_debit_row(boe, item, cif_fc=Decimal("20000.00"), qty=Decimal("500.000"))

        _, expected_summary = LicenseBalanceLedgerBuilder.build_customs_ledger(license_obj)

        response = build_balance_excel(license_obj)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Customs Ledger"]

        found = {}
        for r in range(1, 10):
            label = ws.cell(row=r, column=1).value
            if label in (
                "Original Licence CIF", "Total BOE CIF", "Pending Allotment CIF",
                "Available Balance", "Balance Engine",
            ):
                found[label] = ws.cell(row=r, column=2).value

        self.assertEqual(Decimal(str(found["Original Licence CIF"])), expected_summary["opening_balance"])
        self.assertEqual(Decimal(str(found["Total BOE CIF"])), expected_summary["total_boe_cif"])
        self.assertEqual(
            Decimal(str(found["Pending Allotment CIF"])), expected_summary["total_pending_allotment_cif"],
        )
        self.assertEqual(Decimal(str(found["Available Balance"])), expected_summary["computed_balance"])
        self.assertEqual(Decimal(str(found["Balance Engine"])), expected_summary["engine_balance"])


class BalanceExcelReconciliationSheetTests(LicenseBalanceLedgerFixtureMixin, TestCase):
    """The Reconciliation sheet's three-way comparison must carry exactly
    the numbers `build_reconciliation_summary()` computes."""

    def test_reconciliation_numbers_match_builder(self):
        company = self.make_company()
        license_obj = self.make_license(company)
        LicenseExportItemModel.objects.create(license=license_obj, cif_fc=Decimal("75000.00"))

        item = self.make_item(license_obj, 1)
        boe = self.make_boe(company, number="7651111")
        self.make_debit_row(boe, item, cif_fc=Decimal("15000.00"), qty=Decimal("300.000"))

        data = LicenseBalanceLedgerBuilder.build(license_obj)
        expected = data["reconciliation"]

        response = build_balance_excel(license_obj)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Reconciliation"]

        found = {}
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label in (
                "Financial Ledger Balance", "Customs Ledger Balance",
                "Licence Balance Engine", "Difference", "Tolerance",
            ):
                found[label] = ws.cell(row=r, column=2).value

        self.assertEqual(
            Decimal(str(found["Financial Ledger Balance"])), expected["financial_ledger_balance"],
        )
        self.assertEqual(
            Decimal(str(found["Customs Ledger Balance"])), expected["customs_ledger_balance"],
        )
        self.assertEqual(Decimal(str(found["Licence Balance Engine"])), expected["balance_engine"])
        self.assertEqual(Decimal(str(found["Difference"])), expected["difference"])
        self.assertEqual(Decimal(str(found["Tolerance"])), expected["tolerance"])
