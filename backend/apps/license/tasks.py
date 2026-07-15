# license/tasks.py
"""
Celery tasks for the License module.

Balance recomputation is always asynchronous — never called inline in a
request/response cycle.  PDF generation is a placeholder that will delegate
to a report-generation service in Phase 8.

Item pivot Excel generation and all-license balance update are async per
OQ-6 approval (never converted to synchronous).
"""
import logging
import os
from datetime import datetime

from celery import shared_task

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3, acks_late=True, reject_on_worker_lost=True)
def recompute_license_balance_task(self, license_id: int):
    """
    Recompute the CIF balance for *license_id*.

    Retries up to 3 times with a 60-second countdown on failure.
    """
    try:
        from apps.license.services.balance_service import recompute_license_balance

        recompute_license_balance(license_id)
    except Exception as exc:
        logger.exception(
            "Balance recompute failed for license %s (attempt %s/%s): %s",
            license_id,
            self.request.retries + 1,
            self.max_retries + 1,
            exc,
        )
        raise self.retry(exc=exc, countdown=60)


@shared_task(bind=True, max_retries=2)
def generate_license_pdf_task(self, license_id: int, report_type: str, user_id: int):
    """
    Placeholder task for async PDF generation.

    Phase 8 will implement actual PDF rendering; this task stores the
    submitted job in CeleryTaskTracker so the frontend can poll for completion.

    Returns: None (result stored externally in CeleryTaskTracker).
    """
    task_id = self.request.id
    logger.info(
        "PDF generation requested: license=%s report_type=%s user=%s task=%s",
        license_id,
        report_type,
        user_id,
        task_id,
    )
    try:
        from apps.core.models import CeleryTaskTracker

        CeleryTaskTracker.objects.filter(task_id=task_id).update(
            status=CeleryTaskTracker.STATUS_PENDING,
        )
    except Exception:
        # CeleryTaskTracker may not be populated yet — non-fatal
        pass


@shared_task(bind=True, acks_late=True, name="license.process_single_license")
def process_single_license_task(self, dict_data: dict) -> dict:
    """
    Process a single license dict from the ledger upload parser.

    Each license is dispatched as an independent task so the entire file is
    processed in parallel across available workers.  The dict_data payload has
    already been serialized to JSON-safe types by LedgerUploadView._serialize_for_celery().

    Returns:
        dict with status, license_number, and lic_no on success.
        Re-raises on failure so Celery marks the task as FAILURE.
    """
    from scripts.parse_ledger import create_object

    license_no = dict_data.get('lic_no', 'Unknown')
    logger.info("Processing single license: %s (task=%s)", license_no, self.request.id)

    try:
        license_number = create_object(dict_data)
        logger.info("Successfully processed license: %s", license_number)
        return {
            'status': 'SUCCESS',
            'license_number': license_number,
            'lic_no': license_no,
        }
    except Exception as exc:
        logger.error("Error processing license %s: %s", license_no, exc, exc_info=True)
        raise


# ---------------------------------------------------------------------------
# Item Pivot — async Excel generation (OQ-6: MUST stay async Celery)
# ---------------------------------------------------------------------------

@shared_task(bind=True)
def generate_item_pivot_task(
    self,
    days: int = 30,
    sion_norm: str = None,
    company_ids: str = None,
    exclude_company_ids: str = None,
    min_balance: int = 200,
    license_status: str = 'active',
):
    """
    Generate item pivot report Excel file as a background Celery task.

    OQ-6 approved: this MUST remain async. Never convert to synchronous.

    Returns: dict with file_path, filename, download_url, and metadata.
    """
    from django.conf import settings
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell import WriteOnlyCell
    from apps.license.views.item_pivot_report import _ItemPivotReportEngine, _xlsx_safe_row

    logger.info("Starting generate_item_pivot_task: task_id=%s", self.request.id)
    self.update_state(
        state='PROGRESS',
        meta={'current': 0, 'total': 100, 'status': 'Generating report data...'},
    )

    try:
        engine = _ItemPivotReportEngine()
        report_data = engine.generate_report(
            days=days,
            sion_norm=sion_norm,
            company_ids=company_ids,
            exclude_company_ids=exclude_company_ids,
            min_balance=min_balance,
            license_status=license_status,
        )

        if isinstance(report_data, dict) and 'error' in report_data:
            raise ValueError(report_data['error'])

        self.update_state(
            state='PROGRESS',
            meta={'current': 50, 'total': 100, 'status': 'Creating Excel file...'},
        )

        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"item_pivot_report_{timestamp}_{self.request.id}.xlsx"
        file_path = os.path.join(exports_dir, filename)

        workbook = openpyxl.Workbook(write_only=True)
        licenses_by_norm_notification = report_data.get('licenses_by_norm_notification', {})

        if not licenses_by_norm_notification:
            raise ValueError('No data found matching the filters. Try adjusting the parameters.')

        total_sheets = sum(len(nd) for nd in licenses_by_norm_notification.values())
        current_sheet = 0

        for norm_class in sorted(licenses_by_norm_notification.keys()):
            notifications_dict = licenses_by_norm_notification[norm_class]
            for notification, licenses_list in sorted(notifications_dict.items()):
                current_sheet += 1
                progress = 50 + int((current_sheet / max(total_sheets, 1)) * 40)
                self.update_state(
                    state='PROGRESS',
                    meta={
                        'current': progress,
                        'total': 100,
                        'status': (
                            f'Creating sheet {current_sheet}/{total_sheets}: '
                            f'{norm_class} - {notification}'
                        ),
                    },
                )

                sheet_name = (
                    f"{norm_class}_{notification}"[:31]
                    .replace('/', '-').replace('\\', '-').replace('*', '-')
                    .replace('[', '(').replace(']', ')')
                )
                worksheet = workbook.create_sheet(title=sheet_name)

                title_cell = WriteOnlyCell(
                    worksheet,
                    value=f"Item Pivot Report - {norm_class} - {notification}",
                )
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
                worksheet.append([title_cell] + [None] * 25)
                worksheet.append([])

                # Filter items to those with data in this sheet.
                items_with_data = [
                    item for item in report_data['items']
                    if any(
                        lic['items'].get(item['name'], {}).get('quantity', 0) > 0
                        for lic in licenses_list
                    )
                ]

                base_headers = [
                    'Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt', 'Exporter',
                    'Total CIF', 'Balance CIF',
                ]
                item_headers = []
                for item in items_with_data:
                    item_name = item['name']
                    has_restriction = item.get('has_restriction', False)
                    headers = [
                        f"{item_name} HSN Code",
                        f"{item_name} Product Description",
                        f"{item_name} Total QTY",
                        f"{item_name} Allotted QTY",
                        f"{item_name} Debited QTY",
                        f"{item_name} Balance QTY",
                    ]
                    if has_restriction:
                        headers.extend([
                            f"{item_name} Restriction %",
                            f"{item_name} Restriction Value",
                        ])
                    item_headers.extend(headers)

                all_headers = base_headers + item_headers
                header_row = []
                for header in all_headers:
                    cell = WriteOnlyCell(worksheet, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    header_row.append(cell)
                worksheet.append(header_row)

                for idx, license_data in enumerate(licenses_list, 1):
                    row_values = [
                        idx,
                        license_data['license_number'],
                        license_data['license_date'],
                        license_data['license_expiry_date'],
                        license_data['exporter'],
                        license_data['total_cif'],
                        license_data['balance_cif'],
                    ]
                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        item_data = license_data['items'].get(item_name, {
                            'hs_code': '', 'description': '', 'quantity': 0,
                            'allotted_quantity': 0, 'debited_quantity': 0,
                            'available_quantity': 0, 'restriction': None, 'restriction_value': 0,
                        })
                        row_values.extend([
                            item_data.get('hs_code', ''),
                            item_data.get('description', ''),
                            item_data.get('quantity', 0),
                            item_data.get('allotted_quantity', 0),
                            item_data.get('debited_quantity', 0),
                            item_data.get('available_quantity', 0),
                        ])
                        if has_restriction:
                            restriction_val = item_data.get('restriction')
                            row_values.append(restriction_val if restriction_val else '')
                            row_values.append(
                                item_data.get('restriction_value', 0)
                                if item_data.get('restriction_value') else ''
                            )
                    worksheet.append(row_values)

                # Totals row.
                totals_row = [WriteOnlyCell(worksheet, value='TOTAL')]
                totals_row[0].font = Font(bold=True)
                totals_row.extend([None, None, None, None])

                total_cif_cell = WriteOnlyCell(
                    worksheet, value=sum(l['total_cif'] for l in licenses_list)
                )
                total_cif_cell.font = Font(bold=True)
                totals_row.append(total_cif_cell)

                balance_cif_cell = WriteOnlyCell(
                    worksheet, value=sum(l['balance_cif'] for l in licenses_list)
                )
                balance_cif_cell.font = Font(bold=True)
                totals_row.append(balance_cif_cell)

                for item in items_with_data:
                    item_name = item['name']
                    has_restriction = item.get('has_restriction', False)
                    totals_row.extend([None, None])
                    for qty_type in ['quantity', 'allotted_quantity', 'debited_quantity', 'available_quantity']:
                        total = sum(
                            l['items'].get(item_name, {}).get(qty_type, 0) for l in licenses_list
                        )
                        cell = WriteOnlyCell(worksheet, value=total)
                        cell.font = Font(bold=True)
                        totals_row.append(cell)
                    if has_restriction:
                        totals_row.append(None)
                        total_restriction = sum(
                            l['items'].get(item_name, {}).get('restriction_value', 0)
                            for l in licenses_list
                        )
                        cell = WriteOnlyCell(worksheet, value=total_restriction)
                        cell.font = Font(bold=True)
                        totals_row.append(cell)
                worksheet.append(totals_row)

        self.update_state(
            state='PROGRESS',
            meta={'current': 95, 'total': 100, 'status': 'Saving file...'},
        )
        workbook.save(file_path)
        workbook.close()

        file_size = os.path.getsize(file_path)
        logger.info("Item pivot Excel generated: %s (%s bytes)", filename, file_size)

        return {
            'status': 'SUCCESS',
            'file_path': file_path,
            'filename': filename,
            'file_size': file_size,
            'download_url': f'/media/exports/{filename}',
            'generated_at': datetime.now().isoformat(),
            'task_id': self.request.id,
        }

    except Exception as exc:
        logger.error("Error generating item pivot Excel: %s", exc, exc_info=True)
        self.update_state(
            state='FAILURE',
            meta={'error': str(exc), 'exc_type': type(exc).__name__},
        )
        raise


# ---------------------------------------------------------------------------
# Balance update task (triggered from Item Pivot "Update Balance" action)
# ---------------------------------------------------------------------------

@shared_task(bind=True)
def update_all_license_balances_task(self, license_status: str = 'all'):
    """
    High-priority task to update balance_cif, is_active, is_expired, and
    restrictions for all (or filtered) licenses.

    Triggered manually from the Item Pivot Report's "Update Balance" action.

    Args:
        license_status: 'active', 'inactive', or 'all'

    Returns:
        dict with status, counts, and timing.
    """
    from django.utils import timezone
    from decimal import Decimal
    from apps.license.models import LicenseDetailsModel, LicenseBalance, LicenseFlags

    logger.info(
        "Starting update_all_license_balances_task: task_id=%s license_status=%s",
        self.request.id,
        license_status,
    )
    start_time = datetime.now()

    try:
        licenses = LicenseDetailsModel.objects.all()
        if license_status == 'active':
            licenses = licenses.filter(flags__is_active=True)
        elif license_status == 'inactive':
            licenses = licenses.filter(flags__is_active=False)

        total_licenses = licenses.count()
        self.update_state(
            state='PROGRESS',
            meta={'current': 0, 'total': total_licenses, 'status': 'Updating license balances...'},
        )

        updated_count = 0
        skipped_count = 0
        error_count = 0
        today = timezone.now().date()
        batch_size = 50

        for i, license_obj in enumerate(licenses.iterator(chunk_size=batch_size)):
            try:
                from apps.license.services.balance_service import recompute_license_balance
                recompute_license_balance(license_obj.id)

                is_expired = (
                    license_obj.license_expiry_date < today
                    if license_obj.license_expiry_date else False
                )
                is_null = False
                try:
                    bal = LicenseBalance.objects.get(license_id=license_obj.id)
                    is_null = bal.balance_cif < Decimal('500')
                except LicenseBalance.DoesNotExist:
                    pass

                is_active = not is_expired

                LicenseFlags.objects.filter(license_id=license_obj.id).update(
                    is_expired=is_expired,
                    is_null=is_null,
                    is_active=is_active,
                )
                updated_count += 1

                if (i + 1) % batch_size == 0:
                    self.update_state(
                        state='PROGRESS',
                        meta={
                            'current': i + 1,
                            'total': total_licenses,
                            'status': (
                                f'Updated {updated_count} licenses, '
                                f'skipped {skipped_count}...'
                            ),
                        },
                    )

            except Exception as exc:
                error_count += 1
                logger.error(
                    "Error updating license %s: %s",
                    getattr(license_obj, 'license_number', license_obj.pk),
                    exc,
                )

        elapsed = (datetime.now() - start_time).total_seconds()
        result = {
            'status': 'success',
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': error_count,
            'total_licenses': total_licenses,
            'elapsed_seconds': elapsed,
            'timestamp': datetime.now().isoformat(),
        }
        logger.info("Balance update completed: %s", result)
        return result

    except Exception as exc:
        logger.error("Failed to update license balances: %s", exc)
        return {
            'status': 'error',
            'error': str(exc),
            'timestamp': datetime.now().isoformat(),
        }
