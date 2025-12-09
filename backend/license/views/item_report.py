"""
Item Report - List all License Import Items with filters and inline editing support
"""

from django.db.models import Q, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views import View
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from core.models import ItemNameModel
from license.models import LicenseImportItemsModel, LicenseDetailsModel


class ItemReportView(View):
    """
    Report showing all License Import Items with filters.

    GET parameters:
        - item_names: Comma-separated item name IDs for filtering (multiselect)
        - format: 'json' or 'excel' (default: json)
    """

    def get(self, request, *args, **kwargs):
        output_format = request.GET.get('format', 'json').lower()
        item_names = request.GET.get('item_names')  # Comma-separated item name IDs
        company_ids = request.GET.get('company_ids')  # Comma-separated company IDs
        exclude_company_ids = request.GET.get('exclude_company_ids')  # Comma-separated company IDs to exclude
        min_balance = int(request.GET.get('min_balance', 200))
        min_avail_qty = float(request.GET.get('min_avail_qty', 0))
        license_status = request.GET.get('license_status', 'active')
        is_restricted = request.GET.get('is_restricted')  # 'true', 'false', or None for all
        purchase_status = request.GET.get('purchase_status')  # Comma-separated purchase status codes

        if output_format == 'excel':
            try:
                return self.export_to_excel(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return JsonResponse({'error': str(e)}, status=500)

        # For JSON, generate full report
        try:
            report_data = self.generate_report(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(report_data, safe=False)

    def generate_report(self, item_names=None, company_ids=None, exclude_company_ids=None, min_balance=200, min_avail_qty=0, license_status='active', is_restricted=None, purchase_status=None):
        """
        Generate item report with all license import items.

        Args:
            item_names: Comma-separated item name IDs for filtering
            company_ids: Comma-separated company IDs to include (optional)
            exclude_company_ids: Comma-separated company IDs to exclude (optional)
            min_balance: Minimum balance CIF to include (default 200)
            min_avail_qty: Minimum available quantity to include (default 0)
            license_status: Filter by status - 'active', 'expired', 'expiring_soon', 'all' (default 'active')
            is_restricted: Filter by restriction status - 'true', 'false', or None for all
            purchase_status: Comma-separated purchase status codes (e.g., 'GE,MI,SM')

        Returns:
            Dictionary with report data
        """
        from datetime import date, timedelta
        from core.constants import GE, MI, IP, SM, CO

        today = date.today()

        # Base query - all import items with licenses
        items = LicenseImportItemsModel.objects.select_related(
            'license',
            'license__exporter',
            'hs_code'
        ).prefetch_related('items')

        # Apply license status filter
        if license_status == 'active':
            items = items.filter(
                license__is_active=True,
                license__license_expiry_date__gt=today - timedelta(days=30)
            )
        elif license_status == 'expired':
            items = items.filter(license__license_expiry_date__lt=today)
        elif license_status == 'expiring_soon':
            items = items.filter(
                license__is_active=True,
                license__license_expiry_date__gte=today,
                license__license_expiry_date__lte=today + timedelta(days=30)
            )
        # If 'all', no date or is_active filter applied

        # Filter by min_balance using stored available_value field (can be done in query)
        # This pre-filters before iteration for better performance
        items = items.filter(available_value__gte=min_balance)

        # Filter by min_avail_qty (can be done in query)
        if min_avail_qty > 0:
            items = items.filter(available_quantity__gte=min_avail_qty)

        # Filter by company IDs if specified
        if company_ids:
            company_id_list = [int(cid.strip()) for cid in company_ids.split(',') if cid.strip()]
            items = items.filter(license__exporter_id__in=company_id_list)

        # Exclude company IDs if specified
        if exclude_company_ids:
            exclude_id_list = [int(cid.strip()) for cid in exclude_company_ids.split(',') if cid.strip()]
            items = items.exclude(license__exporter_id__in=exclude_id_list)

        # Filter by item names if specified
        if item_names:
            item_name_ids = [int(id.strip()) for id in item_names.split(',') if id.strip()]
            items = items.filter(items__id__in=item_name_ids).distinct()

        # Filter by is_restricted if specified
        if is_restricted is not None:
            if is_restricted == 'true':
                items = items.filter(is_restricted=True)
            elif is_restricted == 'false':
                items = items.filter(is_restricted=False)

        # Filter by purchase_status if specified
        if purchase_status:
            purchase_status_list = [ps.strip() for ps in purchase_status.split(',') if ps.strip()]
            items = items.filter(license__purchase_status__in=purchase_status_list)

        # Order by license number and serial number
        items = items.order_by('license__license_number', 'serial_number')

        # Build report data
        report_items = []
        for item in items:
            # Get item names
            item_names_list = [{"id": i.id, "name": i.name} for i in item.items.all()]

            # Use the stored available_value field (updated by balance update task)
            # This field already contains the correct value:
            # - For restricted items: restriction-based calculated value
            # - For non-restricted items: license balance_cif
            # Note: Make sure to run "Update Balance" in Item Pivot Report to refresh these values
            available_balance = float(item.available_value or 0)

            report_items.append({
                'id': item.id,
                'license_id': item.license.id,
                'license_number': item.license.license_number,
                'license_date': item.license.license_date.isoformat() if item.license.license_date else None,
                'license_expiry_date': item.license.license_expiry_date.isoformat() if item.license.license_expiry_date else None,
                'exporter_name': item.license.exporter.name if item.license.exporter else None,
                'hs_code': item.hs_code.hs_code if item.hs_code else None,
                'product_description': item.description or '',
                'item_names': item_names_list,
                'quantity': float(item.quantity or 0),
                'available_quantity': float(item.available_quantity or 0),
                'available_balance': available_balance,
                'is_restricted': item.is_restricted,
                'notes': item.license.balance_report_notes or '',
                'condition_sheet': item.license.condition_sheet or '',
                'unit': item.unit,
                'serial_number': item.serial_number,
            })

        return {
            'report_date': date.today().isoformat(),
            'total_items': len(report_items),
            'items': report_items
        }

    def export_to_excel(self, item_names=None, company_ids=None, exclude_company_ids=None, min_balance=200, min_avail_qty=0, license_status='active', is_restricted=None, purchase_status=None):
        """Export item report to Excel with separate sheets for Restricted and Not Restricted items"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        # Generate report data
        report_data = self.generate_report(item_names, company_ids, exclude_company_ids, min_balance, min_avail_qty, license_status, is_restricted, purchase_status)
        items = report_data['items']

        # Separate items into restricted and not restricted
        restricted_items = [item for item in items if item['is_restricted']]
        not_restricted_items = [item for item in items if not item['is_restricted']]

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Headers
        headers = [
            'Sr No', 'License No', 'License Date', 'License Expiry Date', 'Serial Number',
            'HSN Code', 'Product Description', 'Item Name',
            'Available Quantity', 'Available Balance', 'Notes', 'Condition Sheet'
        ]

        def create_sheet(workbook, sheet_name, items_list):
            """Helper function to create a sheet with given items"""
            ws = workbook.create_sheet(title=sheet_name)

            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 40
            ws.column_dimensions['H'].width = 25
            ws.column_dimensions['I'].width = 18
            ws.column_dimensions['J'].width = 18
            ws.column_dimensions['K'].width = 30
            ws.column_dimensions['L'].width = 30

            # Add data rows
            for idx, item in enumerate(items_list, start=2):
                item_names_str = ', '.join([i['name'] for i in item['item_names']])

                ws.cell(row=idx, column=1, value=idx - 1)
                ws.cell(row=idx, column=2, value=item['license_number'])
                ws.cell(row=idx, column=3, value=item['license_date'])
                ws.cell(row=idx, column=4, value=item['license_expiry_date'])
                ws.cell(row=idx, column=5, value=item['serial_number'])
                ws.cell(row=idx, column=6, value=item['hs_code'])
                ws.cell(row=idx, column=7, value=item['product_description'])
                ws.cell(row=idx, column=8, value=item_names_str)
                ws.cell(row=idx, column=9, value=item['available_quantity'])
                ws.cell(row=idx, column=10, value=item['available_balance'])
                ws.cell(row=idx, column=11, value=item['notes'])
                ws.cell(row=idx, column=12, value=item['condition_sheet'])

            return ws

        # Create sheets (Restricted first, then Not Restricted)
        if restricted_items:
            create_sheet(wb, "Restricted", restricted_items)

        if not_restricted_items:
            create_sheet(wb, "Not Restricted", not_restricted_items)

        # If no items at all, create an empty sheet with message
        if not restricted_items and not not_restricted_items:
            ws = wb.create_sheet(title="No Data")
            ws.cell(row=1, column=1, value="No items found matching the filter criteria")

        # Save to bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=item_report.xlsx'
        return response


class ItemReportViewSet(viewsets.ViewSet):
    """
    ViewSet for Item Report actions like getting available items for filter.
    """
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='available-items')
    def available_items(self, request):
        """
        Get list of all active item names that have license import items.
        Returns: List of item names with id and name
        """
        # Get all item names (active ones)
        item_names = ItemNameModel.objects.filter(
            is_active=True
        ).order_by('name').values('id', 'name')

        return Response(list(item_names))
