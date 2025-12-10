"""
Item-wise Pivot Report

Shows licenses with items as column headers, displaying quantities and values per item.
Similar to the GE DFIA report format.
"""

from collections import defaultdict
from decimal import Decimal
from typing import Dict, List, Any

from django.db.models import Sum, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views import View
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from core.constants import DEC_0, DEC_000, GE, MI, IP, SM, CO
from core.models import ItemNameModel
from license.models import LicenseDetailsModel, LicenseImportItemsModel, LicenseExportItemModel


class ItemPivotReportView(View):
    """
    Report showing licenses with items as columns (pivot format).

    GET parameters:
        - format: 'json' or 'excel' (default: json)
        - days: Number of days to look back (default: 30)
    """

    def get(self, request, *args, **kwargs):
        output_format = request.GET.get('format', 'json').lower()
        days = int(request.GET.get('days', 30))
        sion_norm = request.GET.get('sion_norm')
        company_ids = request.GET.get('company_ids')  # Comma-separated company IDs
        exclude_company_ids = request.GET.get('exclude_company_ids')  # Comma-separated company IDs to exclude
        min_balance = int(request.GET.get('min_balance', 200))
        license_status = request.GET.get('license_status', 'active')

        # For Excel export, use streaming approach to avoid timeout
        if output_format == 'excel':
            try:
                return self.export_to_excel_streaming(days, sion_norm, company_ids, exclude_company_ids, min_balance, license_status)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return JsonResponse({
                    'error': str(e)
                }, status=500)

        # For JSON, generate full report
        try:
            report_data = self.generate_report(days, sion_norm, company_ids, exclude_company_ids, min_balance,
                                               license_status)
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)

        return JsonResponse(report_data, safe=False)

    def generate_report(self, days: int = 30, sion_norm: str = None,
                        company_ids: str = None, exclude_company_ids: str = None,
                        min_balance: int = 200, license_status: str = 'active') -> Dict[str, Any]:
        """
        Generate item-wise pivot report.

        Args:
            days: Number of days to look back for active licenses
            sion_norm: Filter by specific SION norm class (optional)
            company_ids: Comma-separated company IDs to include (optional)
            exclude_company_ids: Comma-separated company IDs to exclude (optional)
            min_balance: Minimum balance CIF to include (default 200)
            license_status: Filter by status - 'active', 'expired', 'expiring_soon', 'all' (default 'active')

        Returns:
            Dictionary with report data
        """
        from datetime import date, timedelta
        today = date.today()
        start_date = today - timedelta(days=days)

        # Base query - licenses with required purchase status
        # Note: Don't filter by is_active here as it may exclude expired licenses
        licenses = LicenseDetailsModel.objects.filter(
            purchase_status__in=[GE, MI, IP, SM, CO]
        )

        # Apply license status filter
        if license_status == 'active':
            # Active: expiry date > today - 30 days (not expired more than 30 days ago)
            licenses = licenses.filter(
                is_active=True,
                license_expiry_date__gt=today - timedelta(days=30)
            )
        elif license_status == 'expired':
            # Expired: expiry date < today (don't filter by is_active to include all expired licenses)
            licenses = licenses.filter(license_expiry_date__lt=today)
        elif license_status == 'expiring_soon':
            # Expiring soon: expiry within next 30 days
            licenses = licenses.filter(
                is_active=True,
                license_expiry_date__gte=today,
                license_expiry_date__lte=today + timedelta(days=30)
            )
        # If 'all', no date or is_active filter applied - shows everything

        # Filter by SION norm if specified (optional)
        if sion_norm:
            licenses = licenses.filter(export_license__norm_class__norm_class=sion_norm).distinct()

        # Filter by company IDs if specified
        if company_ids:
            company_id_list = [int(cid.strip()) for cid in company_ids.split(',') if cid.strip()]
            licenses = licenses.filter(exporter_id__in=company_id_list)

        # Exclude company IDs if specified
        if exclude_company_ids:
            exclude_id_list = [int(cid.strip()) for cid in exclude_company_ids.split(',') if cid.strip()]
            licenses = licenses.exclude(exporter_id__in=exclude_id_list)

        # Filter by min_balance at database level using stored balance_cif field
        # This dramatically reduces the number of licenses we need to process
        licenses = licenses.filter(balance_cif__gte=min_balance)

        # Build filtered prefetch querysets based on sion_norm
        import_items_qs = LicenseImportItemsModel.objects.select_related('hs_code')
        export_items_qs = LicenseExportItemModel.objects.select_related('norm_class')
        item_names_qs = ItemNameModel.objects.filter(is_active=True).select_related('sion_norm_class')

        # If sion_norm specified, filter prefetch queries to only that norm
        if sion_norm:
            item_names_qs = item_names_qs.filter(sion_norm_class__norm_class=sion_norm)
            export_items_qs = export_items_qs.filter(norm_class__norm_class=sion_norm)

        # Optimize with select_related and prefetch_related to reduce queries
        licenses = licenses.select_related(
            'exporter',
            'port',
            'current_owner'
        ).prefetch_related(
            Prefetch('import_license',
                     queryset=import_items_qs.prefetch_related(
                         Prefetch('items', queryset=item_names_qs)
                     ).only('id', 'license_id', 'hs_code_id', 'quantity', 'allotted_quantity',
                            'debited_quantity', 'available_quantity', 'debited_value', 'cif_fc', 'description')),
            Prefetch('export_license',
                     queryset=export_items_qs.only('id', 'license_id', 'norm_class_id', 'cif_fc')),
            'license_documents',
            'transfers'
        ).only('id', 'license_number', 'license_date', 'license_expiry_date', 'exporter_id',
               'port_id', 'notification_number', 'purchase_status', 'balance_cif',
               'balance_report_notes', 'condition_sheet', 'current_owner_id'
        ).order_by('license_expiry_date', 'license_date')

        # Collect all unique items across all licenses
        # Use list() with prefetch_related for optimal performance (iterator breaks prefetch)
        all_items = {}  # Changed to dict to store item object for sorting
        valid_licenses = list(licenses)  # Licenses already filtered by balance_cif at DB level

        for license_obj in valid_licenses:
            for import_item in license_obj.import_license.all():
                for item in import_item.items.all():
                    # Only add items with valid names and that are active
                    if item and item.name and item.is_active:
                        # If filtering by norm, only include items matching that norm
                        if sion_norm:
                            if item.sion_norm_class and item.sion_norm_class.norm_class == sion_norm:
                                all_items[item.id] = item
                        else:
                            all_items[item.id] = item

        # Sort items by display_order first, then by name for consistent column order
        sorted_items = sorted(
            [(item.id, item.name) for item in all_items.values()],
            key=lambda x: (all_items[x[0]].display_order, x[1] or '')
        )

        # Build license data with item columns, grouped by norm first, then notification
        from collections import defaultdict
        licenses_by_norm_notification = defaultdict(lambda: defaultdict(list))

        for license_obj in valid_licenses:
            license_row = self._build_license_row(license_obj, sorted_items)

            if license_row:
                # Handle blank/empty notification numbers
                notification = (license_obj.notification_number or '').strip()
                if not notification:
                    notification = 'Unknown'

                # Get norm class from license
                norm_class = 'Unknown'
                if license_obj.export_license.exists():
                    first_export = license_obj.export_license.first()
                    if first_export and first_export.norm_class:
                        norm_class = first_export.norm_class.norm_class

                # Define conversion norms
                conversion_norms = ['E1', 'E5', 'E126', 'E132']
                is_conversion = license_obj.purchase_status == CO

                # Build notification key based on norm class and purchase status
                if norm_class in conversion_norms and is_conversion:
                    # For conversion licenses in E1, E5, E126, E132
                    if norm_class == 'E5':
                        # E5 Conversion: split by Parle vs Others
                        exporter_name = license_obj.exporter.name if license_obj.exporter else ''
                        if 'PARLE' in exporter_name.upper():
                            notification_key = f"{notification} - Conversion - Parle"
                        else:
                            notification_key = f"{notification} - Conversion"
                    else:
                        # E1, E126, E132 Conversion
                        notification_key = f"{notification} - Conversion"

                elif norm_class == 'E5':
                    # E5 non-conversion: split by Parle vs Others
                    exporter_name = license_obj.exporter.name if license_obj.exporter else ''
                    if 'PARLE' in exporter_name.upper():
                        notification_key = f"{notification} - Parle"
                    else:
                        notification_key = f"{notification} - Others"

                else:
                    # Regular grouping by notification for other norms
                    notification_key = notification

                licenses_by_norm_notification[norm_class][notification_key].append(license_row)

        # Determine which items have restrictions
        items_with_restrictions = set()
        for norm_dict in licenses_by_norm_notification.values():
            for licenses_list in norm_dict.values():
                for license_row in licenses_list:
                    for item_id, item_name in sorted_items:
                        item_data = license_row.get('items', {}).get(item_name, {})
                        if item_data.get('restriction') is not None:
                            items_with_restrictions.add(item_id)

        # Convert nested defaultdict to regular dict
        result_dict = {}
        for norm, notification_dict in licenses_by_norm_notification.items():
            result_dict[norm] = dict(notification_dict)

        # Fetch notes and conditions for all norms in a single query
        from core.models import SionNormClassModel
        norm_classes_list = list(result_dict.keys())
        sion_norms = SionNormClassModel.objects.filter(
            norm_class__in=norm_classes_list
        ).prefetch_related('notes', 'conditions')

        # Build dict from fetched norms
        norm_notes_conditions = {}
        sion_norms_dict = {sn.norm_class: sn for sn in sion_norms}

        for norm_class in norm_classes_list:
            if norm_class in sion_norms_dict:
                sion_norm = sion_norms_dict[norm_class]
                norm_notes_conditions[norm_class] = {
                    'notes': [
                        {'note_text': note.note_text, 'display_order': note.display_order}
                        for note in sion_norm.notes.all()
                    ],
                    'conditions': [
                        {'condition_text': cond.condition_text, 'display_order': cond.display_order}
                        for cond in sion_norm.conditions.all()
                    ]
                }
            else:
                norm_notes_conditions[norm_class] = {'notes': [], 'conditions': []}

        return {
            'items': [
                {
                    'id': item_id,
                    'name': item_name,
                    'has_restriction': item_id in items_with_restrictions
                }
                for item_id, item_name in sorted_items
            ],
            'licenses_by_norm_notification': result_dict,
            'norm_notes_conditions': norm_notes_conditions,
            'report_date': today.isoformat(),
        }

    def _build_license_row(self, license_obj: LicenseDetailsModel, all_items: List[tuple]) -> Dict[str, Any]:
        """
        Build a single license row with item columns.

        Args:
            license_obj: LicenseDetailsModel instance
            all_items: List of (item_id, item_name) tuples

        Returns:
            Dictionary with license data and item quantities
        """
        # Calculate total CIF from export license items (already prefetched)
        # Start with Decimal('0') to ensure result is always Decimal type
        total_cif = Decimal('0')
        for item in license_obj.export_license.all():
            # Convert to Decimal to handle cases where database returns float
            cif_value = Decimal(str(item.cif_fc)) if item.cif_fc is not None else Decimal('0')
            total_cif += cif_value

        # Aggregate quantities by item (sum across all serial numbers)
        item_quantities = defaultdict(lambda: {
            'quantity': Decimal('0.000'),
            'allotted_quantity': Decimal('0.000'),
            'debited_quantity': Decimal('0.000'),
            'available_quantity': Decimal('0.000'),
            'debited_value': Decimal('0.00'),
            'cif_value': Decimal('0.00'),
            'hs_code': '',
            'description': '',
            'sion_norm_class': None,
            'restriction_percentage': None,
        })

        # Group items by (sion_norm_class, restriction_percentage) to calculate shared restriction limits
        restriction_groups = defaultdict(lambda: {
            'total_cif': Decimal('0.00'),
            'debited_cif': Decimal('0.00'),
            'available_cif': Decimal('0.00'),
            'restriction_percentage': None,
            'sion_norm_class': None,
            'item_ids': []
        })

        for import_item in license_obj.import_license.all():
            for item in import_item.items.all():
                # Convert all numeric fields to Decimal to handle potential float values from database
                item_quantities[item.id]['quantity'] += Decimal(str(import_item.quantity)) if import_item.quantity is not None else DEC_000
                item_quantities[item.id]['allotted_quantity'] += Decimal(str(import_item.allotted_quantity)) if import_item.allotted_quantity is not None else DEC_000
                item_quantities[item.id]['debited_quantity'] += Decimal(str(import_item.debited_quantity)) if import_item.debited_quantity is not None else DEC_000
                item_quantities[item.id]['available_quantity'] += Decimal(str(import_item.available_quantity)) if import_item.available_quantity is not None else DEC_000
                item_quantities[item.id]['debited_value'] += Decimal(str(import_item.debited_value)) if import_item.debited_value is not None else DEC_0
                item_quantities[item.id]['cif_value'] += Decimal(str(import_item.cif_fc)) if import_item.cif_fc is not None else DEC_0

                if import_item.hs_code and not item_quantities[item.id]['hs_code']:
                    item_quantities[item.id]['hs_code'] = import_item.hs_code.hs_code

                if import_item.description and not item_quantities[item.id]['description']:
                    item_quantities[item.id]['description'] = import_item.description

                # Get restriction from item's sion_norm_class and restriction_percentage
                if item and hasattr(item, 'sion_norm_class') and item.sion_norm_class:
                    sion_norm = item.sion_norm_class.norm_class
                    restriction_pct = item.restriction_percentage

                    item_quantities[item.id]['sion_norm_class'] = sion_norm
                    item_quantities[item.id]['restriction_percentage'] = restriction_pct

                    # Group by (sion_norm_class, restriction_percentage) for shared restriction calculation
                    # Items in same SION norm with same restriction % share the restriction limit
                    restriction_key = f"{sion_norm}_{restriction_pct}"
                    restriction_groups[restriction_key]['sion_norm_class'] = sion_norm
                    restriction_groups[restriction_key]['restriction_percentage'] = restriction_pct
                    # Convert to Decimal to handle potential float values from database
                    restriction_groups[restriction_key]['total_cif'] += Decimal(str(import_item.cif_fc)) if import_item.cif_fc is not None else DEC_0
                    restriction_groups[restriction_key]['debited_cif'] += Decimal(str(import_item.debited_value)) if import_item.debited_value is not None else DEC_0
                    if item.id not in restriction_groups[restriction_key]['item_ids']:
                        restriction_groups[restriction_key]['item_ids'].append(item.id)

        # Calculate available CIF within restriction for each group
        # Use stored balance_cif field instead of property to avoid extra queries
        # Convert to Decimal to handle potential float value from database
        balance_cif = Decimal(str(license_obj.balance_cif)) if license_obj.balance_cif is not None else Decimal('0')
        for group_name, group_data in restriction_groups.items():
            if group_data['restriction_percentage'] and total_cif > 0:
                # Convert restriction_percentage to Decimal to avoid float * Decimal error
                restriction_pct_decimal = Decimal(str(group_data['restriction_percentage']))
                # Maximum allowed CIF for this restriction group
                max_allowed_cif = (total_cif * restriction_pct_decimal) / Decimal('100')
                # Available CIF = max_allowed - debited
                available_cif = max_allowed_cif - group_data['debited_cif']
                # Cap at balance_cif - restriction cannot exceed available balance
                available_cif = min(available_cif, balance_cif)
                group_data['available_cif'] = max(available_cif, Decimal('0'))

        # Build row data
        # Handle blank/empty notification numbers
        notification_display = (license_obj.notification_number or '').strip()
        if not notification_display:
            notification_display = 'Unknown'

        # Check for document types
        has_tl = license_obj.license_documents.filter(type='TRANSFER LETTER').exists()
        has_copy = license_obj.license_documents.filter(type='LICENSE COPY').exists()

        # Get latest transfer
        latest_transfer_text = ''
        transfer_qs = license_obj.transfers.order_by("-transfer_date", "-id")
        if transfer_qs.exists():
            transfer = transfer_qs.first()
            latest_transfer_text = str(transfer)
        elif license_obj.current_owner:
            latest_transfer_text = f"Current Owner is {license_obj.current_owner.name}"
        else:
            latest_transfer_text = "Data Not Found"

        row_data = {
            'id': license_obj.id,
            'license_number': license_obj.license_number,
            'license_date': license_obj.license_date.isoformat() if license_obj.license_date else None,
            'license_expiry_date': license_obj.license_expiry_date.isoformat(),
            'exporter': str(license_obj.exporter) if license_obj.exporter else '',
            'port': str(license_obj.port) if license_obj.port else '',
            'notification_number': notification_display,
            'total_cif': float(total_cif),
            'balance_cif': float(balance_cif),  # Reuse already calculated balance
            'balance_report_notes': license_obj.balance_report_notes or '',
            'condition_sheet': license_obj.condition_sheet or '',
            'latest_transfer': latest_transfer_text,
            'has_tl': has_tl,
            'has_copy': has_copy,
            'items': {}
        }

        # Add item columns
        for item_id, item_name in all_items:
            if item_id in item_quantities:
                item_data = item_quantities[item_id]
                sion_norm = item_data['sion_norm_class']
                restriction_pct = item_data['restriction_percentage']

                # Get restriction percentage only (as number)
                restriction_value = None
                available_cif = Decimal('0')

                if sion_norm and restriction_pct:
                    restriction_key = f"{sion_norm}_{restriction_pct}"
                    if restriction_key in restriction_groups:
                        group = restriction_groups[restriction_key]
                        restriction_value = float(restriction_pct)
                        available_cif = group['available_cif']

                row_data['items'][item_name] = {
                    'hs_code': item_data['hs_code'],
                    'description': item_data['description'],
                    'quantity': float(item_data['quantity']),
                    'allotted_quantity': float(item_data['allotted_quantity']),
                    'debited_quantity': float(item_data['debited_quantity']),
                    'available_quantity': float(item_data['available_quantity']),
                    'restriction': restriction_value,
                    'restriction_value': float(available_cif),
                }
            else:
                row_data['items'][item_name] = {
                    'hs_code': '',
                    'description': '',
                    'quantity': 0,
                    'allotted_quantity': 0,
                    'debited_quantity': 0,
                    'available_quantity': 0,
                    'restriction': None,
                    'restriction_value': 0,
                }

        return row_data

    def export_to_excel(self, report_data: Dict[str, Any]) -> HttpResponse:
        """
        Export report to Excel format with items as columns, split by norm then notification.
        Uses streaming to handle large datasets efficiently.

        Args:
            report_data: Report data dictionary

        Returns:
            StreamingHttpResponse with Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.cell import WriteOnlyCell
        from django.http import StreamingHttpResponse
        import tempfile
        import os

        # Create a temporary file for the workbook
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()

        try:
            # Use write_only mode for streaming
            workbook = openpyxl.Workbook(write_only=True)

            licenses_by_norm_notification = report_data.get('licenses_by_norm_notification', {})

            # Create a sheet for each norm-notification combination
            for norm_class in sorted(licenses_by_norm_notification.keys()):
                notifications_dict = licenses_by_norm_notification[norm_class]
                for notification, licenses_list in sorted(notifications_dict.items()):
                    # Sanitize sheet name (Excel has 31 char limit and doesn't allow certain chars)
                    sheet_name = f"{norm_class}_{notification}"[:31].replace('/', '-').replace('\\', '-').replace('*',
                                                                                                                  '-').replace(
                        '[', '(').replace(']', ')')
                    worksheet = workbook.create_sheet(title=sheet_name)

                    # Title row
                    title = f"Item Pivot Report - {norm_class} - {notification}"
                    title_cell = WriteOnlyCell(worksheet, value=title)
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center')
                    worksheet.append([title_cell] + [None] * 25)  # Span across columns
                    worksheet.append([])  # Empty row

                    # Build headers
                    base_headers = [
                        'Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt', 'Exporter',
                        'Total CIF', 'Balance CIF'
                    ]

                    # Add item columns (HSN Code, Product Description, Total QTY, Debited QTY, Available QTY, Restriction %, Restriction Value)
                    item_headers = []
                    for item in report_data['items']:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)

                        headers = [
                            f"{item_name} HSN Code",
                            f"{item_name} Product Description",
                            f"{item_name} Total QTY",
                            f"{item_name} Allotted QTY",
                            f"{item_name} Debited QTY",
                            f"{item_name} Balance QTY",
                        ]

                        if has_restriction:
                            headers.extend([
                                f"{item_name} Restriction %",
                                f"{item_name} Restriction Value"
                            ])

                        item_headers.extend(headers)

                    all_headers = base_headers + item_headers

                    # Write headers with styling
                    header_row = []
                    for header in all_headers:
                        cell = WriteOnlyCell(worksheet, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                        header_row.append(cell)
                    worksheet.append(header_row)

                    # Write data rows for this norm-notification combination
                    for idx, license_data in enumerate(licenses_list, 1):
                        row_data = []

                        # Base columns
                        row_data.append(idx)
                        row_data.append(license_data['license_number'])
                        row_data.append(license_data['license_date'])
                        row_data.append(license_data['license_expiry_date'])
                        row_data.append(license_data['exporter'])
                        row_data.append(license_data['total_cif'])
                        row_data.append(license_data['balance_cif'])

                        # Item columns
                        for item in report_data['items']:
                            item_name = item['name']
                            has_restriction = item.get('has_restriction', False)
                            item_data = license_data['items'].get(item_name, {
                                'hs_code': '',
                                'description': '',
                                'quantity': 0,
                                'allotted_quantity': 0,
                                'debited_quantity': 0,
                                'available_quantity': 0,
                                'restriction': None,
                                'restriction_value': 0
                            })

                            row_data.append(item_data.get('hs_code', ''))
                            row_data.append(item_data.get('description', ''))
                            row_data.append(item_data.get('quantity', 0))
                            row_data.append(item_data.get('allotted_quantity', 0))
                            row_data.append(item_data.get('debited_quantity', 0))
                            row_data.append(item_data.get('available_quantity', 0))

                            # Only write restriction columns if item has restrictions
                            if has_restriction:
                                restriction_val = item_data.get('restriction')
                                row_data.append(restriction_val if restriction_val else '')
                                restriction_value = item_data.get('restriction_value', 0)
                                row_data.append(restriction_value if restriction_value else '')

                        # Append row to worksheet
                        worksheet.append(row_data)

                    # Add totals row for this norm-notification
                    totals_row = []

                    # Total label
                    total_cell = WriteOnlyCell(worksheet, value='TOTAL')
                    total_cell.font = Font(bold=True)
                    totals_row.append(total_cell)

                    # Skip columns 2-5 (DFIA No, DFIA Dt, Expiry Dt, Exporter)
                    totals_row.extend([None, None, None, None])

                    # Calculate totals for CIF columns
                    total_cif = sum(lic['total_cif'] for lic in licenses_list)
                    balance_cif = sum(lic['balance_cif'] for lic in licenses_list)

                    total_cif_cell = WriteOnlyCell(worksheet, value=total_cif)
                    total_cif_cell.font = Font(bold=True)
                    totals_row.append(total_cif_cell)

                    balance_cif_cell = WriteOnlyCell(worksheet, value=balance_cif)
                    balance_cif_cell.font = Font(bold=True)
                    totals_row.append(balance_cif_cell)

                    # Calculate totals for each item
                    for item in report_data['items']:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)

                        total_qty = sum(
                            lic['items'].get(item_name, {}).get('quantity', 0)
                            for lic in licenses_list
                        )
                        total_allotted = sum(
                            lic['items'].get(item_name, {}).get('allotted_quantity', 0)
                            for lic in licenses_list
                        )
                        total_debited = sum(
                            lic['items'].get(item_name, {}).get('debited_quantity', 0)
                            for lic in licenses_list
                        )
                        total_avail = sum(
                            lic['items'].get(item_name, {}).get('available_quantity', 0)
                            for lic in licenses_list
                        )
                        total_restriction_val = sum(
                            lic['items'].get(item_name, {}).get('restriction_value', 0)
                            for lic in licenses_list
                        )

                        # Skip HSN and Description columns in totals
                        totals_row.extend([None, None])

                        # Add quantity totals with bold font
                        qty_cell = WriteOnlyCell(worksheet, value=total_qty)
                        qty_cell.font = Font(bold=True)
                        totals_row.append(qty_cell)

                        allotted_cell = WriteOnlyCell(worksheet, value=total_allotted)
                        allotted_cell.font = Font(bold=True)
                        totals_row.append(allotted_cell)

                        debited_cell = WriteOnlyCell(worksheet, value=total_debited)
                        debited_cell.font = Font(bold=True)
                        totals_row.append(debited_cell)

                        avail_cell = WriteOnlyCell(worksheet, value=total_avail)
                        avail_cell.font = Font(bold=True)
                        totals_row.append(avail_cell)

                        # Only add restriction columns if item has restrictions
                        if has_restriction:
                            totals_row.append(None)  # Skip Restriction % column

                            restriction_cell = WriteOnlyCell(worksheet, value=total_restriction_val)
                            restriction_cell.font = Font(bold=True)
                            totals_row.append(restriction_cell)

                    worksheet.append(totals_row)

            # Save workbook to temp file
            workbook.save(temp_file.name)
            workbook.close()

            # Create streaming response
            def file_iterator(file_path, chunk_size=8192):
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
                # Clean up temp file after streaming
                try:
                    os.unlink(file_path)
                except:
                    pass

            response = StreamingHttpResponse(
                file_iterator(temp_file.name),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="item_pivot_report.xlsx"'
            return response

        except Exception as e:
            # Clean up temp file in case of error
            try:
                os.unlink(temp_file.name)
            except:
                pass
            raise e

    def export_to_excel_streaming(self, days=30, sion_norm=None, company_ids=None,
                                  exclude_company_ids=None, min_balance=200, license_status='active'):
        """
        Export report to Excel - uses existing generate_report for data, then formats as Excel.
        This ensures consistency with JSON output.

        Returns:
            StreamingHttpResponse with Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.cell import WriteOnlyCell
        from django.http import StreamingHttpResponse
        import tempfile
        import os

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()

        try:
            # Use the working generate_report method
            report_data = self.generate_report(days, sion_norm, company_ids, exclude_company_ids, min_balance, license_status)

            workbook = openpyxl.Workbook(write_only=True)
            licenses_by_norm_notif = report_data.get('licenses_by_norm_notification', {})

            for norm_class in sorted(licenses_by_norm_notif.keys()):
                notifications_dict = licenses_by_norm_notif[norm_class]
                for notification, licenses_list in sorted(notifications_dict.items()):
                    # Filter items to only those with data in THIS norm-notification
                    items_with_data = []
                    for item in report_data['items']:
                        item_name = item['name']
                        has_data = any(
                            lic['items'].get(item_name, {}).get('quantity', 0) > 0
                            for lic in licenses_list
                        )
                        if has_data:
                            items_with_data.append(item)

                    # Create sheet
                    sheet_name = f"{norm_class}_{notification}"[:31].replace('/', '-').replace('\\', '-').replace('*', '-')
                    worksheet = workbook.create_sheet(title=sheet_name)

                    # Title row
                    title_cell = WriteOnlyCell(worksheet, value=f"Item Pivot Report - {norm_class} - {notification}")
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center')
                    worksheet.append([title_cell] + [None] * 25)
                    worksheet.append([])

                    # Headers
                    base_headers = ['Sr no', 'DFIA No', 'DFIA Dt', 'Expiry Dt', 'Exporter', 'Total CIF', 'Balance CIF']
                    item_headers = []
                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        headers = [
                            f"{item_name} HSN Code",
                            f"{item_name} Product Description",
                            f"{item_name} Total QTY",
                            f"{item_name} Allotted QTY",
                            f"{item_name} Debited QTY",
                            f"{item_name} Balance QTY"
                        ]
                        if has_restriction:
                            headers.extend([
                                f"{item_name} Restriction %",
                                f"{item_name} Restriction Value"
                            ])
                        item_headers.extend(headers)

                    all_headers = base_headers + item_headers
                    header_row = []
                    for header in all_headers:
                        cell = WriteOnlyCell(worksheet, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                        header_row.append(cell)
                    worksheet.append(header_row)

                    # Data rows
                    for idx, lic in enumerate(licenses_list, 1):
                        row_data = [
                            idx,
                            lic['license_number'],
                            lic['license_date'],
                            lic['license_expiry_date'],
                            lic['exporter'],
                            lic['total_cif'],
                            lic['balance_cif']
                        ]

                        for item in items_with_data:
                            item_name = item['name']
                            has_restriction = item.get('has_restriction', False)
                            item_data = lic['items'].get(item_name, {})
                            row_data.extend([
                                item_data.get('hs_code', ''),
                                item_data.get('description', ''),
                                item_data.get('quantity', 0),
                                item_data.get('allotted_quantity', 0),
                                item_data.get('debited_quantity', 0),
                                item_data.get('available_quantity', 0)
                            ])
                            if has_restriction:
                                row_data.extend([
                                    item_data.get('restriction'),
                                    item_data.get('restriction_value', 0)
                                ])

                        worksheet.append(row_data)

                    # Totals row
                    totals_row = [WriteOnlyCell(worksheet, value='TOTAL')]
                    totals_row[0].font = Font(bold=True)
                    totals_row.extend([None, None, None, None])

                    total_cif_cell = WriteOnlyCell(worksheet, value=sum(l['total_cif'] for l in licenses_list))
                    total_cif_cell.font = Font(bold=True)
                    totals_row.append(total_cif_cell)

                    balance_cif_cell = WriteOnlyCell(worksheet, value=sum(l['balance_cif'] for l in licenses_list))
                    balance_cif_cell.font = Font(bold=True)
                    totals_row.append(balance_cif_cell)

                    for item in items_with_data:
                        item_name = item['name']
                        has_restriction = item.get('has_restriction', False)
                        totals_row.extend([None, None])  # HSN, Description
                        for qty_type in ['quantity', 'allotted_quantity', 'debited_quantity', 'available_quantity']:
                            total = sum(l['items'].get(item_name, {}).get(qty_type, 0) for l in licenses_list)
                            cell = WriteOnlyCell(worksheet, value=total)
                            cell.font = Font(bold=True)
                            totals_row.append(cell)
                        if has_restriction:
                            totals_row.append(None)  # Restriction %
                            total_restriction = sum(l['items'].get(item_name, {}).get('restriction_value', 0) for l in licenses_list)
                            cell = WriteOnlyCell(worksheet, value=total_restriction)
                            cell.font = Font(bold=True)
                            totals_row.append(cell)

                    worksheet.append(totals_row)

            # Save workbook
            workbook.save(temp_file.name)
            workbook.close()

            # Stream file
            def file_iterator(file_path, chunk_size=8192):
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
                try:
                    os.unlink(file_path)
                except:
                    pass

            response = StreamingHttpResponse(
                file_iterator(temp_file.name),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="item_pivot_report.xlsx"'
            return response

        except Exception as e:
            try:
                os.unlink(temp_file.name)
            except:
                pass
            raise e


class ItemPivotViewSet(viewsets.ViewSet):
    """
    ViewSet for Item Pivot Report.

    Permissions: AllowAny - accessible to all users
    """
    permission_classes = [AllowAny]

    def list(self, request):
        """
        Get item pivot report.

        Query Parameters:
            days: Number of days to look back (default: 30)
        """
        view = ItemPivotReportView()
        return view.get(request)

    @action(detail=False, methods=['get'], url_path='available-norms')
    def available_norms(self, request):
        """
        Get list of all active norm classes with their descriptions.
        Returns array of objects with norm_class and description, including conversion norms (E1, E5, E126, E132).
        """
        try:
            # Define conversion norm classes
            CONVERSION_NORMS = ['E1', 'E5', 'E126', 'E132']

            # Get all active SION norm classes from the database
            from core.models import SionNormClassModel
            active_norms_data = SionNormClassModel.objects.filter(
                is_active=True
            ).values('norm_class', 'description')

            # Build result with norm_class and description
            norms_dict = {}
            for norm in active_norms_data:
                norms_dict[norm['norm_class']] = {
                    'norm_class': norm['norm_class'],
                    'description': norm['description'] or ''
                }

            # Add conversion norms if not already present
            for conv_norm in CONVERSION_NORMS:
                if conv_norm not in norms_dict:
                    norms_dict[conv_norm] = {
                        'norm_class': conv_norm,
                        'description': ''
                    }

            # Sort by norm_class and return as array
            result = sorted(norms_dict.values(), key=lambda x: x['norm_class'])

            return Response(result)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=['post', 'get'], url_path='generate-async')
    def generate_async(self, request):
        """
        Generate Excel report asynchronously using Celery.

        Query Parameters / POST Body:
            days: Number of days to look back (default: 30)
            sion_norm: Filter by SION norm (REQUIRED)
            company_ids: Comma-separated company IDs (optional)
            exclude_company_ids: Comma-separated company IDs to exclude (optional)
            min_balance: Minimum balance CIF (default: 200)
            license_status: Filter by status (default: 'active')

        Returns:
            task_id: ID to check status and download file
        """
        from license.tasks import generate_item_pivot_excel

        # Get parameters from request (support both GET and POST)
        params = request.data if request.method == 'POST' else request.GET
        days = int(params.get('days', 30))
        sion_norm = params.get('sion_norm')  # Optional - if not provided, exports ALL norms
        company_ids = params.get('company_ids')
        exclude_company_ids = params.get('exclude_company_ids')
        min_balance = int(params.get('min_balance', 200))
        license_status = params.get('license_status', 'active')

        # Start the Celery task
        task = generate_item_pivot_excel.delay(
            days=days,
            sion_norm=sion_norm,
            company_ids=company_ids,
            exclude_company_ids=exclude_company_ids,
            min_balance=min_balance,
            license_status=license_status
        )

        return Response({
            'task_id': task.id,
            'status': 'PENDING',
            'message': 'Report generation started. Use the task_id to check status.'
        }, status=202)

    @action(detail=False, methods=['get'], url_path='task-status/(?P<task_id>[^/.]+)')
    def task_status(self, request, task_id=None):
        """
        Check the status of an async Excel generation task.

        Returns:
            state: Task state (PENDING, PROGRESS, SUCCESS, FAILURE)
            current: Current progress (0-100)
            total: Total progress (100)
            status: Status message
            result: Result data (if completed)
        """
        from celery.result import AsyncResult

        task = AsyncResult(task_id)

        if task.state == 'PENDING':
            response = {
                'state': task.state,
                'current': 0,
                'total': 100,
                'status': 'Pending...'
            }
        elif task.state == 'PROGRESS':
            response = {
                'state': task.state,
                'current': task.info.get('current', 0),
                'total': task.info.get('total', 100),
                'status': task.info.get('status', '')
            }
        elif task.state == 'SUCCESS':
            response = {
                'state': task.state,
                'current': 100,
                'total': 100,
                'status': 'Completed!',
                'result': task.info
            }
        else:
            # Something went wrong
            response = {
                'state': task.state,
                'current': 100,
                'total': 100,
                'status': str(task.info) if task.info else 'Unknown error'
            }

        return Response(response)

    @action(detail=False, methods=['post'], url_path='update-balance')
    def update_balance(self, request):
        """
        Trigger high-priority task to update balance_cif, is_active, is_expired, and restrictions.

        This task:
        1. Updates balance_cif for all licenses using LicenseBalanceCalculator
        2. Updates is_expired based on license_expiry_date
        3. Updates is_null based on balance < $500
        4. Updates is_active based on expiry (mark inactive if expired)
        5. Checks and updates restriction flags on import items

        Returns:
            task_id: ID to check status using task-status endpoint
        """
        from license.tasks import update_all_license_balances

        # Get license_status parameter from request body
        license_status = request.data.get('license_status', 'all')

        # Start the Celery task with high priority
        task = update_all_license_balances.apply_async(
            args=[license_status],
            priority=9  # High priority (0-9, 9 is highest)
        )

        return Response({
            'task_id': task.id,
            'status': 'PENDING',
            'license_status': license_status,
            'message': f'Balance update started for {license_status} licenses. Use the task_id to check status.'
        }, status=202)
