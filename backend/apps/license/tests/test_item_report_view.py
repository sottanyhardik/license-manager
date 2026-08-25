"""
Tests for the Item Report (backend/apps/license/views/item_report.py).

Regression coverage for the 962f17af revert: the report must be rooted on
LicenseImportItemsModel (every import item, plan or no plan), and the
`item_names` filter must match the import item's own M2M tag — not a
LicenseItemPlan row.
"""
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

import pytest

from apps.core.models import CompanyModel, HSCodeModel, ItemNameModel, NotificationNumber
from apps.license.models import LicenseDetailsModel, LicenseImportItemsModel, LicenseItemPlan


User = get_user_model()

REPORT_URL = "/api/reports/item-report/"
AVAILABLE_ITEMS_URL = "/api/item-report/available-items/"


@pytest.fixture
def report_viewer_client(db):
    user = User.objects.create_user(
        username="item-report-viewer",
        email="item-report-viewer@example.com",
        password="RoleP@ssw0rd123",
    )
    group, _ = Group.objects.get_or_create(name="REPORT_VIEWER")
    user.groups.add(group)
    token = RefreshToken.for_user(user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token.access_token}")
    return client


@pytest.fixture
def item_report_masters(db):
    return {
        "parle": CompanyModel.objects.create(iec="3111111111", name="Item Report Parle"),
        "other": CompanyModel.objects.create(iec="3222222222", name="Item Report Other"),
        "hs_code": HSCodeModel.objects.create(hs_code="19053100", product_description="Biscuits"),
        "wheat": ItemNameModel.objects.create(name="Wheat Flour - Item Report"),
        "milk": ItemNameModel.objects.create(name="Milk Powder - Item Report"),
    }


def _make_license(number, exporter, *, expiry_days=30):
    return LicenseDetailsModel.objects.create(
        license_number=number,
        license_date=date.today() - timedelta(days=60),
        license_expiry_date=date.today() + timedelta(days=expiry_days),
        exporter=exporter,
    )


def _make_import_item(
    license_obj,
    hs_code,
    *,
    serial=1,
    available_value=Decimal("500.00"),
    available_quantity=Decimal("50.000"),
    condition_type="",
    item_names=None,
    quantity=Decimal("100.000"),
    cif_fc=Decimal("0.00"),
):
    item = LicenseImportItemsModel.objects.create(
        license=license_obj,
        serial_number=serial,
        description=f"Import item {serial}",
        hs_code=hs_code,
        quantity=quantity,
        available_quantity=available_quantity,
        available_value=available_value,
        condition_type=condition_type,
        cif_fc=cif_fc,
    )
    if item_names:
        item.items.set(item_names)
    return item


@pytest.mark.django_db
def test_item_report_includes_import_item_with_no_plan_and_filters_by_own_tag(
    report_viewer_client, item_report_masters
):
    license_obj = _make_license("ITEM-REPORT-001", item_report_masters["parle"])
    item = _make_import_item(
        license_obj,
        item_report_masters["hs_code"],
        item_names=[item_report_masters["wheat"]],
    )
    # Deliberately no LicenseItemPlan row for this item.

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert response.status_code == 200
    rows = {row["id"]: row for row in response.json()["items"]}

    assert item.id in rows, "import item with no plan must still appear in the report"
    row = rows[item.id]
    assert row["planned_quantity"] == 0
    assert row["planned_cif"] == 0
    assert row["plan_source"] == ""
    assert row["planned_splits"] == []
    assert row["item_names"] == [
        {"id": item_report_masters["wheat"].id, "name": item_report_masters["wheat"].name}
    ]

    # item_names filter matches the import item's own M2M tag (not a plan).
    matching = report_viewer_client.get(
        REPORT_URL, {"min_balance": 0, "item_names": str(item_report_masters["wheat"].id)}
    )
    assert [r["id"] for r in matching.json()["items"]] == [item.id]

    non_matching = report_viewer_client.get(
        REPORT_URL, {"min_balance": 0, "item_names": str(item_report_masters["milk"].id)}
    )
    assert non_matching.json()["items"] == []


@pytest.mark.django_db
def test_item_report_filters_min_balance_and_company_ids(report_viewer_client, item_report_masters):
    lic_a = _make_license("ITEM-REPORT-A", item_report_masters["parle"])
    lic_b = _make_license("ITEM-REPORT-B", item_report_masters["other"])
    item_a = _make_import_item(lic_a, item_report_masters["hs_code"], serial=1, available_value=Decimal("500.00"))
    item_b = _make_import_item(lic_b, item_report_masters["hs_code"], serial=1, available_value=Decimal("50.00"))

    # Default min_balance=200 excludes item_b (available_value=50).
    response = report_viewer_client.get(REPORT_URL)
    ids = {r["id"] for r in response.json()["items"]}
    assert item_a.id in ids
    assert item_b.id not in ids

    # With min_balance=0 both appear, but company_ids narrows to Parle only.
    response = report_viewer_client.get(
        REPORT_URL, {"min_balance": 0, "company_ids": str(item_report_masters["parle"].id)}
    )
    ids = {r["id"] for r in response.json()["items"]}
    assert ids == {item_a.id}


@pytest.mark.django_db
def test_item_report_balance_cif_is_live_not_stale_stored_value(report_viewer_client, item_report_masters):
    """Regression test: the report's `balance_cif` field must be computed
    LIVE via `LicenseBalanceCalculator.calculate_financial_balance_for_
    licenses` (the same batched calculator `item_pivot_report.py` uses),
    never read from the denormalized `LicenseBalance.balance_cif` column —
    that column is only refreshed by a background task/manual "Update
    Balance" trigger and can go stale. `.update()` bypasses the post_save
    recalculation signal, matching how the column can legitimately go
    stale in production (e.g. right after a Balance Engine formula change).
    """
    from apps.license.models import LicenseBalance, LicenseExportItemModel
    from apps.license.services.balance_calculator import LicenseBalanceCalculator

    lic = _make_license("ITEM-REPORT-LIVE-BAL", item_report_masters["parle"])
    LicenseExportItemModel.objects.create(license=lic, cif_fc=Decimal("9000.00"))
    item = _make_import_item(lic, item_report_masters["hs_code"], available_value=Decimal("500.00"))

    live_balance = LicenseBalanceCalculator.calculate_financial_balance(lic)
    assert live_balance == Decimal("9000.00")

    # Desync the stored column from the live figure without triggering the
    # recalculation signal, simulating a stale cache.
    LicenseBalance.objects.filter(license=lic).update(balance_cif=Decimal("1.00"))
    lic.refresh_from_db()
    assert lic.balance_cif != live_balance, "fixture setup must actually desync stored vs. live to prove the fix"

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert response.status_code == 200
    rows = {row["id"]: row for row in response.json()["items"]}

    assert rows[item.id]["balance_cif"] == float(live_balance)
    assert rows[item.id]["balance_cif"] != float(Decimal("1.00"))


@pytest.mark.django_db
def test_item_report_available_quantity_is_total_less_boe_debit_not_stored_operational_balance(
    report_viewer_client, item_report_masters, monkeypatch,
):
    """Paired trades must not hide the Customs Ledger's remaining quantity."""
    licence = _make_license("ITEM-REPORT-ACTUAL-AVAILABLE", item_report_masters["parle"])
    item = _make_import_item(
        licence,
        item_report_masters["hs_code"],
        quantity=Decimal("100.000"),
        # Represents an operational field reduced by a direct linked sale.
        available_quantity=Decimal("0.000"),
    )

    from apps.license.services.balance_calculator import ItemBalanceCalculator
    monkeypatch.setattr(
        ItemBalanceCalculator,
        "calculate_debited_quantity_for_items",
        staticmethod(lambda item_ids: {item.id: Decimal("40.000")}),
    )

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert response.status_code == 200
    rows = {row["id"]: row for row in response.json()["items"]}
    assert rows[item.id]["available_quantity"] == 60.0


@pytest.mark.django_db
def test_item_report_filters_is_restricted_combined_with_item_names(report_viewer_client, item_report_masters):
    lic = _make_license("ITEM-REPORT-RESTRICT", item_report_masters["parle"])
    restricted = _make_import_item(
        lic,
        item_report_masters["hs_code"],
        serial=1,
        condition_type="AU",
        item_names=[item_report_masters["wheat"]],
    )
    not_restricted = _make_import_item(
        lic,
        item_report_masters["hs_code"],
        serial=2,
        item_names=[item_report_masters["wheat"]],
    )

    response = report_viewer_client.get(
        REPORT_URL,
        {
            "min_balance": 0,
            "is_restricted": "true",
            "item_names": str(item_report_masters["wheat"].id),
        },
    )
    ids = [r["id"] for r in response.json()["items"]]
    assert ids == [restricted.id]
    assert not_restricted.id not in ids


@pytest.mark.django_db
def test_item_report_filters_by_notification_number_code_not_pk(report_viewer_client, item_report_masters):
    """`notification_numbers` is a comma-separated list of NotificationNumber
    *codes* (e.g. "025/2023") sent by the frontend's Notification dropdown —
    `license.notification_number` is a ForeignKey, so filtering must compare
    against its `code` field, not the FK's own numeric PK. Filtering by PK
    previously raised a 500 (Postgres/Django rejects comparing an integer PK
    column to a string like "025/2023"), so this also guards against that
    regression rather than just checking for a silently-empty result."""
    notif_a = NotificationNumber.objects.create(code="025/2023", label="Deemed Exports")
    notif_b = NotificationNumber.objects.create(code="019/2015", label="Physical Exports")
    lic_a = _make_license("ITEM-REPORT-NOTIF-A", item_report_masters["parle"])
    lic_a.notification_number = notif_a
    lic_a.save(update_fields=["notification_number"])
    lic_b = _make_license("ITEM-REPORT-NOTIF-B", item_report_masters["other"])
    lic_b.notification_number = notif_b
    lic_b.save(update_fields=["notification_number"])
    item_a = _make_import_item(lic_a, item_report_masters["hs_code"], available_value=Decimal("500.00"))
    item_b = _make_import_item(lic_b, item_report_masters["hs_code"], available_value=Decimal("500.00"))

    response = report_viewer_client.get(
        REPORT_URL, {"min_balance": 0, "notification_numbers": notif_a.code}
    )
    assert response.status_code == 200, getattr(response, "data", response.content[:300])
    ids = {r["id"] for r in response.json()["items"]}
    assert ids == {item_a.id}
    assert item_b.id not in ids


@pytest.mark.django_db
def test_item_report_unit_price_from_cif_fc_with_available_balance_fallback(
    report_viewer_client, item_report_masters
):
    """Unit Price = cif_fc / quantity when a usable cif_fc exists; falls back
    to available_balance / actual available quantity only when it doesn't (e.g. an
    older/incomplete record with cif_fc still at its zero default)."""
    lic = _make_license("ITEM-REPORT-UNITPRICE", item_report_masters["parle"])
    with_cif = _make_import_item(
        lic, item_report_masters["hs_code"], serial=1,
        quantity=Decimal("100.000"), cif_fc=Decimal("250.00"),
        available_value=Decimal("999.00"), available_quantity=Decimal("10.000"),
    )
    without_cif = _make_import_item(
        lic, item_report_masters["hs_code"], serial=2,
        quantity=Decimal("100.000"), cif_fc=Decimal("0.00"),
        available_value=Decimal("40.00"), available_quantity=Decimal("20.000"),
    )

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert response.status_code == 200
    rows = {row["id"]: row for row in response.json()["items"]}

    # 250.00 / 100 = 2.50 — uses cif_fc, ignores available_balance entirely.
    assert rows[with_cif.id]["unit_price"] == pytest.approx(2.5)
    # cif_fc is 0 (unusable) — falls back to 40.00 / 100 = 0.40.  The
    # report's quantity is the customs figure (Total − BOE Debited), not
    # the stale operational available_quantity column.
    assert rows[without_cif.id]["unit_price"] == pytest.approx(0.4)


@pytest.mark.django_db
def test_item_report_sorted_by_license_expiry_date_ascending(report_viewer_client, item_report_masters):
    """Business-report ordering: soonest-expiring license first."""
    lic_soon = _make_license("ITEM-REPORT-EXP-SOON", item_report_masters["parle"], expiry_days=10)
    lic_later = _make_license("ITEM-REPORT-EXP-LATER", item_report_masters["other"], expiry_days=200)
    item_later = _make_import_item(lic_later, item_report_masters["hs_code"], available_value=Decimal("500.00"))
    item_soon = _make_import_item(lic_soon, item_report_masters["hs_code"], available_value=Decimal("500.00"))

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert response.status_code == 200
    ids_in_order = [row["id"] for row in response.json()["items"]]

    assert ids_in_order.index(item_soon.id) < ids_in_order.index(item_later.id)


@pytest.mark.django_db
def test_item_report_available_items_includes_name_with_no_plan(report_viewer_client, item_report_masters):
    lic = _make_license("ITEM-REPORT-AVAIL", item_report_masters["parle"])
    _make_import_item(
        lic,
        item_report_masters["hs_code"],
        available_value=Decimal("500.00"),
        item_names=[item_report_masters["wheat"]],
    )
    # No LicenseItemPlan created anywhere for this item name.

    response = report_viewer_client.get(AVAILABLE_ITEMS_URL)
    assert response.status_code == 200
    names = {row["name"] for row in response.json()}
    assert item_report_masters["wheat"].name in names


@pytest.mark.django_db
def test_item_report_excel_export_renders_planning_split_sub_rows(
    report_viewer_client, item_report_masters
):
    """A manually-planned, multi-split import item must render one indented
    sub-row per split (Planning Item Name / Unit Price / Planned Qty /
    Planned CIF / "Split N" badge) in the Excel export — the same per-split
    breakdown license_balance_excel.py renders, sourced from the same
    plan_map_for_import_items() map (no second query / divergent source)."""
    from io import BytesIO

    from openpyxl import load_workbook

    lic = _make_license("ITEM-REPORT-SPLIT", item_report_masters["parle"])
    item = _make_import_item(
        lic,
        item_report_masters["hs_code"],
        available_value=Decimal("500.00"),
        item_names=[item_report_masters["wheat"]],
    )
    # Two manual plan lines splitting this import item across two planning
    # item names — mirrors how milk gets split into WPC / SWP in production.
    LicenseItemPlan.objects.create(
        license=lic,
        import_item=item,
        item_name=item_report_masters["wheat"],
        planned_quantity=Decimal("20.000"),
        unit_price=Decimal("5.00"),
        planned_cif_fc=Decimal("100.00"),
    )
    LicenseItemPlan.objects.create(
        license=lic,
        import_item=item,
        item_name=item_report_masters["milk"],
        planned_quantity=Decimal("10.000"),
        unit_price=Decimal("7.50"),
        planned_cif_fc=Decimal("75.00"),
    )

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0, "format": "excel"})
    assert response.status_code == 200, getattr(response, "data", response.content[:300])

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert "Item Report" in workbook.sheetnames
    ws = workbook["Item Report"]

    # Row 2 = the import item's own row; rows 3-4 = one sub-row per split.
    item_name_col, price_col, badge_col, qty_col, cif_col = 11, 10, 8, 15, 16
    split_row_1 = [ws.cell(row=3, column=c).value for c in
                    (badge_col, price_col, item_name_col, qty_col, cif_col)]
    split_row_2 = [ws.cell(row=4, column=c).value for c in
                    (badge_col, price_col, item_name_col, qty_col, cif_col)]

    assert split_row_1 == [
        "Split 1", "@ $5.00/unit", f"  └ {item_report_masters['wheat'].name}", 20.0, 100.0,
    ]
    assert split_row_2 == [
        "Split 2", "@ $7.50/unit", f"  └ {item_report_masters['milk'].name}", 10.0, 75.0,
    ]

    # The item's own row (row 2) keeps the aggregated Plan Qty / Plan CIF
    # totals across both splits — the per-split rows are additive detail,
    # not a replacement of the existing summary values.
    assert ws.cell(row=2, column=qty_col).value == pytest.approx(30.0)
    assert ws.cell(row=2, column=cif_col).value == pytest.approx(175.0)


@pytest.mark.django_db
def test_item_report_excel_export_merges_shared_description_rows(
    report_viewer_client, item_report_masters
):
    """
    Import items that share a description (the same `plan_group_key` group
    `plan_utilization_rows()` uses elsewhere) collapse into ONE Excel row
    per licence, with the Serial Number cell listing every merged serial
    and the split sub-rows carrying the UNION of every member's splits —
    not just the representative (lowest-serial) member's.

    The JSON `items` list is untouched (still one row per raw import item);
    only the Excel writer merges.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    lic = _make_license("ITEM-REPORT-MERGE", item_report_masters["parle"])
    item_23 = _make_import_item(
        lic, item_report_masters["hs_code"], serial=23,
        available_value=Decimal("200.00"), available_quantity=Decimal("20.000"),
    )
    item_23.description = "Refined Cane Sugar"
    item_23.save(update_fields=["description"])
    item_3 = _make_import_item(
        lic, item_report_masters["hs_code"], serial=3,
        available_value=Decimal("300.00"), available_quantity=Decimal("30.000"),
    )
    item_3.description = "refined cane sugar"
    item_3.save(update_fields=["description"])
    item_13 = _make_import_item(
        lic, item_report_masters["hs_code"], serial=13,
        available_value=Decimal("50.00"), available_quantity=Decimal("5.000"),
    )
    item_13.description = " REFINED CANE SUGAR "
    item_13.save(update_fields=["description"])

    # Manual plan saved against the group's representative (lowest serial —
    # item_3), the real convention `bulk_upsert`/`PlanningEditor.tsx` use.
    LicenseItemPlan.objects.create(
        license=lic, import_item=item_3, item_name=item_report_masters["wheat"],
        planned_quantity=Decimal("15.000"), unit_price=Decimal("2.00"),
        planned_cif_fc=Decimal("30.00"),
    )

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0, "format": "excel"})
    assert response.status_code == 200, getattr(response, "data", response.content[:300])

    # The JSON path (unaffected) still returns 3 raw rows.
    json_response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    assert len(json_response.json()["items"]) == 3

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    ws = workbook["Item Report"]

    serial_col, hsn_col, desc_col, avail_qty_col = 7, 9, 10, 12
    item_name_col, price_col, badge_col, qty_col, cif_col = 11, 10, 8, 15, 16

    # One merged row (not 3) for the group; comma-joined, ascending serials.
    assert ws.cell(row=2, column=serial_col).value == "3, 13, 23"
    assert ws.cell(row=2, column=desc_col).value == "refined cane sugar"
    assert ws.cell(row=2, column=hsn_col).value == item_report_masters["hs_code"].hs_code
    # Available Quantity summed across all 3 merged serials.
    assert ws.cell(row=2, column=avail_qty_col).value == pytest.approx(300.0)
    # Plan Qty/CIF aggregated across the group (only the representative had
    # a LicenseItemPlan row; the merge must not lose or double-count it).
    assert ws.cell(row=2, column=qty_col).value == pytest.approx(15.0)
    assert ws.cell(row=2, column=cif_col).value == pytest.approx(30.0)

    # Split sub-row (row 3) reflects the group's unioned splits.
    split_row = [ws.cell(row=3, column=c).value for c in
                 (badge_col, price_col, item_name_col, qty_col, cif_col)]
    assert split_row == [
        "Split 1", "@ $2.00/unit", f"  └ {item_report_masters['wheat'].name}", 15.0, 30.0,
    ]

    # No leftover rows for the other 2 merged serials.
    assert ws.cell(row=4, column=serial_col).value is None


@pytest.mark.django_db
def test_item_report_excel_export_single_sheet_with_totals_row(report_viewer_client, item_report_masters):
    """Restricted and non-restricted items used to land on separate sheets
    ("Restricted" / "Not Restricted") — the export is now a single sheet
    (matching the View, which never had that split), with a bold totals
    row summing Available Quantity / Available Balance / Plan Qty / Plan
    CIF beneath the data."""
    from io import BytesIO

    from openpyxl import load_workbook

    lic_a = _make_license("ITEM-REPORT-ONESHEET-A", item_report_masters["parle"])
    lic_b = _make_license("ITEM-REPORT-ONESHEET-B", item_report_masters["other"])
    restricted = _make_import_item(
        lic_a, item_report_masters["hs_code"], condition_type="AU",
        available_value=Decimal("500.00"), available_quantity=Decimal("50.000"),
    )
    not_restricted = _make_import_item(
        lic_b, item_report_masters["hs_code"],
        available_value=Decimal("300.00"), available_quantity=Decimal("20.000"),
    )
    LicenseItemPlan.objects.create(
        license=lic_b, import_item=not_restricted, item_name=item_report_masters["wheat"],
        planned_quantity=Decimal("5.000"), unit_price=Decimal("1.00"), planned_cif_fc=Decimal("5.00"),
    )

    response = report_viewer_client.get(REPORT_URL, {"min_balance": 0, "format": "excel"})
    assert response.status_code == 200, getattr(response, "data", response.content[:300])

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Item Report"]
    ws = workbook["Item Report"]

    # Header matches the View's exact column order.
    headers = [c.value for c in ws[1]]
    assert headers == [
        'Sr No', 'License No', 'License Date', 'License Expiry Date', 'Ledger Date', 'Exporter Name',
        'Serial Number', 'Condition', 'HSN Code', 'Product Description', 'Item Name',
        'Available Quantity', 'Unit Price', 'Available Balance', 'Plan Qty', 'Plan CIF',
        'Balance CIF', 'Is Restricted', 'Notes', 'Condition Sheet', 'Transfer Status',
    ]

    # Both items (one restricted, one not) present on the one sheet.
    license_nos = {ws.cell(row=r, column=2).value for r in range(2, ws.max_row)}
    license_nos.discard(None)
    assert license_nos == {lic_a.license_number, lic_b.license_number}

    # Totals row: bold "TOTAL" label, correct sums in the 4 numeric columns.
    total_row = ws.max_row
    assert ws.cell(row=total_row, column=1).value == "TOTAL"
    assert ws.cell(row=total_row, column=1).font.bold is True
    assert ws.cell(row=total_row, column=12).value == pytest.approx(200.0)  # Actual Available: 100 + 100
    assert ws.cell(row=total_row, column=14).value == pytest.approx(800.0)  # Available Balance: 500 + 300
    assert ws.cell(row=total_row, column=15).value == pytest.approx(5.0)    # Plan Qty
    assert ws.cell(row=total_row, column=16).value == pytest.approx(5.0)    # Plan CIF
    # Unit Price (13) and the kept-extra columns (17-21) are never totaled.
    assert ws.cell(row=total_row, column=13).value is None

    # Freeze header + AutoFilter over the data range (excluding totals row).
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == f"A1:U{total_row - 1}"


@pytest.mark.django_db
def test_item_report_generate_report_called_exactly_once_per_request(report_viewer_client, item_report_masters):
    """Phase 2B.1 regression: the view must compute the Display Dataset
    exactly once and thread the same dict to both the JSON response and the
    Excel exporter — `export_to_excel` must never call `generate_report`
    itself. Wraps the real method (not a stub) so this also still exercises
    genuine report generation, just counts the calls."""
    from unittest.mock import patch

    from apps.license.views.item_report import ItemReportView

    lic = _make_license("ITEM-REPORT-ONCE-A", item_report_masters["parle"])
    _make_import_item(lic, item_report_masters["hs_code"], available_value=Decimal("500.00"))

    with patch.object(ItemReportView, "generate_report", autospec=True, side_effect=ItemReportView.generate_report) as mocked:
        json_response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
        assert json_response.status_code == 200
        assert mocked.call_count == 1

    with patch.object(ItemReportView, "generate_report", autospec=True, side_effect=ItemReportView.generate_report) as mocked:
        excel_response = report_viewer_client.get(REPORT_URL, {"min_balance": 0, "format": "excel"})
        assert excel_response.status_code == 200
        assert mocked.call_count == 1


@pytest.mark.django_db
def test_item_report_excel_rows_match_json_rows(report_viewer_client, item_report_masters):
    """Phase 2B.1 regression: JSON and Excel must render the exact same
    per-item figures for the same request — using the shared
    apps.core.tests.report_assertions helper introduced in Phase 2A."""
    from io import BytesIO

    from openpyxl import load_workbook

    from apps.core.tests.report_assertions import assert_excel_rows_match_json_rows

    lic_a = _make_license("ITEM-REPORT-MATCH-A", item_report_masters["parle"])
    lic_b = _make_license("ITEM-REPORT-MATCH-B", item_report_masters["other"])
    _make_import_item(lic_a, item_report_masters["hs_code"], available_value=Decimal("500.00"), available_quantity=Decimal("50.000"))
    _make_import_item(lic_b, item_report_masters["hs_code"], available_value=Decimal("300.00"), available_quantity=Decimal("20.000"))

    json_response = report_viewer_client.get(REPORT_URL, {"min_balance": 0})
    json_items = json_response.json()["items"]
    assert len(json_items) == 2

    excel_response = report_viewer_client.get(REPORT_URL, {"min_balance": 0, "format": "excel"})
    workbook = load_workbook(BytesIO(excel_response.content), data_only=True)

    assert_excel_rows_match_json_rows(
        workbook, json_items, "Item Report",
        column_map={"available_quantity": 12, "available_balance": 14},
        header_row=1, key_field="license_number", key_column=2,
    )
