# bill_of_entry/views_export.py
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from rest_framework.decorators import action

from apps.core.utils.pdf_utils import create_pdf_exporter

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def add_grouped_export_action(viewset_class):
    """
    Decorator to add grouped export functionality to BillOfEntryViewSet
    """

    @action(detail=False, methods=['get'], url_path='export')
    def export_bill_of_entries(self, request):
        """
        Export bill of entries grouped by company.
        URL: /api/bill-of-entries/export/?_export=pdf

        Query params:
            - _export: 'pdf' or 'xlsx' (default: pdf)
            - company: filter by company ID
            - bill_of_entry_date_after: filter by date
            - bill_of_entry_date_before: filter by date
        """
        export_format = request.query_params.get('_export', 'pdf').lower()

        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Filter only BOEs with item details and order by date first, then company, item, port
        queryset = queryset.filter(
            item_details__isnull=False
        ).distinct().order_by(
            'bill_of_entry_date', 'company__name', 'product_name', 'port__code'
        ).prefetch_related(
            'item_details__sr_number__license',
            'item_details__sr_number__hs_code',
            'company',
            'port'
        )

        if export_format == 'pdf':
            return self._export_grouped_pdf(queryset)
        elif export_format == 'xlsx':
            return self._export_grouped_xlsx(queryset)
        elif export_format == 'port_xlsx':
            return self._export_port_xlsx(queryset)
        else:
            return HttpResponse("Invalid export format. Use 'pdf', 'xlsx', or 'port_xlsx'.", status=400)

    def _export_grouped_pdf(self, queryset):
        """Export grouped bill of entries to PDF grouped by Company → Item → Port"""
        from reportlab.lib.units import inch
        from reportlab.platypus import TableStyle, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.lib import colors

        pdf_exporter = create_pdf_exporter(
            title="Bill of Entry Report",
            filename_prefix="Bill_of_Entries",
            orientation='landscape'
        )

        if not pdf_exporter:
            return HttpResponse("PDF export not available", status=500)

        # Group data
        grouped_data = self._group_boe(queryset)

        # Active exchange rate for the mini-table in the header
        from apps.core.models import ExchangeRateModel
        active_rate = ExchangeRateModel.get_active_rate()

        # Calculate grand totals
        total_inr = sum(boe['total_inr']
                        for company in grouped_data.values()
                        for license_dict in company.values()
                        for product in license_dict.values()
                        for port in product.values()
                        for boe in port)
        total_fc = sum(boe['total_fc']
                       for company in grouped_data.values()
                       for license_dict in company.values()
                       for product in license_dict.values()
                       for port in product.values()
                       for boe in port)

        def shorten_exporter(name, max_words=2):
            """Return first N words of an exporter name followed by '…'."""
            if not name or name == '--':
                return name or '--'
            words = name.split()
            if len(words) <= max_words:
                return name
            return ' '.join(words[:max_words]) + '…'  # …

        # Create PDF response - inline display for new tab
        response = HttpResponse(content_type='application/pdf')
        today = datetime.now().strftime('%d-%m-%Y')
        filename = f'BOE Report - {today}.pdf'
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        buffer = BytesIO()
        doc = pdf_exporter.create_document(buffer)

        elements = []

        # Item sub-header style
        item_style = ParagraphStyle(
            'ItemHeader',
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=HexColor('#1e40af'),
            spaceAfter=2,
            spaceBefore=3,
            leftIndent=0,
        )

        # Header: title + totals on left, exchange rate mini-table on right
        subtitle_text = (f"Total CIF (FC): ${pdf_exporter.format_number(total_fc)} | "
                         f"Total CIF (INR): {pdf_exporter.format_number(total_inr)}")

        if active_rate:
            from reportlab.platypus import Table as RLTable
            from reportlab.lib.styles import ParagraphStyle as PSArg
            from reportlab.lib.enums import TA_CENTER as TA_C

            exch_data = [
                ['Exch. Rate', active_rate.date.strftime('%d-%m-%Y')],
                ['USD', str(active_rate.usd)],
                ['EUR', str(active_rate.euro)],
                ['GBP', str(active_rate.pound_sterling)],
                ['CNY', str(active_rate.chinese_yuan)],
            ]
            exch_table = RLTable(exch_data, colWidths=[0.7 * inch, 0.8 * inch])
            exch_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), pdf_exporter.HEADER_BG),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            title_para = Paragraph("Bill of Entry Report", pdf_exporter.title_style)
            subtitle_para = Paragraph(subtitle_text, pdf_exporter.subtitle_style)
            ts_style = PSArg('TS', parent=pdf_exporter.styles['Normal'],
                             fontSize=9, textColor=colors.grey, alignment=TA_C, spaceAfter=10)
            ts_para = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", ts_style)

            header_layout = RLTable([[title_para, exch_table]], colWidths=[8.85 * inch, 1.5 * inch])
            header_layout.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ]))
            elements.append(header_layout)
            elements.append(subtitle_para)
            elements.append(ts_para)
        else:
            pdf_exporter.add_title(elements, subtitle=subtitle_text)

        # 18 columns — full landscape A4 (~11.35" usable after 0.15" margins).
        # Qty(KGS) and Value($) restored.  Column widths are set so that
        # Port, Lic.Date, Lic.Port, BOE $., BOE CIF never split across 2 lines.
        col_widths = [
            0.25 * inch,  # Sr No
            0.78 * inch,  # BOE Number   (+0.20)
            0.75 * inch,  # BOE Date     (+0.10)
            0.54 * inch,  # Port
            0.50 * inch,  # Qty (KGS)
            0.45 * inch,  # Unit Price ($)
            0.73 * inch,  # Value ($)    (+0.08)
            0.45 * inch,  # Exchange Rate
            0.93 * inch,  # Item Name
            0.50 * inch,  # Invoice
            0.90 * inch,  # Exporter — 0.90" = 60pt usable, "PGP GLASS…" ≈ 51pt, fits 1 line
            0.98 * inch,  # License No.  (+0.20)
            0.65 * inch,  # Lic. Date
            0.54 * inch,  # Lic. Port
            0.28 * inch,  # Item Sr.
            0.50 * inch,  # BOE Qty.
            0.70 * inch,  # BOE $.       (+0.12)
            0.92 * inch,  # BOE CIF      (+0.10)
        ]  # total = 11.35"

        table_header = [
            'Sr\nNo', 'BOE\nNumber', 'BOE\nDate', 'Port',
            'Qty\n(KGS)', 'Unit\nPrice ($)', 'Value\n($)', 'Exch.\nRate',
            'Item\nName', 'Invoice',
            'Exporter', 'License\nNo.', 'Lic. Date', 'Lic. Port', 'Item\nSr.',
            'BOE\nQty.', 'BOE $.', 'BOE CIF'
        ]

        # Process each company (sorted alphabetically, None values last)
        for company_name, license_dict in sorted(grouped_data.items(), key=lambda x: (x[0] is None, x[0] or '')):
            pdf_exporter.add_company_header(elements, company_name)

            company_total_qty = 0
            company_total_inr = 0
            company_total_fc = 0

            # Merge license_serial level → group by product → port (same as xlsx)
            merged_products = defaultdict(lambda: defaultdict(list))
            for license_serial, products_dict in license_dict.items():
                for product_name, ports_dict in products_dict.items():
                    for port_code, boe_list in ports_dict.items():
                        merged_products[product_name][port_code].extend(boe_list)

            # Process each item (product) within this company
            for product_name in sorted(merged_products.keys()):
                ports_dict = merged_products[product_name]

                elements.append(Paragraph(f"Item: {product_name}", item_style))

                table_data = [table_header]
                sr_no = 1

                for port_code in sorted(ports_dict.keys()):
                    port_total_qty = 0
                    port_total_value = 0
                    port_total_inr = 0

                    for boe in ports_dict[port_code]:
                        if boe['license_details']:
                            first_detail = boe['license_details'][0]
                            unit_price = (
                                boe['total_fc'] / boe['total_quantity']
                                if boe['total_quantity'] > 0 else 0
                            )

                            # Main BOE row (first licence detail) — 18 columns
                            table_data.append([
                                str(sr_no),
                                boe['boe_number'],
                                boe['boe_date'],
                                port_code,
                                pdf_exporter.format_number(boe['total_quantity'], decimals=0),
                                pdf_exporter.format_number(unit_price),
                                pdf_exporter.format_number(boe['total_fc']),
                                pdf_exporter.format_number(boe['exchange_rate']),
                                boe['product_name'],
                                boe['invoice_no'],
                                shorten_exporter(first_detail['exporter_name']),
                                first_detail['license_no'],
                                first_detail['license_date'],
                                first_detail['license_port'],
                                first_detail['item_sr_no'],
                                pdf_exporter.format_number(first_detail['qty'], decimals=0),
                                pdf_exporter.format_number(first_detail['cif_fc']),
                                pdf_exporter.format_number(first_detail['cif_inr'])
                            ])

                            # Additional licence detail rows — 18 columns (10 empty leading)
                            for detail in boe['license_details'][1:]:
                                table_data.append([
                                    '', '', '', '', '', '', '', '', '', '',
                                    shorten_exporter(detail['exporter_name']),
                                    detail['license_no'],
                                    detail['license_date'],
                                    detail['license_port'],
                                    detail['item_sr_no'],
                                    pdf_exporter.format_number(detail['qty'], decimals=0),
                                    pdf_exporter.format_number(detail['cif_fc']),
                                    pdf_exporter.format_number(detail['cif_inr'])
                                ])

                        sr_no += 1
                        port_total_qty += boe['total_quantity']
                        port_total_value += boe['total_fc']
                        port_total_inr += boe['total_inr']

                    # Port total row — 18 cols; totals in Qty, Value$, and BOE CIF cols
                    table_data.append([
                        '',
                        f'Port Total ({port_code})',
                        '', '',
                        pdf_exporter.format_number(port_total_qty, decimals=0),   # Qty
                        '',
                        pdf_exporter.format_number(port_total_value),              # Value $
                        '', '', '', '', '', '', '', '', '', '',
                        pdf_exporter.format_number(port_total_inr)                 # BOE CIF
                    ])

                    company_total_qty += port_total_qty
                    company_total_inr += port_total_inr
                    company_total_fc += port_total_value

                # col 10 = Exporter — left-align for readable name wrapping
                table = pdf_exporter.create_table(table_data, col_widths=col_widths, repeating_rows=1, left_align_cols={10})

                # Right-align numeric cols (0-indexed) for 18-col layout:
                # Qty(4), UnitPrice(5), Value$(6), ExchRate(7), BOEQty(15), BOE$(16), BOECIF(17)
                additional_styles = []
                for col_idx in [4, 5, 6, 7, 15, 16, 17]:
                    additional_styles.append(
                        ('ALIGN', (col_idx, 1), (col_idx, len(table_data) - 1), 'RIGHT')
                    )

                # Port Total rows — bold + subtle header background
                for row_idx, row in enumerate(table_data):
                    if row[1] and 'Port Total' in str(row[1]):
                        additional_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                        additional_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), pdf_exporter.HEADER_BG))

                table.setStyle(TableStyle(additional_styles))
                elements.append(table)
                pdf_exporter.add_spacer(elements, 0.15)

            # Company grand total
            summary_data = [[
                'Grand Total:',
                f"Qty: {pdf_exporter.format_number(company_total_qty, decimals=0)} KGS",
                f"CIF (FC): ${pdf_exporter.format_number(company_total_fc)}",
                f"CIF (INR): {pdf_exporter.format_number(company_total_inr)}"
            ]]
            summary_table = pdf_exporter.create_summary_table(summary_data)
            elements.append(summary_table)
            pdf_exporter.add_spacer(elements, 0.3)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    def _export_grouped_xlsx(self, queryset):
        """Export grouped bill of entries to Excel"""
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("Excel export not available", status=500)

        # Group data
        grouped_data = self._group_boe(queryset)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bill of Entries"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # Title
        ws.merge_cells(f'A{row}:R{row}')
        cell = ws[f'A{row}']
        cell.value = "Bill of Entry Report"
        cell.font = Font(bold=True, size=16, color="1e3a8a")
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Total and date
        total_inr = sum(boe['total_inr']
                        for company in grouped_data.values()
                        for license_dict in company.values()
                        for product in license_dict.values()
                        for port in product.values()
                        for boe in port)
        total_fc = sum(boe['total_fc']
                       for company in grouped_data.values()
                       for license_dict in company.values()
                       for product in license_dict.values()
                       for port in product.values()
                       for boe in port)
        ws.merge_cells(f'A{row}:R{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total INR: ₹{total_inr:,.2f}    Total FC: ${total_fc:,.2f}    {datetime.now().strftime('%d-%m-%Y')}"
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Process each company (sorted alphabetically, None values last)
        for company_name, license_dict in sorted(grouped_data.items(), key=lambda x: (x[0] is None, x[0] or '')):
            # Company header
            ws.merge_cells(f'A{row}:R{row}')
            cell = ws[f'A{row}']
            cell.value = company_name
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

            # Track company totals
            company_total_qty = 0
            company_total_inr = 0
            company_total_fc = 0

            # Merge all products into a single table (group by product name only, ignore license serial)
            merged_products = defaultdict(lambda: defaultdict(list))
            for license_serial, products_dict in license_dict.items():
                for product_name, ports_dict in products_dict.items():
                    for port_code, boe_list in ports_dict.items():
                        merged_products[product_name][port_code].extend(boe_list)

            # Process each product (Item) within company
            for product_name, ports_dict in merged_products.items():
                # Product subheader
                ws.merge_cells(f'A{row}:S{row}')
                cell = ws[f'A{row}']
                cell.value = f"Item: {product_name}"
                cell.font = Font(bold=True, size=14, color="3b82f6")
                row += 1

                # Table headers (18 columns - added Exporter and Exchange Rate)
                headers = ['Sr No', 'BOE Number', 'BOE Date', 'Port', 'Quantity (KGS)', 'Unit Price ($)', 'Value ($)',
                           'Exchange Rate', 'Total CIF INR', 'Item Name', 'Invoice', 'Exporter', 'License No.', 'License Date',
                           'License Port', 'Item Sr.', 'BOE Qty.', 'BOE $.', 'BOE CIF']

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                    cell.border = border
                row += 1

                sr_no = 1
                product_total_qty = 0
                product_total_value = 0
                product_total_inr = 0

                # Process each port within product
                for port_code, boe_list in ports_dict.items():
                    for boe in boe_list:
                        if boe['license_details']:
                            first_detail = boe['license_details'][0]
                            unit_price = boe['total_fc'] / boe['total_quantity'] if boe['total_quantity'] > 0 else 0

                            # Main BOE row (18 columns - added Exporter)
                            data = [
                                sr_no,
                                boe['boe_number'],
                                boe['boe_date'],
                                port_code,
                                int(boe['total_quantity']),
                                round(unit_price, 2),
                                round(boe['total_fc'], 2),
                                round(boe['exchange_rate'], 2),  # Exchange Rate
                                round(boe['total_inr'], 2),  # Total CIF INR
                                product_name,
                                boe['invoice_no'],
                                first_detail['exporter_name'],  # Exporter
                                first_detail['license_no'],     # License No.
                                first_detail['license_date'],
                                first_detail['license_port'],
                                first_detail['item_sr_no'],
                                int(first_detail['qty']),
                                round(first_detail['cif_fc'], 2),
                                round(first_detail['cif_inr'], 2)
                            ]

                            for col_idx, value in enumerate(data, 1):
                                cell = ws.cell(row=row, column=col_idx, value=value)
                                cell.border = border
                                cell.font = Font(size=12)
                                if col_idx in [5, 6, 7, 8, 9, 17, 18, 19]:  # Number columns
                                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                                    if col_idx in [6, 7, 18]:  # USD columns
                                        cell.number_format = '#,##0.00'
                                    elif col_idx in [9, 19]:  # INR columns
                                        cell.number_format = '₹#,##0.00'
                                    elif col_idx == 8:  # Exchange Rate column
                                        cell.number_format = '#,##0.00'
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                            # Allow 2 lines by default; grow to 3 lines if the
                            # longest cell content suggests more wrapping is needed.
                            longest = max(
                                (len(str(v)) for v in data if v not in (None, '')),
                                default=0,
                            )
                            target_lines = 3 if longest > 30 else 2
                            ws.row_dimensions[row].height = max(28, target_lines * 16)

                            row += 1

                            # Additional license detail rows
                            for detail in boe['license_details'][1:]:
                                data = [
                                    '', '', '', '', '', '', '', '', '',  # Empty main BOE columns
                                    '', '',  # Item Name, Invoice
                                    detail['exporter_name'],  # Exporter
                                    detail['license_no'],     # License No.
                                    detail['license_date'],
                                    detail['license_port'],
                                    detail['item_sr_no'],
                                    int(detail['qty']),
                                    round(detail['cif_fc'], 2),
                                    round(detail['cif_inr'], 2)
                                ]

                                for col_idx, value in enumerate(data, 1):
                                    cell = ws.cell(row=row, column=col_idx, value=value)
                                    cell.border = border
                                    cell.font = Font(size=12)
                                    if col_idx in [17, 18, 19]:
                                        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                                        if col_idx in [18, 19]:
                                            cell.number_format = '#,##0.00'
                                    else:
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                                longest_detail = max(
                                    (len(str(v)) for v in data if v not in (None, '')),
                                    default=0,
                                )
                                target_lines_d = 3 if longest_detail > 30 else 2
                                ws.row_dimensions[row].height = max(28, target_lines_d * 16)

                                row += 1

                        sr_no += 1
                        product_total_qty += boe['total_quantity']
                        product_total_value += boe['total_fc']
                        product_total_inr += boe['total_inr']
                        company_total_qty += boe['total_quantity']
                        company_total_inr += boe['total_inr']
                        company_total_fc += boe['total_fc']

                # Product totals
                ws.cell(row=row, column=3, value="Total").font = Font(bold=True)
                ws.cell(row=row, column=5, value=int(product_total_qty)).font = Font(bold=True)
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=7, value=round(product_total_value, 2)).font = Font(bold=True)
                ws.cell(row=row, column=7).number_format = '#,##0.00'
                ws.cell(row=row, column=9, value=round(product_total_inr, 2)).font = Font(bold=True)
                ws.cell(row=row, column=9).number_format = '₹#,##0.00'
                row += 2

            # Add grand total after all products
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws.cell(row=row, column=1, value="Grand Total:")
            cell.font = Font(bold=True, size=14, color="1e3a8a")
            cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=row, column=5, value=int(company_total_qty)).font = Font(bold=True, size=14)
            ws.cell(row=row, column=5).fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=5).number_format = '#,##0'

            ws.cell(row=row, column=7, value=round(company_total_fc, 2)).font = Font(bold=True, size=14)
            ws.cell(row=row, column=7).fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=7).number_format = '#,##0.00'

            ws.merge_cells(f'H{row}:I{row}')
            cell = ws.cell(row=row, column=8, value=f"CIF INR: ₹{company_total_inr:,.2f}")
            cell.font = Font(bold=True, size=14, color="1e3a8a")
            cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 3

        # Size columns with a friendly cap so wide cells wrap to 2–3 lines
        # instead of forcing the sheet to be enormously wide.
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            if column_letter:
                # Floor 10 (so short headers like "Sr No" stay readable),
                # cap 22 (so long text wraps onto 2–3 lines instead of stretching).
                adjusted_width = max(10, min(max_length + 2, 22))
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        today = datetime.now().strftime('%d-%m-%Y')
        response['Content-Disposition'] = f'attachment; filename="BOE Report - {today}.xlsx"'
        wb.save(response)

        return response

    def _export_port_xlsx(self, queryset):
        """Export simplified flat BOE list: single sheet, one header, no grouping."""
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("Excel export not available", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Port BOE List"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title row
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Port BOE Report — {datetime.now().strftime('%d-%m-%Y')}"
        title_cell.font = Font(bold=True, size=14, color="1e3a8a")
        title_cell.alignment = Alignment(horizontal='center')

        # Single header row
        headers = ['BOE Number', 'BOE Date', 'Port', 'Company', 'Quantity (KGS)', 'Total CIF INR', 'Item Name']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        row = 3
        for boe in queryset:
            boe_number = boe.bill_of_entry_number or '--'
            boe_date = boe.bill_of_entry_date.strftime('%d-%m-%Y') if boe.bill_of_entry_date else '--'
            port_code = boe.port.code if boe.port else '--'
            company_name = boe.company.name if boe.company else '--'
            total_quantity = float(boe.get_total_quantity or 0)
            total_inr = float(boe.get_total_inr or 0)
            product_name = (boe.product_name or '').strip().upper() or '--'

            data = [boe_number, boe_date, port_code, company_name, int(total_quantity), round(total_inr, 2), product_name]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                if col_idx == 5:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '₹#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        today = datetime.now().strftime('%d-%m-%Y')
        response['Content-Disposition'] = f'attachment; filename="Port BOE Report - {today}.xlsx"'
        wb.save(response)
        return response

    def _group_boe(self, queryset):
        """Group bill of entries by company → license_serial_number → product_name → port"""
        # Structure: {company: {license_serial: {product_name: {port: [boe_list]}}}}
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))

        for boe in queryset:
            company_name = boe.company.name if boe.company else "Unknown"
            port_code = boe.port.code if boe.port else "Unknown Port"
            # Normalize product name: uppercase, strip spaces for grouping
            raw_product_name = boe.product_name or "Unknown Product"
            product_name = raw_product_name.strip().upper()

            # Get first license serial number from item details
            first_item = boe.item_details.first()
            license_serial = "Unknown License"
            if first_item and first_item.sr_number:
                license_obj = first_item.sr_number.license if first_item.sr_number else None
                if license_obj:
                    license_serial = f"{license_obj.license_number} (Sr. {first_item.sr_number.serial_number})"

            # Collect license details for this BOE
            license_details = []
            for detail in boe.item_details.all():
                license_obj = detail.sr_number.license if detail.sr_number else None

                # Store exporter name and license number separately
                if license_obj:
                    exporter_name = license_obj.exporter.name if license_obj.exporter else '--'
                    license_number = license_obj.license_number
                else:
                    exporter_name = '--'
                    license_number = '--'

                license_details.append({
                    'exporter_name': exporter_name,
                    'license_no': license_number,
                    'license_port': license_obj.port.code if (license_obj and license_obj.port) else '--',
                    'license_date': license_obj.license_date.strftime('%d-%m-%Y') if license_obj else '--',
                    'item_sr_no': str(detail.sr_number.serial_number) if detail.sr_number else '--',
                    'qty': float(detail.qty or 0),
                    'cif_fc': float(detail.cif_fc or 0),
                    'cif_inr': float(detail.cif_inr or 0)
                })

            # Calculate exchange rate: use boe.exchange_rate if exists, otherwise calculate from total_inr / total_fc
            total_fc = float(boe.get_total_fc or 0)
            total_inr = float(boe.get_total_inr or 0)

            if boe.exchange_rate and float(boe.exchange_rate) > 0:
                exchange_rate = float(boe.exchange_rate)
            elif total_fc > 0:
                exchange_rate = total_inr / total_fc
            else:
                exchange_rate = 0

            boe_data = {
                'boe_number': boe.bill_of_entry_number or '--',
                'boe_date': boe.bill_of_entry_date.strftime('%d-%m-%Y') if boe.bill_of_entry_date else '--',
                'port': port_code,
                'product_name': product_name,
                'invoice_no': boe.invoice_no or '--',
                'total_quantity': float(boe.get_total_quantity or 0),
                'total_fc': total_fc,
                'total_inr': total_inr,
                'exchange_rate': exchange_rate,  # Add exchange rate
                'license_details': license_details,  # List of license entries
            }

            grouped_data[company_name][license_serial][product_name][port_code].append(boe_data)

        return grouped_data

    # Add methods to the viewset class
    viewset_class.export_bill_of_entries = export_bill_of_entries
    viewset_class._export_grouped_pdf = _export_grouped_pdf
    viewset_class._export_grouped_xlsx = _export_grouped_xlsx
    viewset_class._export_port_xlsx = _export_port_xlsx
    viewset_class._group_boe = _group_boe

    return viewset_class
