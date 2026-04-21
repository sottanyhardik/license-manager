# license/views/license.py
from rest_framework.decorators import action
from rest_framework.response import Response

from accounts.permissions import LicensePermission
from core.constants import LICENCE_PURCHASE_CHOICES, LICENCE_PURCHASE_CHOICES_ACTIVE, SCHEME_CODE_CHOICES, \
    NOTIFICATION_NORM_CHOICES, UNIT_CHOICES, \
    CURRENCY_CHOICES
from core.filters import CombinedFilterBackend, EnhancedSearchFilter, AdvancedOrderingFilter
from core.filtersets import LicenseFilterSet
from core.views.master_view import MasterViewSet
from license.models import LicenseDetailsModel
from license.serializers import LicenseDetailsSerializer, LicenseExportItemSerializer, LicenseImportItemSerializer, \
    LicenseDocumentSerializer
from license.views.active_dfia_report import add_active_dfia_report_action
from license.views.license_report import add_license_report_action


# Helper function to get default purchase status IDs from codes
def get_default_purchase_status_ids():
    """Convert default purchase status codes to IDs"""
    from core.models import PurchaseStatus
    default_codes = ['GE', 'MI', 'CO']  # GE Purchase, GE Operating, Conversion
    ids = list(PurchaseStatus.objects.filter(code__in=default_codes).values_list('id', flat=True))
    return ','.join(map(str, ids)) if ids else ''

# Nested field definitions for LicenseDetails
license_nested_field_defs = {
    "export_license": [
        {"name": "id", "type": "text", "label": "ID", "read_only": True, "show_in_list": False},
        {"name": "description", "type": "text", "label": "Description"},
        {"name": "norm_class", "type": "fk", "label": "Norm Class", "fk_endpoint": "/masters/sion-classes/",
         "label_field": "norm_class", "display_field": "norm_class_label"},
        {"name": "start_serial_number", "type": "number", "label": "Start Serial Number", "default": 0},
        {"name": "net_quantity", "type": "number", "label": "Net Quantity", "default": 0},
        {"name": "fob_inr", "type": "number", "label": "FOB (INR)", "default": 0},
        {"name": "currency", "type": "select", "label": "Currency", "choices": list(CURRENCY_CHOICES),
         "default": "usd"},
        {"name": "cif_fc", "type": "number", "label": "CIF (FC)", "default": 0},
        {"name": "cif_inr", "type": "number", "label": "CIF (INR)", "default": 0},
    ],
    "import_license": [
        {"name": "id", "type": "text", "label": "ID", "read_only": True, "show_in_list": False},
        {"name": "serial_number", "type": "number", "label": "Serial Number", "default": 0},
        {"name": "hs_code", "type": "fk", "label": "HS Code", "fk_endpoint": "/masters/hs-codes/",
         "label_field": "hs_code", "display_field": "hs_code_label"},
        {"name": "description", "type": "text", "label": "Description"},
        {"name": "items", "type": "fk_multi", "label": "Items", "fk_endpoint": "/masters/item-names/",
         "label_field": "name", "show_in_list": False},
        {"name": "quantity", "type": "number", "label": "Quantity", "default": 0},
        {"name": "unit", "type": "select", "label": "Unit", "choices": list(UNIT_CHOICES), "default": "kg"},
        {"name": "cif_fc", "type": "number", "label": "CIF (FC)", "default": 0},
        {"name": "cif_inr", "type": "number", "label": "CIF (INR)", "default": 0},
        {"name": "is_restricted", "type": "boolean", "label": "Is Restricted"},
    ],
    "license_documents": [
        {"name": "id", "type": "text", "label": "ID", "read_only": True, "show_in_list": False},
        {
            "name": "type",
            "type": "select",
            "label": "Document Type",
            "choices": [
                {"value": "LICENSE COPY", "label": "LICENSE COPY"},
                {"value": "TRANSFER LETTER", "label": "TRANSFER LETTER"},
                {"value": "OTHER", "label": "OTHER"}
            ]
        },
        {"name": "file", "type": "file", "label": "File"},
    ],
}

# Create base viewset
_LicenseDetailsViewSetBase = MasterViewSet.create_viewset(
    LicenseDetailsModel,
    LicenseDetailsSerializer,
    config={
        "search": ["license_number", "file_number", "exporter__name"],
        "filter": {
            "exporter": {"type": "fk", "fk_endpoint": "/masters/companies/", "label_field": "name"},
            "exclude_exporter": {"type": "exclude_fk", "fk_endpoint": "/masters/companies/", "label_field": "name",
                                 "filter_field": "exporter"},
            "port": {"type": "fk", "fk_endpoint": "/masters/ports/", "label_field": "name"},
            "exclude_port": {"type": "exclude_fk", "fk_endpoint": "/masters/ports/", "label_field": "name",
                             "filter_field": "port"},
            "export_license__norm_class": {"type": "fk", "fk_endpoint": "/masters/sion-classes/",
                                           "label_field": "norm_class"},
            "notification_number": {"type": "choice", "choices": list(NOTIFICATION_NORM_CHOICES)},
            "purchase_status": {"type": "fk", "fk_endpoint": "/masters/purchase-statuses/", "label_field": "label", "filter_params": {"is_active": "true"}},
            "license_date": {"type": "date_range"},
            "license_expiry_date": {"type": "date_range"},
            "balance_cif": {"type": "range"},
            "is_expired": {"type": "exact"},
            "is_null": {"type": "exact"},
        },
        "default_filters": {
            "is_expired": "False",
            "is_null": "False",
        },
        "list_display": [
            "license_number",
            "license_date",
            "license_expiry_date",
            "exporter__name",
            "port__name",
            "purchase_status_label",
            "balance_cif",
            "latest_transfer",
            "get_norm_class"
        ],
        "form_fields": [
            "scheme_code",
            "notification_number",
            "license_number",
            "license_date",
            "license_expiry_date",
            "port",
            "registration_number",
            "registration_date",
            "file_number",
            "exporter",
            "purchase_status",
            "condition_sheet",
        ],
        "ordering": ["-license_expiry_date", "license_number"],
        "nested_field_defs": license_nested_field_defs,
        "nested_list_display": {
            "export_license": ["norm_class_label", "fob_inr", "cif_fc", "cif_inr"],
            "import_license": ["serial_number", "hs_code_label", "description", "quantity", "unit", "cif_fc",
                               "cif_inr", "allotted_quantity", "allotted_value", "debited_quantity", "debited_value",
                               "available_quantity", "available_value"],
            "license_documents": ["type", "file"],
        },
        "field_meta": {
            "exporter": {
                "type": "fk",
                "fk_endpoint": "/masters/companies/",
                "label_field": "name"
            },
            "port": {
                "type": "fk",
                "fk_endpoint": "/masters/ports/",
                "label_field": "name"
            },
            "current_owner": {
                "type": "fk",
                "fk_endpoint": "/masters/companies/",
                "label_field": "name"
            },
            "purchase_status": {
                "type": "fk",
                "fk_endpoint": "/masters/purchase-statuses/",
                "label_field": "label",
                "filter_params": {"is_active": "true"}
            },
            "scheme_code": {
                "type": "select",
                "choices": list(SCHEME_CODE_CHOICES),
                "default": "26"
            },
            "notification_number": {
                "type": "select",
                "choices": list(NOTIFICATION_NORM_CHOICES),
                "default": "025/2023"
            },
        }
    }
)


# Extend the base viewset to add custom filter logic for is_expired and is_null
class LicenseDetailsViewSet(_LicenseDetailsViewSetBase):
    """
    Custom license viewset with special handling for is_expired and is_null filters.

    Filter Logic:
    - is_expired: True when license_expiry_date < today, False when license_expiry_date >= today
    - is_null: True when balance_cif < 200, False when balance_cif >= 200
    """
    permission_classes = [LicensePermission]
    lookup_value_regex = '[^/]+'  # Allow both numbers and strings

    # Apply advanced filter backends
    filterset_class = LicenseFilterSet
    filter_backends = [CombinedFilterBackend, EnhancedSearchFilter, AdvancedOrderingFilter]
    search_fields = ['license_number', 'file_number', 'exporter__name']
    ordering_fields = ['license_date', 'license_expiry_date', 'balance_cif', 'exporter__name', 'license_number']

    def list(self, request, *args, **kwargs):
        """Override list to add dynamic purchase_status default to metadata"""
        response = super().list(request, *args, **kwargs)

        # Add purchase_status default to metadata if present
        if isinstance(response.data, dict) and 'default_filters' in response.data:
            default_ps_ids = get_default_purchase_status_ids()
            if default_ps_ids:
                response.data['default_filters']['purchase_status'] = default_ps_ids

        return response

    def get_object(self):
        """
        Override to support lookup by either pk (ID) or license_number.
        """
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field
        filter_kwargs = {self.lookup_field: self.kwargs[lookup_url_kwarg]}

        # Try to get by primary key first (if it's a number)
        lookup_value = self.kwargs[lookup_url_kwarg]
        try:
            # If lookup value is numeric, try ID first
            if lookup_value.isdigit():
                filter_kwargs = {'pk': int(lookup_value)}
                obj = self.get_queryset().get(**filter_kwargs)
            else:
                # Otherwise lookup by license number
                filter_kwargs = {'license_number': lookup_value}
                obj = self.get_queryset().get(**filter_kwargs)
        except (ValueError, self.queryset.model.DoesNotExist):
            # If ID lookup fails, try license_number
            try:
                filter_kwargs = {'license_number': lookup_value}
                obj = self.get_queryset().get(**filter_kwargs)
            except self.queryset.model.DoesNotExist:
                from django.http import Http404
                raise Http404(f"License with identifier '{lookup_value}' not found")

        # May raise a permission denied
        self.check_object_permissions(self.request, obj)
        return obj

    def update(self, request, *args, **kwargs):
        """Override to log incoming request data."""
        import logging
        logger = logging.getLogger(__name__)
        logger.info("=" * 50)
        logger.info("ViewSet.update called")
        logger.info("request.data type: %s", type(request.data))
        logger.info("request.data keys: %s", list(request.data.keys()) if hasattr(request.data, 'keys') else 'N/A')

        # Log license_documents from request.data
        if 'license_documents' in request.data:
            docs = request.data.getlist('license_documents') if hasattr(request.data, 'getlist') else request.data.get(
                'license_documents')
            logger.info("license_documents in request.data: %s", docs)
        else:
            logger.info("license_documents NOT in request.data")

        # Log all keys that contain 'license_documents'
        if hasattr(request.data, 'keys'):
            doc_keys = [k for k in request.data.keys() if 'license_documents' in k]
            logger.info("Keys containing 'license_documents': %s", doc_keys)
            for key in doc_keys[:5]:  # Log first 5
                logger.info("  %s = %s", key, request.data.get(key))

        return super().update(request, *args, **kwargs)

    def get_queryset(self):
        """
        Override to add performance optimizations with select_related and prefetch_related.
        For detail view (retrieve/update/partial_update), don't apply default filters so expired licenses can be edited.
        """
        # For detail view actions, skip default filters by temporarily clearing them
        skip_default_filters = self.action in ['retrieve', 'update', 'partial_update', 'destroy', 'nested_items', 'item_usage', 'balance_pdf', 'balance_excel']

        if skip_default_filters:
            # Save original default filters
            original_default_filters = getattr(self, 'default_filters', {})
            # Temporarily clear default filters
            self.default_filters = {}

        qs = super().get_queryset()

        # Restore original default filters if they were cleared
        if skip_default_filters:
            self.default_filters = original_default_filters

        # Add select_related for ForeignKey fields to avoid N+1 queries
        qs = qs.select_related('exporter', 'port', 'current_owner')

        # Only prefetch nested items for detail view (single object)
        # For list view, this causes massive performance issues
        if self.action == 'retrieve':
            # Prefetch nested relationships for detail view only
            qs = qs.prefetch_related(
                'export_license',
                'export_license__norm_class',
                'import_license',
                'import_license__hs_code',
                'import_license__items',
                'import_license__items__sion_norm_class',
                'license_documents'
            )

        return qs

    def apply_advanced_filters(self, qs, params, filter_config):
        """Override to add custom logic for is_expired and is_null with default values."""
        from datetime import date
        from django.db.models import Q

        # Get default filters
        default_filters = getattr(self, "default_filters", {})

        # Handle is_expired filter - apply default if not provided
        is_expired_value = params.get('is_expired')

        # Special case: if user explicitly selects "all", don't apply any filter
        if is_expired_value and is_expired_value.lower() == "all":
            is_expired_value = None
        elif is_expired_value is None or is_expired_value == "":
            # Apply default if no value provided
            is_expired_value = default_filters.get('is_expired')

        if is_expired_value is not None and is_expired_value != "":
            today = date.today()
            if is_expired_value in ("True", "true", "1", True):
                # Show expired licenses: expiry_date < today
                qs = qs.filter(license_expiry_date__lt=today)
            elif is_expired_value in ("False", "false", "0", False):
                # Show non-expired licenses: expiry_date >= today OR null
                qs = qs.filter(Q(license_expiry_date__gte=today) | Q(license_expiry_date__isnull=True))

        # Handle is_null filter - apply default if not provided
        is_null_value = params.get('is_null')

        # Special case: if user explicitly selects "all", don't apply any filter
        if is_null_value and is_null_value.lower() == "all":
            is_null_value = None
        elif is_null_value is None or is_null_value == "":
            # Apply default if no value provided
            is_null_value = default_filters.get('is_null')

        if is_null_value is not None and is_null_value != "":
            if is_null_value in ("True", "true", "1", True):
                # Show null licenses: balance_cif < 200
                qs = qs.filter(balance_cif__lt=200)
            elif is_null_value in ("False", "false", "0", False):
                # Show non-null licenses: balance_cif >= 200
                qs = qs.filter(balance_cif__gte=200)

        # Call parent method for remaining filters (exclude is_expired and is_null)
        # Create a new QueryDict-like object
        from django.http import QueryDict
        filtered_params = QueryDict(mutable=True)
        for key, value in params.items():
            if key not in ('is_expired', 'is_null'):
                # Handle array format for purchase_status
                if key == 'purchase_status[]':
                    # Frontend sends purchase_status[] for multi-select
                    for val in params.getlist(key):
                        filtered_params.appendlist('purchase_status[]', val)
                else:
                    filtered_params[key] = value

        # Create a copy of filter_config without custom-handled fields
        filtered_config = {k: v for k, v in filter_config.items() if k not in ('is_expired', 'is_null')}

        # Call parent method with filtered params and config
        qs = super().apply_advanced_filters(qs, filtered_params, filtered_config)

        return qs

    @action(detail=True, methods=['get'])
    def nested_items(self, request, pk=None):
        """
        Fetch nested items for a specific license on demand.
        This endpoint is called lazily when user expands a license row.
        """
        license_obj = self.get_object()

        return Response({
            'export_license': LicenseExportItemSerializer(license_obj.export_license.all(), many=True).data,
            'import_license': LicenseImportItemSerializer(license_obj.import_license.all(), many=True,
                                                          context={'request': request}).data,
            'license_documents': LicenseDocumentSerializer(license_obj.license_documents.all(), many=True).data
        })

    @action(detail=True, methods=['get'], url_path='item-usage')
    def item_usage(self, request, pk=None):
        """
        Get real-time usage details for a specific license item (export or import).
        Shows where the item is used in BOEs and Allotments.

        Query params:
        - item_id: ID of the export or import item
        - type: 'export' or 'import'
        """
        from bill_of_entry.models import RowDetails
        from allotment.models import AllotmentItems

        license_obj = self.get_object()
        item_id = request.query_params.get('item_id')
        item_type = request.query_params.get('type')

        if not item_id or not item_type:
            return Response({'error': 'item_id and type parameters are required'}, status=400)

        boes = []
        allotments = []

        if item_type == 'import':
            # Find BOE usage through RowDetails (sr_number points to LicenseImportItemsModel)
            # Only show transaction_type = 'D' (Debited)
            row_details = RowDetails.objects.filter(
                sr_number_id=item_id,
                transaction_type='D'
            ).select_related(
                'bill_of_entry',
                'bill_of_entry__company',
                'bill_of_entry__port'
            ).values(
                'bill_of_entry__id',
                'bill_of_entry__bill_of_entry_number',
                'bill_of_entry__bill_of_entry_date',
                'bill_of_entry__port__name',
                'bill_of_entry__company__name',
                'qty',
                'cif_fc',
                'cif_inr'
            )

            for detail in row_details:
                boes.append({
                    'id': detail['bill_of_entry__id'],
                    'bill_of_entry_number': detail['bill_of_entry__bill_of_entry_number'],
                    'date': detail['bill_of_entry__bill_of_entry_date'],
                    'port': detail['bill_of_entry__port__name'],
                    'company': detail['bill_of_entry__company__name'],
                    'quantity': float(detail['qty']),
                    'cif_fc': float(detail['cif_fc']),
                    'cif_inr': float(detail['cif_inr'])
                })

            # Find Allotment usage through AllotmentItems
            # Only show allotments where bill_of_entry is NULL (not yet converted to BOE)
            allotment_items = AllotmentItems.objects.filter(
                item_id=item_id,
                allotment__bill_of_entry__isnull=True
            ).select_related(
                'allotment',
                'allotment__company'
            ).values(
                'allotment__id',
                'allotment__invoice',
                'allotment__company__name',
                'qty',
                'cif_fc',
                'cif_inr'
            )

            for item in allotment_items:
                allotments.append({
                    'id': item['allotment__id'],
                    'allotment_number': item['allotment__invoice'] or f"Allotment #{item['allotment__id']}",
                    'company': item['allotment__company__name'],
                    'quantity': float(item['qty']),
                    'cif_fc': float(item['cif_fc']),
                    'cif_inr': float(item['cif_inr'])
                })

        elif item_type == 'export':
            # For export items, check allotments (AllotmentItems might link to export items)
            # This depends on your data model structure
            pass

        return Response({
            'boes': boes,
            'allotments': allotments
        })

    @action(detail=True, methods=['get'], url_path='balance-pdf')
    def balance_pdf(self, request, pk=None):
        """
        Generate PDF report for license balance details with all BOEs and Allotments.
        """
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io
        from datetime import date
        from bill_of_entry.models import RowDetails
        from allotment.models import AllotmentItems

        license_obj = self.get_object()

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                              rightMargin=10*mm, leftMargin=10*mm,
                              topMargin=12*mm, bottomMargin=12*mm)

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            alignment=TA_CENTER,
            spaceAfter=8,
            spaceBefore=3,
            fontName='Helvetica-Bold'
        )

        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=10,
            textColor=colors.HexColor('#2c3e50'),
            spaceBefore=5,
            spaceAfter=3,
            fontName='Helvetica-Bold'
        )

        # Add title
        title = Paragraph(f"<b>License Balance Report</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 3))

        # Add license header information split into 2 rows for clarity
        # Check if license has specific document types
        has_tl = license_obj.license_documents.filter(type='TRANSFER LETTER').exists()
        has_copy = license_obj.license_documents.filter(type='LICENSE COPY').exists()

        # Build license number with link if documents exist
        license_number_text = license_obj.license_number or '-'
        if has_tl or has_copy:
            # Get the base URL from request
            base_url = request.build_absolute_uri('/').rstrip('/')
            merge_url = f"{base_url}/api/licenses/{license_obj.id}/merged-documents/"
            license_number_text = f'{license_obj.license_number or "-"} (<link href="{merge_url}" color="blue"><u>Copy</u></link>)'

        header_data = [
            # Row 1: Headers
            ['License Number', 'License Date', 'License Expiry Date', 'Exporter Name', 'Port Name'],
            # Row 1: Values
            [
                Paragraph(license_number_text, styles['Normal']),
                license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-',
                license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-',
                Paragraph(license_obj.exporter.name if license_obj.exporter else '-', styles['Normal']),
                Paragraph(license_obj.port.name if license_obj.port else '-', styles['Normal'])
            ],
            # Row 2: Headers (spanning to match 5 columns)
            ['Purchase Status', 'Balance CIF', 'Get Norm Class', '', 'Latest Transfer'],
            # Row 2: Values
            [
                license_obj.purchase_status or '-',
                f"{float(license_obj.balance_cif or 0):.2f}",
                license_obj.get_norm_class or '-',
                '',
                Paragraph(str(license_obj.latest_transfer) if license_obj.latest_transfer else '-', styles['Normal'])
            ]
        ]

        # Landscape A4 is ~277mm wide, with margins = ~257mm usable width
        # Split into 5 columns: 50mm, 35mm, 40mm, 70mm, 62mm = 257mm
        header_table = Table(header_data, colWidths=[50*mm, 35*mm, 40*mm, 70*mm, 62*mm])
        header_table.setStyle(TableStyle([
            # Row 1 header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            # Row 2 header
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.whitesmoke),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            # Data rows
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#ecf0f1')),
            # Common styles
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 2), (-1, 2), 8),
            ('FONTSIZE', (0, 1), (-1, 1), 7.5),
            ('FONTSIZE', (0, 3), (-1, 3), 7.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 5))

        # Export Items Section
        if license_obj.export_license.exists():
            # Section header as table row
            export_section_header = Table([['Export Items']], colWidths=[275*mm])
            export_section_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(export_section_header)

            export_data = [['Item', 'Total CIF', 'Balance CIF']]
            for item in license_obj.export_license.all():
                item_desc = item.description or (str(item.norm_class) if item.norm_class else None) or 'None'
                export_data.append([
                    Paragraph(item_desc, styles['Normal']),
                    f"{float(item.cif_fc or item.fob_fc or 0):.2f}",
                    f"{float(license_obj.balance_cif or 0):.2f}"
                ])

            export_table = Table(export_data, colWidths=[185*mm, 45*mm, 45*mm])
            export_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')])
            ]))
            elements.append(export_table)
            elements.append(Spacer(1, 8))

        # Import Items Section with Usage Details
        if license_obj.import_license.exists():
            # Section header as table row
            import_section_header = Table([['Import Items']], colWidths=[275*mm])
            import_section_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(import_section_header)

            for item in license_obj.import_license.all():
                # Main item data
                item_names = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else '-'

                # Get HS code label
                hs_code_display = str(item.hs_code.hs_code if item.hs_code else '-')

                item_data = [[
                    'Sr', 'HS Code', 'Description', 'Item', 'Total Qty',
                    'Allotted', 'Debited', 'Available', 'CIF FC', 'Bal CIF'
                ], [
                    str(item.serial_number or '-'),
                    hs_code_display,
                    Paragraph(str(item.description or '-'), styles['Normal']),
                    Paragraph(item_names, styles['Normal']),
                    f"{float(item.quantity or 0):.2f}",
                    f"{float(item.allotted_quantity or 0):.2f}",
                    f"{float(item.debited_quantity or 0):.2f}",
                    f"{float(item.available_quantity or 0):.2f}",
                    f"{float(item.cif_fc or 0):.2f}",
                    f"{float(item.balance_cif_fc or 0):.2f}"
                ]]

                item_table = Table(item_data, colWidths=[12*mm, 25*mm, 60*mm, 50*mm, 23*mm, 21*mm, 21*mm, 21*mm, 21*mm, 21*mm])
                item_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7.5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1'))
                ]))
                elements.append(item_table)
                elements.append(Spacer(1, 3))

                # BOEs
                boes = RowDetails.objects.filter(
                    sr_number_id=item.id,
                    transaction_type='D'
                ).select_related('bill_of_entry', 'bill_of_entry__company', 'bill_of_entry__port')

                if boes.exists():
                    # BOEs header as section row
                    boe_section_header = Table([['BOEs']], colWidths=[275*mm])
                    boe_section_header.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#27ae60')),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#333333')),
                    ]))
                    elements.append(boe_section_header)

                    boe_data = [['BOE Number', 'Date', 'Port', 'Company', 'Qty', 'CIF $', 'CIF INR']]

                    # Calculate totals
                    total_qty = 0
                    total_cif_fc = 0
                    total_cif_inr = 0

                    for detail in boes:
                        total_qty += float(detail.qty or 0)
                        total_cif_fc += float(detail.cif_fc or 0)
                        total_cif_inr += float(detail.cif_inr or 0)

                        boe_data.append([
                            detail.bill_of_entry.bill_of_entry_number,
                            detail.bill_of_entry.bill_of_entry_date.strftime('%d/%m/%Y') if detail.bill_of_entry.bill_of_entry_date else '-',
                            Paragraph(detail.bill_of_entry.port.name if detail.bill_of_entry.port else '-', styles['Normal']),
                            Paragraph(detail.bill_of_entry.company.name if detail.bill_of_entry.company else '-', styles['Normal']),
                            f"{float(detail.qty):.2f}",
                            f"{float(detail.cif_fc):.2f}",
                            f"{float(detail.cif_inr):.2f}"
                        ])

                    # Add total footer row
                    boe_data.append([
                        '', '', '', 'Total',
                        f"{total_qty:.2f}",
                        f"{total_cif_fc:.2f}",
                        f"{total_cif_inr:.2f}"
                    ])

                    boe_table = Table(boe_data, colWidths=[40*mm, 25*mm, 50*mm, 70*mm, 25*mm, 30*mm, 35*mm])
                    boe_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTSIZE', (0, 1), (-1, -2), 7.5),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#e8f5e9')]),
                        # Footer row styling
                        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4edda')),
                        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, -1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                    ]))
                    elements.append(boe_table)
                    elements.append(Spacer(1, 5))

                # Allotments
                # Only show allotments where bill_of_entry is NULL (not yet converted to BOE)
                allotments = AllotmentItems.objects.filter(
                    item_id=item.id,
                    allotment__bill_of_entry__isnull=True
                ).select_related('allotment', 'allotment__company')

                if allotments.exists():
                    # Allotments header as section row
                    allot_section_header = Table([['Allotments']], colWidths=[275*mm])
                    allot_section_header.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e67e22')),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#333333')),
                    ]))
                    elements.append(allot_section_header)

                    allot_data = [['Company', 'Qty', 'CIF $', 'CIF INR']]

                    # Calculate totals
                    total_allot_qty = 0
                    total_allot_cif_fc = 0
                    total_allot_cif_inr = 0

                    for allot in allotments:
                        total_allot_qty += float(allot.qty or 0)
                        total_allot_cif_fc += float(allot.cif_fc or 0)
                        total_allot_cif_inr += float(allot.cif_inr or 0)

                        allot_data.append([
                            Paragraph(allot.allotment.company.name if allot.allotment.company else '-', styles['Normal']),
                            f"{float(allot.qty):.2f}",
                            f"{float(allot.cif_fc):.2f}",
                            f"{float(allot.cif_inr):.2f}"
                        ])

                    # Add total footer row
                    allot_data.append([
                        'Total',
                        f"{total_allot_qty:.2f}",
                        f"{total_allot_cif_fc:.2f}",
                        f"{total_allot_cif_inr:.2f}"
                    ])

                    allot_table = Table(allot_data, colWidths=[155*mm, 40*mm, 40*mm, 40*mm])
                    allot_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e67e22')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTSIZE', (0, 1), (-1, -2), 7.5),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#fef5e7')]),
                        # Footer row styling
                        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fdebd0')),
                        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, -1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                    ]))
                    elements.append(allot_table)
                    elements.append(Spacer(1, 5))

                # Balance calculation as table footer
                balance = float(item.quantity or 0) - float(item.debited_quantity or 0) - float(item.allotted_quantity or 0)
                balance_table = Table([[f'Balance Quantity: {balance:.2f}']], colWidths=[275*mm])
                balance_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e8e8e8')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#e74c3c')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                elements.append(balance_table)
                elements.append(Spacer(1, 8))

        # Add Notes Section if notes exist
        if license_obj.balance_report_notes:
            notes_header = Table([['Notes']], colWidths=[275*mm])
            notes_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(notes_header)

            notes_content = Table([[Paragraph(license_obj.balance_report_notes, styles['Normal'])]], colWidths=[275*mm])
            notes_content.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fffacd')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(notes_content)

        # ── End-of-PDF Summary Table ─────────────────────────────────────────
        # One flat row per BOE/Allotment per item — easy to copy-paste to Excel
        # Light green = BOE rows, Light red = Allotment rows
        COLOR_BOE   = colors.HexColor('#d9ead3')   # light green
        COLOR_ALLOT = colors.HexColor('#fce8e6')   # light red
        COLOR_HDR   = colors.HexColor('#1a1a1a')

        # Paragraph style for wrapping text in summary cells
        wrap_style = ParagraphStyle('wrap', parent=styles['Normal'], fontSize=7.5, leading=10)

        def P(text):
            return Paragraph(str(text), wrap_style)

        # Collect all rows — sort by item name before building table
        # Columns: License No | License Date | Item | Type | Company | Reference | Qty | Rate | CIF Value (FC)
        # BOE Reference  = "BOE number\nDate"
        # Allot Reference = "Invoice\nETA: date" (if available)
        summary_data = [['License No', 'License Date', 'Item', 'Type', 'Company', 'Reference', 'Qty', 'Rate', 'CIF Value (FC)']]
        summary_rows = []  # (sort_key, row_cells, color)
        total_cif    = 0.0

        license_date_str = license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-'
        lic_no = license_obj.license_number or '-'

        # Pre-aggregate by item name across all sr numbers.
        from collections import defaultdict
        from decimal import Decimal as _Dec
        from django.db.models import Sum as _Sum, DecimalField as _DF, Value as _Val
        from django.db.models.functions import Coalesce as _Coalesce

        _bal_agg = defaultdict(lambda: {'qty': 0.0, 'is_restricted': False, 'restriction_pct': None, 'sr_ids': [], 'description': '', 'hs_code': ''})
        for _item in license_obj.import_license.all():
            _key = ', '.join(sorted([i.name for i in _item.items.all()])) if _item.items.exists() else (_item.description or '-')
            _bal_agg[_key]['qty'] += float(_item.available_quantity or 0)
            _bal_agg[_key]['sr_ids'].append(_item.id)
            if _item.is_restricted:
                _bal_agg[_key]['is_restricted'] = True
                if _bal_agg[_key]['restriction_pct'] is None:
                    # Get restriction_percentage directly from linked ItemNameModel
                    # (without sion_norm_class requirement, which may be null)
                    _rpct_val = _item.items.filter(
                        restriction_percentage__gt=0
                    ).values_list('restriction_percentage', flat=True).first()
                    if _rpct_val:
                        _bal_agg[_key]['restriction_pct'] = _Dec(str(_rpct_val))
            if not _bal_agg[_key]['description']:
                _bal_agg[_key]['description'] = _item.description or _key
            if not _bal_agg[_key]['hs_code']:
                _bal_agg[_key]['hs_code'] = str(_item.hs_code.hs_code if _item.hs_code else '-')

        for item in license_obj.import_license.all():
            item_name = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else (item.description or '-')

            boes = RowDetails.objects.filter(
                sr_number_id=item.id, transaction_type='D'
            ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')

            for rd in boes:
                qty     = float(rd.qty or 0)
                cif     = float(rd.cif_fc or 0)
                rate    = cif / qty if qty else 0.0
                total_cif += cif
                boe_company = rd.bill_of_entry.company.name if rd.bill_of_entry.company else '-'
                ref_no  = rd.bill_of_entry.bill_of_entry_number or '-'
                ref_date = rd.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if rd.bill_of_entry.bill_of_entry_date else ''
                ref_str = f"{ref_no}\n{ref_date}" if ref_date else ref_no
                product = rd.bill_of_entry.product_name or item_name
                summary_rows.append((product.lower(), [
                    P(lic_no), P(license_date_str), P(product),
                    P('BOE'), P(boe_company), P(ref_str),
                    P(f"{qty:,.2f}"), P(f"{rate:.2f}"), P(f"{cif:,.2f}"),
                ], COLOR_BOE))

            allotments = AllotmentItems.objects.filter(
                item_id=item.id, allotment__bill_of_entry__isnull=True
            ).select_related('allotment', 'allotment__company')

            for ai in allotments:
                qty     = float(ai.qty or 0)
                cif     = float(ai.cif_fc or 0)
                rate    = cif / qty if qty else 0.0
                total_cif += cif
                company = ai.allotment.company.name if ai.allotment.company else '-'
                invoice = ai.allotment.invoice or '-'
                eta     = ai.allotment.estimated_arrival_date.strftime('%d-%m-%Y') if ai.allotment.estimated_arrival_date else ''
                ref_str = f"{invoice}\nETA: {eta}" if eta else invoice
                product = ai.allotment.item_name or item_name
                summary_rows.append((product.lower(), [
                    P(lic_no), P(license_date_str), P(product),
                    P('Allotment'), P(company), P(ref_str),
                    P(f"{qty:,.2f}"), P(f"{rate:.2f}"), P(f"{cif:,.2f}"),
                ], COLOR_ALLOT))

        # Sort by item name
        summary_rows.sort(key=lambda x: x[0])
        row_colors = []
        for _, row_cells, color in summary_rows:
            summary_data.append(row_cells)
            row_colors.append(color)

        if len(summary_data) > 1:
            # Total row
            summary_data.append([P(''), P(''), P(''), P(''), P('TOTAL'), P(''), P(''), P(''), P(f"{total_cif:,.2f}")])
            row_colors.append(colors.HexColor('#f2f2f2'))

            # ── License info mini-header (License No | License Date | Total CIF) ──
            total_license_cif = total_cif + float(license_obj.balance_cif or 0)
            info_style = ParagraphStyle('info', parent=styles['Normal'], fontSize=8, leading=11,
                                        textColor=colors.white, fontName='Helvetica-Bold')
            def IP(label, value):
                return Paragraph(f"<b>{label}:</b> {value}", info_style)

            info_row = Table([[
                IP('License No', lic_no),
                IP('License Date', license_date_str),
                IP('Total CIF', f"{total_license_cif:,.2f}"),
            ]], colWidths=[92*mm, 92*mm, 93*mm])
            info_row.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLOR_HDR),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(Spacer(1, 10))
            elements.append(info_row)

            # Section header
            summ_hdr = Table([['Summary (BOE & Allotments)']], colWidths=[277*mm])
            summ_hdr.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLOR_HDR),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(summ_hdr)

            # col widths: lic_no, lic_date, item, type, company, reference, qty, rate, cif = 277mm
            col_w = [28*mm, 22*mm, 55*mm, 18*mm, 40*mm, 35*mm, 20*mm, 22*mm, 37*mm]
            summ_table = Table(summary_data, colWidths=col_w)

            style_cmds = [
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_HDR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7.5),
                # Data rows
                ('FONTSIZE', (0, 1), (-1, -1), 7.5),
                ('ALIGN', (6, 0), (-1, -1), 'RIGHT'),   # qty, rate, cif right-aligned
                ('ALIGN', (0, 0), (5, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                # Total row bold
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f2f2f2')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ]
            # Apply per-row background colours (skip header row at index 0)
            for i, bg in enumerate(row_colors, start=1):
                if i < len(summary_data):  # skip total row (handled above)
                    style_cmds.append(('BACKGROUND', (0, i), (-1, i), bg))

            summ_table.setStyle(TableStyle(style_cmds))
            elements.append(summ_table)

        # ── Balance Summary Table ─────────────────────────────────────────────
        if _bal_agg:
            total_bal_cif_fc = float(license_obj.balance_cif or 0)
            COLOR_YELLOW = colors.HexColor('#ffff00')

            # "Summary (Balance Quantity)" section header
            bal_hdr = Table([['Summary (Balance Quantity)']], colWidths=[277*mm])
            bal_hdr.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLOR_HDR),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(Spacer(1, 8))
            elements.append(bal_hdr)

            # col widths: hsn, item_name, bal_qty, unit_price, cif_fc = 277mm
            # 30 + 87 + 45 + 45 + 70 = 277mm
            bal_col_w = [30*mm, 87*mm, 45*mm, 45*mm, 70*mm]

            Pb      = ParagraphStyle('balwrap',     parent=styles['Normal'], fontSize=8, leading=11)
            Pb_hdr  = ParagraphStyle('balwrap_hdr', parent=styles['Normal'], fontSize=8, leading=11,
                                     textColor=colors.white, fontName='Helvetica-Bold')
            Pb_yel  = ParagraphStyle('balwrap_yel', parent=styles['Normal'], fontSize=9, leading=12,
                                     fontName='Helvetica-Bold')

            def BP(text):
                return Paragraph(str(text), Pb)
            def BH(text):   # white bold header cell
                return Paragraph(str(text), Pb_hdr)
            def BY(text):   # yellow-cell (black bold)
                return Paragraph(str(text), Pb_yel)

            _license_balance = float(license_obj.balance_cif or 0)
            _total_export_cif = _Dec(str(license_obj._calculate_license_credit() or 0))

            bal_table_data = [
                # Row 0: cols 0-3 merged "BALANCE CIF $" | col 4 = total (yellow)
                [BH('BALANCE CIF $'), '', '', '', BY(f"{total_bal_cif_fc:,.2f}")],
                # Row 1: column headers
                [BH('HSN Code'), BH('Item Name'), BH('Bal Qty'), BH('Unit Price'), BH('CIF FC')],
            ]
            for item_key in sorted(_bal_agg.keys()):
                b_qty = _bal_agg[item_key]['qty']
                # Restricted items: compute per-group balance = (export_cif × pct/100) − group_debits − group_allotments
                # Non-restricted items: show license-level balance
                if _bal_agg[item_key]['is_restricted'] and _bal_agg[item_key]['restriction_pct'] is not None:
                    _rpct = _bal_agg[item_key]['restriction_pct']
                    _sr_ids = _bal_agg[item_key]['sr_ids']
                    _grp_debits = RowDetails.objects.filter(
                        sr_number_id__in=_sr_ids, transaction_type='D'
                    ).aggregate(
                        total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                    )['total'] or _Dec('0')
                    _grp_allots = AllotmentItems.objects.filter(
                        item_id__in=_sr_ids, allotment__bill_of_entry__isnull=True
                    ).aggregate(
                        total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                    )['total'] or _Dec('0')
                    _allowed = _total_export_cif * _rpct / _Dec('100')
                    b_cif = float(max(_allowed - _Dec(str(_grp_debits)) - _Dec(str(_grp_allots)), _Dec('0')))
                else:
                    b_cif = _license_balance
                unit_price = b_cif / b_qty if b_qty else 0.0
                desc       = _bal_agg[item_key]['description'] or item_key
                hs         = _bal_agg[item_key]['hs_code']
                bal_table_data.append([
                    BP(hs),
                    BP(desc),
                    BP(f"{b_qty:,.2f}"),
                    BP(f"{unit_price:,.2f}"),
                    BP(f"{b_cif:,.2f}"),
                ])

            bal_table = Table(bal_table_data, colWidths=bal_col_w)
            bal_style = TableStyle([
                # Row 0: merge cols 0-3, dark header | col 4 yellow
                ('SPAN', (0, 0), (3, 0)),
                ('BACKGROUND', (0, 0), (3, 0), COLOR_HDR),
                ('BACKGROUND', (4, 0), (4, 0), COLOR_YELLOW),
                ('TEXTCOLOR', (4, 0), (4, 0), colors.black),
                ('ALIGN', (0, 0), (3, 0), 'CENTER'),
                ('ALIGN', (4, 0), (4, 0), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                # Row 1: column headers
                ('BACKGROUND', (0, 1), (-1, 1), COLOR_HDR),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, 1), 8),
                # Data rows: cols 2-4 right-aligned
                ('FONTSIZE', (0, 2), (-1, -1), 8),
                ('ALIGN', (2, 2), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 2), (1, -1), 'LEFT'),
                ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                # All cells
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ])
            bal_table.setStyle(bal_style)
            elements.append(Spacer(1, 8))
            elements.append(bal_table)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{license_obj.license_number}-balance.pdf"'
        response.write(pdf)

        return response

    @action(detail=False, methods=['post'], url_path='bulk-balance-excel')
    def bulk_balance_excel(self, request):
        """
        Generate a multi-sheet Excel with one sheet per license.
        Sheet name = license number. Same layout as balance_excel.
        POST body: {"license_numbers": ["3011007415", "3011007018", ...]}
        """
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        from decimal import Decimal as _Dec
        from collections import defaultdict
        from django.db.models import Sum as _Sum, DecimalField as _DF, Value as _Val
        from django.db.models.functions import Coalesce as _Coalesce
        from bill_of_entry.models import RowDetails
        from allotment.models import AllotmentItems

        license_numbers = request.data.get('license_numbers', [])
        if not license_numbers:
            return Response({'error': 'No license numbers provided.'}, status=400)

        licenses = LicenseDetailsModel.objects.filter(
            license_number__in=license_numbers
        ).prefetch_related('import_license', 'import_license__items')

        if not licenses.exists():
            return Response({'error': 'No matching licenses found.'}, status=404)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # ── Shared styles ──────────────────────────────────────────────────────
        HDR_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
        BOE_FILL   = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        ALLOT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        YEL_FILL   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ALT_FILL   = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        BOLD       = Font(bold=True, size=9)
        NORM       = Font(size=9)
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def _hdr(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            return c

        def _cell(ws, row, col, value, fill=None, bold=False, align='left', num_fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            if fill: c.fill = fill
            c.font = BOLD if bold else NORM
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            if num_fmt: c.number_format = num_fmt
            return c

        def _write_license_sheet(wb, license_obj):
            from datetime import date as _date_cls
            sheet_name = str(license_obj.license_number)[:31]
            ws = wb.create_sheet(title=sheet_name)

            license_date_str = license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-'
            license_expiry_str = license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-'
            lic_no = license_obj.license_number or '-'

            summary_rows = []
            total_cif = 0.0

            for item in license_obj.import_license.all():
                item_name = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else (item.description or '-')

                boes = RowDetails.objects.filter(
                    sr_number_id=item.id, transaction_type='D'
                ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')

                for rd in boes:
                    qty  = float(rd.qty or 0)
                    cif  = float(rd.cif_fc or 0)
                    rate = cif / qty if qty else 0.0
                    total_cif += cif
                    boe_company = rd.bill_of_entry.company.name if rd.bill_of_entry.company else '-'
                    ref_no   = rd.bill_of_entry.bill_of_entry_number or '-'
                    ref_date = rd.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if rd.bill_of_entry.bill_of_entry_date else ''
                    ref_str  = f"{ref_no} / {ref_date}" if ref_date else ref_no
                    product  = rd.bill_of_entry.product_name or item_name
                    _sort_dt = rd.bill_of_entry.bill_of_entry_date or _date_cls.min
                    summary_rows.append((0, _sort_dt, {
                        'item': product, 'type': 'BOE', 'company': boe_company,
                        'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
                    }, True))

                allotments = AllotmentItems.objects.filter(
                    item_id=item.id, allotment__bill_of_entry__isnull=True
                ).select_related('allotment', 'allotment__company')

                for ai in allotments:
                    qty     = float(ai.qty or 0)
                    cif     = float(ai.cif_fc or 0)
                    rate    = cif / qty if qty else 0.0
                    total_cif += cif
                    company = ai.allotment.company.name if ai.allotment.company else '-'
                    invoice = ai.allotment.invoice or '-'
                    eta     = ai.allotment.estimated_arrival_date.strftime('%d-%m-%Y') if ai.allotment.estimated_arrival_date else ''
                    ref_str = f"{invoice} / ETA: {eta}" if eta else invoice
                    product = ai.allotment.item_name or item_name
                    _sort_dt = ai.allotment.estimated_arrival_date or _date_cls.min
                    summary_rows.append((1, _sort_dt, {
                        'item': product, 'type': 'Allotment', 'company': company,
                        'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
                    }, False))

            # BOEs first (sorted by BOE date), then allotments (sorted by allotment date)
            summary_rows.sort(key=lambda x: (x[0], x[1]))

            _bal_agg = defaultdict(lambda: {
                'qty': 0.0, 'is_restricted': False, 'restriction_pct': None,
                'sr_ids': [], 'description': '', 'hs_code': ''
            })
            for _item in license_obj.import_license.all():
                _key = ', '.join(sorted([i.name for i in _item.items.all()])) if _item.items.exists() else (_item.description or '-')
                _bal_agg[_key]['qty'] += float(_item.available_quantity or 0)
                _bal_agg[_key]['sr_ids'].append(_item.id)
                if _item.is_restricted:
                    _bal_agg[_key]['is_restricted'] = True
                    if _bal_agg[_key]['restriction_pct'] is None:
                        _rpct_val = _item.items.filter(
                            restriction_percentage__gt=0
                        ).values_list('restriction_percentage', flat=True).first()
                        if _rpct_val:
                            _bal_agg[_key]['restriction_pct'] = _Dec(str(_rpct_val))
                if not _bal_agg[_key]['description']:
                    _bal_agg[_key]['description'] = _item.description or _key
                if not _bal_agg[_key]['hs_code']:
                    _bal_agg[_key]['hs_code'] = str(_item.hs_code.hs_code if _item.hs_code else '-')

            _license_balance = float(license_obj.balance_cif or 0)
            _total_export_cif = _Dec(str(license_obj._calculate_license_credit() or 0))
            total_license_cif = total_cif + _license_balance

            r = 1
            _today = _date_cls.today()
            INFO_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            INFO_FONT = Font(bold=True, color="FFFFFF", size=9)
            if license_obj.license_expiry_date:
                _days = (license_obj.license_expiry_date - _today).days
                if _days < 0:
                    EXPIRY_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                elif _days <= 90:
                    EXPIRY_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                else:
                    EXPIRY_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            else:
                EXPIRY_FILL = INFO_FILL
            for col, (label, val) in enumerate([
                ('License No', lic_no),
                ('License Date', license_date_str),
                ('Expiry Date', license_expiry_str),
                ('Total CIF', f"{total_license_cif:,.2f}")
            ], 1):
                c = ws.cell(row=r, column=col, value=f"{label}: {val}")
                c.fill = EXPIRY_FILL if col == 3 else INFO_FILL
                c.font = INFO_FONT
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal='left' if col < 4 else 'right', vertical='center')
            r += 1

            ws.merge_cells(f'A{r}:G{r}')
            sh = ws[f'A{r}']
            sh.value = 'Summary (BOE & Allotments)'
            sh.fill = HDR_FILL; sh.font = Font(bold=True, color="FFFFFF", size=10)
            sh.alignment = Alignment(horizontal='center', vertical='center')
            r += 1

            SUMM_COLS = ['Item', 'Type', 'Company', 'Reference', 'Qty', 'Rate', 'CIF Value (FC)']
            for col, h in enumerate(SUMM_COLS, 1):
                _hdr(ws, r, col, h)
            r += 1

            for _s, _sd, row_data, is_boe in summary_rows:
                fill = BOE_FILL if is_boe else ALLOT_FILL
                _cell(ws, r, 1, row_data['item'],      fill=fill)
                _cell(ws, r, 2, row_data['type'],      fill=fill)
                _cell(ws, r, 3, row_data['company'],   fill=fill)
                _cell(ws, r, 4, row_data['reference'], fill=fill)
                _cell(ws, r, 5, row_data['qty'],       fill=fill, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 6, row_data['rate'],      fill=fill, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 7, row_data['cif'],       fill=fill, align='right', num_fmt='#,##0.00')
                r += 1

            if summary_rows:
                _cell(ws, r, 1, '', fill=TOTAL_FILL); _cell(ws, r, 2, '', fill=TOTAL_FILL)
                _cell(ws, r, 3, '', fill=TOTAL_FILL)
                _cell(ws, r, 4, 'TOTAL', fill=TOTAL_FILL, bold=True, align='right')
                _cell(ws, r, 5, '', fill=TOTAL_FILL); _cell(ws, r, 6, '', fill=TOTAL_FILL)
                _cell(ws, r, 7, total_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
                r += 1

            r += 1

            # ── Norm check for utilization planning ──────────────────────────
            _norm_vals = list(license_obj.export_license.values_list('norm_class__norm_class', flat=True))
            _is_e1 = any(n and 'E1' in str(n) and 'E126' not in str(n) and 'E132' not in str(n) for n in _norm_vals)
            _is_e5 = any(n and str(n).strip() == 'E5' for n in _norm_vals)
            if _is_e1:
                _UTIL_PLAN = [
                    ('OTHER CONFECTIONERY / FRUIT & FRUIT PRODUCTS', 2.7, ['0802'], ['confectionery', 'fruit & fruit', 'fruit product']),
                    ('FRUIT JUICE',                                   3.0, [],       ['fruit juice']),
                    ('TARTARIC ACID / CITRIC ACID',                   1.0, [],       ['tartaric', 'citric']),
                    ('POLYPROPYLENE',                                  1.0, [],       ['polypropylene']),
                    ('PAPER & PAPER BOARD',                           0.7, [],       ['paper']),
                ]
                def _cat_match(hs, desc, hs_kw, name_kw):
                    return any(k in (hs or '').lower() for k in hs_kw) or any(k in (desc or '').lower() for k in name_kw)
                _cat_totals = {label: 0.0 for label, *_ in _UTIL_PLAN}
                _unclassified = []
                for _ik in _bal_agg:
                    _bq = _bal_agg[_ik]['qty']
                    _de = _bal_agg[_ik]['description'] or _ik
                    _hs = _bal_agg[_ik]['hs_code']
                    _found = False
                    for _lbl, _rt, _hskw, _nkw in _UTIL_PLAN:
                        if _cat_match(_hs, _de, _hskw, _nkw):
                            _cat_totals[_lbl] += _bq
                            _found = True
                            break
                    if not _found:
                        _unclassified.append((_de, _hs, _bq))
            elif _is_e5:
                _E5_CATS = [
                    ('DIETARY FIBRE',            2.7,   ['0802']),
                    ('VEGETABLE OIL (1513)',      2.26,  ['1513']),
                    ('VEGETABLE OIL (15119020)',  1.2,   ['15119020']),
                    ('MILK (0404)',               3.75,  ['0404']),
                    ('MILK (3502)',               14.0,  ['3502']),
                ]
                _e5_totals = {label: 0.0 for label, *_ in _E5_CATS}
                _e5_unclassified = []
                for _ik in _bal_agg:
                    _bq = _bal_agg[_ik]['qty']
                    _hs = (_bal_agg[_ik]['hs_code'] or '').lower()
                    _de = _bal_agg[_ik]['description'] or _ik
                    _found = False
                    for _lbl, _rt, _hskw in _E5_CATS:
                        if any(k in _hs for k in _hskw):
                            _e5_totals[_lbl] += _bq
                            _found = True
                            break
                    if not _found:
                        _e5_unclassified.append((_de, _bal_agg[_ik]['hs_code'], _bq))

            ws.merge_cells(f'A{r}:E{r}')
            bh = ws[f'A{r}']
            bh.value = 'Utilization Planning' if (_is_e1 or _is_e5) else 'Summary (Balance Quantity)'
            bh.fill = HDR_FILL; bh.font = Font(bold=True, color="FFFFFF", size=10)
            bh.alignment = Alignment(horizontal='center', vertical='center')
            r += 1

            ws.merge_cells(f'A{r}:D{r}')
            bc = ws[f'A{r}']
            bc.value = 'BALANCE CIF $'
            bc.fill = HDR_FILL; bc.font = Font(bold=True, color="FFFFFF", size=9)
            bc.alignment = Alignment(horizontal='center', vertical='center')
            bc.border = THIN_BORDER
            yc = ws.cell(row=r, column=5, value=_license_balance)
            yc.fill = YEL_FILL; yc.font = Font(bold=True, size=9)
            yc.border = THIN_BORDER
            yc.alignment = Alignment(horizontal='right', vertical='center')
            yc.number_format = '#,##0.00'
            r += 1

            if _is_e1:
                for col, h in enumerate(['Item Category', 'Rate ($/unit)', 'Bal Qty', 'Unit Price', 'Planned CIF ($)'], 1):
                    _hdr(ws, r, col, h)
                r += 1

                _total_planned = 0.0
                _e1_remaining = _license_balance
                for _idx, (_lbl, _rt, _hskw, _nkw) in enumerate(_UTIL_PLAN):
                    _bq = _cat_totals[_lbl]
                    _pc = min(_rt * _bq, _e1_remaining)
                    _up = _pc / _bq if _bq else 0.0
                    _e1_remaining -= _pc
                    _total_planned += _pc
                    _rf = None if _idx % 2 == 0 else ALT_FILL
                    _cell(ws, r, 1, _lbl, fill=_rf)
                    _cell(ws, r, 2, _rt,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 3, _bq,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 4, _up,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 5, _pc,  fill=_rf, align='right', num_fmt='#,##0.00')
                    r += 1

                if _unclassified:
                    r += 1
                    ws.merge_cells(f'A{r}:E{r}')
                    _uh = ws[f'A{r}']
                    _uh.value = 'UNCLASSIFIED ITEMS'
                    _uh.fill = HDR_FILL; _uh.font = Font(bold=True, color="FFFFFF", size=9)
                    _uh.alignment = Alignment(horizontal='center', vertical='center')
                    _uh.border = THIN_BORDER
                    r += 1
                    for col, h in enumerate(['Item Name', 'HS Code', 'Bal Qty', '', ''], 1):
                        _hdr(ws, r, col, h)
                    r += 1
                    for _i2, (_de2, _hs2, _bq2) in enumerate(_unclassified):
                        _rf2 = None if _i2 % 2 == 0 else ALT_FILL
                        _cell(ws, r, 1, _de2, fill=_rf2)
                        _cell(ws, r, 2, _hs2, fill=_rf2)
                        _cell(ws, r, 3, _bq2, fill=_rf2, align='right', num_fmt='#,##0.00')
                        _cell(ws, r, 4, '',   fill=_rf2)
                        _cell(ws, r, 5, '',   fill=_rf2)
                        r += 1

                r += 1
                _cell(ws, r, 1, '', fill=TOTAL_FILL)
                _cell(ws, r, 2, '', fill=TOTAL_FILL)
                _cell(ws, r, 3, '', fill=TOTAL_FILL)
                _cell(ws, r, 4, 'TOTAL PLANNED CIF $', fill=TOTAL_FILL, bold=True, align='right')
                _cell(ws, r, 5, _total_planned, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
                r += 1
                _rem = _license_balance - _total_planned
                _RF = PatternFill(start_color="C00000" if _rem < 0 else "1F4E79",
                                  end_color="C00000" if _rem < 0 else "1F4E79", fill_type="solid")
                for _ci in range(1, 5):
                    _cx = ws.cell(row=r, column=_ci)
                    _cx.fill = _RF; _cx.border = THIN_BORDER
                _rc = ws.cell(row=r, column=4, value='REMAINING BALANCE CIF $')
                _rc.fill = _RF; _rc.font = Font(bold=True, color="FFFFFF", size=9)
                _rc.border = THIN_BORDER; _rc.alignment = Alignment(horizontal='right', vertical='center')
                _rc2 = ws.cell(row=r, column=5, value=_rem)
                _rc2.fill = _RF; _rc2.font = Font(bold=True, color="FFFFFF", size=9)
                _rc2.border = THIN_BORDER
                _rc2.alignment = Alignment(horizontal='right', vertical='center')
                _rc2.number_format = '#,##0.00'
                r += 1
            elif _is_e5:
                for col, h in enumerate(['Item Category', 'Rate ($/unit)', 'Bal Qty', 'Unit Price', 'Planned CIF ($)'], 1):
                    _hdr(ws, r, col, h)
                r += 1

                _e5_planned = 0.0
                _e5_remaining = _license_balance
                for _idx, (_lbl, _rt, _hskw) in enumerate(_E5_CATS):
                    _bq = _e5_totals[_lbl]
                    _pc = min(_rt * _bq, _e5_remaining)
                    _up = _pc / _bq if _bq else 0.0
                    _e5_remaining -= _pc
                    _e5_planned += _pc
                    _rf = None if _idx % 2 == 0 else ALT_FILL
                    _cell(ws, r, 1, _lbl, fill=_rf)
                    _cell(ws, r, 2, _rt,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 3, _bq,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 4, _up,  fill=_rf, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 5, _pc,  fill=_rf, align='right', num_fmt='#,##0.00')
                    r += 1

                # WHEAT FLOUR row — last priority, remaining balance
                _wf = _e5_remaining
                _wf_rf = None if len(_E5_CATS) % 2 == 0 else ALT_FILL
                _cell(ws, r, 1, 'WHEAT FLOUR', fill=_wf_rf)
                _cell(ws, r, 2, '-', fill=_wf_rf, align='center')
                _cell(ws, r, 3, '-', fill=_wf_rf, align='center')
                _cell(ws, r, 4, '-', fill=_wf_rf, align='center')
                _cell(ws, r, 5, _wf, fill=_wf_rf, align='right', num_fmt='#,##0.00')
                r += 1

                if _e5_unclassified:
                    r += 1
                    ws.merge_cells(f'A{r}:E{r}')
                    _uh = ws[f'A{r}']
                    _uh.value = 'UNCLASSIFIED ITEMS'
                    _uh.fill = HDR_FILL; _uh.font = Font(bold=True, color="FFFFFF", size=9)
                    _uh.alignment = Alignment(horizontal='center', vertical='center')
                    _uh.border = THIN_BORDER
                    r += 1
                    for col, h in enumerate(['Item Name', 'HS Code', 'Bal Qty', '', ''], 1):
                        _hdr(ws, r, col, h)
                    r += 1
                    for _i2, (_de2, _hs2, _bq2) in enumerate(_e5_unclassified):
                        _rf2 = None if _i2 % 2 == 0 else ALT_FILL
                        _cell(ws, r, 1, _de2, fill=_rf2)
                        _cell(ws, r, 2, _hs2, fill=_rf2)
                        _cell(ws, r, 3, _bq2, fill=_rf2, align='right', num_fmt='#,##0.00')
                        _cell(ws, r, 4, '',   fill=_rf2)
                        _cell(ws, r, 5, '',   fill=_rf2)
                        r += 1

                r += 1
                _cell(ws, r, 1, '', fill=TOTAL_FILL)
                _cell(ws, r, 2, '', fill=TOTAL_FILL)
                _cell(ws, r, 3, '', fill=TOTAL_FILL)
                _cell(ws, r, 4, 'TOTAL ALLOCATED CIF $', fill=TOTAL_FILL, bold=True, align='right')
                _cell(ws, r, 5, _e5_planned + _wf, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
                r += 1
            else:
                BAL_COLS = ['HSN Code', 'Item Name', 'Bal Qty', 'Unit Price', 'CIF FC']
                for col, h in enumerate(BAL_COLS, 1):
                    _hdr(ws, r, col, h)
                r += 1

                for idx, item_key in enumerate(sorted(_bal_agg.keys())):
                    b_qty = _bal_agg[item_key]['qty']
                    if _bal_agg[item_key]['is_restricted'] and _bal_agg[item_key]['restriction_pct'] is not None:
                        _rpct  = _bal_agg[item_key]['restriction_pct']
                        _sr_ids = _bal_agg[item_key]['sr_ids']
                        _grp_debits = RowDetails.objects.filter(
                            sr_number_id__in=_sr_ids, transaction_type='D'
                        ).aggregate(
                            total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                        )['total'] or _Dec('0')
                        _grp_allots = AllotmentItems.objects.filter(
                            item_id__in=_sr_ids, allotment__bill_of_entry__isnull=True
                        ).aggregate(
                            total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                        )['total'] or _Dec('0')
                        _allowed = _total_export_cif * _rpct / _Dec('100')
                        b_cif = float(max(_allowed - _Dec(str(_grp_debits)) - _Dec(str(_grp_allots)), _Dec('0')))
                    else:
                        b_cif = _license_balance

                    unit_price = b_cif / b_qty if b_qty else 0.0
                    desc = _bal_agg[item_key]['description'] or item_key
                    hs   = _bal_agg[item_key]['hs_code']
                    row_fill = None if idx % 2 == 0 else ALT_FILL

                    _cell(ws, r, 1, hs,         fill=row_fill)
                    _cell(ws, r, 2, desc,       fill=row_fill)
                    _cell(ws, r, 3, b_qty,      fill=row_fill, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 4, unit_price, fill=row_fill, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 5, b_cif,      fill=row_fill, align='right', num_fmt='#,##0.00')
                    r += 1

            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 28
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 14
            ws.column_dimensions['I'].width = 16
            ws.freeze_panes = 'A2'

        for license_obj in licenses:
            _write_license_sheet(wb, license_obj)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bulk_license_summary.xlsx"'
        return response

    @action(detail=True, methods=['get'], url_path='balance-excel')
    def balance_excel(self, request, pk=None):
        """
        Generate Excel summary report matching the two bottom tables in balance_pdf:
        1. Summary (BOE & Allotments)
        2. Summary (Balance Quantity)
        """
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from decimal import Decimal as _Dec
        from collections import defaultdict
        from django.db.models import Sum as _Sum, DecimalField as _DF, Value as _Val
        from django.db.models.functions import Coalesce as _Coalesce
        from bill_of_entry.models import RowDetails
        from allotment.models import AllotmentItems

        license_obj = self.get_object()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        # ── Styles ────────────────────────────────────────────────────────────
        HDR_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
        BOE_FILL   = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        ALLOT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        YEL_FILL   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ALT_FILL   = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        BOLD       = Font(bold=True, size=9)
        NORM       = Font(size=9)
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def _hdr(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            return c

        def _cell(ws, row, col, value, fill=None, bold=False, align='left', num_fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            if fill: c.fill = fill
            c.font = BOLD if bold else NORM
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            if num_fmt: c.number_format = num_fmt
            return c

        license_date_str = license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-'
        license_expiry_str = license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-'
        lic_no = license_obj.license_number or '-'

        # ── Collect summary rows ──────────────────────────────────────────────
        from datetime import date as _date_cls
        summary_rows = []   # (group, sort_date, row_data_dict, is_boe)
        total_cif = 0.0

        for item in license_obj.import_license.all():
            item_name = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else (item.description or '-')

            boes = RowDetails.objects.filter(
                sr_number_id=item.id, transaction_type='D'
            ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')

            for rd in boes:
                qty  = float(rd.qty or 0)
                cif  = float(rd.cif_fc or 0)
                rate = cif / qty if qty else 0.0
                total_cif += cif
                boe_company = rd.bill_of_entry.company.name if rd.bill_of_entry.company else '-'
                ref_no   = rd.bill_of_entry.bill_of_entry_number or '-'
                ref_date = rd.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if rd.bill_of_entry.bill_of_entry_date else ''
                ref_str  = f"{ref_no} / {ref_date}" if ref_date else ref_no
                product  = rd.bill_of_entry.product_name or item_name
                _sort_dt = rd.bill_of_entry.bill_of_entry_date or _date_cls.min
                summary_rows.append((0, _sort_dt, {
                    'item': product, 'type': 'BOE', 'company': boe_company,
                    'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
                }, True))

            allotments = AllotmentItems.objects.filter(
                item_id=item.id, allotment__bill_of_entry__isnull=True
            ).select_related('allotment', 'allotment__company')

            for ai in allotments:
                qty     = float(ai.qty or 0)
                cif     = float(ai.cif_fc or 0)
                rate    = cif / qty if qty else 0.0
                total_cif += cif
                company = ai.allotment.company.name if ai.allotment.company else '-'
                invoice = ai.allotment.invoice or '-'
                eta     = ai.allotment.estimated_arrival_date.strftime('%d-%m-%Y') if ai.allotment.estimated_arrival_date else ''
                ref_str = f"{invoice} / ETA: {eta}" if eta else invoice
                product = ai.allotment.item_name or item_name
                _sort_dt = ai.allotment.estimated_arrival_date or _date_cls.min
                summary_rows.append((1, _sort_dt, {
                    'item': product, 'type': 'Allotment', 'company': company,
                    'reference': ref_str, 'qty': qty, 'rate': rate, 'cif': cif
                }, False))

        # BOEs first (sorted by BOE date), then allotments (sorted by allotment date)
        summary_rows.sort(key=lambda x: (x[0], x[1]))

        # ── Pre-aggregate balance data ─────────────────────────────────────────
        _bal_agg = defaultdict(lambda: {
            'qty': 0.0, 'is_restricted': False, 'restriction_pct': None,
            'sr_ids': [], 'description': '', 'hs_code': ''
        })
        for _item in license_obj.import_license.all():
            _key = ', '.join(sorted([i.name for i in _item.items.all()])) if _item.items.exists() else (_item.description or '-')
            _bal_agg[_key]['qty'] += float(_item.available_quantity or 0)
            _bal_agg[_key]['sr_ids'].append(_item.id)
            if _item.is_restricted:
                _bal_agg[_key]['is_restricted'] = True
                if _bal_agg[_key]['restriction_pct'] is None:
                    _rpct_val = _item.items.filter(
                        restriction_percentage__gt=0
                    ).values_list('restriction_percentage', flat=True).first()
                    if _rpct_val:
                        _bal_agg[_key]['restriction_pct'] = _Dec(str(_rpct_val))
            if not _bal_agg[_key]['description']:
                _bal_agg[_key]['description'] = _item.description or _key
            if not _bal_agg[_key]['hs_code']:
                _bal_agg[_key]['hs_code'] = str(_item.hs_code.hs_code if _item.hs_code else '-')

        _license_balance = float(license_obj.balance_cif or 0)
        _total_export_cif = _Dec(str(license_obj._calculate_license_credit() or 0))
        total_license_cif = total_cif + _license_balance

        # ══════════════════════════════════════════════════════════════════════
        # Section 1: License info row
        # ══════════════════════════════════════════════════════════════════════
        r = 1
        _today = _date_cls.today()
        INFO_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        INFO_FONT = Font(bold=True, color="FFFFFF", size=9)
        if license_obj.license_expiry_date:
            _days = (license_obj.license_expiry_date - _today).days
            if _days < 0:
                EXPIRY_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            elif _days <= 90:
                EXPIRY_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            else:
                EXPIRY_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        else:
            EXPIRY_FILL = INFO_FILL
        for col, (label, val) in enumerate([
            ('License No', lic_no),
            ('License Date', license_date_str),
            ('Expiry Date', license_expiry_str),
            ('Total CIF', f"{total_license_cif:,.2f}")
        ], 1):
            c = ws.cell(row=r, column=col, value=f"{label}: {val}")
            c.fill = EXPIRY_FILL if col == 3 else INFO_FILL
            c.font = INFO_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='left' if col < 4 else 'right', vertical='center')
        r += 1

        # ══════════════════════════════════════════════════════════════════════
        # Section 2: Summary (BOE & Allotments)
        # ══════════════════════════════════════════════════════════════════════
        # Section header (merged A:G)
        ws.merge_cells(f'A{r}:G{r}')
        sh = ws[f'A{r}']
        sh.value = 'Summary (BOE & Allotments)'
        sh.fill = HDR_FILL; sh.font = Font(bold=True, color="FFFFFF", size=10)
        sh.alignment = Alignment(horizontal='center', vertical='center')
        r += 1

        # Column headers
        SUMM_COLS = ['Item', 'Type', 'Company', 'Reference', 'Qty', 'Rate', 'CIF Value (FC)']
        for col, h in enumerate(SUMM_COLS, 1):
            _hdr(ws, r, col, h)
        r += 1

        # Data rows
        for _s, _sd, row_data, is_boe in summary_rows:
            fill = BOE_FILL if is_boe else ALLOT_FILL
            _cell(ws, r, 1, row_data['item'],      fill=fill)
            _cell(ws, r, 2, row_data['type'],      fill=fill)
            _cell(ws, r, 3, row_data['company'],   fill=fill)
            _cell(ws, r, 4, row_data['reference'], fill=fill)
            _cell(ws, r, 5, row_data['qty'],       fill=fill, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 6, row_data['rate'],      fill=fill, align='right', num_fmt='#,##0.00')
            _cell(ws, r, 7, row_data['cif'],       fill=fill, align='right', num_fmt='#,##0.00')
            r += 1

        # Total row
        if summary_rows:
            _cell(ws, r, 1, '', fill=TOTAL_FILL)
            _cell(ws, r, 2, '', fill=TOTAL_FILL)
            _cell(ws, r, 3, '', fill=TOTAL_FILL)
            _cell(ws, r, 4, 'TOTAL', fill=TOTAL_FILL, bold=True, align='right')
            _cell(ws, r, 5, '', fill=TOTAL_FILL)
            _cell(ws, r, 6, '', fill=TOTAL_FILL)
            _cell(ws, r, 7, total_cif, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
            r += 1

        r += 1  # blank row

        # ══════════════════════════════════════════════════════════════════════
        # Section 3: Utilization Planning (E1) / Summary (Balance Quantity)
        # ══════════════════════════════════════════════════════════════════════
        _norm_vals = list(license_obj.export_license.values_list('norm_class__norm_class', flat=True))
        _is_e1 = any(n and 'E1' in str(n) and 'E126' not in str(n) and 'E132' not in str(n) for n in _norm_vals)
        _is_e5 = any(n and str(n).strip() == 'E5' for n in _norm_vals)
        if _is_e1:
            _UTIL_PLAN = [
                ('OTHER CONFECTIONERY / FRUIT & FRUIT PRODUCTS', 2.7, ['0802'], ['confectionery', 'fruit & fruit', 'fruit product']),
                ('FRUIT JUICE',                                   3.0, [],       ['fruit juice']),
                ('TARTARIC ACID / CITRIC ACID',                   1.0, [],       ['tartaric', 'citric']),
                ('POLYPROPYLENE',                                  1.0, [],       ['polypropylene']),
                ('PAPER & PAPER BOARD',                           0.7, [],       ['paper']),
            ]
            def _cat_match(hs, desc, hs_kw, name_kw):
                return any(k in (hs or '').lower() for k in hs_kw) or any(k in (desc or '').lower() for k in name_kw)
            _cat_totals = {label: 0.0 for label, *_ in _UTIL_PLAN}
            _unclassified = []
            for _ik in _bal_agg:
                _bq = _bal_agg[_ik]['qty']
                _de = _bal_agg[_ik]['description'] or _ik
                _hs = _bal_agg[_ik]['hs_code']
                _found = False
                for _lbl, _rt, _hskw, _nkw in _UTIL_PLAN:
                    if _cat_match(_hs, _de, _hskw, _nkw):
                        _cat_totals[_lbl] += _bq
                        _found = True
                        break
                if not _found:
                    _unclassified.append((_de, _hs, _bq))

        ws.merge_cells(f'A{r}:E{r}')
        bh = ws[f'A{r}']
        bh.value = 'Utilization Planning' if (_is_e1 or _is_e5) else 'Summary (Balance Quantity)'
        bh.fill = HDR_FILL; bh.font = Font(bold=True, color="FFFFFF", size=10)
        bh.alignment = Alignment(horizontal='center', vertical='center')
        r += 1

        # BALANCE CIF $ row: cols A-D merged + col E yellow
        ws.merge_cells(f'A{r}:D{r}')
        bc = ws[f'A{r}']
        bc.value = 'BALANCE CIF $'
        bc.fill = HDR_FILL; bc.font = Font(bold=True, color="FFFFFF", size=9)
        bc.alignment = Alignment(horizontal='center', vertical='center')
        bc.border = THIN_BORDER
        yc = ws.cell(row=r, column=5, value=_license_balance)
        yc.fill = YEL_FILL; yc.font = Font(bold=True, size=9)
        yc.border = THIN_BORDER
        yc.alignment = Alignment(horizontal='right', vertical='center')
        yc.number_format = '#,##0.00'
        r += 1

        if _is_e1:
            for col, h in enumerate(['Item Category', 'Rate ($/unit)', 'Bal Qty', 'Unit Price', 'Planned CIF ($)'], 1):
                _hdr(ws, r, col, h)
            r += 1

            _total_planned = 0.0
            _e1_remaining = _license_balance
            for _idx, (_lbl, _rt, _hskw, _nkw) in enumerate(_UTIL_PLAN):
                _bq = _cat_totals[_lbl]
                _pc = min(_rt * _bq, _e1_remaining)
                _up = _pc / _bq if _bq else 0.0
                _e1_remaining -= _pc
                _total_planned += _pc
                _rf = None if _idx % 2 == 0 else ALT_FILL
                _cell(ws, r, 1, _lbl, fill=_rf)
                _cell(ws, r, 2, _rt,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 3, _bq,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 4, _up,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 5, _pc,  fill=_rf, align='right', num_fmt='#,##0.00')
                r += 1

            if _unclassified:
                r += 1
                ws.merge_cells(f'A{r}:E{r}')
                _uh = ws[f'A{r}']
                _uh.value = 'UNCLASSIFIED ITEMS'
                _uh.fill = HDR_FILL; _uh.font = Font(bold=True, color="FFFFFF", size=9)
                _uh.alignment = Alignment(horizontal='center', vertical='center')
                _uh.border = THIN_BORDER
                r += 1
                for col, h in enumerate(['Item Name', 'HS Code', 'Bal Qty', '', ''], 1):
                    _hdr(ws, r, col, h)
                r += 1
                for _i2, (_de2, _hs2, _bq2) in enumerate(_unclassified):
                    _rf2 = None if _i2 % 2 == 0 else ALT_FILL
                    _cell(ws, r, 1, _de2, fill=_rf2)
                    _cell(ws, r, 2, _hs2, fill=_rf2)
                    _cell(ws, r, 3, _bq2, fill=_rf2, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 4, '',   fill=_rf2)
                    _cell(ws, r, 5, '',   fill=_rf2)
                    r += 1

            r += 1
            _cell(ws, r, 1, '', fill=TOTAL_FILL)
            _cell(ws, r, 2, '', fill=TOTAL_FILL)
            _cell(ws, r, 3, '', fill=TOTAL_FILL)
            _cell(ws, r, 4, 'TOTAL PLANNED CIF $', fill=TOTAL_FILL, bold=True, align='right')
            _cell(ws, r, 5, _total_planned, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
            r += 1
            _rem = _license_balance - _total_planned
            _RF = PatternFill(start_color="C00000" if _rem < 0 else "1F4E79",
                              end_color="C00000" if _rem < 0 else "1F4E79", fill_type="solid")
            for _ci in range(1, 5):
                _cx = ws.cell(row=r, column=_ci)
                _cx.fill = _RF; _cx.border = THIN_BORDER
            _rc = ws.cell(row=r, column=4, value='REMAINING BALANCE CIF $')
            _rc.fill = _RF; _rc.font = Font(bold=True, color="FFFFFF", size=9)
            _rc.border = THIN_BORDER; _rc.alignment = Alignment(horizontal='right', vertical='center')
            _rc2 = ws.cell(row=r, column=5, value=_rem)
            _rc2.fill = _RF; _rc2.font = Font(bold=True, color="FFFFFF", size=9)
            _rc2.border = THIN_BORDER
            _rc2.alignment = Alignment(horizontal='right', vertical='center')
            _rc2.number_format = '#,##0.00'
            r += 1
        elif _is_e5:
            _E5_CATS = [
                ('DIETARY FIBRE',            2.7,  ['0802']),
                ('VEGETABLE OIL (1513)',      2.26, ['1513']),
                ('VEGETABLE OIL (15119020)',  1.2,  ['15119020']),
                ('MILK (0404)',               3.75, ['0404']),
                ('MILK (3502)',               14.0, ['3502']),
            ]
            _e5_totals = {lbl: 0.0 for lbl, *_ in _E5_CATS}
            _e5_unclassified = []
            for _ik in _bal_agg:
                _bq = _bal_agg[_ik]['qty']
                _hs = (_bal_agg[_ik]['hs_code'] or '').lower()
                _de = _bal_agg[_ik]['description'] or _ik
                _found = False
                for _lbl, _rt, _hskw in _E5_CATS:
                    if any(k in _hs for k in _hskw):
                        _e5_totals[_lbl] += _bq
                        _found = True
                        break
                if not _found:
                    _e5_unclassified.append((_de, _bal_agg[_ik]['hs_code'], _bq))

            for col, h in enumerate(['Item Category', 'Rate ($/unit)', 'Bal Qty', 'Unit Price', 'Planned CIF ($)'], 1):
                _hdr(ws, r, col, h)
            r += 1

            _e5_planned = 0.0
            _e5_remaining = _license_balance
            for _idx, (_lbl, _rt, _hskw) in enumerate(_E5_CATS):
                _bq = _e5_totals[_lbl]
                _pc = min(_rt * _bq, _e5_remaining)
                _up = _pc / _bq if _bq else 0.0
                _e5_remaining -= _pc
                _e5_planned += _pc
                _rf = None if _idx % 2 == 0 else ALT_FILL
                _cell(ws, r, 1, _lbl, fill=_rf)
                _cell(ws, r, 2, _rt,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 3, _bq,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 4, _up,  fill=_rf, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 5, _pc,  fill=_rf, align='right', num_fmt='#,##0.00')
                r += 1

            # WHEAT FLOUR row — last priority, remaining balance
            _wf = _e5_remaining
            _wf_rf = None if len(_E5_CATS) % 2 == 0 else ALT_FILL
            _cell(ws, r, 1, 'WHEAT FLOUR', fill=_wf_rf)
            _cell(ws, r, 2, '-', fill=_wf_rf, align='center')
            _cell(ws, r, 3, '-', fill=_wf_rf, align='center')
            _cell(ws, r, 4, '-', fill=_wf_rf, align='center')
            _cell(ws, r, 5, _wf, fill=_wf_rf, align='right', num_fmt='#,##0.00')
            r += 1

            if _e5_unclassified:
                r += 1
                ws.merge_cells(f'A{r}:E{r}')
                _uh = ws[f'A{r}']
                _uh.value = 'UNCLASSIFIED ITEMS'
                _uh.fill = HDR_FILL; _uh.font = Font(bold=True, color="FFFFFF", size=9)
                _uh.alignment = Alignment(horizontal='center', vertical='center')
                _uh.border = THIN_BORDER
                r += 1
                for col, h in enumerate(['Item Name', 'HS Code', 'Bal Qty', '', ''], 1):
                    _hdr(ws, r, col, h)
                r += 1
                for _i2, (_de2, _hs2, _bq2) in enumerate(_e5_unclassified):
                    _rf2 = None if _i2 % 2 == 0 else ALT_FILL
                    _cell(ws, r, 1, _de2, fill=_rf2)
                    _cell(ws, r, 2, _hs2, fill=_rf2)
                    _cell(ws, r, 3, _bq2, fill=_rf2, align='right', num_fmt='#,##0.00')
                    _cell(ws, r, 4, '',   fill=_rf2)
                    _cell(ws, r, 5, '',   fill=_rf2)
                    r += 1

            r += 1
            _cell(ws, r, 1, '', fill=TOTAL_FILL)
            _cell(ws, r, 2, '', fill=TOTAL_FILL)
            _cell(ws, r, 3, '', fill=TOTAL_FILL)
            _cell(ws, r, 4, 'TOTAL ALLOCATED CIF $', fill=TOTAL_FILL, bold=True, align='right')
            _cell(ws, r, 5, _e5_planned + _wf, fill=TOTAL_FILL, bold=True, align='right', num_fmt='#,##0.00')
            r += 1
        else:
            # Column headers
            BAL_COLS = ['HSN Code', 'Item Name', 'Bal Qty', 'Unit Price', 'CIF FC']
            for col, h in enumerate(BAL_COLS, 1):
                _hdr(ws, r, col, h)
            r += 1

            # Data rows
            for idx, item_key in enumerate(sorted(_bal_agg.keys())):
                b_qty = _bal_agg[item_key]['qty']
                if _bal_agg[item_key]['is_restricted'] and _bal_agg[item_key]['restriction_pct'] is not None:
                    _rpct  = _bal_agg[item_key]['restriction_pct']
                    _sr_ids = _bal_agg[item_key]['sr_ids']
                    _grp_debits = RowDetails.objects.filter(
                        sr_number_id__in=_sr_ids, transaction_type='D'
                    ).aggregate(
                        total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                    )['total'] or _Dec('0')
                    _grp_allots = AllotmentItems.objects.filter(
                        item_id__in=_sr_ids, allotment__bill_of_entry__isnull=True
                    ).aggregate(
                        total=_Coalesce(_Sum('cif_fc'), _Val(_Dec('0')), output_field=_DF())
                    )['total'] or _Dec('0')
                    _allowed = _total_export_cif * _rpct / _Dec('100')
                    b_cif = float(max(_allowed - _Dec(str(_grp_debits)) - _Dec(str(_grp_allots)), _Dec('0')))
                else:
                    b_cif = _license_balance

                unit_price = b_cif / b_qty if b_qty else 0.0
                desc = _bal_agg[item_key]['description'] or item_key
                hs   = _bal_agg[item_key]['hs_code']
                row_fill = None if idx % 2 == 0 else ALT_FILL

                _cell(ws, r, 1, hs,         fill=row_fill)
                _cell(ws, r, 2, desc,       fill=row_fill)
                _cell(ws, r, 3, b_qty,      fill=row_fill, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 4, unit_price, fill=row_fill, align='right', num_fmt='#,##0.00')
                _cell(ws, r, 5, b_cif,      fill=row_fill, align='right', num_fmt='#,##0.00')
                r += 1

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 16

        ws.freeze_panes = 'A2'

        # ── Save ──────────────────────────────────────────────────────────────
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{license_obj.license_number}-summary.xlsx"'
        return response

    @action(detail=True, methods=['get'], url_path='balance-excel-unused')
    def balance_excel_unused(self, request, pk=None):
        """Original full balance Excel — kept for reference, no longer exposed."""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        from datetime import date

        license_obj = self.get_object()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "License Balance"

        # Header styling
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        data_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        section_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=12)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        current_row = 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "License Balance Report"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2

        # License Header Information
        # Row 1 Headers
        headers_row1 = ['License Number', 'License Date', 'License Expiry Date', 'Exporter Name', 'Port Name']
        for col_num, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # Row 1 Values
        values_row1 = [
            license_obj.license_number or '-',
            license_obj.license_date.strftime('%d-%m-%Y') if license_obj.license_date else '-',
            license_obj.license_expiry_date.strftime('%d-%m-%Y') if license_obj.license_expiry_date else '-',
            license_obj.exporter.name if license_obj.exporter else '-',
            license_obj.port.name if license_obj.port else '-'
        ]
        for col_num, value in enumerate(values_row1, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.fill = data_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 2

        # Row 2 Headers
        headers_row2 = ['Purchase Status', 'Balance CIF', 'Get Norm Class', 'Latest Transfer']
        for col_num, header in enumerate(headers_row2, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # Row 2 Values
        values_row2 = [
            str(license_obj.purchase_status) if license_obj.purchase_status else '-',
            f"{float(license_obj.balance_cif or 0):.2f}",
            license_obj.get_norm_class or '-',
            str(license_obj.latest_transfer) if license_obj.latest_transfer else '-'
        ]
        for col_num, value in enumerate(values_row2, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.fill = data_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 2

        # Export Items Section
        if license_obj.export_license.exists():
            # Section header
            ws.merge_cells(f'A{current_row}:C{current_row}')
            section_cell = ws[f'A{current_row}']
            section_cell.value = "Export Items"
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1

            # Export items headers
            export_headers = ['Item', 'Total CIF', 'Balance CIF']
            for col_num, header in enumerate(export_headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            current_row += 1

            # Export items data
            for item in license_obj.export_license.all():
                item_desc = item.description or (str(item.norm_class) if item.norm_class else None) or 'None'
                values = [
                    item_desc,
                    f"{float(item.cif_fc or item.fob_fc or 0):.2f}",
                    f"{float(license_obj.balance_cif or 0):.2f}"
                ]
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.border = thin_border
                current_row += 1

            current_row += 1

        # Import Items Section
        if license_obj.import_license.exists():
            from bill_of_entry.models import RowDetails
            from allotment.models import AllotmentItems

            # Section header
            ws.merge_cells(f'A{current_row}:J{current_row}')
            section_cell = ws[f'A{current_row}']
            section_cell.value = "Import Items"
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1

            for item in license_obj.import_license.all():
                # Item headers
                item_headers = ['Sr', 'HS Code', 'Description', 'Item', 'Total Qty',
                               'Allotted', 'Debited', 'Available', 'CIF FC', 'Bal CIF']
                for col_num, header in enumerate(item_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                current_row += 1

                # Item data
                item_names = ', '.join([i.name for i in item.items.all()]) if item.items.exists() else '-'
                hs_code_display = str(item.hs_code.hs_code if item.hs_code else '-')

                item_values = [
                    str(item.serial_number or '-'),
                    hs_code_display,
                    str(item.description or '-'),
                    item_names,
                    f"{float(item.quantity or 0):.2f}",
                    f"{float(item.allotted_quantity or 0):.2f}",
                    f"{float(item.debited_quantity or 0):.2f}",
                    f"{float(item.available_quantity or 0):.2f}",
                    f"{float(item.cif_fc or 0):.2f}",
                    f"{float(item.balance_cif_fc or 0):.2f}"
                ]
                for col_num, value in enumerate(item_values, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.fill = data_fill
                    cell.border = thin_border
                current_row += 1

                # BOE Details
                boes = RowDetails.objects.filter(
                    sr_number_id=item.id,
                    transaction_type='D'
                ).select_related('bill_of_entry', 'bill_of_entry__port', 'bill_of_entry__company')
                if boes.exists():
                    current_row += 1
                    ws.merge_cells(f'A{current_row}:G{current_row}')
                    boe_header_cell = ws[f'A{current_row}']
                    boe_header_cell.value = "BOEs"
                    boe_header_cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                    boe_header_cell.font = Font(bold=True, color="FFFFFF")
                    current_row += 1

                    boe_headers = ['BOE Number', 'Date', 'Port', 'Company', 'Qty', 'CIF $', 'CIF INR']
                    for col_num, header in enumerate(boe_headers, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=header)
                        cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.border = thin_border
                    current_row += 1

                    for boe in boes:
                        boe_values = [
                            boe.bill_of_entry.bill_of_entry_number if boe.bill_of_entry else '-',
                            boe.bill_of_entry.bill_of_entry_date.strftime('%d-%m-%Y') if boe.bill_of_entry and boe.bill_of_entry.bill_of_entry_date else '-',
                            boe.bill_of_entry.port.name if boe.bill_of_entry and boe.bill_of_entry.port else '-',
                            boe.bill_of_entry.company.name if boe.bill_of_entry and boe.bill_of_entry.company else '-',
                            f"{float(boe.qty or 0):.2f}",
                            f"{float(boe.cif_fc or 0):.2f}",
                            f"{float(boe.cif_inr or 0):.2f}"
                        ]
                        for col_num, value in enumerate(boe_values, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=value)
                            cell.border = thin_border
                        current_row += 1

                # Allotment Details
                # Only show allotments where bill_of_entry is NULL (not yet converted to BOE)
                allotments = AllotmentItems.objects.filter(
                    item=item,
                    allotment__bill_of_entry__isnull=True
                ).select_related('allotment', 'allotment__company')
                if allotments.exists():
                    current_row += 1
                    ws.merge_cells(f'A{current_row}:D{current_row}')
                    allot_header_cell = ws[f'A{current_row}']
                    allot_header_cell.value = "Allotments"
                    allot_header_cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                    allot_header_cell.font = Font(bold=True, color="FFFFFF")
                    current_row += 1

                    allot_headers = ['Company', 'Qty', 'CIF $', 'CIF INR']
                    for col_num, header in enumerate(allot_headers, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=header)
                        cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.border = thin_border
                    current_row += 1

                    for allot in allotments:
                        allot_values = [
                            allot.allotment.company.name if allot.allotment and allot.allotment.company else '-',
                            f"{float(allot.qty or 0):.2f}",
                            f"{float(allot.cif_fc or 0):.2f}",
                            f"{float(allot.cif_inr or 0):.2f}"
                        ]
                        for col_num, value in enumerate(allot_values, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=value)
                            cell.border = thin_border
                        current_row += 1

                # Balance calculation
                current_row += 1
                balance = float(item.quantity or 0) - float(item.debited_quantity or 0) - float(item.allotted_quantity or 0)
                ws.merge_cells(f'A{current_row}:J{current_row}')
                balance_cell = ws[f'A{current_row}']
                balance_cell.value = f"Balance Quantity: {balance:.2f}"
                balance_cell.fill = PatternFill(start_color="e8e8e8", end_color="e8e8e8", fill_type="solid")
                balance_cell.font = Font(bold=True, color="e74c3c")
                balance_cell.border = thin_border
                current_row += 2

        # Notes Section
        if license_obj.balance_report_notes:
            current_row += 1
            ws.merge_cells(f'A{current_row}:J{current_row}')
            notes_header_cell = ws[f'A{current_row}']
            notes_header_cell.value = "Notes"
            notes_header_cell.fill = section_fill
            notes_header_cell.font = section_font
            notes_header_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1

            ws.merge_cells(f'A{current_row}:J{current_row}')
            notes_cell = ws[f'A{current_row}']
            notes_cell.value = license_obj.balance_report_notes
            notes_cell.fill = PatternFill(start_color="fffacd", end_color="fffacd", fill_type="solid")
            notes_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            notes_cell.border = thin_border
            ws.row_dimensions[current_row].height = 60

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15

        # Save to bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{license_obj.license_number}-balance.xlsx"'
        return response

    @action(detail=True, methods=['get'], url_path='merged-documents')
    def merged_documents(self, request, pk=None):
        """
        Merge all license documents (LICENSE COPY + TRANSFER LETTER) into one PDF.
        Converts images to PDF if needed.
        """
        from django.http import FileResponse, HttpResponse
        import io
        import os
        import logging

        logger = logging.getLogger(__name__)
        license_obj = LicenseDetailsModel.objects.get(pk=pk)
        documents = license_obj.license_documents.all()

        if not documents.exists():
            return HttpResponse("No documents found for this license", status=404)

        # Check if required libraries are installed
        try:
            from PyPDF2 import PdfMerger, PdfReader
            from PIL import Image
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.utils import ImageReader
        except ImportError as e:
            return HttpResponse(f"Missing required library: {str(e)}. Please install PyPDF2 and Pillow.", status=500)

        try:
            merger = PdfMerger()

            # Sort documents: TRANSFER LETTER first, then LICENSE COPY, then OTHER
            type_order = {'TRANSFER LETTER': 0, 'LICENSE COPY': 1, 'OTHER': 2}
            sorted_documents = sorted(documents, key=lambda doc: type_order.get(doc.type, 3))

            for doc in sorted_documents:
                if not doc.file:
                    continue

                file_path = doc.file.path
                file_ext = os.path.splitext(file_path)[1].lower()

                if file_ext == '.pdf':
                    # Add PDF directly
                    merger.append(file_path)
                    logger.info(f"Added PDF: {file_path}")
                elif file_ext in ['.doc', '.docx']:
                    # Convert DOCX/DOC to PDF
                    try:
                        from docx import Document
                        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.pagesizes import A4
                        from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
                        import tempfile

                        # Read DOCX content
                        doc = Document(file_path)

                        # Create temporary PDF file
                        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
                            tmp_pdf_path = tmp_pdf.name

                        # Create PDF
                        pdf = SimpleDocTemplate(tmp_pdf_path, pagesize=A4)
                        story = []
                        styles = getSampleStyleSheet()

                        # Add custom style
                        normal_style = ParagraphStyle(
                            'CustomNormal',
                            parent=styles['Normal'],
                            fontSize=10,
                            leading=14,
                            alignment=TA_LEFT
                        )

                        # Convert paragraphs to PDF
                        for paragraph in doc.paragraphs:
                            if paragraph.text.strip():
                                # Escape special characters for reportlab
                                text = paragraph.text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                p = Paragraph(text, normal_style)
                                story.append(p)
                                story.append(Spacer(1, 6))

                        # Build PDF
                        pdf.build(story)

                        # Add converted PDF to merger
                        merger.append(tmp_pdf_path)
                        logger.info(f"Converted and added DOCX/DOC: {file_path}")

                        # Clean up temp file
                        try:
                            os.remove(tmp_pdf_path)
                        except OSError:
                            pass

                    except Exception as e:
                        logger.error(f"Error converting DOCX file {file_path}: {str(e)}")
                        continue
                elif file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                    # Convert image to PDF
                    img = Image.open(file_path)

                    # Convert to RGB if necessary
                    if img.mode != 'RGB':
                        img = img.convert('RGB')

                    # Create PDF from image using ReportLab
                    img_buffer = io.BytesIO()
                    img_width, img_height = img.size

                    # Use A4 size for PDF
                    pdf_canvas = canvas.Canvas(img_buffer, pagesize=A4)
                    a4_width, a4_height = A4

                    # Calculate scaling to fit image on A4 page
                    scale = min(a4_width / img_width, a4_height / img_height)
                    new_width = img_width * scale
                    new_height = img_height * scale

                    # Center image on page
                    x = (a4_width - new_width) / 2
                    y = (a4_height - new_height) / 2

                    # Use ImageReader for PIL Image object
                    img_reader = ImageReader(img)
                    pdf_canvas.drawImage(img_reader, x, y, width=new_width, height=new_height)
                    pdf_canvas.save()

                    # Rewind buffer and create PdfReader from it
                    img_buffer.seek(0)
                    pdf_reader = PdfReader(img_buffer)
                    merger.append(pdf_reader)
                    logger.info(f"Converted and added image: {file_path}")

            # Write merged PDF to buffer
            output_buffer = io.BytesIO()
            merger.write(output_buffer)
            merger.close()
            output_buffer.seek(0)

            # Return merged PDF
            response = FileResponse(
                output_buffer,
                content_type='application/pdf',
                as_attachment=False,
                filename=f'license_{license_obj.license_number}_documents.pdf'
            )
            return response

        except Exception as e:
            logger.error(f"Error merging documents: {str(e)}", exc_info=True)
            return HttpResponse(f"Error merging documents: {str(e)}", status=500)


# Add license report actions to viewset
LicenseDetailsViewSet = add_license_report_action(LicenseDetailsViewSet)
LicenseDetailsViewSet = add_active_dfia_report_action(LicenseDetailsViewSet)
